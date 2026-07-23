from django.shortcuts import render

# from django.db.models import Sum
from acceso_empresas.decorators import login_o_portal_required
from caja_chica.models import FondeoCajaChica, GastoCajaChica, ValeCaja
from clientes.models import Cliente
from facturacion.models import CobroOtrosIngresos, Factura, FacturaOtrosIngresos, Pago
from gastos.models import  PagoGasto
from empresas.models import CuentaBancaria, Empresa
from collections import OrderedDict

# from django.db.models import Case, When, Value, CharField
#import calendar
import datetime
import locale
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models.functions import ExtractMonth, ExtractYear
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import F, Value, CharField, Sum, Case, When, IntegerField
from django.db.models import  OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
from openpyxl.utils import get_column_letter


@login_o_portal_required 
def reporte_ingresos_vs_gastos(request):
    empresas = Empresa.objects.all()
    empresa_id = request.GET.get("empresa")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    periodo = request.GET.get("periodo")

    # if not request.user.is_superuser:
    #     empresa_id = str(request.user.perfilusuario.empresa.id)
    # else:
    #     empresa_id = request.GET.get("empresa") or ""
    if not request.user.is_superuser:
        # Verificar si viene del portal de acceso
        if getattr(request, 'is_portal_acceso', False):
            empresa_id = str(request.session.get('empresa_id', ''))
        else:
            empresa_id = str(request.user.perfilusuario.empresa.id)
    else:
        empresa_id = request.GET.get("empresa") or request.session.get("empresa_id")
        if not empresa_id or not str(empresa_id).isdigit():
            empresa_id = None

    # Si no hay ningún filtro, mostrar periodo actual por default
    if not periodo and not fecha_inicio and not fecha_fin and not mes and not anio:
        periodo = "periodo_actual"

    hoy = datetime.date.today()
    # Prioridad: periodo > mes/año > fechas manuales
    if periodo == "mes_actual":
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = (hoy.replace(day=1) + datetime.timedelta(days=32)).replace(
            day=1
        ) - datetime.timedelta(days=1)
        mes = hoy.month
        anio = hoy.year
    elif periodo == "periodo_actual":
        fecha_inicio = hoy.replace(month=1, day=1)
        fecha_fin = hoy
        mes = ""
        anio = ""
    elif mes and anio:
        try:
            mes = int(mes)
            anio = int(anio)
            fecha_inicio = datetime.date(anio, mes, 1)
            if mes == 12:
                fecha_fin = datetime.date(anio, 12, 31)
            else:
                fecha_fin = datetime.date(anio, mes + 1, 1) - datetime.timedelta(days=1)
        except Exception:
            fecha_inicio = None
            fecha_fin = None
    elif fecha_inicio and fecha_fin:
        # Ya vienen del formulario
        pass
    else:
        fecha_inicio = None
        fecha_fin = None

    # Convierte a date si es string
    if isinstance(fecha_inicio, str):
        try:
            fecha_inicio_dt = datetime.datetime.strptime(
                fecha_inicio, "%Y-%m-%d"
            ).date()
        except Exception:
            fecha_inicio_dt = None
    else:
        fecha_inicio_dt = fecha_inicio

    if isinstance(fecha_fin, str):
        try:
            fecha_fin_dt = datetime.datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except Exception:
            fecha_fin_dt = None
    else:
        fecha_fin_dt = fecha_fin

    # Para mostrar el mes y año en letras

    try:
        locale.setlocale(locale.LC_TIME, "es_MX.UTF-8")
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
        except locale.Error:
            locale.setlocale(locale.LC_TIME, "C")  # Fallback seguro

    mes_letra = ""
    if (
        fecha_inicio_dt
        and fecha_fin_dt
        and fecha_inicio_dt == fecha_fin_dt.replace(day=1)
    ):
        mes_letra = fecha_inicio_dt.strftime("%B %Y").capitalize()
    elif fecha_inicio_dt and fecha_fin_dt:
        mes_letra = f"{fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')}"

    pagos = Pago.objects.exclude(forma_pago="nota_credito")
    pagos_gastos = PagoGasto.objects.all()
    cobros_otros = CobroOtrosIngresos.objects.select_related(
        "factura", "factura__empresa"
    )
    gastos_caja_chica = GastoCajaChica.objects.all()
    vales_caja_chica = ValeCaja.objects.all()

    # PAGOS POR IDENTIFICAR
    pagos_por_identificar = Pago.objects.filter(
        factura__isnull=True, identificado=False
    )

    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id)
        pagos_gastos = pagos_gastos.filter(gasto__empresa_id=empresa_id)
        cobros_otros = cobros_otros.filter(factura__empresa_id=empresa_id)
        gastos_caja_chica = gastos_caja_chica.filter(fondeo__empresa_id=empresa_id)
        vales_caja_chica = vales_caja_chica.filter(fondeo__empresa_id=empresa_id)
        pagos_por_identificar = pagos_por_identificar.filter(empresa_id=empresa_id)
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio)
        pagos_gastos = pagos_gastos.filter(fecha_pago__gte=fecha_inicio)
        cobros_otros = cobros_otros.filter(fecha_cobro__gte=fecha_inicio)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__gte=fecha_inicio)
        vales_caja_chica = vales_caja_chica.filter(fecha__gte=fecha_inicio)
        pagos_por_identificar = pagos_por_identificar.filter(
            fecha_pago__gte=fecha_inicio
        )
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin)
        pagos_gastos = pagos_gastos.filter(fecha_pago__lte=fecha_fin)
        cobros_otros = cobros_otros.filter(fecha_cobro__lte=fecha_fin)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__lte=fecha_fin)
        vales_caja_chica = vales_caja_chica.filter(fecha__lte=fecha_fin)
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__lte=fecha_fin)

    total_ingresos = pagos.aggregate(total=Sum("monto"))["total"] or 0
    total_otros_ingresos = cobros_otros.aggregate(total=Sum("monto"))["total"] or 0
    total_pagos_por_identificar = (
        pagos_por_identificar.aggregate(total=Sum("monto"))["total"] or 0
    )
    total_ingresos_cobrados = total_ingresos + total_otros_ingresos + total_pagos_por_identificar
    
    total_gastos_pagados = pagos_gastos.aggregate(total=Sum("monto"))["total"] or 0
    total_gastos_caja_chica = (
        gastos_caja_chica.aggregate(total=Sum("importe"))["total"] or 0
    )
    total_vales_caja_chica = (
        vales_caja_chica.aggregate(total=Sum("importe"))["total"] or 0
    )
    total_egresos = (
        total_gastos_pagados + total_gastos_caja_chica + total_vales_caja_chica
    )

    # Agrupar por tipo de origen (Local/Área)
    ingresos_qs = (
        pagos.annotate(
            origen=Case(
                When(factura__local__isnull=False, then=Value("Propiedades")),
                When(factura__area_comun__isnull=False, then=Value("Áreas Comunes")),
                default=Value("Sin origen"),
                output_field=CharField(),
            )
        )
        .values("origen")
        .annotate(total=Sum("monto"))
        .order_by("origen")
    )

    otros_ingresos_qs = (
        cobros_otros.select_related("factura__tipo_ingreso")
        .values("factura__tipo_ingreso__nombre")
        .annotate(total=Sum("monto"))
        .order_by("factura__tipo_ingreso")
    )

    # Agrupar y sumar todos los gastos por tipo (gastos normales, caja chica y vales)
    gastos_por_tipo_dict = {}
    # Gastos normales
    for g in pagos_gastos.values("gasto__tipo_gasto__nombre").annotate(
        total=Sum("monto")
    ):
        tipo = g["gasto__tipo_gasto__nombre"] or "Sin tipo"
        gastos_por_tipo_dict[tipo] = gastos_por_tipo_dict.get(tipo, 0) + float(
            g["total"]
        )
    # Caja chica
    for g in gastos_caja_chica.values("tipo_gasto__nombre").annotate(
        total=Sum("importe")
    ):
        tipo = g["tipo_gasto__nombre"] or "Sin tipo"
        gastos_por_tipo_dict[tipo] = gastos_por_tipo_dict.get(tipo, 0) + float(
            g["total"]
        )
    # Vales de caja chica agrupados por tipo real
    for g in vales_caja_chica.values("tipo_gasto__nombre").annotate(
        total=Sum("importe")
    ):
        tipo = g["tipo_gasto__nombre"] or "Sin tipo"
        gastos_por_tipo_dict[tipo] = gastos_por_tipo_dict.get(tipo, 0) + float(
            g["total"]
        )

    gastos_por_tipo = [
        {"tipo": tipo, "total": total} for tipo, total in gastos_por_tipo_dict.items()
    ]

    # Crear un diccionario ordenado para los ingresos por origen
    ingresos_por_origen = OrderedDict()
    for x in ingresos_qs:
        ingresos_por_origen[x["origen"]] = float(x["total"])
    for x in otros_ingresos_qs:
        tipo = x["factura__tipo_ingreso__nombre"] or "Otros ingresos"
        ingresos_por_origen[f" {tipo}"] = float(x["total"])
    ingresos_por_origen["Depositos no identificados"] = float(
        total_pagos_por_identificar
    )

    gastos_agregados = {}

    # Define gastos_qs similar to the commented-out section above
    gastos_qs = PagoGasto.objects.select_related("gasto__tipo_gasto__subgrupo__grupo")
    if empresa_id:
        gastos_qs = gastos_qs.filter(gasto__empresa_id=empresa_id)
    if fecha_inicio:
        gastos_qs = gastos_qs.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        gastos_qs = gastos_qs.filter(fecha_pago__lte=fecha_fin)

    # Gastos normales
    for g in gastos_qs.values(
        "gasto__tipo_gasto__subgrupo__grupo__nombre",
        "gasto__tipo_gasto__subgrupo__nombre",
        "gasto__tipo_gasto__nombre",
    ).annotate(total=Sum("monto")):
        grupo = (
            (g["gasto__tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
            .strip()
            .title()
        )
        subgrupo = (
            (g["gasto__tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
        )
        tipo = (g["gasto__tipo_gasto__nombre"] or "Sin tipo").strip().title()
        key = (grupo, subgrupo, tipo)
        gastos_agregados[key] = gastos_agregados.get(key, 0) + float(g["total"])

    # Caja chica
    for g in gastos_caja_chica.values(
        "tipo_gasto__subgrupo__grupo__nombre",
        "tipo_gasto__subgrupo__nombre",
        "tipo_gasto__nombre",
    ).annotate(total=Sum("importe")):
        grupo = (
            (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo").strip().title()
        )
        subgrupo = (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
        tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
        key = (grupo, subgrupo, tipo)
        gastos_agregados[key] = gastos_agregados.get(key, 0) + float(g["total"])

    # Vales de caja chica
    for g in vales_caja_chica.values(
        "tipo_gasto__subgrupo__grupo__nombre",
        "tipo_gasto__subgrupo__nombre",
        "tipo_gasto__nombre",
    ).annotate(total=Sum("importe")):
        grupo = (
            (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo").strip().title()
        )
        subgrupo = (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
        tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
        key = (grupo, subgrupo, tipo)
        gastos_agregados[key] = gastos_agregados.get(key, 0) + float(g["total"])

    # Ahora arma la estructura final agrupada
    estructura_gastos = OrderedDict()
    for (grupo, subgrupo, tipo), total in gastos_agregados.items():
        if grupo not in estructura_gastos:
            estructura_gastos[grupo] = OrderedDict()
        if subgrupo not in estructura_gastos[grupo]:
            estructura_gastos[grupo][subgrupo] = []
        estructura_gastos[grupo][subgrupo].append({"tipo": tipo, "total": total})

    saldo = total_ingresos_cobrados - total_egresos

    return render(
        request,
        "informes_financieros/ingresos_vs_gastos.html",
        {
            "empresas": empresas,
            "total_ingresos": total_ingresos_cobrados,
            "total_otros_ingresos": total_otros_ingresos,
            "total_pagos_por_identificar": total_pagos_por_identificar,
            "total_gastos_pagados": total_gastos_pagados,
            "total_gastos_caja_chica": total_gastos_caja_chica,
            "total_vales_caja_chica": total_vales_caja_chica,
            "total_egresos": total_egresos,
            "empresa_id": empresa_id,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "ingresos_por_origen": ingresos_por_origen,
            "periodo": periodo,
            "mes_letra": mes_letra,
            "mes": mes,
            "anio": anio,
            "gastos_por_tipo": gastos_por_tipo,
            "saldo": saldo,
            "estructura_gastos": estructura_gastos,
        },
    )


@login_o_portal_required
def estado_resultados(request):
    empresas = Empresa.objects.all()
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    periodo = request.GET.get("periodo")
    hoy = datetime.date.today()

    if not request.user.is_superuser:
        # Verificar si viene del portal de acceso
        if getattr(request, 'is_portal_acceso', False):
            empresa_id = str(request.session.get('empresa_id', ''))
        else:
            empresa_id = str(request.user.perfilusuario.empresa.id)
    else:
        empresa_id = request.GET.get("empresa") or request.session.get("empresa_id")
        if not empresa_id or not str(empresa_id).isdigit():
            empresa_id = None

    # --- Meses y años disponibles ---
    if empresa_id:
        meses_anios = (
            Factura.objects.filter(empresa_id=empresa_id)
            .annotate(mes=ExtractMonth("fecha_vencimiento"), anio=ExtractYear("fecha_vencimiento"))
            .values("mes", "anio").distinct()
        )
        meses_anios_otros = (
            FacturaOtrosIngresos.objects.filter(empresa_id=empresa_id)
            .annotate(mes=ExtractMonth("fecha_vencimiento"), anio=ExtractYear("fecha_vencimiento"))
            .values("mes", "anio").distinct()
        )
    else:
        meses_anios = (
            Factura.objects
            .annotate(mes=ExtractMonth("fecha_vencimiento"), anio=ExtractYear("fecha_vencimiento"))
            .values("mes", "anio").distinct()
        )
        meses_anios_otros = (
            FacturaOtrosIngresos.objects
            .annotate(mes=ExtractMonth("fecha_vencimiento"), anio=ExtractYear("fecha_vencimiento"))
            .values("mes", "anio").distinct()
        )

    meses_anios_set = set((x["mes"], x["anio"]) for x in list(meses_anios) + list(meses_anios_otros))
    meses_anios_list = sorted(
        [t for t in meses_anios_set if t[0] is not None and t[1] is not None],
        key=lambda x: (x[1], x[0])
    )
    meses_unicos = sorted(set(m for m, y in meses_anios_list if m))
    anios_unicos = sorted(set(y for m, y in meses_anios_list if y))

    # --- Periodo por defecto ---
    if not periodo and not fecha_inicio and not fecha_fin and not mes and not anio:
        periodo = "periodo_actual"

    if periodo == "mes_actual":
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = (hoy.replace(day=1) + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)
        mes = hoy.month
        anio = hoy.year
    elif periodo == "periodo_actual":
        fecha_inicio = hoy.replace(month=1, day=1)
        fecha_fin = hoy
        mes = ""
        anio = ""
    elif mes and anio:
        try:
            mes = int(mes)
            anio = int(anio)
            fecha_inicio = datetime.date(anio, mes, 1)
            fecha_fin = datetime.date(anio, mes + 1, 1) - datetime.timedelta(days=1) if mes < 12 else datetime.date(anio, 12, 31)
        except Exception:
            fecha_inicio = None
            fecha_fin = None
    elif fecha_inicio and fecha_fin:
        pass
    else:
        fecha_inicio = None
        fecha_fin = None

    # Convertir strings a date
    if isinstance(fecha_inicio, str):
        try:
            fecha_inicio = datetime.datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        except Exception:
            fecha_inicio = None
    if isinstance(fecha_fin, str):
        try:
            fecha_fin = datetime.datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except Exception:
            fecha_fin = None

    # --- Etiqueta del periodo ---
    try:
        locale.setlocale(locale.LC_TIME, "es_MX.UTF-8")
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
        except locale.Error:
            locale.setlocale(locale.LC_TIME, "C")

    mes_letra = ""
    if fecha_inicio and fecha_fin:
        if fecha_inicio == fecha_fin.replace(day=1) and fecha_inicio.month == fecha_fin.month:
            mes_letra = fecha_inicio.strftime("%B %Y").capitalize()
        else:
            mes_letra = f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"

    # --- Empresa ---
    empresa = None
    if empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            pass

    # Saldo inicial = suma de saldos iniciales de cuentas bancarias activas
    saldo_inicial_empresa = float(
        CuentaBancaria.objects.filter(empresa_id=empresa_id, activa=True)
        .aggregate(t=Sum("saldo_inicial"))["t"] or 0
    ) if empresa_id else 0.0

    # --- Querysets base ---
    pagos = Pago.objects.exclude(forma_pago="nota_credito")
    cobros_otros = CobroOtrosIngresos.objects.select_related("factura", "factura__empresa")
    gastos_modo = PagoGasto.objects.all()
    gastos_caja_chica = GastoCajaChica.objects.all()
    vales_caja_chica = ValeCaja.objects.all()

    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id)
        cobros_otros = cobros_otros.filter(factura__empresa_id=empresa_id)
        gastos_modo = gastos_modo.filter(gasto__empresa_id=empresa_id)
        gastos_caja_chica = gastos_caja_chica.filter(fondeo__empresa_id=empresa_id)
        vales_caja_chica = vales_caja_chica.filter(fondeo__empresa_id=empresa_id)

    # --- Saldo inicial dinámico acumulado ---
    saldo_inicial = saldo_inicial_empresa
    if empresa and empresa_id and fecha_inicio:
        anio_inicio = empresa.fecha_creacion.year if hasattr(empresa, 'fecha_creacion') and empresa.fecha_creacion else fecha_inicio.year

        if mes and anio:
            anio_tope = int(anio)
            mes_tope = int(mes) - 1
            if mes_tope == 0:
                anio_tope -= 1
                mes_tope = 12
        else:
            if fecha_inicio.month == 1:
                anio_tope = fecha_inicio.year - 1
                mes_tope = 12
            else:
                anio_tope = fecha_inicio.year
                mes_tope = fecha_inicio.month - 1

        if anio_tope >= anio_inicio:
            for y in range(anio_inicio, anio_tope + 1):
                mes_fin_loop = mes_tope if y == anio_tope else 12
                for m in range(1, mes_fin_loop + 1):
                    fi = datetime.date(y, m, 1)
                    ff = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1) if m < 12 else datetime.date(y, 12, 31)
                    ing = float(
                        Pago.objects.exclude(forma_pago="nota_credito")
                        .filter(factura__empresa_id=empresa_id, fecha_pago__gte=fi, fecha_pago__lte=ff)
                        .aggregate(t=Sum("monto"))["t"] or 0
                    ) + float(
                        CobroOtrosIngresos.objects
                        .filter(factura__empresa_id=empresa_id, fecha_cobro__gte=fi, fecha_cobro__lte=ff)
                        .aggregate(t=Sum("monto"))["t"] or 0
                    ) + float(
                        Pago.objects.filter(factura__isnull=True, identificado=False, empresa_id=empresa_id, fecha_pago__gte=fi, fecha_pago__lte=ff)
                        .aggregate(t=Sum("monto"))["t"] or 0
                    )
                    gto = float(
                        PagoGasto.objects
                        .filter(gasto__empresa_id=empresa_id, fecha_pago__gte=fi, fecha_pago__lte=ff)
                        .aggregate(t=Sum("monto"))["t"] or 0
                    ) + float(
                        FondeoCajaChica.objects  # ← usar fondeos, NO gastos individuales
                        .filter(empresa_id=empresa_id, fecha__gte=fi, fecha__lte=ff)
                        .aggregate(t=Sum("importe_cheque"))["t"] or 0
                    )
                    saldo_inicial += ing - gto
        # # Temporal — después del loop del saldo inicial
        # print(f"DEBUG saldo_inicial calculado: {saldo_inicial}")
        
    # --- Filtro por fechas ---
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio)
        cobros_otros = cobros_otros.filter(fecha_cobro__gte=fecha_inicio)
        gastos_modo = gastos_modo.filter(fecha_pago__gte=fecha_inicio)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__gte=fecha_inicio)
        vales_caja_chica = vales_caja_chica.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin)
        cobros_otros = cobros_otros.filter(fecha_cobro__lte=fecha_fin)
        gastos_modo = gastos_modo.filter(fecha_pago__lte=fecha_fin)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__lte=fecha_fin)
        vales_caja_chica = vales_caja_chica.filter(fecha__lte=fecha_fin)

    # --- Depósitos no identificados ---
    pagos_por_identificar = Pago.objects.filter(factura__isnull=True, identificado=False)
    if empresa_id:
        pagos_por_identificar = pagos_por_identificar.filter(empresa_id=empresa_id)
    if fecha_inicio:
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__lte=fecha_fin)
    total_pagos_por_identificar = float(pagos_por_identificar.aggregate(t=Sum("monto"))["t"] or 0)

    # --- INGRESOS por origen ---
    ingresos_por_origen = OrderedDict()
    ingresos_qs = (
        pagos.annotate(
            origen=Case(
                When(factura__local__isnull=False, then=Value("Locales")),
                When(factura__area_comun__isnull=False, then=Value("Áreas Comunes")),
                default=Value("Sin origen"),
                output_field=CharField(),
            )
        )
        .values("origen")
        .annotate(total=Sum("monto"))
        .order_by("origen")
    )
    for x in ingresos_qs:
        ingresos_por_origen[(x["origen"] or "Sin origen").strip().title()] = float(x["total"])

    otros_qs = (
        cobros_otros.select_related("factura__tipo_ingreso")
        .values("factura__tipo_ingreso__nombre")
        .annotate(total=Sum("monto"))
    )
    for x in otros_qs:
        tipo = (x["factura__tipo_ingreso__nombre"] or "Otros ingresos").strip().title()
        ingresos_por_origen[f"Otros ingresos - {tipo}"] = float(x["total"])

    ingresos_por_origen["Depósitos no identificados"] = total_pagos_por_identificar
    total_ingresos = float(sum(ingresos_por_origen.values()))

    # --- GASTOS por grupo/subgrupo/tipo ---
    estructura_gastos = OrderedDict()

    def agregar_a_estructura(qs, grupo_key, subgrupo_key, tipo_key, monto_key):
        for g in qs.values(grupo_key, subgrupo_key, tipo_key).annotate(total=Sum(monto_key)):
            grupo = (g[grupo_key] or "Sin grupo").strip().title()
            subgrupo = (g[subgrupo_key] or "Sin subgrupo").strip().title()
            tipo = (g[tipo_key] or "Sin tipo").strip().title()
            total = float(g["total"])
            estructura_gastos.setdefault(grupo, OrderedDict()).setdefault(subgrupo, {})
            estructura_gastos[grupo][subgrupo][tipo] = estructura_gastos[grupo][subgrupo].get(tipo, 0) + total

    # --- CAJA CHICA ---
    filtros_fondeo = {}
    if empresa_id:
        filtros_fondeo["empresa_id"] = empresa_id
    if fecha_inicio:
        filtros_fondeo["fecha__gte"] = fecha_inicio
    if fecha_fin:
        filtros_fondeo["fecha__lte"] = fecha_fin



    agregar_a_estructura(gastos_modo,
        "gasto__tipo_gasto__subgrupo__grupo__nombre",
        "gasto__tipo_gasto__subgrupo__nombre",
        "gasto__tipo_gasto__nombre", "monto")
    
    # Por esto — fondeos como una sola línea de egreso:
    fondeos_periodo = FondeoCajaChica.objects.filter(**filtros_fondeo)
    total_fondeos = float(fondeos_periodo.aggregate(t=Sum("importe_cheque"))["t"] or 0)
    if total_fondeos > 0:
        estructura_gastos.setdefault("Caja Chica", OrderedDict())
        estructura_gastos["Caja Chica"]["Fondeos"] = [{"tipo": "Fondeo de caja chica", "total": total_fondeos}]

    # Convertir a listas para el template
    for grupo in estructura_gastos:
        for subgrupo in estructura_gastos[grupo]:
            tipos_dict = estructura_gastos[grupo][subgrupo]
            if isinstance(tipos_dict, dict):
                estructura_gastos[grupo][subgrupo] = [
                    {"tipo": tipo, "total": total} for tipo, total in tipos_dict.items()
            ]

    total_gastos = sum(
    sum(t["total"] for t in tipos)
    for subgrupos in estructura_gastos.values()
    for tipos in subgrupos.values()
    if isinstance(tipos, list)
    )

    # Justo después de calcular total_gastos en la vista
    # print(f"DEBUG total_gastos: {total_gastos}")
    # print(f"DEBUG total_ingresos: {total_ingresos}")
    # print(f"DEBUG saldo_inicial: {saldo_inicial}")
    # print(f"DEBUG saldo_final_flujo: {saldo_inicial + total_ingresos - total_gastos}")

    # --- SALDO FINAL FLUJO ---
    saldo_final_flujo = saldo_inicial + total_ingresos - total_gastos
    #saldo_final_bancos_menos_caja = saldo_final_flujo - saldo_caja_chica
    saldo = total_ingresos - total_gastos


    return render(
        request,
        "informes_financieros/estado_resultados.html",
        {
            "empresas": empresas,
            "empresa_id": str(empresa_id or ""),
            "empresa": empresa,
            "ingresos_por_origen": ingresos_por_origen,
            "estructura_gastos": estructura_gastos,
            "total_ingresos": total_ingresos,
            "total_gastos": total_gastos,
            "saldo": saldo,
            "saldo_inicial": saldo_inicial,
            "saldo_final_flujo": saldo_final_flujo,
            #"saldo_caja_chica": saldo_caja_chica,
            #"saldo_final_bancos_menos_caja": saldo_final_bancos_menos_caja,
            "fecha_inicio": fecha_inicio.strftime('%Y-%m-%d') if fecha_inicio else '',
            "fecha_fin": fecha_fin.strftime('%Y-%m-%d') if fecha_fin else '',
            "mes": str(mes or ""),
            "anio": str(anio or ""),
            "periodo": periodo,
            "meses_unicos": meses_unicos,
            "anios_unicos": anios_unicos,
            "mes_letra": mes_letra,
            "total_pagos_por_identificar": total_pagos_por_identificar,
        },
    )


@login_o_portal_required 
def exportar_estado_resultados_excel(request):
    from collections import OrderedDict

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    periodo = request.GET.get("periodo")
    hoy = datetime.date.today()

    if not request.user.is_superuser:
        empresa_id = str(request.user.perfilusuario.empresa.id)
    else:
        empresa_id = request.GET.get("empresa") or ""

    empresa_nombre = ""
    empresa = None

    if empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
            empresa_nombre = empresa.nombre
        except Empresa.DoesNotExist:
            pass

    # Saldo inicial = suma de saldos iniciales de cuentas bancarias activas
    saldo_inicial_empresa = float(
        CuentaBancaria.objects.filter(empresa_id=empresa_id, activa=True)
        .aggregate(t=Sum("saldo_inicial"))["t"] or 0
    ) if empresa_id else 0.0

    # --- Querysets base ---
    pagos = Pago.objects.exclude(forma_pago="nota_credito")
    cobros_otros = CobroOtrosIngresos.objects.select_related("factura", "factura__empresa")
    gastos_caja_chica = GastoCajaChica.objects.all()
    vales_caja_chica = ValeCaja.objects.all()

    # --- Periodo ---
    if not periodo and not fecha_inicio and not fecha_fin and not mes and not anio:
        periodo = "periodo_actual"

    if periodo == "mes_actual":
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = (hoy.replace(day=1) + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)
        mes = hoy.month
        anio = hoy.year
    elif periodo == "periodo_actual":
        fecha_inicio = hoy.replace(month=1, day=1)
        fecha_fin = hoy
        mes = ""
        anio = ""
    elif mes and anio:
        try:
            mes = int(mes)
            anio = int(anio)
            fecha_inicio = datetime.date(anio, mes, 1)
            fecha_fin = datetime.date(anio, mes + 1, 1) - datetime.timedelta(days=1) if mes < 12 else datetime.date(anio, 12, 31)
        except Exception:
            fecha_inicio = None
            fecha_fin = None

    if isinstance(fecha_inicio, str):
        try:
            fecha_inicio = datetime.datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        except Exception:
            fecha_inicio = None
    if isinstance(fecha_fin, str):
        try:
            fecha_fin = datetime.datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except Exception:
            fecha_fin = None

    # --- Filtros por empresa ---
    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id)
        cobros_otros = cobros_otros.filter(factura__empresa_id=empresa_id)
        gastos_caja_chica = gastos_caja_chica.filter(fondeo__empresa_id=empresa_id)
        vales_caja_chica = vales_caja_chica.filter(fondeo__empresa_id=empresa_id)

    gastos_modo = PagoGasto.objects.all()
    if empresa_id:
        gastos_modo = gastos_modo.filter(gasto__empresa_id=empresa_id)

    # --- Saldo inicial dinámico acumulado ---
    saldo_inicial = saldo_inicial_empresa
    if empresa and empresa_id and fecha_inicio:
        anio_inicio = empresa.fecha_creacion.year if hasattr(empresa, 'fecha_creacion') and empresa.fecha_creacion else fecha_inicio.year

        if mes and anio:
            anio_tope = int(anio)
            mes_tope = int(mes) - 1
            if mes_tope == 0:
                anio_tope -= 1
                mes_tope = 12
        else:
            if fecha_inicio.month == 1:
                anio_tope = fecha_inicio.year - 1
                mes_tope = 12
            else:
                anio_tope = fecha_inicio.year
                mes_tope = fecha_inicio.month - 1

        if anio_tope >= anio_inicio:
            for y in range(anio_inicio, anio_tope + 1):
                mes_fin_loop = mes_tope if y == anio_tope else 12
                for m in range(1, mes_fin_loop + 1):
                    fi = datetime.date(y, m, 1)
                    ff = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1) if m < 12 else datetime.date(y, 12, 31)
                    ing = float(
                        Pago.objects.exclude(forma_pago="nota_credito")
                        .filter(factura__empresa_id=empresa_id, fecha_pago__gte=fi, fecha_pago__lte=ff)
                        .aggregate(t=Sum("monto"))["t"] or 0
                    ) + float(
                        CobroOtrosIngresos.objects
                        .filter(factura__empresa_id=empresa_id, fecha_cobro__gte=fi, fecha_cobro__lte=ff)
                        .aggregate(t=Sum("monto"))["t"] or 0
                    )
                    gto = float(
                    PagoGasto.objects
                    .filter(gasto__empresa_id=empresa_id, fecha_pago__gte=fi, fecha_pago__lte=ff)
                    .aggregate(t=Sum("monto"))["t"] or 0
                    ) + float(
                        FondeoCajaChica.objects  # ← usar fondeos, NO gastos individuales
                        .filter(empresa_id=empresa_id, fecha__gte=fi, fecha__lte=ff)
                        .aggregate(t=Sum("importe_cheque"))["t"] or 0
                    )
                    saldo_inicial += ing - gto

    # --- Filtros por fecha ---
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio)
        cobros_otros = cobros_otros.filter(fecha_cobro__gte=fecha_inicio)
        gastos_modo = gastos_modo.filter(fecha_pago__gte=fecha_inicio)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__gte=fecha_inicio)
        vales_caja_chica = vales_caja_chica.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin)
        cobros_otros = cobros_otros.filter(fecha_cobro__lte=fecha_fin)
        gastos_modo = gastos_modo.filter(fecha_pago__lte=fecha_fin)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__lte=fecha_fin)
        vales_caja_chica = vales_caja_chica.filter(fecha__lte=fecha_fin)

    # --- Depósitos no identificados ---
    pagos_por_identificar = Pago.objects.filter(factura__isnull=True, identificado=False)
    if empresa_id:
        pagos_por_identificar = pagos_por_identificar.filter(empresa_id=empresa_id)
    if fecha_inicio:
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__lte=fecha_fin)
    total_pagos_por_identificar = float(pagos_por_identificar.aggregate(t=Sum("monto"))["t"] or 0)

    # --- INGRESOS ---
    ingresos_por_origen = OrderedDict()
    ingresos_qs = (
        pagos.annotate(
            origen=Case(
                When(factura__local__isnull=False, then=Value("Locales")),
                When(factura__area_comun__isnull=False, then=Value("Áreas Comunes")),
                default=Value("Sin origen"),
                output_field=CharField(),
            )
        )
        .values("origen")
        .annotate(total=Sum("monto"))
        .order_by("origen")
    )
    for x in ingresos_qs:
        ingresos_por_origen[(x["origen"] or "Sin origen").strip().title()] = float(x["total"])

    otros_qs = (
        cobros_otros.values("factura__tipo_ingreso__nombre")
        .annotate(total=Sum("monto"))
    )
    for x in otros_qs:
        tipo = (x["factura__tipo_ingreso__nombre"] or "Otros ingresos").strip().title()
        ingresos_por_origen[f"Otros ingresos - {tipo}"] = float(x["total"])

    ingresos_por_origen["Depósitos no identificados"] = total_pagos_por_identificar
    total_ingresos = float(sum(ingresos_por_origen.values()))

    # --- GASTOS ---
    def agregar_a_estructura_excel(estructura, qs, grupo_key, subgrupo_key, tipo_key, monto_key):
        for g in qs.values(grupo_key, subgrupo_key, tipo_key).annotate(total=Sum(monto_key)):
            grupo = (g[grupo_key] or "Sin grupo").strip().title()
            subgrupo = (g[subgrupo_key] or "Sin subgrupo").strip().title()
            tipo = (g[tipo_key] or "Sin tipo").strip().title()
            total = float(g["total"])
            estructura.setdefault(grupo, OrderedDict()).setdefault(subgrupo, {})
            estructura[grupo][subgrupo][tipo] = estructura[grupo][subgrupo].get(tipo, 0) + total

    estructura_gastos = OrderedDict()
    agregar_a_estructura_excel(estructura_gastos, gastos_modo,
        "gasto__tipo_gasto__subgrupo__grupo__nombre",
        "gasto__tipo_gasto__subgrupo__nombre",
        "gasto__tipo_gasto__nombre", "monto")
    agregar_a_estructura_excel(estructura_gastos, gastos_caja_chica,
        "tipo_gasto__subgrupo__grupo__nombre",
        "tipo_gasto__subgrupo__nombre",
        "tipo_gasto__nombre", "importe")
    agregar_a_estructura_excel(estructura_gastos, vales_caja_chica,
        "tipo_gasto__subgrupo__grupo__nombre",
        "tipo_gasto__subgrupo__nombre",
        "tipo_gasto__nombre", "importe")

    for grupo in estructura_gastos:
        for subgrupo in estructura_gastos[grupo]:
            tipos_dict = estructura_gastos[grupo][subgrupo]
            estructura_gastos[grupo][subgrupo] = [
                {"tipo": tipo, "total": total} for tipo, total in tipos_dict.items()
            ]

    total_gastos = sum(
        sum(t["total"] for t in tipos)
        for subgrupos in estructura_gastos.values()
        for tipos in subgrupos.values()
    )
    saldo_final_flujo = float(saldo_inicial) + float(total_ingresos) - float(total_gastos)
    saldo = float(total_ingresos) - float(total_gastos)

    # --- Excel ---
    bold = Font(bold=True)
    title_font = Font(size=14, bold=True)
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    group_fill = PatternFill("solid", fgColor="DDEBF7")
    sub_fill = PatternFill("solid", fgColor="F7F7F7")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Estado de Resultados — Flujo de Efectivo"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Empresa: {empresa_nombre}"
    ws["A2"].font = bold
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:B3")
    ws["A3"] = f"Periodo: {fecha_inicio} a {fecha_fin}"
    ws["A3"].font = bold
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.append([])

    row = ws.max_row + 1
    ws.append(["Saldo inicial bancos", saldo_inicial])
    ws[f"A{row}"].font = bold
    ws[f"B{row}"].font = bold
    ws[f"B{row}"].number_format = '#,##0.00'

    ws.append(["Ingresos", "Importe"])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
    for origen, monto in ingresos_por_origen.items():
        ws.append([origen, monto])
        ws[f"A{ws.max_row}"].fill = group_fill
        ws[f"A{ws.max_row}"].border = border
        ws[f"B{ws.max_row}"].border = border
        ws[f"B{ws.max_row}"].number_format = '#,##0.00'
    ws.append(["Total Ingresos", total_ingresos])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.number_format = '#,##0.00'
    ws.append([])

    ws.append(["Gastos", "Importe"])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
    for grupo, subgrupos in estructura_gastos.items():
        ws.append([grupo, ""])
        ws[f"A{ws.max_row}"].fill = group_fill
        ws[f"A{ws.max_row}"].font = bold
        ws[f"A{ws.max_row}"].border = border
        ws[f"B{ws.max_row}"].border = border
        for subgrupo, tipos in subgrupos.items():
            ws.append(["  " + subgrupo, ""])
            ws[f"A{ws.max_row}"].fill = sub_fill
            ws[f"A{ws.max_row}"].font = bold
            ws[f"A{ws.max_row}"].border = border
            ws[f"B{ws.max_row}"].border = border
            for tipo in tipos:
                ws.append(["    " + tipo["tipo"], tipo["total"]])
                ws[f"A{ws.max_row}"].border = border
                ws[f"B{ws.max_row}"].border = border
                ws[f"B{ws.max_row}"].number_format = '#,##0.00'
    ws.append(["Total Gastos", total_gastos])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.number_format = '#,##0.00'
    ws.append([])

    ws.append(["Resultado del periodo", saldo])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.number_format = '#,##0.00'

    ws.append(["Saldo final bancos", saldo_final_flujo])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=estado_resultados.xlsx"
    wb.save(response)
    return response


# reporte detalle adeudos vencidos por origen (local, area comun, tipo de ingreso)
@login_required
def cartera_vencida_por_origen(request):
    empresa_id = request.GET.get("empresa")
    hoy = timezone.now().date()
    filtro_origen = request.GET.get("filtro_origen", "")
    cliente_id = request.GET.get("cliente")

    if not request.user.is_superuser:
        empresa = request.user.perfilusuario.empresa
        empresa_id = str(empresa.id)
    else:
        empresa = None
        if empresa_id:
            try:
                empresa = Empresa.objects.get(id=empresa_id)
            except Empresa.DoesNotExist:
                empresa = None

    if empresa:
        clientes = Cliente.objects.filter(empresa=empresa)
    else:
        clientes = Cliente.objects.all()

    # ── Subquery para total pagado por factura ──
    total_pagado_sq = Pago.objects.filter(
        factura=OuterRef('pk')
    ).values('factura').annotate(
        total=Sum('monto')
    ).values('total')

    facturas = Factura.objects.filter(
        estatus="pendiente",
        fecha_vencimiento__lt=hoy,
        monto__gt=0
    ).select_related('local', 'area_comun', 'cliente').annotate(
        origen_id=Case(
            When(local__isnull=False, then=F("local__id")),
            When(area_comun__isnull=False, then=F("area_comun__id")),
            default=Value(0),
            output_field=IntegerField(),
        ),
        origen_tipo=Case(
            When(local__isnull=False, then=Value("local")),
            When(area_comun__isnull=False, then=Value("area")),
            default=Value("sin_origen"),
            output_field=CharField(),
        ),
        origen_nombre=Case(
            When(local__isnull=False, then=F("local__numero")),
            When(area_comun__isnull=False, then=F("area_comun__numero")),
            default=Value("Sin origen"),
            output_field=CharField(),
        ),
        total_pagado_ann=Coalesce(
            Subquery(total_pagado_sq, output_field=DecimalField()),
            Value(0, output_field=DecimalField())
        )
    ).order_by("origen_tipo", "origen_nombre", "fecha_vencimiento")

    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
    if cliente_id and cliente_id.isdigit():
        facturas = facturas.filter(cliente_id=cliente_id)

    # ── Subquery para saldo de otros ingresos ──
    total_cobrado_oi_sq = CobroOtrosIngresos.objects.filter(
        factura=OuterRef('pk')
    ).values('factura').annotate(
        total=Sum('monto')
    ).values('total')

    facturas_oi = FacturaOtrosIngresos.objects.filter(
        estatus="pendiente",
        fecha_vencimiento__lt=hoy,
        monto__gt=0
    ).select_related('tipo_ingreso', 'cliente').annotate(
        origen_id=F("tipo_ingreso__id"),
        origen_tipo=Value("tipoingreso", output_field=CharField()),
        origen_nombre=F("tipo_ingreso__nombre"),
        total_cobrado_ann=Coalesce(
            Subquery(total_cobrado_oi_sq, output_field=DecimalField()),
            Value(0, output_field=DecimalField())
        )
    ).order_by("origen_nombre", "fecha_vencimiento")

    if empresa_id:
        facturas_oi = facturas_oi.filter(empresa_id=empresa_id)
    if cliente_id and cliente_id.isdigit():
        facturas_oi = facturas_oi.filter(cliente_id=cliente_id)

    # ── Construcción del resultado usando anotaciones ──
    origenes_dict = {}

    for f in facturas:
        if f.origen_tipo == "local":
            origen_id = f"local_{f.origen_id}"
            origen_nombre = f"Local: {f.origen_nombre}"
        elif f.origen_tipo == "area":
            origen_id = f"area_{f.origen_id}"
            origen_nombre = f"Área Común: {f.origen_nombre}"
        else:
            origen_id = "sin_origen"
            origen_nombre = "Sin origen"

        if origen_id not in origenes_dict:
            origenes_dict[origen_id] = {"origen_nombre": origen_nombre, "total_vencido": 0, "facturas": []}

        # Usar anotación en lugar del property
        saldo = float(f.monto) - float(f.total_pagado_ann)
        origenes_dict[origen_id]["facturas"].append({
            "cliente": f.cliente.nombre if f.cliente else "Desconocido",
            "factura_id": f.id,
            "folio": f.folio,
            "fecha_vencimiento": f.fecha_vencimiento,
            "dias_vencidos": (hoy - f.fecha_vencimiento).days,
            "monto": float(f.monto),
            "saldo": saldo,
            "concepto": f.observaciones,
        })
        origenes_dict[origen_id]["total_vencido"] += saldo

    for f in facturas_oi:
        origen_id = f"tipoingreso_{f.origen_id}"
        origen_nombre = f"Tipo de Ingreso: {f.origen_nombre}"
        if origen_id not in origenes_dict:
            origenes_dict[origen_id] = {"origen_nombre": origen_nombre, "total_vencido": 0, "facturas": []}

        # Usar anotación en lugar del property
        saldo = float(f.monto) - float(f.total_cobrado_ann)
        origenes_dict[origen_id]["facturas"].append({
            "cliente": f.cliente.nombre if f.cliente else "Desconocido",
            "factura_id": f.id,
            "folio": f.folio,
            "fecha_vencimiento": f.fecha_vencimiento,
            "dias_vencidos": (hoy - f.fecha_vencimiento).days,
            "monto": float(f.monto),
            "saldo": saldo,
            "concepto": f.observaciones,
        })
        origenes_dict[origen_id]["total_vencido"] += saldo

    if filtro_origen in ("local", "area", "tipoingreso"):
        resultado = [v for k, v in origenes_dict.items() if k.startswith(filtro_origen + "_")]
    else:
        resultado = list(origenes_dict.values())

    total_cartera = sum(o["total_vencido"] for o in resultado)

    return render(request, "informes_financieros/cartera_vencida_x_origen.html", {
        "cartera_vencida": resultado,
        "total_cartera": total_cartera,
        "cliente_id": cliente_id,
        "clientes": clientes,
        "empresas": Empresa.objects.all() if request.user.is_superuser else None,
    })


@login_required
def exportar_cartera_vencida_excel(request):
    hoy = timezone.now().date()
    filtro_origen = request.GET.get("filtro_origen", "")

    if not request.user.is_superuser:
        empresa_id = str(request.user.perfilusuario.empresa.id)
    else:
        empresa_id = request.GET.get("empresa") or ""

    # ── Subquery saldo facturas cuotas ──
    total_pagado_sq = Pago.objects.filter(
        factura=OuterRef('pk')
    ).values('factura').annotate(total=Sum('monto')).values('total')

    facturas = Factura.objects.filter(
        estatus="pendiente", fecha_vencimiento__lt=hoy, monto__gt=0
    ).select_related("local", "area_comun", "cliente").annotate(
        total_pagado_ann=Coalesce(
            Subquery(total_pagado_sq, output_field=DecimalField()),
            Value(0, output_field=DecimalField())
        )
    ).order_by("local__numero", "area_comun__numero", "fecha_vencimiento")

    # ── Subquery saldo otros ingresos ──
    total_cobrado_oi_sq = CobroOtrosIngresos.objects.filter(
        factura=OuterRef('pk')
    ).values('factura').annotate(total=Sum('monto')).values('total')

    facturas_oi = FacturaOtrosIngresos.objects.filter(
        estatus="pendiente", fecha_vencimiento__lt=hoy, monto__gt=0
    ).select_related("tipo_ingreso", "cliente").annotate(
        total_cobrado_ann=Coalesce(
            Subquery(total_cobrado_oi_sq, output_field=DecimalField()),
            Value(0, output_field=DecimalField())
        )
    ).order_by("tipo_ingreso__nombre", "fecha_vencimiento")

    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
        facturas_oi = facturas_oi.filter(empresa_id=empresa_id)

    # ── Agrupación ──
    origenes_dict = {}

    for f in facturas:
        if f.local:
            origen_id = f"local_{f.local.id}"
            origen_nombre = f"Local: {f.local.numero}"
        elif f.area_comun:
            origen_id = f"area_{f.area_comun.id}"
            origen_nombre = f"Área Común: {f.area_comun.numero}"
        else:
            origen_id = "sin_origen"
            origen_nombre = "Sin origen"

        if origen_id not in origenes_dict:
            origenes_dict[origen_id] = {"origen_nombre": origen_nombre, "total_vencido": 0, "facturas": []}

        saldo = float(f.monto) - float(f.total_pagado_ann)
        origenes_dict[origen_id]["facturas"].append({
            "cliente": f.cliente.nombre if f.cliente else "Desconocido",
            "folio": f.folio,
            "fecha_vencimiento": f.fecha_vencimiento,
            "dias_vencidos": (hoy - f.fecha_vencimiento).days,
            "monto": float(f.monto),
            "saldo": saldo,
            "concepto": f.observaciones or "",
        })
        origenes_dict[origen_id]["total_vencido"] += saldo

    for f in facturas_oi:
        origen_id = f"tipoingreso_{f.tipo_ingreso.id}"
        origen_nombre = f"Tipo de Ingreso: {f.tipo_ingreso.nombre}"

        if origen_id not in origenes_dict:
            origenes_dict[origen_id] = {"origen_nombre": origen_nombre, "total_vencido": 0, "facturas": []}

        saldo = float(f.monto) - float(f.total_cobrado_ann)
        origenes_dict[origen_id]["facturas"].append({
            "cliente": f.cliente.nombre if f.cliente else "Desconocido",
            "folio": f.folio,
            "fecha_vencimiento": f.fecha_vencimiento,
            "dias_vencidos": (hoy - f.fecha_vencimiento).days,
            "monto": float(f.monto),
            "saldo": saldo,
            "concepto": f.observaciones or "",
        })
        origenes_dict[origen_id]["total_vencido"] += saldo

    if filtro_origen in ("local", "area", "tipoingreso"):
        resultado = [v for k, v in origenes_dict.items() if k.startswith(filtro_origen + "_")]
    else:
        resultado = list(origenes_dict.values())

    total_cartera = sum(o["total_vencido"] for o in resultado)

    # ── EXCEL PROFESIONAL ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera Vencida"

    # Estilos
    fmt_moneda = '$#,##0.00'
    fmt_num = '#,##0'

    def estilo_header_origen():
        return {
            'font': Font(name='Arial', bold=True, color='FFFFFF', size=11),
            'fill': PatternFill('solid', fgColor='8B0000'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }

    def estilo_header_col():
        return {
            'font': Font(name='Arial', bold=True, color='FFFFFF', size=10),
            'fill': PatternFill('solid', fgColor='1F4E79'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }

    def estilo_celda(align='left'):
        return {
            'font': Font(name='Arial', size=10),
            'alignment': Alignment(horizontal=align, vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }

    def aplicar_estilos(cell, estilos):
        for attr, val in estilos.items():
            setattr(cell, attr, val)

    # TÍTULO
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A1:G1')
    ws['A1'] = f'Cartera Vencida por Origen — Al {hoy.strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='5A0000')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 20
    ws['A2'] = f'Total cartera: ${total_cartera:,.2f}'
    ws['A2'].font = Font(name='Arial', bold=True, size=12, color='8B0000')
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    row = 4

    for origen in resultado:
        # Header del origen
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=origen['origen_nombre'])
        ws.cell(row=row, column=7, value=origen['total_vencido'])
        ws.cell(row=row, column=7).number_format = fmt_moneda

        for col in range(1, 8):
            aplicar_estilos(ws.cell(row=row, column=col), estilo_header_origen())
        row += 1

        # Headers columnas
        ws.row_dimensions[row].height = 18
        headers = ['Cliente', 'Folio', 'Vencimiento', 'Días Vencido', 'Monto', 'Saldo', 'Concepto']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            aplicar_estilos(cell, estilo_header_col())
        row += 1

        # Filas
        for f in origen['facturas']:
            ws.row_dimensions[row].height = 16
            dias = f['dias_vencidos']

            # Color semáforo por días
            if dias <= 30:
                fill_dias = PatternFill('solid', fgColor='EAFAF1')
                color_dias = '1E8449'
            elif dias <= 60:
                fill_dias = PatternFill('solid', fgColor='FEF9E7')
                color_dias = 'C8A000'
            elif dias <= 90:
                fill_dias = PatternFill('solid', fgColor='FEF0E7')
                color_dias = 'E67E22'
            elif dias <= 180:
                fill_dias = PatternFill('solid', fgColor='FDE8D8')
                color_dias = 'D35400'
            else:
                fill_dias = PatternFill('solid', fgColor='FDF0F0')
                color_dias = 'C0392B'

            valores = [
                f['cliente'],
                f['folio'],
                f['fecha_vencimiento'].strftime('%d/%m/%Y'),
                f['dias_vencidos'],
                f['monto'],
                f['saldo'],
                f['concepto'],
            ]
            formatos = ['@', '@', '@', fmt_num, fmt_moneda, fmt_moneda, '@']
            alineaciones = ['left', 'center', 'center', 'center', 'right', 'right', 'left']

            for col, (val, fmt, aln) in enumerate(zip(valores, formatos, alineaciones), 1):
                cell = ws.cell(row=row, column=col, value=val)
                aplicar_estilos(cell, estilo_celda(aln))
                cell.number_format = fmt
                if col == 4:  # Días vencido con semáforo
                    cell.fill = fill_dias
                    cell.font = Font(name='Arial', size=10, bold=True, color=color_dias)
                if col == 6 and f['saldo'] > 0:  # Saldo en rojo
                    cell.font = Font(name='Arial', size=10, bold=True, color='C0392B')
            row += 1

        # Subtotal origen
        ws.cell(row=row, column=6, value=origen['total_vencido'])
        ws.cell(row=row, column=6).number_format = fmt_moneda
        ws.cell(row=row, column=6).font = Font(name='Arial', bold=True, color='8B0000')
        ws.cell(row=row, column=6).fill = PatternFill('solid', fgColor='FDF0F0')
        ws.cell(row=row, column=5, value='Subtotal:')
        ws.cell(row=row, column=5).font = Font(name='Arial', bold=True)
        row += 2  # Espacio entre orígenes

    # GRAN TOTAL
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='GRAN TOTAL CARTERA VENCIDA')
    ws.cell(row=row, column=6, value=total_cartera)
    ws.cell(row=row, column=6).number_format = fmt_moneda

    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center' if col == 1 else 'right', vertical='center')
        cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )

    # Ancho columnas
    anchos = [30, 14, 14, 13, 14, 14, 35]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Congelar primera fila
    ws.freeze_panes = 'A3'

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=cartera_vencida_{hoy.strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response