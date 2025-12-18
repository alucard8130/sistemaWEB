
# Create your views here.
from django.forms import CharField
from django.shortcuts import render, redirect,get_object_or_404
from areas.models import AreaComun
from clientes.models import Cliente
from empresas.models import Empresa
from locales.models import LocalComercial
from .forms import FacturaEditForm, FacturaForm, FacturaOtrosIngresosForm, MotivoReversaCobroForm, PagoForm,FacturaCargaMasivaForm, CobroForm, PagoPorIdentificarForm, TipoOtroIngresoForm
from .models import CobroOtrosIngresos, Factura, FacturaOtrosIngresos, Pago, TipoOtroIngreso
from principal.models import AuditoriaCambio, PerfilUsuario
from django.utils.timezone import now
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from datetime import date, timedelta
from django.db.models import Q,  Value, Case, When,  CharField,FloatField
from django.db.models import F, OuterRef, Subquery, Sum, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from facturacion.models import Pago
from decimal import Decimal, InvalidOperation
from unidecode import unidecode
from django.db.models.functions import TruncMonth, TruncYear
from datetime import datetime
from django.db import transaction
from django.contrib.auth import authenticate
from django.core.paginator import Paginator
import io
from django.utils.dateformat import DateFormat
from presupuestos.models import PresupuestoIngreso
from collections import defaultdict
import json
from django.http import JsonResponse
from django.db.models import Max,Min
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Sum

@login_required
def crear_factura(request):
    conflicto = False
    conflicto_tipo = ""
    superuser_auth_ok = False

    
    if request.method == 'POST':
        form = FacturaForm(request.POST, request.FILES, user=request.user)
        superuser_username = request.POST.get('superuser_username')
        superuser_password = request.POST.get('superuser_password')

        if form.is_valid():
            factura = form.save(commit=False)
            tipo = form.cleaned_data.get('tipo_origen')
            if tipo == 'local':
                factura.area_comun = None
            elif tipo == 'area_comun':
                factura.local = None
            cliente = factura.cliente
            
            #validar observaciones
            if factura.observaciones is None:
                factura.observaciones = ""
            
            # ---- Validación local ----
            if factura.local:
                local_cliente = factura.local.cliente
                if local_cliente and local_cliente != cliente:
                    conflicto = True
                    conflicto_tipo = "local"

            # ---- Validación área ----
            if factura.area_comun:
                area_cliente = factura.area_comun.cliente
                if area_cliente and area_cliente != cliente:
                    conflicto = True
                    conflicto_tipo = "area"

            # ---- Lógica de autorización ----
            if conflicto:
         
                if request.user.is_superuser:
                    superuser_auth_ok = True
                    #print("[DEBUG] Usuario actual es superuser, pasa conflicto.")

                elif superuser_username and superuser_password:
                    from django.contrib.auth import authenticate
                    superuser = authenticate(username=superuser_username, password=superuser_password)
                    if superuser and superuser.is_superuser:
                        superuser_auth_ok = True
                       # print("[DEBUG] Autenticación por superuser exitosa.")
                    else:
                        form.add_error(None, "usuario sin permisos. No se creó la factura.")
                        #print("[DEBUG] Autenticación superuser fallida.")
                else:
                    form.add_error(None, f"El cliente del {conflicto_tipo} seleccionado no coincide. Contacta al administrador del sistema para autorizar el cambio.")
                    #print("[DEBUG] Conflicto detectado sin autorización.")
            else:
                superuser_auth_ok = True  # Si no hay conflicto, siempre debe pasar

            if superuser_auth_ok:
                try:
                    with transaction.atomic():
                        # Asignar empresa
                        if request.user.is_superuser:
                            factura.empresa = factura.cliente.empresa
                        else:
                            factura.empresa = request.user.perfilusuario.empresa
                        
                        factura.estatus = 'pendiente'
                        
                        if not factura.fecha_emision:
                            factura.fecha_emision = timezone.now().date()
                        factura.save()
                        
                        # Folio único
                        count = Factura.objects.filter(fecha_emision__year=now().year).count() + 1
                        if tipo == 'local':
                            factura.folio = f"CM-F{count:05d}"
                        elif tipo == 'area_comun':
                            factura.folio = f"AC-F{count:05d}"
                        factura.save()

                        # Asignar cliente a local/área si está vacío o si hay conflicto autorizado
                        if factura.local and (factura.local.cliente is None or request.user.is_superuser or superuser_auth_ok):
                            factura.local.cliente = cliente
                            factura.local.save()
                        if factura.area_comun and (factura.area_comun.cliente is None or request.user.is_superuser or superuser_auth_ok):
                            factura.area_comun.cliente = cliente
                            factura.area_comun.save()

                        messages.success(request, "Registro Exitoso.")
                       
                        return redirect('lista_facturas')
                except Exception as e:
                    form.add_error(None, f"Error al guardar: {e}")
                    
        else:
            ''
    else:
        form = FacturaForm(user=request.user)
   
    return render(request, 'facturacion/crear_factura.html', {
        'form': form,
        'pedir_superuser': conflicto and not superuser_auth_ok and request.method == 'POST',
        
    })

@login_required
def lista_facturas(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')
    tipo_cuota = request.GET.get('tipo_cuota')
    query = request.GET.get('q', '')
    anio = request.GET.get('anio')

    if request.user.is_superuser:
        facturas = Factura.objects.all().order_by('-fecha_vencimiento')
        empresas = Empresa.objects.all()
        locales = LocalComercial.objects.filter(activo=True).order_by('numero')
        areas = AreaComun.objects.filter(activo=True).order_by('numero')
    else:
        empresa = request.user.perfilusuario.empresa
        facturas = Factura.objects.filter(empresa=empresa).order_by('-fecha_vencimiento')
        empresas = None
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True).order_by('numero')
        areas = AreaComun.objects.filter(empresa=empresa, activo=True).order_by('numero')

    # Si no hay ningún filtro, no mostrar nada
    if not (local_id or area_id or empresa_id  or query):
        facturas = Factura.objects.none()

    else:    
        if empresa_id:
            facturas = facturas.filter(empresa_id=empresa_id)
        if local_id:
            facturas = facturas.filter(local_id=local_id)
        if area_id:
            facturas = facturas.filter(area_comun_id=area_id)
        if tipo_cuota:
             facturas = facturas.filter(tipo_cuota=tipo_cuota)
        if query:
            facturas = facturas.filter(
                Q(folio__icontains=query) | Q(cliente__nombre__icontains=query)
            )
        if anio:
            facturas = facturas.filter(fecha_vencimiento__year=anio)

    facturas = facturas.select_related('cliente', 'empresa', 'local', 'area_comun').prefetch_related('pagos').order_by('-fecha_vencimiento')

    # Opciones únicas de tipo_cuota para el filtro
    tipos_cuota = Factura.objects.values_list('tipo_cuota', flat=True).distinct().order_by('tipo_cuota')
    
    # Paginación
    paginator = Paginator(facturas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    anios_disponibles = Factura.objects.dates('fecha_vencimiento', 'year').distinct()
    total_saldo = sum(f.saldo_pendiente for f in facturas)

    return render(request, 'facturacion/lista_facturas.html', {
        'facturas': page_obj,
        'empresas': empresas,
        'empresa_seleccionada': int(empresa_id) if empresa_id else None,
        'locales': locales,
        'areas': areas,
        'local_id': local_id,
        'area_id': area_id,
        'tipos_cuota': tipos_cuota,
        'tipo_cuota_seleccionada': tipo_cuota,
        'q': query,
        'anios_disponibles': anios_disponibles,
        'anio_seleccionado': int(anio) if anio else None,
        'total_saldo': total_saldo,
    })

@login_required
def facturar_mes_actual(request, facturar_locales=True, facturar_areas=True):
    # Permitir seleccionar año y mes por GET o POST (solo superusuario puede facturar meses anteriores)
    if request.method == 'POST':
        año = int(request.POST.get('anio', datetime.now().year))
        mes = int(request.POST.get('mes', datetime.now().month))
    else:
        año = int(request.GET.get('anio', datetime.now().year))
        mes = int(request.GET.get('mes', datetime.now().month))
    
    hoy = date.today()


    facturas_creadas = 0
    facturas_a_crear = []

    if request.user.is_superuser:
        locales = LocalComercial.objects.filter(activo=True, cliente__isnull=False) if facturar_locales else []
        locales_anuales = LocalComercial.objects.filter(activo=True, cliente__isnull=False, es_cuota_anual=True) if facturar_locales else []
        areas = AreaComun.objects.filter(activo=True, cliente__isnull=False) if facturar_areas else []
        areas_anuales = AreaComun.objects.filter(activo=True, cliente__isnull=False, es_cuota_anual=True) if facturar_areas else []
    else:
        empresa = request.user.perfilusuario.empresa
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True, cliente__isnull=False) if facturar_locales else []
        locales_anuales = LocalComercial.objects.filter(empresa=empresa, activo=True, cliente__isnull=False, es_cuota_anual=True) if facturar_locales else []
        areas = AreaComun.objects.filter(empresa=empresa, activo=True, cliente__isnull=False) if facturar_areas else []
        areas_anuales = AreaComun.objects.filter(empresa=empresa, activo=True, cliente__isnull=False, es_cuota_anual=True) if facturar_areas else []

    fecha_factura = date(año, mes, 1)

    # Pre-carga los folios existentes por empresa y tipo para evitar consultas repetidas
    folios_locales = set(
        Factura.objects.filter(
            empresa__in=[l.empresa for l in locales],
            folio__startswith="CM-F"
        ).values_list('empresa_id', 'folio')
    )
    folios_areas = set(
        Factura.objects.filter(
            empresa__in=[a.empresa for a in areas],
            folio__startswith="AC-F"
        ).values_list('empresa_id', 'folio')
    )
    folios_deposito = set(
        Factura.objects.filter(
            empresa__in=[a.empresa for a in areas],
            folio__startswith="DG-F"
        ).values_list('empresa_id', 'folio')
    )

    # Locales Mensuales
    max_folio_local = {}
    for local in locales:
        existe = Factura.objects.filter(
            cliente=local.cliente,
            local=local,
            fecha_emision__year=año,
            fecha_emision__month=mes
        ).exists()
        if not existe:
            empresa_id = local.empresa_id
            if empresa_id not in max_folio_local:
                max_folio = Factura.objects.filter(
                    empresa_id=empresa_id,
                    folio__startswith="CM-F"
                ).aggregate(max_f=Max('folio'))['max_f']
                if max_folio:
                    try:
                        last_num = int(max_folio.replace("CM-F", ""))
                    except Exception:
                        last_num = 0
                else:
                    last_num = 0
                max_folio_local[empresa_id] = last_num
            max_folio_local[empresa_id] += 1
            folio = f"CM-F{max_folio_local[empresa_id]:05d}"
            facturas_a_crear.append(Factura(
                empresa=local.empresa,
                cliente=local.cliente,
                local=local,
                folio=folio,
                fecha_emision=fecha_factura,
                fecha_vencimiento=fecha_factura,
                monto=local.cuota,
                tipo_cuota='mantenimiento',
                estatus='pendiente',
                observaciones='emision mensual'
            ))
            facturas_creadas += 1
    # Locales Anuales
    for local in locales_anuales:
        existe = Factura.objects.filter(
            cliente=local.cliente,
            local=local,
            fecha_emision__year=año,
            fecha_emision__month=1,
            observaciones='cuota anual'
        ).exists()
        if not existe and mes == 1: # Solo facturar anuales en enero
            empresa_id = local.empresa_id
            if empresa_id not in max_folio_local:
                max_folio = Factura.objects.filter(
                    empresa_id=empresa_id,
                    folio__startswith="CM-F"
                ).aggregate(max_f=Max('folio'))['max_f']
                if max_folio:
                    try:
                        last_num = int(max_folio.replace("CM-F", ""))
                    except Exception:
                        last_num = 0
                else:
                    last_num = 0
                max_folio_local[empresa_id] = last_num
            max_folio_local[empresa_id] += 1
            folio = f"CM-F{max_folio_local[empresa_id]:05d}"
            facturas_a_crear.append(Factura(
                empresa=local.empresa,
                cliente=local.cliente,
                local=local,
                folio=folio,
                fecha_emision=fecha_factura,
                fecha_vencimiento=fecha_factura,
                monto=local.cuota * 12,  # Monto anual
                tipo_cuota='mantenimiento',
                estatus='pendiente',
                observaciones='cuota anual'
            ))
            facturas_creadas += 1        

    # Áreas
    max_folio_area = {}
    max_folio_deposito = {}
    for area in areas:
        existe = Factura.objects.filter(
            cliente=area.cliente,
            area_comun=area,
            fecha_emision__year=año,
            fecha_emision__month=mes
        ).exists()
        if not existe:
            empresa_id = area.empresa_id
            if empresa_id not in max_folio_area:
                max_folio = Factura.objects.filter(
                    empresa_id=empresa_id,
                    folio__startswith="AC-F"
                ).aggregate(max_f=Max('folio'))['max_f']
                if max_folio:
                    try:
                        last_num = int(max_folio.replace("AC-F", ""))
                    except Exception:
                        last_num = 0
                else:
                    last_num = 0
                max_folio_area[empresa_id] = last_num
            max_folio_area[empresa_id] += 1
            folio = f"AC-F{max_folio_area[empresa_id]:05d}"
            facturas_a_crear.append(Factura(
                empresa=area.empresa,
                cliente=area.cliente,
                area_comun=area,
                folio=folio,
                fecha_emision=fecha_factura,
                fecha_vencimiento=fecha_factura,
                monto=area.cuota,
                tipo_cuota='renta',
                estatus='pendiente',
                observaciones='emision mensual'
            ))
            facturas_creadas += 1
    # Áreas Anuales
    for area in areas_anuales:
        existe = Factura.objects.filter(
            cliente=area.cliente,
            area_comun=area,
            fecha_emision__year=año,
            fecha_emision__month=1,
            observaciones='cuota anual'
        ).exists()
        if not existe and mes == 1: # Solo facturar anuales en enero
            empresa_id = area.empresa_id
            if empresa_id not in max_folio_area:
                max_folio = Factura.objects.filter(
                    empresa_id=empresa_id,
                    folio__startswith="AC-F"
                ).aggregate(max_f=Max('folio'))['max_f']
                if max_folio:
                    try:
                        last_num = int(max_folio.replace("AC-F", ""))
                    except Exception:
                        last_num = 0
                else:
                    last_num = 0
                max_folio_area[empresa_id] = last_num
            max_folio_area[empresa_id] += 1
            folio = f"AC-F{max_folio_area[empresa_id]:05d}"
            facturas_a_crear.append(Factura(
                empresa=area.empresa,
                cliente=area.cliente,
                area_comun=area,
                folio=folio,
                fecha_emision=fecha_factura,
                fecha_vencimiento=fecha_factura,
                monto=area.cuota * 12,  # Monto anual
                tipo_cuota='renta',
                estatus='pendiente',
                observaciones='cuota anual'
            ))
            facturas_creadas += 1

        # Depósito en garantía por única vez
        if area.deposito and area.deposito > 0:
            existe_deposito = Factura.objects.filter(
                cliente=area.cliente,
                area_comun=area,
                tipo_cuota='deposito',
            ).exists()
            if not existe_deposito:
                empresa_id = area.empresa_id
                if empresa_id not in max_folio_deposito:
                    max_folio = Factura.objects.filter(
                        empresa_id=empresa_id,
                        folio__startswith="DG-F"
                    ).aggregate(max_f=Max('folio'))['max_f']
                    if max_folio:
                        try:
                            last_num = int(max_folio.replace("DG-F", ""))
                        except Exception:
                            last_num = 0
                    else:
                        last_num = 0
                    max_folio_deposito[empresa_id] = last_num
                max_folio_deposito[empresa_id] += 1
                folio_deposito = f"DG-F{max_folio_deposito[empresa_id]:05d}"
                facturas_a_crear.append(Factura(
                    empresa=area.empresa,
                    cliente=area.cliente,
                    area_comun=area,
                    folio=folio_deposito,
                    fecha_emision=fecha_factura,
                    fecha_vencimiento=fecha_factura,
                    monto=area.deposito,
                    tipo_cuota='deposito',
                    estatus='pendiente',
                    observaciones='Depósito en garantía'
                ))

    # Bulk create
    if facturas_a_crear:
        Factura.objects.bulk_create(facturas_a_crear, batch_size=50)

    messages.success(request, f"{facturas_creadas} facturas generadas para {fecha_factura.strftime('%B %Y')}")
    return redirect('lista_facturas')

@login_required
def confirmar_facturacion(request):
    hoy = now().date()
    año, mes = hoy.year, hoy.month

    empresa = None
    if not request.user.is_superuser:
        empresa = request.user.perfilusuario.empresa

    # FILTRAR locales y áreas activos con cliente
    if request.user.is_superuser:
        locales = LocalComercial.objects.filter(activo=True, cliente__isnull=False)
        areas = AreaComun.objects.filter(activo=True, cliente__isnull=False)
    else:
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True, cliente__isnull=False)
        areas = AreaComun.objects.filter(empresa=empresa, activo=True, cliente__isnull=False)

    # CONTAR locales no facturados aún este mes
    total_locales = sum(
        not Factura.objects.filter(cliente=l.cliente, local=l,
                                   fecha_emision__year=año, fecha_emision__month=mes).exists()
        for l in locales
    )

    total_areas = sum(
        not Factura.objects.filter(cliente=a.cliente, area_comun=a,
                                   fecha_emision__year=año, fecha_emision__month=mes).exists()
        for a in areas
    )

    if request.method == 'POST':
        facturar_locales = 'locales' in request.POST
        facturar_areas = 'areas' in request.POST
        return facturar_mes_actual(request, facturar_locales, facturar_areas)

    return render(request, 'facturacion/confirmar_facturacion.html', {
        'total_locales': total_locales,
        'total_areas': total_areas,
        'año': año,
        'mes': mes
    })

@login_required
@transaction.atomic
def registrar_pago(request, factura_id):
    factura = get_object_or_404(Factura, pk=factura_id)

    if factura.estatus == "pagada" or factura.saldo_pendiente <= 0:
        messages.warning(
            request,
            "La factura ya está completamente pagada. No se pueden registrar más pagos.",
        )
        return redirect("lista_facturas")

    if request.method == "POST":
        form = PagoForm(request.POST, request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.factura = factura  
            pago.registrado_por = request.user

            if pago.forma_pago == "nota_credito":
                pago.save()
                factura.estatus = "cancelada"
                factura.monto = 0  # Saldo pendiente a 0
                factura.save()
                messages.success(request,"La factura ha sido cancelada por nota de crédito. el saldo pendiente es $0.00",)
                next_url = request.GET.get('next')
                if next_url:
                     return redirect(next_url)
                return redirect("lista_facturas")

            # Permitir pagar hasta el saldo pendiente, considerando decimales
            monto_pago = Decimal(str(pago.monto)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            saldo_pendiente = Decimal(str(factura.saldo_pendiente)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            if monto_pago > saldo_pendiente:
                form.add_error(
                    "monto",
                    f"El monto excede el saldo pendiente (${saldo_pendiente:.2f}).",
                )
            else:
                pago.monto = monto_pago
                pago.save()
                pagos_validos = factura.pagos.exclude(forma_pago="nota_credito")
                total_pagado = sum([p.monto for p in pagos_validos])
                if total_pagado >= float(factura.monto):
                    factura.estatus = "pagada"
                else:
                    factura.estatus = "pendiente"
                factura.save()
                factura.actualizar_estatus()  
                messages.success(request,f"Cobro registrado Factura:{factura.folio}. Saldo restante: ${factura.saldo_pendiente:.2f}",)
                next_url = request.GET.get('next')
                if next_url:
                     return redirect(next_url)
                return redirect("lista_facturas")
    else:
        form = PagoForm()

    return render(
        request,
        "facturacion/registrar_pago.html",
        {
            "form": form,
            "factura": factura,
            "saldo": factura.saldo_pendiente,
        },
    )

@login_required
def facturas_detalle(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    cobros = factura.pagos.all().order_by('fecha_pago')
    reversados_ids = set()
    for cobro in cobros:
        if cobro.monto < 0 and cobro.observaciones and "Reversa de pago ID" in cobro.observaciones:
            # Extrae el ID del pago original
            try:
                id_original = int(cobro.observaciones.split("Reversa de pago ID")[1].split(".")[0].strip())
                reversados_ids.add(id_original)
            except Exception:
                pass
    return render(request, 'facturacion/facturas_detalle.html', {
        'factura': factura,
        'cobros': cobros,
        'reversados_ids': reversados_ids,
    })

#reversa cobros erroneos
@login_required
def reversa_cobro_erroneo(request, pago_id, factura_id):
    pago = get_object_or_404(Pago, id=pago_id)
    factura = get_object_or_404(Factura, id=factura_id)
    next_url = request.GET.get('next')

    if request.method == "POST":
        form = MotivoReversaCobroForm(request.POST)
        if form.is_valid():
            motivo = form.cleaned_data["motivo"]
            # Registrar un pago negativo
            Pago.objects.create(
                factura=factura,
                monto=-pago.monto,
                forma_pago=pago.forma_pago,
                fecha_pago=pago.fecha_pago,
                observaciones=f"Reversa de pago ID {pago.id}. Motivo: {motivo}",
                registrado_por=request.user,
            )
            factura.actualizar_estatus()
            factura.save()

            messages.success(request, "Pago reversado correctamente.")
            return redirect(next_url or 'facturas_detalle', pk=factura.id)
    else:
        form = MotivoReversaCobroForm()

    return render(request, "facturacion/reversa_cobro_erroneo.html", {
        "form": form,
        "pago": pago,
        "factura": factura,
        "next": next_url,
    })

#registro depositos x identificar
@login_required
def registrar_deposito_por_identificar(request):
    if request.method == 'POST':
        form = PagoPorIdentificarForm(request.POST, request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.registrado_por = request.user
            pago.identificado = False
            pago.factura = None
            pago.empresa= request.user.perfilusuario.empresa
            pago.save()
            return redirect('lista_depositos_por_identificar')
    else:
        form = PagoPorIdentificarForm()
    return render(request, 'facturacion/registrar_deposito_por_identificar.html', {'form': form})    

@login_required
def identificar_deposito(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id, factura__isnull=True, identificado=False)
    if request.method == 'POST':
        factura_id = request.POST.get('factura_id')
        factura = get_object_or_404(Factura, id=factura_id)
        saldo_pendiente = Decimal(str(factura.saldo_pendiente))

        if pago.monto > saldo_pendiente:
            # Asignación parcial: descuenta del depósito y crea un nuevo pago para la factura
            pago.monto -= saldo_pendiente
            pago.observaciones = (pago.observaciones or '') + f'Deposito parcialmente asignado a {factura.folio} el {timezone.now().date()}'
            pago.save()
            Pago.objects.create(
                factura=factura,
                monto=saldo_pendiente,
                fecha_pago=pago.fecha_pago,
                forma_pago=pago.forma_pago,
                observaciones=f'Depósito identificado asignado el {timezone.now().date()}',
                registrado_por=pago.registrado_por,
                identificado=True,
                empresa=pago.empresa,
                #comprobante=pago.comprobante
            )
            messages.success(request, f'Se asignaron ${saldo_pendiente:.2f} a la factura {factura.folio}. El depósito sigue con saldo disponible.')
        else:
            # Asignación total
            pago.factura = factura
            pago.identificado = True
            pago.observaciones = (pago.observaciones or '') + f'Deposito identificado el {timezone.now().date()}'
            pago.save()
            messages.success(request, f'El depósito fue asignado completamente a la factura {factura.folio}.')
        factura.actualizar_estatus()
        return redirect('lista_depositos_por_identificar')
    facturas = Factura.objects.filter(empresa=pago.empresa, estatus='pendiente').order_by('cliente__nombre', 'fecha_vencimiento')
    return render(request, 'facturacion/identificar_deposito.html', {'pago': pago, 'facturas': facturas})


@login_required
def lista_depositos_por_identificar(request):
    empresa= request.user.perfilusuario.empresa
    pagos = Pago.objects.filter(identificado=False, factura__isnull=True, empresa=empresa).order_by('-fecha_pago')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            pagos = pagos.filter(fecha_pago__gte=fecha_i)
        except Exception:
            pass
    if fecha_fin:
        try:
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            pagos = pagos.filter(fecha_pago__lte=fecha_f)
        except Exception:
            pass
    total_depositos = pagos.aggregate(total=Sum('monto'))['total'] or 0        
    return render(request, 'facturacion/lista_depositos_por_identificar.html', 
                  {'pagos': pagos,
                   'fecha_inicio': fecha_inicio,
                   'fecha_fin': fecha_fin,
                    'total_depositos': total_depositos,
                   })


#pagos_por_origen.html
#reporte depositos cuotas
@login_required
def pagos_por_origen(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_cuota= request.GET.get('tipo_cuota')
    
    if request.user.is_superuser:
        pagos = Pago.objects.select_related('factura', 'factura__empresa', 'factura__local', 'factura__area_comun', 'factura__cliente').all().order_by('-fecha_pago')
        empresas = Empresa.objects.all()
        locales = LocalComercial.objects.filter(activo=True).order_by('numero')
        areas = AreaComun.objects.filter(activo=True).order_by('numero')
    else:
        empresa = request.user.perfilusuario.empresa
        pagos = Pago.objects.select_related('factura').filter(factura__empresa=empresa).order_by('-fecha_pago')
        empresas = None
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True).order_by('numero')
        areas = AreaComun.objects.filter(empresa=empresa, activo=True).order_by('numero')

    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id).order_by('fecha_pago')
    if local_id:
        pagos = pagos.filter(factura__local_id=local_id).order_by('fecha_pago')
    if area_id:
        pagos = pagos.filter(factura__area_comun_id=area_id).order_by('fecha_pago')

    if fecha_inicio and fecha_fin:
        try:
            from datetime import datetime
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            pagos = pagos.filter(fecha_pago__range=[fecha_i, fecha_f])
        except Exception:
            pass
    if tipo_cuota:
        pagos = pagos.filter(factura__tipo_cuota=tipo_cuota).order_by('fecha_pago')

    pagos_validos = pagos.exclude(forma_pago='nota_credito')
    total_pagos = pagos_validos.aggregate(total=Sum('monto'))['total'] or 0

    #paginacion
    paginator = Paginator(pagos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'facturacion/pagos_por_origen.html', {
        'pagos': pagos,
        'total_pagos': total_pagos,
        'empresas': empresas,
        'empresa_seleccionada': int(empresa_id) if empresa_id else None,
        'locales': locales,
        'areas': areas,
        'local_id': local_id,
        'area_id': area_id,
        'pagos': page_obj,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo_cuota': tipo_cuota,
        'TIPO_CUOTA_CHOICES': Factura.TIPO_CUOTA_CHOICES,
    })

#saldos.html
#dashboard cartera vencida
@login_required
def dashboard_saldos(request):
    hoy = timezone.now().date()
    cliente_id = request.GET.get('cliente')
    origen = request.GET.get('origen')
    tipo_cuota = request.GET.get('tipo_cuota')
    
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    mes = int(mes) if mes and mes.isdigit() else None
    anio = int(anio) if anio and anio.isdigit() else None
        
    es_super = request.user.is_superuser

    # Empresa según usuario
    if es_super:
        empresas = Empresa.objects.all()
        if not empresas.exists():
            return render(request, 'dashboard/saldos.html', {'empresas': [], 'facturas': []})
        empresa_id = request.GET.get('empresa')
        if not empresa_id or empresa_id == "todas":
            filtro_empresa = Q()
        else:
            filtro_empresa = Q(empresa_id=empresa_id)
    else:
        try:
            empresa = request.user.perfilusuario.empresa
            empresas = Empresa.objects.filter(id=empresa.id)
            empresa_id = empresa.id
            filtro_empresa = Q(empresa_id=empresa_id)
        except Exception:
            messages.error(request, "No tienes una empresa asignada. Contacta al administrador.")
            return render(request, 'dashboard/saldos.html', {'empresas': [], 'facturas': []})   

    # Filtrado base de facturas pendientes
    facturas = Factura.objects.filter(estatus='pendiente').filter(filtro_empresa)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if origen == 'local':
        facturas = facturas.filter(local__isnull=False)
        
    elif origen == 'area':
        facturas = facturas.filter(area_comun__isnull=False)
        
    if tipo_cuota:
        facturas = facturas.filter(tipo_cuota=tipo_cuota)

    if anio:
        facturas = facturas.filter(fecha_vencimiento__year=anio)
    if mes:
        facturas = facturas.filter(fecha_vencimiento__month=mes)


    # Subconsulta: total pagado por factura
    pagos_subquery = Pago.objects.filter(factura=OuterRef('pk')) \
        .values('factura') \
        .annotate(total_pagado_dash=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))) \
        .values('total_pagado_dash')
    facturas = facturas.annotate(
        total_pagado_dash=Coalesce(Subquery(pagos_subquery), Value(0, output_field=DecimalField())),
        saldo_pendiente_dash=ExpressionWrapper(
            F('monto') - Coalesce(Subquery(pagos_subquery), Value(0, output_field=DecimalField())),
            output_field=DecimalField()
        )
    )
    # --- Facturas otros ingresos ---
    facturas_otros = FacturaOtrosIngresos.objects.filter(estatus='pendiente', activo=True)
    if not es_super:
        facturas_otros = facturas_otros.filter(empresa=empresa)
    if empresa_id:
        facturas_otros = facturas_otros.filter(empresa_id=empresa_id)
    if cliente_id:
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)
    if anio:
        facturas_otros = facturas_otros.filter(fecha_vencimiento__year=anio)
    if mes:
        facturas_otros = facturas_otros.filter(fecha_vencimiento__month=mes)

    cobros_subquery = CobroOtrosIngresos.objects.filter(factura=OuterRef('pk')) \
        .values('factura') \
        .annotate(total_cobrado_dash=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))) \
        .values('total_cobrado_dash')
    facturas_otros = facturas_otros.annotate(
        total_cobrado_dash=Coalesce(Subquery(cobros_subquery), Value(0, output_field=DecimalField())),
        saldo_pendiente_dash=ExpressionWrapper(
            F('monto') - Coalesce(Subquery(cobros_subquery), Value(0, output_field=DecimalField())),
            output_field=DecimalField()
        )   
    )
    
    saldo_0_30 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0 
    saldo_31_60 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=60), fecha_vencimiento__lte=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0 
    saldo_61_90 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=90), fecha_vencimiento__lte=hoy - timedelta(days=60)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_91_180 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=180), fecha_vencimiento__lte=hoy - timedelta(days=90)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_181_mas = facturas.filter(fecha_vencimiento__lte=hoy - timedelta(days=180)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    
    saldo_0_30_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_31_60_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=60), fecha_vencimiento__lte=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_61_90_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=90), fecha_vencimiento__lte=hoy - timedelta(days=60)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_91_180_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=180), fecha_vencimiento__lte=hoy - timedelta(days=90)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_181_mas_otros = facturas_otros.filter(fecha_vencimiento__lte=hoy - timedelta(days=180)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0


    if origen == "todos":
        saldo_0_30_total = (saldo_0_30 or 0) + (saldo_0_30_otros or 0)
        saldo_31_60_total = (saldo_31_60 or 0) + (saldo_31_60_otros or 0)
        saldo_61_90_total = (saldo_61_90 or 0) + (saldo_61_90_otros or 0)
        saldo_91_180_total = (saldo_91_180 or 0) + (saldo_91_180_otros or 0)
        saldo_181_mas_total = (saldo_181_mas or 0) + (saldo_181_mas_otros or 0)
    else:
        saldo_0_30_total = saldo_0_30 if origen != "otros" else saldo_0_30_otros
        saldo_31_60_total = saldo_31_60 if origen != "otros" else saldo_31_60_otros
        saldo_61_90_total = saldo_61_90 if origen != "otros" else saldo_61_90_otros
        saldo_91_180_total = saldo_91_180 if origen != "otros" else saldo_91_180_otros
        saldo_181_mas_total = saldo_181_mas if origen != "otros" else saldo_181_mas_otros

        # --- Top 10 adeudos por local/área/otros ingresos ---
    if origen == "otros":
        top_adeudos = (
            facturas_otros
            .annotate(
                nombre_otro=Coalesce(
                    F('tipo_ingreso__nombre'),
                    Value('Otro ingreso'),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_otro', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
            .order_by('-total')[:10]
        )
        top_labels = [x['nombre_otro'] for x in top_adeudos]
        top_data = [float(x['total']) for x in top_adeudos]
        top_clientes = [x['nombre_cliente'] for x in top_adeudos]
    elif origen == "todos":
        # Top locales/áreas
        top_local_area = (
            facturas
            .annotate(
                nombre_local_area=Coalesce(
                    F('local__numero'),
                    F('area_comun__numero'),
                    output_field=CharField()
                ),
                tipo_origen=Case(
                    When(local__isnull=False, then=Value('Local')),
                    When(area_comun__isnull=False, then=Value('Área')),
                    default=Value(''),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_local_area', 'tipo_origen', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
        )
        # Top otros ingresos
        top_otros = (
            facturas_otros
            .annotate(
                nombre_otro=Coalesce(
                    F('tipo_ingreso__nombre'),
                    Value('Otro ingreso'),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_otro', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
        )
        # Unir ambos y ordenar
        top_combined = [
            {'label': f"{x['tipo_origen']} {x['nombre_local_area']}".strip(), 'total': float(x['total']), 'cliente': x['nombre_cliente']}
            for x in top_local_area
        ] + [
            {'label': x['nombre_otro'], 'total': float(x['total']), 'cliente': x['nombre_cliente']}
            for x in top_otros
        ]
        top_combined = sorted(top_combined, key=lambda x: x['total'], reverse=True)[:10]
        top_labels = [x['label'] for x in top_combined]
        top_data = [x['total'] for x in top_combined]
        top_clientes = [x['cliente'] for x in top_combined]
    else:
        top_adeudos = (
            facturas
            .annotate(
                nombre_local_area=Coalesce(
                    F('local__numero'),
                    F('area_comun__numero'),
                    output_field=CharField()
                ),
                tipo_origen=Case(
                    When(local__isnull=False, then=Value('Local')),
                    When(area_comun__isnull=False, then=Value('Área')),
                    default=Value(''),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_local_area', 'tipo_origen', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
            .order_by('-total')[:10]
        )
        top_labels = [
            f"{x['tipo_origen']} {x['nombre_local_area']}" if x['nombre_local_area'] else x['tipo_origen']
            for x in top_adeudos
        ]
        top_data = [float(x['total']) for x in top_adeudos]
        top_clientes = [x['nombre_cliente'] for x in top_adeudos]
    
    #clientes = Cliente.objects.filter(empresa__in=empresas)
    if origen == "otros":
        clientes = Cliente.objects.filter(
            id__in=facturas_otros.values_list('cliente_id', flat=True).distinct()
        )
    else:
        clientes = Cliente.objects.filter(empresa__in=empresas)

    # Obtén el rango de años de las facturas y facturas_otros
    min_year = Factura.objects.aggregate(min=Min('fecha_vencimiento'))['min']
    max_year = Factura.objects.aggregate(max=Max('fecha_vencimiento'))['max']
    min_year_otros = FacturaOtrosIngresos.objects.aggregate(min=Min('fecha_vencimiento'))['min']
    max_year_otros = FacturaOtrosIngresos.objects.aggregate(max=Max('fecha_vencimiento'))['max']

    years = set()
    if min_year and max_year:
        years.update(range(min_year.year, max_year.year + 1))
    if min_year_otros and max_year_otros:
        years.update(range(min_year_otros.year, max_year_otros.year + 1))
    anios_facturas = sorted(years)

    tipos_cuota = Factura.objects.values_list('tipo_cuota', flat=True).distinct().order_by('tipo_cuota')

    return render(request, 'dashboard/saldos.html', {
        'facturas': facturas,
        'clientes': clientes,
        'empresas': empresas,
        'empresa_id': str(empresa_id) if empresa_id else "",
        'cliente_id': int(cliente_id) if cliente_id else None,
        'saldo_0_30': saldo_0_30,
        'saldo_31_60': saldo_31_60,
        'saldo_61_90': saldo_61_90,
        'saldo_91_180': saldo_91_180,
        'saldo_181_mas': saldo_181_mas,
        'saldo_0_30_otros': saldo_0_30_otros,
        'saldo_31_60_otros': saldo_31_60_otros,
        'saldo_61_90_otros': saldo_61_90_otros,
        'saldo_91_180_otros': saldo_91_180_otros,
        'saldo_181_mas_otros': saldo_181_mas_otros,
        'origen': origen,
        'es_super': es_super,
        'top_labels': top_labels,
        'top_data': top_data,
        'facturas_otros': facturas_otros,
        'top_clientes': top_clientes,
        'saldo_0_30_total': saldo_0_30_total,
        'saldo_31_60_total': saldo_31_60_total,
        'saldo_61_90_total': saldo_61_90_total,
        'saldo_91_180_total': saldo_91_180_total,
        'saldo_181_mas_total': saldo_181_mas_total,
        'tipos_cuota': tipos_cuota,
        'tipo_cuota_seleccionada': tipo_cuota,
        'meses': range(1,13),
        'anios_facturas': anios_facturas,
        'mes': mes,
        'anio': anio,
    })

#pagos.html
#dashboard cuotas
@login_required
def dashboard_pagos(request):
    anio_actual = datetime.now().year
    anio = request.GET.get('anio')
    anio_seleccionado = request.GET.get('anio', anio_actual)
    if not anio:
        anio = anio_actual
    es_super = request.user.is_superuser

    if es_super:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get('empresa')
        if not empresa_id or empresa_id == "todas":
            filtro_empresa = Q()
            empresa = None
        else:
            filtro_empresa = Q(factura__empresa_id=empresa_id)
            empresa = Empresa.objects.get(pk=empresa_id)
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = Empresa.objects.filter(id=empresa.id)
        empresa_id = empresa.id
        filtro_empresa = Q(factura__empresa_id=empresa_id)

    cliente_id = request.GET.get('cliente')
    origen = request.GET.get('origen')
    anio = request.GET.get('anio')
    mes = request.GET.get('mes')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

   

    filtro = Q(factura__activo=True) & filtro_empresa

    if cliente_id:
        filtro &= Q(factura__cliente_id=cliente_id)
    if origen == 'local':
        filtro &= Q(factura__local__isnull=False)
    elif origen == 'area':
        filtro &= Q(factura__area_comun__isnull=False)

    pagos = Pago.objects.exclude(forma_pago='nota_credito').filter(filtro)

    # Cobros de otros ingresos
    otros_cobros = CobroOtrosIngresos.objects.select_related(
    'factura', 'factura__empresa', 'factura__cliente', 'factura__tipo_ingreso').all()
    #otros_cobros = CobroOtrosIngresos.objects.all()
    if not request.user.is_superuser:
        otros_cobros = otros_cobros.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        otros_cobros = otros_cobros.filter(factura__empresa_id=empresa_id)
    if cliente_id:
        otros_cobros = otros_cobros.filter(factura__cliente_id=cliente_id)
    if anio:
        otros_cobros = otros_cobros.filter(fecha_cobro__year=anio)
    if mes:
        otros_cobros = otros_cobros.filter(fecha_cobro__month=mes)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            otros_cobros = otros_cobros.filter(fecha_cobro__range=[fecha_i, fecha_f])
        except:
            pass

    
                
    # --- AJUSTE CLAVE ---
    if origen == 'local' or origen == 'area':
        otros_cobros = CobroOtrosIngresos.objects.none()
        otros_por_mes = []
        otros_por_anio = []
    else:
        otros_por_mes = otros_cobros.annotate(mes=TruncMonth('fecha_cobro')).values('mes').annotate(total=Sum('monto')).order_by('mes')
        otros_por_anio = otros_cobros.annotate(anio=TruncYear('fecha_cobro')).values('anio').annotate(total=Sum('monto')).order_by('anio')    

    # Filtros de fechas para pagos
    if anio:
        pagos = pagos.filter(fecha_pago__year=anio)
    if mes:
        pagos = pagos.filter(fecha_pago__month=mes)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            pagos = pagos.filter(fecha_pago__range=[fecha_i, fecha_f])
        except:
            pass

    if origen == 'otros':
        pagos = Pago.objects.none() 
        pagos_por_mes = []
        pagos_por_anio = []
    else:
        pagos_por_mes = pagos.annotate(mes=TruncMonth('fecha_pago')).values('mes').annotate(total=Sum('monto')).order_by('mes')
        pagos_por_anio = pagos.annotate(anio=TruncYear('fecha_pago')).values('anio').annotate(total=Sum('monto')).order_by('anio')

    # Otros ingresos por mes y año para los gráficos
    otros_por_mes = otros_cobros.annotate(mes=TruncMonth('fecha_cobro')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    otros_por_anio = otros_cobros.annotate(anio=TruncYear('fecha_cobro')).values('anio').annotate(total=Sum('monto')).order_by('anio')

    clientes = Cliente.objects.filter(empresa__in=empresas)

    # Suma total de pagos y otros ingresos
    total_pagos = pagos.aggregate(total=Sum('monto'))['total'] or 0
    total_otros = otros_cobros.aggregate(total=Sum('monto'))['total'] or 0
    total_general = total_pagos + total_otros

    # Unifica meses de ambos queryset
    meses_cuotas = {p['mes']: p['total'] for p in pagos_por_mes}
    meses_otros = {o['mes']: o['total'] for o in otros_por_mes}
    todos_los_meses = sorted(set(list(meses_cuotas.keys()) + list(meses_otros.keys())))

    # Prepara datos alineados
    labels_meses = [DateFormat(m).format('F Y') for m in todos_los_meses]
    data_cuotas = [meses_cuotas.get(m, 0) for m in todos_los_meses]
    data_otros = [meses_otros.get(m, 0) for m in todos_los_meses]

    # --- PRESUPUESTO DE INGRESOS POR MES ---
    
    presup_qs = PresupuestoIngreso.objects.all()
    if anio:
        presup_qs = presup_qs.filter(anio=anio)
    if empresa:
        presup_qs = presup_qs.filter(empresa=empresa)
    
    presup_dict = {}
    for p in presup_qs:
        key = (p.anio, p.mes)
        presup_dict.setdefault(key, 0)
        if origen == 'local' and p.origen == 'local':
            presup_dict[key] += float(p.monto_presupuestado)
        elif origen == 'area' and p.origen == 'area':
            presup_dict[key] += float(p.monto_presupuestado)
        elif origen == 'otros' and p.origen == 'otros':
            presup_dict[key] += float(p.monto_presupuestado)
        elif origen in (None, '', 'todos', 'Todo', 'Todos'):  # Todos los orígenes
            presup_dict[key] += float(p.monto_presupuestado)
    # Prepara datos de presupuesto alineados
    data_presupuesto = []
    for m in todos_los_meses:
        key = (m.year, m.month)
        data_presupuesto.append(presup_dict.get(key, 0))

        # Obtén todos los años presentes en pagos y presupuesto
    anios_pagos = pagos.values_list('fecha_pago__year', flat=True).distinct()
    anios_otros = otros_cobros.values_list('fecha_cobro__year', flat=True).distinct()
    anios_presupuesto = PresupuestoIngreso.objects.values_list('anio', flat=True).distinct()
    todos_los_anios = sorted(set(list(anios_pagos) + list(anios_otros) + list(anios_presupuesto)))
    #todos_los_anios = sorted(set(anios_pagos + anios_otros + anios_presupuesto))

    # Suma ingresos reales por año
    data_cuotas_anio = { (p['anio'].year if hasattr(p['anio'], 'year') else p['anio']): float(p['total']) for p in pagos_por_anio }
    data_otros_anio = { (o['anio'].year if hasattr(o['anio'], 'year') else o['anio']): float(o['total']) for o in otros_por_anio }

    # Suma presupuesto anual por año
    presup_anual_qs = PresupuestoIngreso.objects.all()
    if empresa:
        presup_anual_qs = presup_anual_qs.filter(empresa=empresa)
    presup_anual_dict = {}
    for p in presup_anual_qs:
        key = (p.anio, p.origen, p.tipo_otro or "")
        presup_anual_dict.setdefault(p.anio, {}).setdefault((p.origen, p.tipo_otro or ""), 0)
        presup_anual_dict[p.anio][(p.origen, p.tipo_otro or "")] += float(p.monto_presupuestado)

    data_presupuesto_anio = []
    for anio_ in todos_los_anios:
        if origen == 'local':
            total_presup = presup_anual_dict.get(anio_, {}).get(('local', ''), 0)
        elif origen == 'area':
            total_presup = presup_anual_dict.get(anio_, {}).get(('area', ''), 0)
        elif origen == 'otros':
            total_presup = sum(
                v for (o, _), v in presup_anual_dict.get(anio_, {}).items() if o == 'otros'
            )
        else:  
            presup_local = presup_anual_dict.get(anio_, {}).get(('local', ''), 0)
            presup_area = presup_anual_dict.get(anio_, {}).get(('area', ''), 0)
            presup_otros = sum(
                v for (o, _), v in presup_anual_dict.get(anio_, {}).items() if o == 'otros'
            )
            total_presup = presup_local + presup_area + presup_otros
        data_presupuesto_anio.append(total_presup)

    labels_anios = [str(a) for a in todos_los_anios]
    data_cuotas_anio_list = [data_cuotas_anio.get(a, 0) for a in todos_los_anios]
    data_otros_anio_list = [data_otros_anio.get(a, 0) for a in todos_los_anios]    

        # --- FACTURAS POR COBRAR POR MES ---
    facturas_pendientes = Factura.objects.filter(estatus='pendiente', activo=True)
    if empresa:
        facturas_pendientes = facturas_pendientes.filter(empresa=empresa)
    if cliente_id:
        facturas_pendientes = facturas_pendientes.filter(cliente_id=cliente_id)
    if anio:
        facturas_pendientes = facturas_pendientes.filter(fecha_emision__year=anio)
    if mes:
        facturas_pendientes = facturas_pendientes.filter(fecha_emision__month=mes)
    if origen == 'local':
        facturas_pendientes = facturas_pendientes.filter(local__isnull=False)
    elif origen == 'area':
        facturas_pendientes = facturas_pendientes.filter(area_comun__isnull=False)
        
        # Anota el total pagado por factura
    facturas_pendientes = facturas_pendientes.annotate(
    pagado=Coalesce(Sum('pagos__monto'), 0.0, output_field=DecimalField())
).annotate(
    saldo_pendiente_db=ExpressionWrapper(
        F('monto') - F('pagado'),
        output_field=DecimalField()
    )
)
        
        # Obtén todos los meses/años presentes
    # meses_facturas = sorted(set(f.fecha_vencimiento.replace(day=1) for f in facturas_pendientes))
    # anios_facturas = sorted(set(f.fecha_vencimiento.year for f in facturas_pendientes))
    
    meses_facturas = sorted(set(
    f.fecha_vencimiento.replace(day=1)
    for f in facturas_pendientes
    if f.fecha_vencimiento is not None
    ))
    anios_facturas = sorted(set(
        f.fecha_vencimiento.year
        for f in facturas_pendientes
        if f.fecha_vencimiento is not None
    ))


        # Por mes: solo facturas emitidas en ese mes, no vencidas
    meses_por_cobrar = {}
    for mes in todos_los_meses:
        total = sum(
            float(f.monto) - float(f.pagado)
            for f in facturas_pendientes
            if f.fecha_vencimiento.year == mes.year and f.fecha_vencimiento.month == mes.month
        )
        meses_por_cobrar[mes] = total
    data_por_cobrar = [meses_por_cobrar.get(m, 0) for m in todos_los_meses]

    # Por año: solo facturas emitidas en ese año
    anios_por_cobrar = {}
    for anio in todos_los_anios:
        total = sum(
            float(f.monto) - float(f.pagado)
            for f in facturas_pendientes
            if f.fecha_vencimiento.year == anio
        )
        anios_por_cobrar[anio] = total
    data_por_cobrar_anio = [anios_por_cobrar.get(a, 0) for a in todos_los_anios]

    otros_tipos_por_mes = [defaultdict(float) for _ in range(len(todos_los_meses))]
    for cobro in otros_cobros:
        for idx, mes in enumerate(todos_los_meses):
            if cobro.fecha_cobro.year == mes.year and cobro.fecha_cobro.month == mes.month:

                tipo = getattr(cobro.factura, 'tipo_ingreso', None) or 'Otro'
              
                if hasattr(cobro.factura, 'get_tipo_ingreso_display'):
                    tipo = cobro.factura.get_tipo_ingreso_display()
                otros_tipos_por_mes[idx][tipo] += float(cobro.monto)
                break

    # Convierte a lista de listas de strings para el tooltip
    otros_tipos_tooltip = [
        [f"{tipo}: ${monto:,.2f}" for tipo, monto in tipos.items()]
        for tipos in otros_tipos_por_mes
]

    return render(request, 'dashboard/pagos.html', {
        'pagos': pagos,
        'empresas': empresas,
        'empresa_id': str(empresa_id) if empresa_id else "",
        'clientes': clientes,
        'cliente_id': int(cliente_id) if cliente_id else None,
        'origen': origen,
        'es_super': es_super,
        'pagos_por_mes': pagos_por_mes,
        'pagos_por_anio': pagos_por_anio,
        'otros_por_mes': otros_por_mes,
        'otros_por_anio': otros_por_anio,
        'anio': anio,
        'mes': mes,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'otros_cobros': otros_cobros,
        'total_pagos': total_pagos,
        'total_otros': total_otros,
        'total_general': total_general,
        'labels_meses': labels_meses,
        'data_cuotas': data_cuotas, 
        'data_otros': data_otros,
        'data_presupuesto': data_presupuesto,
        'data_cuotas_anio': data_cuotas_anio_list,
        'data_otros_anio': data_otros_anio_list,
        'data_presupuesto_anio': data_presupuesto_anio,
        'labels_anios': labels_anios,
        'data_por_cobrar': data_por_cobrar,
        'data_por_cobrar_anio': data_por_cobrar_anio,
        'meses_facturas': meses_facturas,
        'anios_facturas': anios_facturas,
        'otros_tipos_tooltip': json.dumps(otros_tipos_tooltip),
        'anio_actual': anio_actual,
        'anio_seleccionado': anio_seleccionado,
    })


@login_required
def cartera_vencida(request):
    hoy = timezone.now().date()
    filtro = request.GET.get('rango')
    origen = request.GET.get('origen')
    tipo_cuota = request.GET.get('tipo_cuota')  

    # Optimiza con select_related para evitar N+1 queries
    facturas = Factura.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    ).select_related('cliente', 'empresa', 'local', 'area_comun')

    facturas_otros = FacturaOtrosIngresos.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    ).select_related('cliente', 'empresa', 'tipo_ingreso')
   
    # Filtrar por empresa
    if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
        empresa = request.user.perfilusuario.empresa
        facturas = facturas.filter(empresa=empresa)
        facturas_otros = facturas_otros.filter(empresa=empresa)
        clientes = Cliente.objects.filter(empresa=empresa).order_by('nombre')
    else:
        if request.GET.get('empresa'):
            facturas = facturas.filter(empresa_id=request.GET['empresa'])
            facturas_otros = facturas_otros.filter(empresa_id=request.GET['empresa'])
            clientes = Cliente.objects.filter(empresa_id=request.GET['empresa']).order_by('nombre')
        else:
            clientes = Cliente.objects.all().order_by('nombre')
    
    # Filtrar por cliente
    cliente_id = request.GET.get('cliente')
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)

    # Filtrar por tipo de cuota
    if tipo_cuota:
        facturas = facturas.filter(tipo_cuota=tipo_cuota)

    # Filtrar por origen
    if origen == 'local':
        facturas = facturas.filter(local__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'area':
        facturas = facturas.filter(area_comun__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'otros':
        facturas = facturas.none()

    # Agrupar por antigüedad
    rangos = [
        ('0-30 días', 0, 30),
        ('31-60 días', 31, 60),
        ('61-90 días', 61, 90),
        ('91-180 días', 91, 180),
        ('181+ días', 181, 10000),
    ]
    agrupado = []
    gran_total = 0

    for nombre, desde, hasta in rangos:
        grupo_facturas = []
        subtotal = 0
        # Facturas cuotas
        for f in facturas:
            dias = (hoy - f.fecha_vencimiento).days
            if desde <= dias <= hasta:
                f.dias_vencidos = dias
                f.es_otro = False
                grupo_facturas.append(f)
                subtotal += float(getattr(f, 'saldo_pendiente', f.monto))
        # Facturas otros ingresos
        for f in facturas_otros:
            dias = (hoy - f.fecha_vencimiento).days
            if desde <= dias <= hasta:
                f.dias_vencidos = dias
                f.es_otro = True
                grupo_facturas.append(f)
                subtotal += float(getattr(f, 'saldo', f.monto))
        agrupado.append({
            'rango': nombre,
            'facturas': grupo_facturas,
            'subtotal': subtotal,
        })
        gran_total += subtotal

    tipos_cuota = Factura.objects.values_list('tipo_cuota', flat=True).distinct().order_by('tipo_cuota')

    return render(request, 'facturacion/cartera_vencida.html', {
        'agrupado': agrupado,
        'gran_total': gran_total,
        'hoy': hoy,
        'empresas': Empresa.objects.all(),
        'clientes': clientes,
        'rango_seleccionado': filtro,
        'tipo_cuota_seleccionado': tipo_cuota,
        'tipos_cuota': tipos_cuota,
    })

@login_required
def exportar_cartera_excel(request):
    hoy = timezone.now().date()
    filtro = request.GET.get('rango')
    origen = request.GET.get('origen')
    tipo_cuota = request.GET.get('tipo_cuota')
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente')

    facturas = Factura.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    )
    facturas_otros = FacturaOtrosIngresos.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    )

    if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
        empresa = request.user.perfilusuario.empresa
        facturas = facturas.filter(empresa=empresa)
        facturas_otros = facturas_otros.filter(empresa=empresa)
    elif empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
        facturas_otros = facturas_otros.filter(empresa_id=empresa_id)

    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)

    if tipo_cuota:
        facturas = facturas.filter(tipo_cuota=tipo_cuota)

    if origen == 'local':
        facturas = facturas.filter(local__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'area':
        facturas = facturas.filter(area_comun__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'otros':
        facturas = facturas.none()

    # Agrupación por rangos de antigüedad
    rangos = [
        ('0-30 días', 0, 30),
        ('31-60 días', 31, 60),
        ('61-90 días', 61, 90),
        ('91-180 días', 91, 180),
        ('181+ días', 181, 10000),
    ]
    agrupado = []
    gran_total = 0

    for nombre, desde, hasta in rangos:
        grupo_facturas = []
        subtotal = 0
        for f in facturas:
            dias = (hoy - f.fecha_vencimiento).days
            if desde <= dias <= hasta:
                f.dias_vencidos = dias
                grupo_facturas.append(f)
                subtotal += float(getattr(f, 'saldo_pendiente', f.monto))
        for f in facturas_otros:
            dias = (hoy - f.fecha_vencimiento).days
            if desde <= dias <= hasta:
                f.dias_vencidos = dias
                grupo_facturas.append(f)
                subtotal += float(getattr(f, 'saldo', f.monto))
        agrupado.append({
            'rango': nombre,
            'facturas': grupo_facturas,
            'subtotal': subtotal,
        })
        gran_total += subtotal

    # Crear libro y hoja
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera Vencida"

    bold = Font(bold=True)
    ws.append(["Reporte de Cartera Vencida por Antigüedad"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 2
    for grupo in agrupado:
        ws.append([grupo['rango'], f"Subtotal: ${grupo['subtotal']:.2f}"])
        ws[f"A{row}"].font = bold
        ws[f"B{row}"].font = bold
        row += 1
        ws.append([
            'Folio', 'Tipo de Cuota', 'Cliente', 'Origen',
            'Monto', 'Saldo Pendiente', 'Fecha Vencimiento', 'Días Vencidos'
        ])
        for col in "ABCDEFGH":
            ws[f"{col}{row}"].font = bold
        row += 1
        for f in grupo['facturas']:
            if isinstance(f, FacturaOtrosIngresos):
                # Es un "otro ingreso"
                tipo_cuota = f.tipo_ingreso.nombre if f.tipo_ingreso else 'Otro ingreso'
                origen_str = tipo_cuota
                saldo = float(getattr(f, 'saldo', f.monto))
            else:
                # Es una Factura normal
                tipo_cuota = f.tipo_cuota
                if f.local:
                    origen_str = f"Local {f.local.numero}"
                elif f.area_comun:
                    origen_str = f"Área {f.area_comun.numero}"
                else:
                    origen_str = "-"
                saldo = float(getattr(f, 'saldo_pendiente', f.monto))
            ws.append([
                f.folio,
                tipo_cuota,
                f.cliente.nombre,
                #f.empresa.nombre,
                origen_str,
                float(f.monto),
                saldo,
                str(f.fecha_vencimiento),
                f.dias_vencidos
            ])
            row += 1
        # Línea vacía después de cada grupo
        ws.append([])
        row += 1

    # Gran total
    ws.append(["Gran Total", "", "", "", "", "", f"${gran_total:.2f}"])
    ws[f"A{row}"].font = bold
    ws[f"G{row}"].font = bold

    # Ajustar anchos de columna
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 18

    # Respuesta HTTP
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cartera_vencida.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_pagos_excel(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_cuota = request.GET.get('tipo_cuota')

    pagos = Pago.objects.select_related('factura', 'factura__empresa', 'factura__local', 'factura__area_comun', 'factura__cliente').all()
    if not request.user.is_superuser:
        pagos = pagos.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id)
    if local_id:
        pagos = pagos.filter(factura__local_id=local_id)
    if area_id:
        pagos = pagos.filter(factura__area_comun_id=area_id)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            pagos = pagos.filter(fecha_pago__range=[fecha_i, fecha_f])
        except Exception:
            pass
    if tipo_cuota:
        pagos = pagos.filter(factura__tipo_cuota=tipo_cuota)

    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresos"

    # Encabezados
    ws.append([
        'Local/Área','Cliente','Monto Cobro','Tipo Cuota','Forma de Cobro','Folio Factura', 'Empresa',  
        'Fecha Cobro', 'Observaciones'
    ])

    # Contenido
    for pago in pagos:
        factura = pago.factura
        local_area = factura.local.numero if factura.local else factura.area_comun.numero if factura.area_comun else '-'
        ws.append([
            local_area,
            factura.cliente.nombre,
            float(pago.monto),
            factura.get_tipo_cuota_display() if hasattr(factura, 'get_tipo_cuota_display') else factura.tipo_cuota,
            pago.forma_pago,
            factura.folio,
            factura.empresa.nombre,
            pago.fecha_pago,
            pago.observaciones or '',
        ])

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pagos.xlsx'
    wb.save(response)
    return response


def buscar_por_id_o_nombre(modelo, valor, campo='nombre', empresa=None):
    """Busca por ID, si falla busca por nombre (sin acentos, insensible a mayúsculas y espacios)."""
    if not valor:
        return None
    val = unidecode(str(valor)).strip().lower()

    # intento por PK
    try:
        obj = modelo.objects.get(pk=int(val))
        if empresa is not None:
            obj_emp_id = getattr(obj, 'empresa_id', None)
            if obj_emp_id is None and getattr(obj, 'empresa', None):
                obj_emp_id = getattr(obj, 'empresa').id
            empresa_id = empresa.id if hasattr(empresa, 'id') else empresa
            if obj_emp_id != empresa_id:
                raise Exception(f"No se encontró '{valor}' en {modelo.__name__} para la empresa seleccionada")
        return obj
    except ValueError:
        # no era un entero -> seguir buscando por campo
        pass
    except modelo.DoesNotExist:
        # seguir buscando por campo
        pass

    # normalizar y buscar entre todos los registros
    def _norm(v):
        return unidecode(str(v)).strip().lower() if v is not None else ''

    todos = modelo.objects.all()
    candidatos = [obj for obj in todos if _norm(getattr(obj, campo)) == val]

    if empresa is not None:
        # obtener id de empresa buscada
        empresa_id = empresa.id if hasattr(empresa, 'id') else empresa

        def _obj_empresa_id(o):
            eid = getattr(o, 'empresa_id', None)
            if eid is None and getattr(o, 'empresa', None):
                try:
                    return getattr(o, 'empresa').id
                except Exception:
                    return None
            return eid

        candidatos_same = [o for o in candidatos if _obj_empresa_id(o) == empresa_id]

        if len(candidatos_same) == 1:
            return candidatos_same[0]
        if len(candidatos_same) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos_same])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__} para la misma empresa: {conflicto}")
        # no hay coincidencias en la misma empresa:
        if candidatos:
            # existen coincidencias en otras empresas -> no considerarlo conflicto para esta empresa
            return None
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__} para la empresa seleccionada")
    else:
        # comportamiento global (sin empresa)
        if len(candidatos) == 1:
            return candidatos[0]
        if len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")


@login_required
def plantilla_facturas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Facturas"

    # Encabezados (ajusta según tu modelo)
    ws.append([
        'folio', 'condominio', 'cliente', 'Num.local', 'Num.area',  'tipo cuota',
        'monto','fecha emision', 'fecha vencimiento', 'observaciones'
    ])
    # Fila de ejemplo (puedes poner valores ficticios)
    ws.append([
        'FAC001', 'Condominio Torre Reforma AC', 'Juan Pérez', 'L-101', '','mantenimiento', '1500.00', '2025-06-10', '2025-07-10', 'carga inicial'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_facturas.xlsx'
    wb.save(response)
    return response        

@login_required
def editar_factura(request, factura_id):
    factura = get_object_or_404(Factura, pk=factura_id)
    
     # Bloqueo si la factura está pagada
    if factura.estatus == 'pagada':
        messages.warning(request, "Esta factura ya está pagada y no puede ser editada.")
        return redirect('lista_facturas')    

    if request.method == 'POST':
        form = FacturaEditForm(request.POST, instance=factura)
        if form.is_valid():
            factura_original = Factura.objects.get(pk=factura_id)
            factura_modificada = form.save(commit=False)
            # Si la fecha viene vacía, conserva la original
            if not factura_modificada.fecha_vencimiento or str(factura_modificada.fecha_vencimiento).strip() == "":
                factura_modificada.fecha_vencimiento = factura.fecha_vencimiento
            # Comparar y guardar auditoría
            for field in form.changed_data:
                valor_anterior = getattr(factura_original, field)
                valor_nuevo = getattr(factura_modificada, field)
                if str(valor_anterior) != str(valor_nuevo):
                    AuditoriaCambio.objects.create(
                        modelo='factura',
                        objeto_id=factura.pk,
                        campo=field,
                        valor_anterior=valor_anterior,
                        valor_nuevo=valor_nuevo,
                        usuario=request.user,
                    )
            factura_modificada.save()
            messages.success(request, "Factura actualizada correctamente.")
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            
            return redirect('lista_facturas')
    else:
        form = FacturaEditForm(instance=factura)
    return render(request, 'facturacion/editar_factura.html', {
        'form': form,
        'factura': factura,
    })

@login_required
def exportar_lista_facturas_excel(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')
    tipo_cuota = request.GET.get('tipo_cuota')
    query=request.GET.get('q','')
    
    facturas = Factura.objects.all().select_related('cliente', 'empresa', 'local', 'area_comun')
    
    if not request.user.is_superuser:
        facturas = facturas.filter(empresa=request.user.perfilusuario.empresa)  
    
    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
    if local_id:
        facturas = facturas.filter(local_id=local_id)
    if area_id:
        facturas = facturas.filter(area_comun_id=area_id)
    if tipo_cuota:
        facturas = facturas.filter(tipo_cuota=tipo_cuota)
    if query:
        facturas = facturas.filter(
            Q(folio__icontains=query) |
            Q(cliente__nombre__icontains=query) 
        )
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Facturas"
    # Encabezados
    ws.append([
        'Folio', 'Empresa', 'Cliente', "Tipo Cuota",'Local/Área', 'Monto',
        'Saldo', 'Periodo', 'Estatus', 'Observaciones'
    ])
    # Contenido
    for factura in facturas:
        local_area = factura.local.numero if factura.local else (factura.area_comun.numero if factura.area_comun else '-')
        ws.append([
            factura.folio,
            factura.empresa.nombre,
            factura.cliente.nombre,
            factura.get_tipo_cuota_display() if hasattr(factura, 'get_tipo_cuota_display') else factura.tipo_cuota,
            local_area,
            float(factura.monto),
            float(factura.saldo_pendiente),
            factura.fecha_vencimiento,
            factura.estatus,
            factura.observaciones or ''
        ])
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=lista_facturas.xlsx'
    wb.save(response)
    return response

@staff_member_required
def carga_masiva_facturas_cobradas(request):
    if request.method == 'POST':
        form = FacturaCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            COLUMNAS_ESPERADAS = 10  # Cambia según tus columnas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})")
                    continue
                folio, empresa_val, cliente_val, local_val, area_val, tipo_cuota, monto, fecha_emision, fecha_vencimiento, observaciones = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                        continue

                    # Validar folio único por empresa
                    if Factura.objects.filter(folio=str(folio), empresa=empresa).exists():
                        errores.append(f"Fila {i}: El folio '{folio}' ya existe para la empresa '{empresa}'.")
                        continue

                    # Validar local o área
                    local = buscar_por_id_o_nombre(LocalComercial, local_val, campo='numero') if local_val else None
                    area = buscar_por_id_o_nombre(AreaComun, area_val, campo='numero') if area_val else None
                    if local_val and not local:
                        errores.append(f"Fila {i}: No se encontró el local '{local_val}'")
                        continue
                    if area_val and not area:
                        errores.append(f"Fila {i}: No se encontró el área '{area_val}'")
                        continue

                    # Buscar o crear cliente
                    #cliente, _ = Cliente.objects.get_or_create(
                     #   nombre=cliente_val,
                     # 3  empresa=empresa
                    #)

                    # Buscar o crear cliente (manejo de duplicados)
                    clientes = Cliente.objects.filter(nombre=cliente_val, empresa=empresa)
                    #if clientes.count() == 1:
                    if clientes.exists():    
                        cliente = clientes.first()
                    #elif clientes.count() == 0:
                     #   cliente = Cliente.objects.create(nombre=cliente_val, empresa=empresa)
                    else:
                        cliente = Cliente.objects.create(nombre=cliente_val, empresa=empresa)

                    # Validar y convertir monto
                    try:
                        cuota_decimal = Decimal(monto)
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de monto '{monto}' no es válido.")
                        continue

                    factura=Factura.objects.create(
                        folio=str(folio),
                        empresa=empresa,
                        cliente=cliente,
                        local=local,
                        area_comun=area,
                        tipo_cuota=tipo_cuota,
                        monto=cuota_decimal,
                        fecha_emision=fecha_emision,
                        fecha_vencimiento=fecha_vencimiento,
                        observaciones=observaciones or "",
                        estatus='cobrada',  # Establecer estatus como 'cobrada'
                    )
                    Pago.objects.create(
                        factura=Factura.objects.get(folio=str(folio), empresa=empresa),
                        monto=cuota_decimal,
                        forma_pago='transferencia', 
                        fecha_pago=fecha_emision,
                        registrado_por=request.user,
                        observaciones=observaciones or "",
                    )
                    
                    #factura.actualizar_estatus()  # ✅ Correcto

                    print(f"[DEBUG] Factura creada: {folio} para {cliente.nombre} ({empresa.nombre})")
                except ValueError as ve:
                    errores.append(f"Fila {i}: Error de valor - {str(ve)}")     
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}")

            if errores:
                messages.error(request, "Algunas facturas no se cargaron:<br>" + "<br>".join(errores))
            else:
                messages.success(request, "¡Facturas cargadas exitosamente!")
            return redirect('carga_masiva_facturas_cobradas')
    else:
        form = FacturaCargaMasivaForm()
    return render(request, 'facturacion/carga_masiva_facturas_cobradas.html', {'form': form})

#carga masiva cuentas x cobrar (cartera vencida)
@login_required
def carga_masiva_facturas(request):
    if request.method == 'POST':
        form = FacturaCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            errores = []
            exitos = 0

            # Intentar leer encabezado y mapear columnas por nombre (si existe)
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers_map = {}
            if header:
                hdrs = [str(h).strip().lower() if h is not None else "" for h in header]
                for idx, h in enumerate(hdrs):
                    if h in ('folio', 'folio factura', 'número', 'numero'):
                        headers_map['folio'] = idx
                    if h in ('empresa', 'condominio'):
                        headers_map['empresa'] = idx
                    if h in ('cliente', 'cliente nombre', 'nombre'):
                        headers_map['cliente'] = idx
                    if h in ('rfc', 'rfc cliente', 'cliente rfc'):
                        headers_map['rfc'] = idx
                    if h in ('local', 'num.local', 'num local', 'local numero', 'número local'):
                        headers_map['local'] = idx
                    if h in ('area', 'area comun', 'área', 'num.area', 'num area', 'numero area'):
                        headers_map['area'] = idx
                    if h in ('tipo_cuota', 'tipo cuota', 'tipo'):
                        headers_map['tipo_cuota'] = idx
                    if h in ('monto', 'importe', 'cantidad'):
                        headers_map['monto'] = idx
                    if h in ('fecha_emision', 'fecha emision', 'fecha emisión'):
                        headers_map['fecha_emision'] = idx
                    if h in ('fecha_vencimiento', 'fecha vencimiento', 'vencimiento'):
                        headers_map['fecha_vencimiento'] = idx
                    if h in ('observaciones', 'obs', 'comentarios'):
                        headers_map['observaciones'] = idx

            # helper para leer celda con fallback positional si no hay encabezado
            def cell(row, key, pos):
                if key in headers_map:
                    i = headers_map[key]
                    return row[i] if i < len(row) else None
                return row[pos] if pos < len(row) else None

            # Iterar filas a partir de la 2
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # saltar filas vacías
                if not row or all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                    continue

                try:
                    # obtener valores (posicional por defecto)
                    folio = cell(row, 'folio', 0)
                    empresa_val = cell(row, 'empresa', 1)
                    cliente_val = cell(row, 'cliente', 2)
                    rfc_val = cell(row, 'rfc', 10)
                    local_val = cell(row, 'local', 3)
                    area_val = cell(row, 'area', 4)
                    tipo_cuota = cell(row, 'tipo_cuota', 5)
                    monto = cell(row, 'monto', 6)
                    fecha_emision = cell(row, 'fecha_emision', 7)
                    fecha_vencimiento = cell(row, 'fecha_vencimiento', 8)
                    observaciones = cell(row, 'observaciones', 9)

                    # Normalizaciones simples
                    folio = str(folio).strip() if folio is not None else ""
                    cliente_nombre = str(cliente_val).strip() if cliente_val is not None else ""
                    tipo_cuota = str(tipo_cuota).strip() if tipo_cuota is not None else ""
                    observaciones = str(observaciones).strip() if observaciones is not None else ""
                    rfc_val = str(rfc_val).strip().upper() if rfc_val not in (None, "") else None

                    # Determinar empresa
                    if request.user.is_superuser:
                        empresa = buscar_por_id_o_nombre(Empresa, empresa_val) if empresa_val else None
                        if not empresa:
                            raise Exception(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                    else:
                        perfil = getattr(request.user, 'perfilusuario', None)
                        if not perfil or not getattr(perfil, 'empresa', None):
                            raise Exception("No se pudo determinar la empresa del usuario")
                        empresa = perfil.empresa

                    # Validar folio único por empresa (si se proporcionó)
                    if folio:
                        if Factura.objects.filter(folio=str(folio), empresa=empresa).exists():
                            raise Exception(f"El folio '{folio}' ya existe para la empresa '{empresa.nombre}'.")

                    # Validar local/area si se informaron
                    local = buscar_por_id_o_nombre(LocalComercial, local_val, campo='numero', empresa=empresa) if local_val else None
                    area = buscar_por_id_o_nombre(AreaComun, area_val, campo='numero', empresa=empresa) if area_val else None
                    # local = buscar_por_id_o_nombre(LocalComercial, local_val, campo='numero') if local_val else None
                    # area = buscar_por_id_o_nombre(AreaComun, area_val, campo='numero') if area_val else None
                    if local_val and not local:
                        raise Exception(f"No se encontró el local '{local_val}'")
                    if area_val and not area:
                        raise Exception(f"No se encontró el área '{area_val}'")

                    # ===== BÚSQUEDA/CREACIÓN DE CLIENTE - PRIORIDAD RFC =====
                    if not cliente_nombre and not rfc_val:
                        raise Exception("Nombre de cliente vacío y no se proporcionó RFC")

                    cliente = None

                    # 1) Si RFC viene en columna, buscar por RFC primero
                    if rfc_val:
                        cliente = Cliente.objects.filter(empresa=empresa, rfc__iexact=rfc_val).first()
                        if cliente:
                            # actualizar nombre si está vacío
                            if not cliente.nombre and cliente_nombre:
                                cliente.nombre = cliente_nombre
                                cliente.save()
                    # 2) Si no se encontró por RFC, intentar extraer RFC del campo cliente (ej. "Nombre | RFC")
                    if not cliente and isinstance(cliente_val, str):
                        potential_rfc = None
                        if '|' in cliente_val:
                            parts = [p.strip() for p in cliente_val.split('|') if p.strip()]
                            if len(parts) >= 2:
                                cliente_nombre = parts[0]
                                potential_rfc = parts[1]
                        elif ',' in cliente_val:
                            parts = [p.strip() for p in cliente_val.split(',') if p.strip()]
                            if len(parts) >= 2:
                                cliente_nombre = parts[0]
                                potential_rfc = parts[1]
                        if potential_rfc:
                            potential_rfc = potential_rfc.upper()
                            cliente = Cliente.objects.filter(empresa=empresa, rfc__iexact=potential_rfc).first()
                            if cliente and not rfc_val:
                                rfc_val = potential_rfc

                    # 3) Si sigue sin cliente, buscar por nombre (preferir registros con RFC)
                    if not cliente:
                        clientes_qs = Cliente.objects.filter(nombre__iexact=cliente_nombre, empresa=empresa)
                        if clientes_qs.exists():
                            cliente = clientes_qs.filter(rfc__isnull=False).exclude(rfc='').first() or clientes_qs.first()

                    # 4) Si no existe cliente, crear uno nuevo (incluye RFC si lo tenemos)
                    if not cliente:
                        cliente = Cliente.objects.create(
                            empresa=empresa,
                            nombre=cliente_nombre,
                            rfc=rfc_val if rfc_val else None,
                            activo=True,
                        )
                    else:
                        # Si encontramos cliente por nombre pero tenemos RFC en archivo y cliente no tiene RFC, actualizarlo
                        if rfc_val and (not getattr(cliente, 'rfc', None) or cliente.rfc.strip() == ''):
                            cliente.rfc = rfc_val
                            cliente.save()

                    # Validar monto
                    try:
                        cuota_decimal = Decimal(str(monto)) if monto is not None and str(monto).strip() != "" else Decimal('0.00')
                    except (InvalidOperation, TypeError, ValueError):
                        raise Exception(f"El valor de monto '{monto}' no es válido.")

                    # Parsear fechas (acepta date o string)
                    def parse_date(v):
                        if v is None or (isinstance(v, str) and v.strip() == ""):
                            return None
                        if isinstance(v, (datetime, date)):
                            return v.date() if isinstance(v, datetime) else v
                        s = str(v).strip()
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                            try:
                                return datetime.strptime(s, fmt).date()
                            except Exception:
                                pass
                        # intento con dateutil si está disponible
                        try:
                            from dateutil.parser import parse as _parse
                            return _parse(s, dayfirst=True).date()
                        except Exception:
                            raise Exception(f"Fecha inválida: '{v}'")

                    fecha_emision_parsed = parse_date(fecha_emision) or timezone.now().date()
                    fecha_vencimiento_parsed = parse_date(fecha_vencimiento)

                    # Crear factura dentro de transacción por fila
                    with transaction.atomic():
                        factura = Factura.objects.create(
                            folio=str(folio) if folio else None,
                            empresa=empresa,
                            cliente=cliente,
                            local=local,
                            area_comun=area,
                            tipo_cuota=tipo_cuota or '',
                            monto=cuota_decimal,
                            fecha_emision=fecha_emision_parsed,
                            fecha_vencimiento=fecha_vencimiento_parsed,
                            observaciones=observaciones or "",
                            estatus='pendiente',
                        )
                    exitos += 1
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e)}")
            # mensajes
            if exitos:
                messages.success(request, f"¡{exitos} facturas cargadas correctamente!")
            if errores:
                from django.utils.safestring import mark_safe
                msg = "<br>".join(errores[:80])
                if len(errores) > 80:
                    msg += f"<br>...y {len(errores)-80} errores más."
                messages.error(request, mark_safe("Algunas facturas no se cargaron:<br>" + msg))
            return redirect('carga_masiva_facturas')
    else:
        form = FacturaCargaMasivaForm()
    return render(request, 'facturacion/carga_masiva_facturas.html', {'form': form})

@login_required
def crear_factura_otros_ingresos(request):
    if request.method == 'POST':
        form = FacturaOtrosIngresosForm(request.POST,request.FILES,user=request.user)
        if form.is_valid():
            factura = form.save(commit=False)
            # Asignar empresa según el cliente seleccionado
            factura.empresa = request.user.perfilusuario.empresa
            # Generar folio único
            count = FacturaOtrosIngresos.objects.filter(fecha_emision__year=now().year).count() + 1
            factura.folio = f"OI-F{count:05d}"
            #validar observaciones
            if factura.observaciones is None:
                factura.observaciones = ""

            factura.save()
            messages.success(request, "Registro Exitoso.")
            return redirect('lista_facturas_otros_ingresos')
    else:
        form = FacturaOtrosIngresosForm(user=request.user)
    try:
        tipos_ingreso = TipoOtroIngreso.objects.filter(empresa=request.user.perfilusuario.empresa)
    except Exception:
        
        tipos_ingreso = []

    return render(request, 'otros_ingresos/crear_factura.html', {'form': form, 'tipos_ingreso': tipos_ingreso})

@login_required
def lista_facturas_otros_ingresos(request):
    empresa_id = request.session.get("empresa_id")
    tipo_ingreso=request.GET.get('tipo_ingreso')
    cliente_id=request.GET.get('cliente_id')
    
    facturas = FacturaOtrosIngresos.objects.select_related('cliente', 'empresa', 'tipo_ingreso').all().order_by('-fecha_emision')
    # Filtrar por empresa si no es superusuario 
    if request.user.is_superuser and empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
    elif not request.user.is_superuser:
        facturas = facturas.filter(empresa=request.user.perfilusuario.empresa)
   
    # Filtros adicionales
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if tipo_ingreso:
        facturas = facturas.filter(tipo_ingreso__nombre=tipo_ingreso)

    # Para el filtro de clientes y tipos en el formulario
    clientes = Cliente.objects.all()
    tipos_ingreso = FacturaOtrosIngresos.objects.values_list('tipo_ingreso__nombre', flat=True).distinct()
    
    # Paginación
    paginator = Paginator(facturas, 25)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'otros_ingresos/lista_facturas.html', {
        'facturas': page_obj,
        'clientes': clientes,
        'tipos_ingreso': tipos_ingreso,
        'cliente_id': cliente_id,
        'tipo_ingreso_seleccionado': tipo_ingreso,
    })

@login_required
def registrar_cobro_otros_ingresos(request, factura_id):
    factura = get_object_or_404(FacturaOtrosIngresos, pk=factura_id)

    if request.method == "POST":
        form = CobroForm(request.POST, request.FILES)
        if form.is_valid():
            cobro = form.save(commit=False)
            cobro.factura = factura
            cobro.registrado_por = request.user

            if cobro.forma_cobro == "nota_credito":
                    cobro.save()
                    factura.estatus = "cancelada"
                    factura.monto = 0
                    factura.save()
                    messages.success(
                        request,
                        "La factura ha sido cancelada por nota de crédito. el saldo pendiente es $0.00",
                    )
                    return redirect("lista_facturas_otros_ingresos")
            
            monto_cobro = Decimal(str(cobro.monto)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            saldo_pendiente = Decimal(str(factura.saldo)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            if monto_cobro > saldo_pendiente:
                    messages.error(
                        request,
                        f"El monto del cobro no puede ser mayor al saldo pendiente de la factura (${saldo_pendiente:.2f}).",
                    )
            else:
                    cobro.monto = monto_cobro
                    cobro.save()
                    total_cobrado = sum([c.monto for c in factura.cobros.all()])
                    if total_cobrado >= float(factura.monto):
                        factura.estatus = "cobrada"
                        factura.save()
                    messages.success(request, "Cobro registrado correctamente.")
                    return redirect("lista_facturas_otros_ingresos")
    else:
        form = CobroForm()

    return render(
        request,
        "otros_ingresos/registrar_cobro.html",
        {
            "form": form,
            "factura": factura,
        },
    )

@login_required
def detalle_factura_otros_ingresos(request, factura_id):
    factura = get_object_or_404(FacturaOtrosIngresos, pk=factura_id)
    cobros = factura.cobros.all().order_by('fecha_cobro')
    reversados_ids = set()
    for cobro in cobros:
        if cobro.monto < 0 and cobro.observaciones and "Reversa de pago ID" in cobro.observaciones:
            # Extrae el ID del pago original
            try:
                id_original = int(cobro.observaciones.split("Reversa de pago ID")[1].split(".")[0].strip())
                reversados_ids.add(id_original)
            except Exception:
                pass
    return render(request, 'otros_ingresos/detalle_factura.html', {
        'factura': factura,
        'cobros': cobros,
        'reversados_ids': reversados_ids,
    })

#reversar cobro otros ingresos
@login_required
def reversa_cobro_erroneo_otros_ingresos(request, pago_id, factura_id):
    cobro = get_object_or_404(CobroOtrosIngresos, id=pago_id)
    factura = get_object_or_404(FacturaOtrosIngresos, id=factura_id)
    next_url = request.GET.get('next')

    if request.method == "POST":
        form = MotivoReversaCobroForm(request.POST)
        if form.is_valid():
            motivo = form.cleaned_data["motivo"]
            # Registrar un pago negativo
            CobroOtrosIngresos.objects.create(
                factura=factura,
                monto=-cobro.monto,
                forma_cobro=cobro.forma_cobro,
                fecha_cobro=cobro.fecha_cobro,
                observaciones=f"Reversa de pago ID {cobro.id}. Motivo: {motivo}",
                registrado_por=request.user,
            )
            factura.actualizar_estatus()
            factura.save()

            messages.success(request, "Cobro cancelado correctamente.")
            return redirect(next_url or 'detalle_factura_otros_ingresos', pk=factura.id)
    else:
        form = MotivoReversaCobroForm()

    return render(request, "otros_ingresos/reversa_cobro_erroneo_oi.html", {
        "form": form,
        "cobro": cobro,
        "factura": factura,
        "next": next_url,
    })    

@login_required
def reporte_cobros_otros_ingresos(request):    
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_ingreso_id = request.GET.get('tipo_ingreso')    

    cobros = CobroOtrosIngresos.objects.select_related('factura', 'factura__empresa', 'factura__cliente', 'factura__tipo_ingreso')

    # Filtros
    if not request.user.is_superuser:
        cobros = cobros.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        cobros = cobros.filter(factura__empresa_id=empresa_id).order_by('-fecha_cobro')
    if cliente_id:
        cobros = cobros.filter(factura__cliente_id=cliente_id).order_by('-fecha_cobro')
    if fecha_inicio and fecha_fin:
        cobros = cobros.filter(fecha_cobro__range=[fecha_inicio, fecha_fin]).order_by('-fecha_cobro')
        
    if tipo_ingreso_id and tipo_ingreso_id.isdigit():
        cobros = cobros.filter(factura__tipo_ingreso_id=tipo_ingreso_id)

    total_cobrado = cobros.aggregate(total=Sum('monto'))['total'] or 0

    # Paginación
    paginator = Paginator(cobros.order_by('-fecha_cobro'), 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    empresas = Empresa.objects.all() if request.user.is_superuser else Empresa.objects.filter(id=request.user.perfilusuario.empresa.id)
    clientes = Cliente.objects.filter(empresa__in=empresas).order_by('nombre')
    tipos_ingreso = TipoOtroIngreso.objects.filter(empresa__in=empresas).order_by('nombre')

    return render(request, 'otros_ingresos/reporte_cobros.html', {
        'cobros': page_obj,
        'empresas': empresas,
        'clientes': clientes,
        'empresa_id': empresa_id,
        'cliente_id': cliente_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_cobrado': total_cobrado,
        'tipo_ingreso_id': tipo_ingreso_id,
        'tipos_ingreso': tipos_ingreso,
    })


@login_required
def exportar_cobros_otros_ingresos_excel(request):
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_ingreso_id = request.GET.get('tipo_ingreso')

    cobros = CobroOtrosIngresos.objects.select_related('factura', 'factura__empresa', 'factura__cliente','factura__tipo_ingreso')

    if not request.user.is_superuser:
        cobros = cobros.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id and empresa_id.isdigit():
        cobros = cobros.filter(factura__empresa_id=empresa_id)
    if cliente_id and cliente_id.isdigit():
        cobros = cobros.filter(factura__cliente_id=cliente_id)

    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            cobros = cobros.filter(fecha_cobro__range=[fecha_i, fecha_f])
        except ValueError:
            pass  # Ignora el filtro si las fechas no son válidas

    if tipo_ingreso_id and tipo_ingreso_id.isdigit():
        cobros = cobros.filter(factura__tipo_ingreso_id=tipo_ingreso_id)


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobros Otros Ingresos"

    headers = [
        "Fecha cobro", "Empresa", "Cliente", "Tipo ingreso", "Monto", "Forma cobro",
        "Factura", "Comprobante", "Observaciones"
    ]
    ws.append(headers)

    for cobro in cobros:
        ws.append([
            cobro.fecha_cobro,
            cobro.factura.empresa.nombre,
            cobro.factura.cliente.nombre,
            cobro.factura.tipo_ingreso.nombre if cobro.factura.tipo_ingreso else '',
            float(cobro.monto),
            cobro.get_forma_cobro_display(),
            cobro.factura.folio,
            cobro.comprobante.url if cobro.comprobante else '',
            cobro.observaciones or ''
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cobros_otros_ingresos.xlsx"'
    return response

@login_required
def exportar_lista_facturas_otros_ingresos_excel(request):
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente_id')  # <-- usa cliente_id para consistencia
    tipo_ingreso = request.GET.get('tipo_ingreso')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    facturas = FacturaOtrosIngresos.objects.select_related('empresa', 'cliente', 'tipo_ingreso').all()

    if not request.user.is_superuser:
        facturas = facturas.filter(empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if tipo_ingreso:
        facturas = facturas.filter(tipo_ingreso__nombre=tipo_ingreso)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            facturas = facturas.filter(fecha_emision__range=[fecha_i, fecha_f])
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas Otros Ingresos"

    ws.append([
        'Folio', 'Empresa', 'Cliente', 'Tipo ingreso', 'Monto', 'Saldo',
        'Periodo', 'Estatus', 'Observaciones'
    ])

    for factura in facturas:
        ws.append([
            factura.folio,
            factura.empresa.nombre,
            factura.cliente.nombre,
            factura.tipo_ingreso.nombre if factura.tipo_ingreso else '',
            float(factura.monto),
            float(factura.saldo),
            factura.fecha_vencimiento if factura.fecha_vencimiento else '',
            factura.estatus,
            factura.observaciones or ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=lista_facturas_otros_ingresos.xlsx'
    wb.save(response)
    return response


@login_required
def crear_tipo_otro_ingreso(request):
    if request.method == 'POST':
        form = TipoOtroIngresoForm(request.POST)
        if form.is_valid():
            tipo_ingreso = form.save(commit=False)
            tipo_ingreso.empresa = request.user.perfilusuario.empresa
            tipo_ingreso.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'id': tipo_ingreso.id, 'nombre': tipo_ingreso.nombre})
            messages.success(request, "Tipo de ingreso creado correctamente.")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Error al crear el tipo de ingreso.'})
            messages.error(request, "Error al crear el tipo de ingreso.")
    return redirect(request.META.get('HTTP_REFERER', 'crear_factura_otros_ingresos'))


@login_required
def tipos_otro_ingreso_json(request):
    tipos = TipoOtroIngreso.objects.filter(empresa=request.user.perfilusuario.empresa)
    data = [{'id': t.id, 'nombre': t.nombre} for t in tipos]
    return JsonResponse({'tipos': data})


@login_required
def recibo_factura(request, factura_id):
    factura = get_object_or_404(Factura, pk=factura_id)
    cliente = factura.cliente
    empresa = factura.empresa
    return render(
        request,
        "facturacion/recibo_factura.html",
        {
            "factura": factura,
            "cliente": cliente,
            "empresa": empresa,
        },
    )

@login_required
def recibo_pago(request, pago_id):
    pago = get_object_or_404(Pago, pk=pago_id)
    factura = pago.factura
    cliente = factura.cliente
    empresa = factura.empresa
    return render(
        request,
        "facturacion/recibo_pago.html",
        {
            "pago": pago,
            "factura": factura,
            "cliente": cliente,
            "empresa": empresa,
        },
    )

@login_required
def recibo_factura_otras_cuotas(request, factura_id):
    factura = get_object_or_404(FacturaOtrosIngresos, pk=factura_id)
    cliente = factura.cliente
    empresa = factura.empresa
    return render(
        request,
        "otros_ingresos/recibo_factura_otras_cuotas.html",
        {
            "factura": factura,
            "cliente": cliente,
            "empresa": empresa,
        },
    )

@login_required
def recibo_pago_otras_cuotas(request, pago_id):
    pago = get_object_or_404(CobroOtrosIngresos, pk=pago_id)
    factura = pago.factura
    cliente = factura.cliente
    empresa = factura.empresa
    return render(
        request,
        "otros_ingresos/recibo_pago_otras_cuotas.html",
        {
            "pago": pago,
            "factura": factura,
            "cliente": cliente,
            "empresa": empresa,
        },
    )


@login_required
def consulta_facturas(request):
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')

    # Obtén los locales y áreas asignados al visitante
    locales = LocalComercial.objects.filter(visitantes=request.user, activo=True).order_by('numero')
    areas = AreaComun.objects.filter(visitantes=request.user, activo=True).order_by('numero')

    empresa = None
    if locales.exists():
        empresa = locales.first().empresa
        facturas = Factura.objects.filter(local__in=locales, empresa=empresa).order_by('-fecha_vencimiento')
    elif areas.exists():
        empresa = areas.first().empresa
        facturas = Factura.objects.filter(area_comun__in=areas, empresa=empresa).order_by('-fecha_vencimiento')
    else:
        facturas = Factura.objects.none()

    # Filtros adicionales
    if local_id:
        facturas = facturas.filter(local_id=local_id)
    if area_id:
        facturas = facturas.filter(area_comun_id=area_id)

    facturas = facturas.select_related('cliente', 'empresa', 'local', 'area_comun').prefetch_related('pagos')

    total_pendiente = sum(f.saldo_pendiente for f in facturas if f.estatus == 'pendiente')
    total_cobrado = sum(f.monto for f in facturas if f.estatus == 'cobrada')

    # Paginación
    paginator = Paginator(facturas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'facturacion/consulta_facturas.html', {
        'facturas': page_obj,
        'locales': locales,
        'areas': areas,
        'local_id': local_id,
        'area_id': area_id,
        'total_pendiente': total_pendiente,
        'total_cobrado': total_cobrado,
        'empresa': empresa,
    })

@login_required
def exportar_consulta_facturas_excel(request):
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')

    if request.user.is_superuser:
        facturas = Factura.objects.all()
    else:
        empresa = request.user.perfilusuario.empresa
        facturas = Factura.objects.filter(empresa=empresa)

    if local_id:
        facturas = facturas.filter(local_id=local_id)
    if area_id:
        facturas = facturas.filter(area_comun_id=area_id)

    facturas = facturas.select_related('cliente', 'empresa', 'local', 'area_comun')

    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consulta Facturas"
    # Encabezados
    ws.append([
        'Folio', 'Empresa', 'Cliente', 'Local', 'Área común', 'Monto',
        'Saldo', 'Fecha emisión', 'Fecha vencimiento', 'Estatus', 'Observaciones'
    ])
    # Contenido
    for factura in facturas:
        ws.append([
            factura.folio,
            factura.empresa.nombre,
            factura.cliente.nombre,
            factura.local.numero if factura.local else '',
            factura.area_comun.numero if factura.area_comun else '',
            float(factura.monto),
            float(factura.saldo_pendiente),
            factura.fecha_emision,
            factura.fecha_vencimiento,
            factura.estatus,
            factura.observaciones or ''
        ])
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=consulta_facturas.xlsx'
    wb.save(response)
    return response    



