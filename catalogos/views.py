import difflib
from datetime import date
from decimal import Decimal

import openpyxl
from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from clientes.models import Cliente  # ajusta el import real de tu app de clientes
from empresas.models import Empresa
from facturacion.models import Factura, TipoCuotaHomologacion, TipoOtroIngreso
from gastos.forms import CargaMasivaCuentasForm
from gastos.models import (
    CargaCatalogoFila,
    CargaCatalogoSesion,
    CuentaContable,
    GrupoGasto,
    SubgrupoGasto,
    TipoGasto,
)
from locales.models import LocalComercial
from proveedores.models import (
    Proveedor,  # ajusta el import real de tu app de proveedores
)


@login_required
def plantilla_catalogo_cuentas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo de Cuentas"
    ws.append(['codigo', 'nombre_cuenta', 'codigo_padre', 'naturaleza', 'grupo_gasto', 'subgrupo_gasto', 'uso_especial'])
    ws.append(['500', 'Gastos de Operación', '', 'deudora', '', '', ''])
    ws.append(['500-01', 'Mantenimiento y Conservación', '500', 'deudora', '', '', ''])
    ws.append(['500-01-001', 'Mantenimiento de Áreas Comunes', '500-01', 'deudora', 'Mantenimiento', 'Áreas Comunes', ''])
    ws.append(['400', 'Ingresos por Cuotas', '', 'acreedora', '', '', ''])
    ws.append(['400-01', 'Cuotas de Mantenimiento', '400', 'acreedora', '', '', 'cuota_mantenimiento'])
    ws.append(['400-02', 'Cuotas de Áreas Comunes', '400', 'acreedora', '', '', 'cuota_renta'])
    ws.append(['400-03', 'Depósitos en Garantía', '400', 'acreedora', '', '', 'cuota_deposito'])
    ws.append(['400-04', 'Cuotas Extraordinarias', '400', 'acreedora', '', '', 'cuota_extraordinaria'])
    ws.append(['700', 'Otros Ingresos', '', 'acreedora', '', '', ''])
    ws.append(['700-01', 'Intereses', '700', 'acreedora', '', '', 'cuota_intereses'])
    ws.append(['700-02', 'Multas', '700', 'acreedora', '', '', 'cuota_penalidad'])
    ws.append(['700-03', 'Renta de Estacionamiento', '700', 'acreedora', '', '', 'otro_ingreso'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_catalogo_cuentas.xlsx'
    wb.save(response)
    return response


def _empresa_del_usuario(request):
    if request.user.is_superuser:
        empresa_id = request.session.get('empresa_id')
        return Empresa.objects.filter(id=empresa_id).first() if empresa_id else None
    perfil = getattr(request.user, 'perfilusuario', None)
    return perfil.empresa if perfil else None


class CargaMasivaClientesProveedoresForm(forms.Form):
    archivo = forms.FileField(label='Archivo Excel (.xlsx)')

@login_required
def plantilla_clientes_proveedores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes y Proveedores"
    ws.append(['tipo', 'nombre', 'rfc', 'email', 'telefono'])
    ws.append(['cliente', 'Juan Pérez López', 'PELJ800101ABC', 'juan@ejemplo.com', '5512345678'])
    ws.append(['cliente', 'María González', '', '', ''])
    ws.append(['proveedor', 'Suministros del Valle SA de CV', 'SVA050101XYZ', 'contacto@suministros.com', ''])
    ws.append(['proveedor', 'Mantenimientos Integrales', '', '', ''])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_clientes_proveedores.xlsx'
    wb.save(response)
    return response


@login_required
def carga_masiva_clientes_proveedores(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    if request.method == 'POST':
        form = CargaMasivaClientesProveedoresForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            errores = []
            clientes_creados = 0
            clientes_reutilizados = 0
            proveedores_creados = 0
            proveedores_reutilizados = 0

            with transaction.atomic():
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                        continue

                    try:
                        tipo = str(row[0]).strip().lower() if row[0] not in (None, "") else None
                        nombre = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else None
                        rfc = str(row[2]).strip().upper() if len(row) > 2 and row[2] not in (None, "") else None
                        email = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else None
                        telefono = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "") else None

                        if tipo not in ('cliente', 'proveedor'):
                            raise Exception(f"Tipo '{tipo}' inválido. Usa 'cliente' o 'proveedor'.")
                        if not nombre:
                            raise Exception("Nombre vacío.")

                        if tipo == 'cliente':
                            if rfc:
                                cliente, creado = Cliente.objects.get_or_create(
                                    empresa=empresa, rfc=rfc,
                                    defaults={'nombre': nombre, 'email': email or None}
                                )
                            else:
                                cliente = Cliente.objects.filter(
                                    empresa=empresa, rfc__isnull=True, nombre__iexact=nombre
                                ).first()
                                creado = cliente is None
                                if creado:
                                    cliente = Cliente.objects.create(
                                        empresa=empresa, nombre=nombre, email=email or None
                                    )
                            if creado:
                                clientes_creados += 1
                            else:
                                clientes_reutilizados += 1
                                if telefono and not getattr(cliente, 'telefono', None):
                                    cliente.telefono = telefono
                                    cliente.save(update_fields=['telefono'])

                        else:
                            if rfc:
                                proveedor, creado = Proveedor.objects.get_or_create(
                                    empresa=empresa, rfc=rfc,
                                    defaults={'nombre': nombre, 'email': email or None, 'telefono': telefono or None}
                                )
                            else:
                                proveedor = Proveedor.objects.filter(
                                    empresa=empresa, rfc__isnull=True, nombre__iexact=nombre
                                ).first()
                                creado = proveedor is None
                                if creado:
                                    proveedor = Proveedor.objects.create(
                                        empresa=empresa, nombre=nombre, email=email or None, telefono=telefono or None
                                    )
                            if creado:
                                proveedores_creados += 1
                            else:
                                proveedores_reutilizados += 1

                    except Exception as e:
                        errores.append(f"Fila {i}: {str(e)}")

            resumen = (
                f"✅ Clientes: {clientes_creados} nuevos, {clientes_reutilizados} ya existían. "
                f"Proveedores: {proveedores_creados} nuevos, {proveedores_reutilizados} ya existían."
            )
            messages.success(request, resumen)
            if errores:
                from django.utils.safestring import mark_safe
                msg = "<br>".join(errores[:80])
                if len(errores) > 80:
                    msg += f"<br>...y {len(errores)-80} errores más."
                messages.error(request, mark_safe("Algunas filas tuvieron problemas:<br>" + msg))

            return redirect('carga_masiva_clientes_proveedores')
    else:
        form = CargaMasivaClientesProveedoresForm()

    return render(request, 'catalogos/carga_masiva_clientes_proveedores.html', {'form': form})


@login_required
def carga_inicial_completa(request):
    if not request.user.is_superuser:
        messages.error(request, "No tienes permiso para acceder a esta sección.")
        return redirect("dashboard_inicio")

    empresas = Empresa.objects.all().order_by("nombre")
    empresa_id = request.GET.get("empresa") or request.POST.get("empresa")
    empresa = empresas.filter(id=empresa_id).first() if empresa_id else None

    if request.method == "POST":
        if not empresa:
            messages.error(request, "Selecciona la empresa/plaza antes de procesar la carga.")
            return redirect("carga_inicial_completa")
        
        archivo_clientes = request.FILES.get("archivo_clientes")
        archivo_propiedades = request.FILES.get("archivo_propiedades")
        archivo_adeudos = request.FILES.get("archivo_adeudos")

        if not archivo_clientes or not archivo_propiedades:
            messages.error(request, "Los archivos de Clientes y Propiedades son obligatorios. Adeudos es opcional.")
            return redirect(f"{request.path}?empresa={empresa.id}")

        resumen = {
            "clientes_creados": 0, "clientes_reutilizados": 0,
            "propiedades_creadas": 0, "adeudos_creados": 0,
        }
        errores = []

        try:
            with transaction.atomic():
                # ── PASO 1: CLIENTES ──
                clientes_dict = {}  # nombre normalizado -> objeto Cliente
                wb1 = openpyxl.load_workbook(archivo_clientes, data_only=True)
                ws1 = wb1.active
                for i, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                        continue
                    try:
                        nombre = str(row[0]).strip()
                        rfc = str(row[1]).strip().upper() if len(row) > 1 and row[1] else None
                        email = str(row[2]).strip() if len(row) > 2 and row[2] else None
                        telefono = str(row[3]).strip() if len(row) > 3 and row[3] else None

                        if not nombre:
                            raise Exception("Nombre vacío.")

                        if rfc:
                            cliente, creado = Cliente.objects.get_or_create(
                                empresa=empresa, rfc=rfc,
                                defaults={"nombre": nombre, "email": email, "telefono": telefono},
                            )
                        else:
                            cliente = Cliente.objects.filter(
                                empresa=empresa, rfc__isnull=True, nombre__iexact=nombre
                            ).first()
                            creado = cliente is None
                            if creado:
                                cliente = Cliente.objects.create(
                                    empresa=empresa, nombre=nombre, email=email, telefono=telefono,
                                )

                        clientes_dict[nombre.strip().lower()] = cliente
                        if creado:
                            resumen["clientes_creados"] += 1
                        else:
                            resumen["clientes_reutilizados"] += 1

                    except Exception as e:
                        errores.append(f"[Clientes] Fila {i}: {str(e)}")

                # ── PASO 2: PROPIEDADES ──
                locales_dict = {}  # numero normalizado -> objeto LocalComercial
                wb2 = openpyxl.load_workbook(archivo_propiedades, data_only=True)
                ws2 = wb2.active
                for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                        continue
                    try:
                        numero = str(row[0]).strip()
                        cliente_nombre = str(row[1]).strip() if len(row) > 1 and row[1] else None
                        cuota = Decimal(str(row[2])) if len(row) > 2 and row[2] not in (None, "") else Decimal("0")
                        superficie = Decimal(str(row[3])) if len(row) > 3 and row[3] not in (None, "") else None
                        giro = str(row[4]).strip() if len(row) > 4 and row[4] else None

                        if not numero:
                            raise Exception("Número de local vacío.")

                        cliente_obj = None
                        if cliente_nombre:
                            cliente_obj = clientes_dict.get(cliente_nombre.strip().lower())
                            if not cliente_obj:
                                raise Exception(  # noqa: TRY002
                                    f"Cliente '{cliente_nombre}' no encontrado -- revisa que su nombre "
                                    f"coincida EXACTO con el archivo de Clientes."
                                )

                        local, creado = LocalComercial.objects.get_or_create(
                            empresa=empresa, numero=numero,
                            defaults={
                                "cliente": cliente_obj, "cuota": cuota,
                                "superficie_m2": superficie, "giro": giro, "activo": True,
                            },
                        )
                        if not creado:
                            local.cliente = cliente_obj or local.cliente
                            local.cuota = cuota
                            local.superficie_m2 = superficie or local.superficie_m2
                            local.giro = giro or local.giro
                            local.save()

                        locales_dict[numero.strip().lower()] = local
                        if creado:
                            resumen["propiedades_creadas"] += 1

                    except Exception as e:
                        errores.append(f"[Propiedades] Fila {i}: {str(e)}")

                # ── PASO 3: ADEUDOS HISTÓRICOS (opcional) ──
                if archivo_adeudos:
                    wb3 = openpyxl.load_workbook(archivo_adeudos, data_only=True)
                    ws3 = wb3.active

                    prefix = "CM-F"
                    last_folio = Factura.objects.filter(
                        empresa=empresa, folio__startswith=prefix
                    ).order_by("-folio").values_list("folio", flat=True).first()
                    last_num = int(last_folio.replace(prefix, "")) if last_folio else 0

                    for i, row in enumerate(ws3.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                            continue
                        try:
                            numero = str(row[0]).strip()
                            mes = int(row[1])
                            anio = int(row[2])
                            monto = Decimal(str(row[3]))

                            if not (1 <= mes <= 12):
                                raise Exception(f"Mes inválido: {mes}")
                            if monto <= 0:
                                raise Exception("El monto debe ser mayor a 0.")

                            local = locales_dict.get(numero.strip().lower())
                            if not local:
                                raise Exception(  # noqa: TRY002
                                    f"Local '{numero}' no encontrado -- revisa que coincida EXACTO "
                                    f"con el archivo de Propiedades."
                                )

                            if not local.cliente:
                                raise Exception(f"El local '{numero}' no tiene cliente asignado, no se puede generar el adeudo.")

                            fecha_venc = date(anio, mes, 1)
                            last_num += 1
                            Factura.objects.create(
                                empresa=empresa, cliente=local.cliente, local=local,
                                folio=f"{prefix}{last_num:05d}",
                                fecha_emision=fecha_venc, fecha_vencimiento=fecha_venc,
                                monto=monto, tipo_cuota="mantenimiento", estatus="pendiente",
                                observaciones="Saldo inicial -- carga histórica",
                            )
                            resumen["adeudos_creados"] += 1

                        except Exception as e:
                            errores.append(f"[Adeudos] Fila {i}: {str(e)}")

        except Exception as e:
            messages.error(request, f"❌ Error crítico, no se guardó nada: {str(e)}")
            return redirect("carga_inicial_completa")

        msg = (
            f"✅ Clientes: {resumen['clientes_creados']} nuevos, {resumen['clientes_reutilizados']} ya existían. "
            f"Propiedades: {resumen['propiedades_creadas']} creadas. "
            f"Adeudos históricos: {resumen['adeudos_creados']} facturas generadas."
        )
        messages.success(request, msg)

        if errores:
            from django.utils.safestring import mark_safe
            texto_errores = "<br>".join(errores[:100])
            if len(errores) > 100:
                texto_errores += f"<br>...y {len(errores)-100} errores más."
            messages.error(request, mark_safe("Algunas filas tuvieron problemas:<br>" + texto_errores))

        return redirect(f"{request.path}?empresa={empresa.id}")

    return render(request, "catalogos/carga_inicial_completa.html", {"empresa": empresa, "empresas": empresas})


@login_required
def plantilla_clientes_carga_inicial_excel(request):
    if not request.user.is_superuser:
        messages.error(request, "No tienes permiso para acceder a esta sección.")
        return redirect("dashboard_inicio")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["nombre", "rfc", "email", "telefono"])
    ws.append(["Tiendas Soriana", "TSO123456AB1", "contacto@soriana.com", "5512345678"])
    ws.append(["Público en General", "", "", ""])
    for col, ancho in [("A", 35), ("B", 16), ("C", 28), ("D", 16)]:
        ws.column_dimensions[col].width = ancho
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=plantilla_1_clientes.xlsx"
    wb.save(response)
    return response


@login_required
def plantilla_propiedades_carga_inicial_excel(request):
    if not request.user.is_superuser:
        messages.error(request, "No tienes permiso para acceder a esta sección.")
        return redirect("dashboard_inicio")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propiedades"
    ws.append(["numero_local", "cliente_nombre", "cuota", "superficie_m2", "giro"])
    ws.append(["A-101", "Tiendas Soriana", 15000.00, 250.5, "Supermercado"])
    ws.append(["A-102", "", 8000.00, 120.0, ""])
    for col, ancho in [("A", 14), ("B", 35), ("C", 12), ("D", 14), ("E", 20)]:
        ws.column_dimensions[col].width = ancho
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=plantilla_2_propiedades.xlsx"
    wb.save(response)
    return response


@login_required
def plantilla_adeudos_carga_inicial_excel(request):
    if not request.user.is_superuser:
        messages.error(request, "No tienes permiso para acceder a esta sección.")
        return redirect("dashboard_inicio")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adeudos Historicos"
    ws.append(["numero_local", "mes", "anio", "monto"])
    ws.append(["A-101", 3, 2026, 15000.00])
    ws.append(["A-101", 4, 2026, 15000.00])
    ws.append(["A-101", 5, 2026, 15000.00])
    for col, ancho in [("A", 14), ("B", 8), ("C", 8), ("D", 12)]:
        ws.column_dimensions[col].width = ancho
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=plantilla_3_adeudos_historicos.xlsx"
    wb.save(response)
    return response



#####catalogos contables gastos e ingresos

def _mejor_coincidencia_tipo_gasto(nombre_cuenta, tipos_existentes):
    """Compara el nombre de la cuenta contra los tipos de gasto ya
    existentes de la empresa, y regresa (tipo_gasto, porcentaje) del
    más parecido, o (None, 0) si ninguno se acerca lo suficiente."""
    mejor_tipo = None
    mejor_pct = 0
    nombre_norm = nombre_cuenta.strip().lower()

    for tipo in tipos_existentes:
        ratio = difflib.SequenceMatcher(None, nombre_norm, tipo.nombre.strip().lower()).ratio()
        pct = round(ratio * 100)
        if pct > mejor_pct:
            mejor_pct = pct
            mejor_tipo = tipo

    if mejor_pct >= 70:  # umbral -- ajustable
        return mejor_tipo, mejor_pct
    return None, 0


@login_required
def carga_masiva_catalogo_cuentas(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    if request.method == 'POST':
        form = CargaMasivaCuentasForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            tipos_existentes = list(TipoGasto.objects.filter(empresa=empresa).select_related('subgrupo'))

            sesion = CargaCatalogoSesion.objects.create(empresa=empresa, registrado_por=request.user)
            errores = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                    continue

                try:
                    codigo = str(row[0]).strip() if row[0] not in (None, "") else None
                    nombre_cuenta = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else None
                    codigo_padre = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else None
                    naturaleza = str(row[3]).strip().lower() if len(row) > 3 and row[3] not in (None, "") else 'deudora'
                    grupo_nombre = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "") else None
                    subgrupo_nombre = str(row[5]).strip() if len(row) > 5 and row[5] not in (None, "") else None
                    uso_especial = str(row[6]).strip().lower() if len(row) > 6 and row[6] not in (None, "") else None

                    if naturaleza not in ('deudora', 'acreedora'):
                        naturaleza = 'deudora'

                    if not codigo or not nombre_cuenta:
                        raise Exception("Código o nombre de cuenta vacío.")

                    tipo_sugerido, pct = (None, 0)
                    if grupo_nombre and subgrupo_nombre:
                        tipo_sugerido, pct = _mejor_coincidencia_tipo_gasto(nombre_cuenta, tipos_existentes)

                    CargaCatalogoFila.objects.create(
                        sesion=sesion,
                        fila_excel=i,
                        codigo=codigo,
                        nombre_cuenta=nombre_cuenta,
                        codigo_padre=codigo_padre,
                        naturaleza=naturaleza,
                        grupo_nombre=grupo_nombre,
                        subgrupo_nombre=subgrupo_nombre,
                        uso_especial=uso_especial,
                        tipo_gasto_sugerido=tipo_sugerido,
                        similitud_pct=pct,
                    )

                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")

            if errores:
                from django.utils.safestring import mark_safe
                msg = "<br>".join(errores[:50])
                messages.warning(request, mark_safe("Algunas filas se omitieron:<br>" + msg))

            return redirect('revisar_carga_catalogo', sesion_id=sesion.id)
    else:
        form = CargaMasivaCuentasForm()

    return render(request, 'catalogos/carga_masiva_catalogo_cuentas.html', {'form': form})


@login_required
def revisar_carga_catalogo(request, sesion_id):
    empresa = _empresa_del_usuario(request)
    sesion = get_object_or_404(CargaCatalogoSesion, pk=sesion_id, empresa=empresa, estado='pendiente_revision')

    tipos_existentes = TipoGasto.objects.filter(empresa=empresa).select_related('subgrupo', 'subgrupo__grupo').order_by('nombre')

    filas_con_tipo = sesion.filas.exclude(grupo_nombre__isnull=True).exclude(grupo_nombre='')
    filas_solo_cuenta = sesion.filas.filter(grupo_nombre__isnull=True) | sesion.filas.filter(grupo_nombre='')

    return render(request, 'catalogos/revisar_carga_catalogo.html', {
        'sesion': sesion,
        'filas_con_tipo': filas_con_tipo,
        'filas_solo_cuenta': filas_solo_cuenta,
        'tipos_existentes': tipos_existentes,
    })


@login_required
def confirmar_carga_catalogo(request, sesion_id):
    empresa = _empresa_del_usuario(request)
    sesion = get_object_or_404(CargaCatalogoSesion, pk=sesion_id, empresa=empresa, estado='pendiente_revision')

    if request.method != 'POST':
        return redirect('revisar_carga_catalogo', sesion_id=sesion_id)

    cuentas_creadas = 0
    tipos_creados = 0
    tipos_homologados = 0

    with transaction.atomic():
        # --- Primera pasada: crear/actualizar TODAS las cuentas contables ---
        cuentas_por_codigo = {}
        for fila in sesion.filas.all():
            cuenta, _ = CuentaContable.objects.update_or_create(
                empresa=empresa, codigo=fila.codigo,
                defaults={'nombre': fila.nombre_cuenta, 'naturaleza': fila.naturaleza, 'activa': True}
            )
            cuentas_por_codigo[fila.codigo] = cuenta
            cuentas_creadas += 1

        # --- Segunda pasada: jerarquía ---
        for fila in sesion.filas.exclude(codigo_padre__isnull=True).exclude(codigo_padre=''):
            padre = cuentas_por_codigo.get(fila.codigo_padre) or CuentaContable.objects.filter(
                empresa=empresa, codigo=fila.codigo_padre
            ).first()
            if padre:
                cuenta = cuentas_por_codigo[fila.codigo]
                cuenta.cuenta_padre = padre
                cuenta.save(update_fields=['cuenta_padre'])

        # --- Tercera pasada: decisiones de tipo de gasto, según lo elegido en el form ---
        for fila in sesion.filas.exclude(grupo_nombre__isnull=True).exclude(grupo_nombre=''):
            accion = request.POST.get(f'accion_{fila.id}')
            cuenta = cuentas_por_codigo[fila.codigo]

            if accion == 'usar_existente':
                tipo_id = request.POST.get(f'tipo_existente_{fila.id}')
                if tipo_id:
                    tipo = TipoGasto.objects.filter(id=tipo_id, empresa=empresa).first()
                    if tipo and not tipo.cuenta_contable_id:
                        tipo.cuenta_contable = cuenta
                        tipo.save(update_fields=['cuenta_contable'])
                        tipos_homologados += 1

            elif accion == 'crear_nuevo':
                grupo, _ = GrupoGasto.objects.get_or_create(nombre=fila.grupo_nombre)
                subgrupo, _ = SubgrupoGasto.objects.get_or_create(grupo=grupo, nombre=fila.subgrupo_nombre)
                tipo, creado = TipoGasto.objects.get_or_create(
                    empresa=empresa, subgrupo=subgrupo, nombre=fila.nombre_cuenta,
                    defaults={'cuenta_contable': cuenta}
                )
                if creado:
                    tipos_creados += 1
                elif not tipo.cuenta_contable_id:
                    tipo.cuenta_contable = cuenta
                    tipo.save(update_fields=['cuenta_contable'])
                    tipos_homologados += 1

            # accion == 'solo_cuenta' -> no se crea ni homologa ningún TipoGasto

        # --- Cuarta pasada: homologaciones especiales (cuotas y otros ingresos) ---
        TIPOS_CUOTA_VALIDOS = dict(Factura.TIPO_CUOTA_CHOICES).keys()
        otros_ingresos_homologados = 0
        cuotas_homologadas = 0

        for fila in sesion.filas.exclude(uso_especial__isnull=True).exclude(uso_especial=''):
            cuenta = cuentas_por_codigo[fila.codigo]

            if fila.uso_especial.startswith('cuota_'):
                tipo_cuota_valor = fila.uso_especial.replace('cuota_', '', 1)
                if tipo_cuota_valor in TIPOS_CUOTA_VALIDOS:
                    homologacion, _ = TipoCuotaHomologacion.objects.get_or_create(
                        empresa=empresa, tipo_cuota=tipo_cuota_valor
                    )
                    if not homologacion.cuenta_contable_id:
                        homologacion.cuenta_contable = cuenta
                        homologacion.save(update_fields=['cuenta_contable'])
                        cuotas_homologadas += 1

            elif fila.uso_especial == 'otro_ingreso':
                tipo_oi, creado = TipoOtroIngreso.objects.get_or_create(
                    empresa=empresa, nombre=fila.nombre_cuenta,
                    defaults={'cuenta_contable': cuenta}
                )
                if not creado and not tipo_oi.cuenta_contable_id:
                    tipo_oi.cuenta_contable = cuenta
                    tipo_oi.save(update_fields=['cuenta_contable'])
                if creado or not tipo_oi.cuenta_contable_id:
                    otros_ingresos_homologados += 1

        sesion.estado = 'aplicada'
        sesion.save(update_fields=['estado'])

    messages.success(
        request,
        f"✅ {cuentas_creadas} cuentas cargadas, {tipos_creados} tipos de gasto nuevos, "
        f"{tipos_homologados} tipos de gasto homologados, {cuotas_homologadas} tipos de cuota "
        f"homologados, {otros_ingresos_homologados} tipos de otro ingreso homologados."
    )
    return redirect('lista_catalogo_cuentas')


@login_required
def lista_catalogo_cuentas(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    cuentas_qs = list(
        CuentaContable.objects.filter(empresa=empresa)
        .prefetch_related('tipos_gasto')
        .order_by('codigo')
    )
    cuentas_por_id = {c.id: c for c in cuentas_qs}

    def calcular_nivel(cuenta):
        nivel = 0
        actual_id = cuenta.cuenta_padre_id
        visitados = set()
        while actual_id and actual_id not in visitados:
            visitados.add(actual_id)
            nivel += 1
            padre = cuentas_por_id.get(actual_id)
            actual_id = padre.cuenta_padre_id if padre else None
        return nivel

    cuentas = [{'obj': c, 'nivel': calcular_nivel(c)} for c in cuentas_qs]

    return render(request, 'catalogos/lista_catalogo_cuentas.html', {'cuentas': cuentas})

