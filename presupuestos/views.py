
# Create your views here.
from pyexpat.errors import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
import openpyxl
from gastos.models import Gasto, GrupoGasto, PagoGasto, TipoGasto, SubgrupoGasto
from .models import Presupuesto, PresupuestoCierre
from .forms import PresupuestoForm
from django.utils.timezone import now
from django.db.models import Sum,F
from empresas.models import Empresa
from calendar import month_name
from collections import defaultdict
from decimal import Decimal
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth import authenticate
import calendar
from openpyxl.styles import Font, Alignment
from django.db.models.functions import ExtractYear,ExtractMonth
from io import BytesIO
from datetime import datetime


@login_required
def presupuesto_nuevo(request):
    if request.method == 'POST':
        form = PresupuestoForm(request.POST,user=request.user)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            # Solo para usuarios normales, asigna empresa directamente
            if not request.user.is_superuser:
                presupuesto.empresa = request.user.perfilusuario.empresa
            presupuesto.save()
            form.save()
            return redirect('presupuesto_lista')
    else:
        form = PresupuestoForm(user=request.user)
    return render(request, 'presupuestos/form.html', {'form': form})

@login_required
def presupuesto_editar(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    if request.method == 'POST':
        form = PresupuestoForm(request.POST, instance=presupuesto,user=request.user)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            if not request.user.is_superuser:
                presupuesto.empresa = request.user.perfilusuario.empresa
            presupuesto.save()
            form.save()
            return redirect('presupuesto_lista')
    else:
        form = PresupuestoForm(instance=presupuesto,user=request.user)
    return render(request, 'presupuestos/form.html', {'form': form, 'presupuesto': presupuesto})

@login_required
def presupuesto_eliminar(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    if request.method == 'POST':
        presupuesto.delete()
        return redirect('presupuesto_lista')
    return render(request, 'presupuestos/confirmar_eliminar.html', {'presupuesto': presupuesto})

@login_required
def dashboard_presupuestal(request):
    es_super = request.user.is_superuser
    anio_actual = now().year
    anio = int(request.GET.get('anio', anio_actual))
    mes = int(request.GET.get('mes', 0))  # 0 = todo el año

    # Filtro empresa
    if es_super:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get('empresa')
        if empresa_id:
            try:
                empresa = Empresa.objects.get(pk=int(empresa_id))
            except (Empresa.DoesNotExist, ValueError):
                empresa = empresas.first() if empresas else None
        else:
            empresa = empresas.first() if empresas else None
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = Empresa.objects.filter(pk=empresa.id)

    if not empresa:
        # Si no hay empresa, devuelve página vacía o mensaje amigable
        meses_esp = [
                "01 Ene", "02 Feb", "03 Mar", "04 Abr", "05 May", "06 Jun",
                "07 Jul", "08 Ago", "09 Sep", "10 Oct", "11 Nov", "12 Dic"
        ]
        contexto = {
            'empresas': empresas,
            'empresa_id': None,
            'anio': anio,
            'mes': mes,
            'labels': [],
            'datos_presupuesto': [],
            'datos_gastado': [],
            'total_presupuestado': 0,
            'total_gastado': 0,
            'es_super': es_super,
            'meses_esp': meses_esp,
        }
        return render(request, 'presupuestos/dashboard.html', contexto)

    # Presupuestos del año y empresa
    pres_qs = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    if mes:
        pres_qs = pres_qs.filter(mes=mes)

    total_presupuestado = pres_qs.aggregate(total=Sum('monto'))['total'] or 0

    # Gastos reales
    gastos_qs = Gasto.objects.filter(empresa=empresa, fecha__year=anio)
    if mes:
        gastos_qs = gastos_qs.filter(fecha__month=mes)
    total_gastado = gastos_qs.aggregate(total=Sum('monto'))['total'] or 0

    # Datos para gráfico (línea presupuestos vs. gastos)
    meses = list(range(1, 13))
    meses_esp = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    labels = meses_esp
    datos_presupuesto = []
    datos_gastado = []

    for m in meses:
        pres_mes = pres_qs.filter(mes=m).aggregate(total=Sum('monto'))['total'] or 0
        gasto_mes = gastos_qs.filter(fecha__month=m).aggregate(total=Sum('monto'))['total'] or 0
        datos_presupuesto.append(float(pres_mes))
        datos_gastado.append(float(gasto_mes))

    contexto = {
        'empresas': empresas,
        'empresa_id': empresa.id,
        'anio': anio,
        'mes': mes,
        'labels': labels,
        'datos_presupuesto': datos_presupuesto,
        'datos_gastado': datos_gastado,
        'total_presupuestado': total_presupuestado,
        'total_gastado': total_gastado,
        'es_super': es_super,
    }
    return render(request, 'presupuestos/dashboard.html', contexto)

# Matriz de presupuestos por tipo de gasto y mes
@login_required
def matriz_presupuesto(request):
    anio = int(request.GET.get('anio', now().year))
    now_year = now().year
    anios = list(range(now_year, 2021, -1))

    # Empresa y permisos
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get('empresa')
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None

    meses = list(range(1, 13))
    meses_nombres = [month_name[m].capitalize() for m in meses]

    # ¿Cerrado?
    cierre = PresupuestoCierre.objects.filter(empresa=empresa, anio=anio).first()
    bloqueado = cierre.cerrado if cierre else False
    pedir_superuser = False

    # ----------- NUEVO: lógica para abrir presupuesto con superuser -------------
    if bloqueado and not request.user.is_superuser:
        if request.method == "POST" and 'superuser_username' in request.POST:
            username = request.POST.get('superuser_username')
            password = request.POST.get('superuser_password')
            superuser = authenticate(username=username, password=password)
            if superuser and superuser.is_superuser:
                # Reabrir presupuesto
                cierre.cerrado = False
                cierre.fecha_cierre = None
                cierre.cerrado_por = None
                cierre.save()
                messages.success(request, "Presupuesto reabierto correctamente. Ahora puedes editarlo.")
                return redirect(request.path + f"?anio={anio}" + (f"&empresa={empresa.id}" if empresa else ""))
            else:
                pedir_superuser = True
                messages.error(request, "Usuario o contraseña de superusuario incorrectos.")
        else:
            pedir_superuser = True

    # Ahora recalcula bloqueado
    cierre = PresupuestoCierre.objects.filter(empresa=empresa, anio=anio).first()
    bloqueado = cierre.cerrado if cierre else False

    # Edición habilitada si no está bloqueado o si eres superusuario
    edicion_habilitada = not bloqueado or request.user.is_superuser

    # --- Resto de la vista igual ---
    tipos = TipoGasto.objects.select_related('subgrupo', 'subgrupo__grupo').order_by(
        'subgrupo__grupo__nombre', 'subgrupo__nombre', 'nombre'
    )
    grupos_lista = []
    grupos_dict = defaultdict(lambda: defaultdict(list))
    for tipo in tipos:
        grupos_dict[tipo.subgrupo.grupo][tipo.subgrupo].append(tipo)
    for grupo, subgrupos in grupos_dict.items():
        subgrupos_lista = []
        for subgrupo, tipos_ in subgrupos.items():
            tipos_lista = []
            for tipo in tipos_:
                tipos_lista.append({'id': tipo.id, 'nombre': str(tipo), 'tipo_obj': tipo})
            subgrupos_lista.append({'id': subgrupo.id, 'nombre': str(subgrupo), 'tipos': tipos_lista, 'subgrupo_obj': subgrupo})
        grupos_lista.append({'id': grupo.id, 'nombre': str(grupo), 'subgrupos': subgrupos_lista, 'grupo_obj': grupo})

    presupuestos = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    presup_dict = {(p.tipo_gasto_id, p.mes): p for p in presupuestos}

    # Totales, subtotales
    totales_mes = []
    for mes in meses:
        total_mes = 0
        for tipo in tipos:
            p = presup_dict.get((tipo.id, mes))
            total_mes += p.monto if p else 0
        totales_mes.append(total_mes)

    subtotales_grupo = {}
    subtotales_subgrupo = {}
    subtotales_tipo = {}
    for grupo, subgrupos in grupos_dict.items():
        subtotal_grupo = [0] * 12
        for subgrupo, tipos_gasto in subgrupos.items():
            subtotal_subgrupo = [0] * 12
            for tipo in tipos_gasto:
                subtotal_tipo = []
                for i, mes in enumerate(meses):
                    p = presup_dict.get((tipo.id, mes))
                    valor = p.monto if p else 0
                    subtotal_tipo.append(valor)
                    subtotal_subgrupo[i] += valor
                    subtotal_grupo[i] += valor
                subtotales_tipo[tipo.id] = subtotal_tipo
            subtotales_subgrupo[subgrupo.id] = subtotal_subgrupo
        subtotales_grupo[grupo.id] = subtotal_grupo

    # Guardado de presupuestos (solo si habilitado)
    if request.method == "POST" and edicion_habilitada:
        for tipo in tipos:
            for mes in meses:
                key = f"presupuesto_{tipo.id}_{mes}"
                monto = request.POST.get(key)
                if monto is not None:
                    monto = float(monto or 0)
                    grupo = tipo.subgrupo.grupo
                    subgrupo = tipo.subgrupo
                    obj, created = Presupuesto.objects.get_or_create(
                        empresa=empresa, tipo_gasto=tipo, anio=anio, mes=mes,
                        defaults={"monto": monto, "grupo": grupo, "subgrupo": subgrupo}
                    )
                    if not created and obj.monto != monto:
                        obj.monto = monto
                        obj.save()
        # Cerrar presupuesto solo si corresponde
        if "cerrar_presupuesto" in request.POST and edicion_habilitada:
            cierre, created = PresupuestoCierre.objects.get_or_create(empresa=empresa, anio=anio)
            cierre.cerrado = True
            cierre.cerrado_por = request.user
            cierre.fecha_cierre = now()
            cierre.save()
            messages.success(request, "¡Presupuesto cerrado! Solo el superusuario puede volver a abrirlo.")
            return redirect(request.path + f"?anio={anio}" + (f"&empresa={empresa.id}" if empresa else ""))
        else:
            messages.success(request, "Presupuestos actualizados")
        return redirect(request.path + f"?anio={anio}" + (f"&empresa={empresa.id}" if empresa else ""))

    return render(request, "presupuestos/matriz.html", {
        "grupos_lista": grupos_lista,
        "meses": meses,
        "meses_nombres": meses_nombres,
        "presup_dict": presup_dict,
        "anio": anio,
        "anios": anios,
        "empresas": empresas,
        "empresa": empresa,
        "totales_mes": totales_mes,
        "subtotales_grupo": subtotales_grupo,
        "subtotales_subgrupo": subtotales_subgrupo,
        "subtotales_tipo": subtotales_tipo,
        "is_super": request.user.is_superuser,
        "edicion_habilitada": edicion_habilitada,
        "pedir_superuser": pedir_superuser,
        "bloqueado": bloqueado,
        "cierre": cierre,
    })


@login_required
def exportar_presupuesto_excel(request):
    anio = int(request.GET.get('anio', now().year))
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get('empresa')
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa

    meses = list(range(1, 13))
    meses_nombres = [month_name[m][:3].capitalize() for m in meses]

    # Construir estructura de grupos/subgrupos/tipos
    tipos = TipoGasto.objects.select_related('subgrupo', 'subgrupo__grupo').order_by(
        'subgrupo__grupo__nombre', 'subgrupo__nombre', 'nombre'
    )
    grupos_dict = defaultdict(lambda: defaultdict(list))
    for tipo in tipos:
        grupos_dict[tipo.subgrupo.grupo][tipo.subgrupo].append(tipo)
    presupuestos = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    presup_dict = {(p.tipo_gasto_id, p.mes): p for p in presupuestos}

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presupuesto {anio}"

    # Escribir encabezados
    ws.append(["Grupo", "Subgrupo", "Tipo de Gasto"] + meses_nombres + ["Total"])

    for grupo, subgrupos in grupos_dict.items():
        for subgrupo, tipos in subgrupos.items():
            for tipo in tipos:
                fila = [
                    str(grupo), str(subgrupo), str(tipo)
                ]
                total = 0
                for mes in meses:
                    monto = presup_dict.get((tipo.id, mes)).monto if presup_dict.get((tipo.id, mes)) else 0
                    fila.append(float(monto))
                    total += monto
                fila.append(float(total))
                ws.append(fila)

    # Opcional: Ajustar anchos de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Generar respuesta HTTP con el archivo
    nombre_archivo = f"Presupuesto_{empresa.nombre}_{anio}.xlsx".replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
    wb.save(response)
    return response


@login_required
def reporte_presupuesto_vs_gasto(request):
    anio = int(request.GET.get('anio', datetime.now().year))
    meses = list(range(1, 13))
    meses_nombres = [calendar.month_name[m] for m in meses]

    # Empresa y permisos (igual que tu flujo)
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get('empresa')
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None

    tipos = TipoGasto.objects.select_related('subgrupo', 'subgrupo__grupo').order_by(
        'subgrupo__grupo__nombre', 'subgrupo__nombre', 'nombre'
    )

    # Diccionario: {(tipo_id, mes): monto}
    presupuestos = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    presup_dict = {(p.tipo_gasto_id, p.mes): float(p.monto) for p in presupuestos}

    # Diccionario: {(tipo_id, mes): monto}
    gastos = (PagoGasto.objects
        .annotate(
            anio_pago=ExtractYear('fecha_pago'),
            mes_pago=ExtractMonth('fecha_pago'),
            tipo_id=F('gasto__tipo_gasto_id')
        )
        .filter(anio_pago=anio, gasto__empresa=empresa)
        .values('tipo_id', 'mes_pago')
        .annotate(total=Sum('monto'))
    )
    gastos_dict = {(g['tipo_id'], g['mes_pago']): float(g['total']) for g in gastos}

    # Estructura: grupos > subgrupos > tipos > meses
    grupos_dict = defaultdict(lambda: defaultdict(list))
    for tipo in tipos:
        grupos_dict[tipo.subgrupo.grupo][tipo.subgrupo].append(tipo)

    comparativo = []
    # Totales generales anuales
    total_general_presup = 0
    total_general_real = 0
    total_general_var = 0

    for grupo, subgrupos in grupos_dict.items():
        grupo_row = {
            'nombre': str(grupo),
            'subgrupos': [],
            'total': [0]*12,
            'total_gasto': [0]*12,
            'total_var': [0]*12,
            'pct_var': [None]*12,
        }
        grupo_total_presup = 0
        grupo_total_gasto = 0
        grupo_total_var = 0

        for subgrupo, tipos_ in subgrupos.items():
            subgrupo_row = {
                'nombre': str(subgrupo),
                'tipos': [],
                'total': [0]*12,
                'total_gasto': [0]*12,
                'total_var': [0]*12,
                'pct_var': [None]*12,
            }
            subgrupo_total_presup = 0
            subgrupo_total_gasto = 0
            subgrupo_total_var = 0

            for tipo in tipos_:
                row = {'nombre': str(tipo), 'meses': []}
                anual_presup = 0
                anual_gasto = 0
                anual_var = 0

                for i, mes in enumerate(meses):
                    presupuesto = presup_dict.get((tipo.id, mes), 0)
                    gasto = gastos_dict.get((tipo.id, mes), 0)
                    variacion = gasto - presupuesto
                    row['meses'].append({'presupuesto': presupuesto, 'gasto': gasto, 'variacion': variacion})
                    subgrupo_row['total'][i] += presupuesto
                    subgrupo_row['total_gasto'][i] += gasto
                    subgrupo_row['total_var'][i] += variacion
                    grupo_row['total'][i] += presupuesto
                    grupo_row['total_gasto'][i] += gasto
                    grupo_row['total_var'][i] += variacion

                    anual_presup += presupuesto
                    anual_gasto += gasto
                    anual_var += variacion

                # Totales anuales por tipo
                row['total_anual_presup'] = anual_presup
                row['total_anual_gasto'] = anual_gasto
                row['total_anual_var'] = anual_var
                row['total_anual_pct'] = int(round(anual_var / anual_presup * 100)) if anual_presup else ''

                subgrupo_row['tipos'].append(row)

                subgrupo_total_presup += anual_presup
                subgrupo_total_gasto += anual_gasto
                subgrupo_total_var += anual_var

            # Totales mensuales y anuales por subgrupo
            for i in range(12):
                if subgrupo_row['total'][i]:
                    subgrupo_row['pct_var'][i] = round(subgrupo_row['total_var'][i] / subgrupo_row['total'][i] * 100)
                else:
                    subgrupo_row['pct_var'][i] = ''
            subgrupo_row['total_anual_presup'] = subgrupo_total_presup
            subgrupo_row['total_anual_gasto'] = subgrupo_total_gasto
            subgrupo_row['total_anual_var'] = subgrupo_total_var
            subgrupo_row['total_anual_pct'] = int(round(subgrupo_total_var / subgrupo_total_presup * 100)) if subgrupo_total_presup else ''

            grupo_row['subgrupos'].append(subgrupo_row)

            grupo_total_presup += subgrupo_total_presup
            grupo_total_gasto += subgrupo_total_gasto
            grupo_total_var += subgrupo_total_var

        # Totales mensuales y anuales por grupo
        for i in range(12):
            if grupo_row['total'][i]:
                grupo_row['pct_var'][i] = round(grupo_row['total_var'][i] / grupo_row['total'][i] * 100)
            else:
                grupo_row['pct_var'][i] = ''
        grupo_row['total_anual_presup'] = grupo_total_presup
        grupo_row['total_anual_gasto'] = grupo_total_gasto
        grupo_row['total_anual_var'] = grupo_total_var
        grupo_row['total_anual_pct'] = int(round(grupo_total_var / grupo_total_presup * 100)) if grupo_total_presup else ''

        comparativo.append(grupo_row)

        total_general_presup += grupo_total_presup
        total_general_real += grupo_total_gasto
        total_general_var += grupo_total_var

    # Exportar a Excel
    if request.GET.get('excel') == '1':
        return exportar_comparativo_excel(anio, empresa, meses_nombres, comparativo)

    # Calcula los totales generales por mes (presupuesto, real, variación)
    tot_gen_presup = [0] * 12
    tot_gen_real = [0] * 12
    tot_gen_var = [0] * 12
    tot_gen_pct = [None] * 12

    for i in range(12):
        for grupo in comparativo:
            tot_gen_presup[i] += grupo['total'][i]
            tot_gen_real[i] += grupo['total_gasto'][i]
            tot_gen_var[i] += grupo['total_var'][i]
        if tot_gen_presup[i]:
            tot_gen_pct[i] = round(tot_gen_var[i] / tot_gen_presup[i] * 100)
        else:
            tot_gen_pct[i] = None

    # Totales anuales generales
    total_general_pct = int(round(total_general_var / total_general_presup * 100)) if total_general_presup else ''

    return render(request, "presupuestos/reporte_comparativo.html", {
        "anio": anio,
        "anios": list(range(datetime.now().year, 2021, -1)),
        "empresa": empresa,
        "empresas": empresas,
        "meses": meses,
        "meses_nombres": meses_nombres,
        "comparativo": comparativo,
        "tot_gen_presup": tot_gen_presup,
        "tot_gen_real": tot_gen_real,
        "tot_gen_var": tot_gen_var,
        "tot_gen_pct": tot_gen_pct,
        # Totales anuales generales
        "total_general_presup": total_general_presup,
        "total_general_real": total_general_real,
        "total_general_var": total_general_var,
        "total_general_pct": total_general_pct,
    })

def exportar_comparativo_excel(anio, empresa, meses_nombres, comparativo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto vs Gasto"

    # Header
    encabezado = ["Grupo", "Subgrupo", "Tipo de Gasto"]
    for m in meses_nombres:
        encabezado += [m, "Real", "Var", "% Var"]
    encabezado += ["Total Presupuesto", "Total Gasto", "Total Variación", "% Variación"]
    ws.append(encabezado)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def porcentaje_var(presupuesto, variacion):
        try:
            return (variacion / presupuesto) * 100 if presupuesto else 0
        except ZeroDivisionError:
            return 0

    total_presupuesto_general = 0
    total_gasto_general = 0
    total_var_general = 0

    for grupo in comparativo:
        for subgrupo in grupo['subgrupos']:
            for tipo in subgrupo['tipos']:
                fila = [grupo['nombre'], subgrupo['nombre'], tipo['nombre']]
                total_presupuesto = 0
                total_gasto = 0
                total_var = 0
                for mes in tipo['meses']:
                    fila.extend([
                        mes['presupuesto'],
                        mes['gasto'],
                        mes['variacion'],
                        porcentaje_var(mes['presupuesto'], mes['variacion'])
                    ])
                    total_presupuesto += mes['presupuesto']
                    total_gasto += mes['gasto']
                    total_var += mes['variacion']
                fila.extend([
                    total_presupuesto,
                    total_gasto,
                    total_var,
                    porcentaje_var(total_presupuesto, total_var)
                ])
                ws.append(fila)
                total_presupuesto_general += total_presupuesto
                total_gasto_general += total_gasto
                total_var_general += total_var
            # Subtotal subgrupo
            fila = [grupo['nombre'], f"Subtotal {subgrupo['nombre']}", ""]
            total_presupuesto = sum(subgrupo['total'])
            total_gasto = sum(subgrupo['total_gasto'])
            total_var = sum(subgrupo['total_var'])
            for i in range(12):
                fila.extend([
                    subgrupo['total'][i],
                    subgrupo['total_gasto'][i],
                    subgrupo['total_var'][i],
                    porcentaje_var(subgrupo['total'][i], subgrupo['total_var'][i])
                ])
            fila.extend([
                total_presupuesto,
                total_gasto,
                total_var,
                porcentaje_var(total_presupuesto, total_var)
            ])
            ws.append(fila)
        # Subtotal grupo
        fila = [f"TOTAL {grupo['nombre']}", "", ""]
        total_presupuesto = sum(grupo['total'])
        total_gasto = sum(grupo['total_gasto'])
        total_var = sum(grupo['total_var'])
        for i in range(12):
            fila.extend([
                grupo['total'][i],
                grupo['total_gasto'][i],
                grupo['total_var'][i],
                porcentaje_var(grupo['total'][i], grupo['total_var'][i])
            ])
        fila.extend([
            total_presupuesto,
            total_gasto,
            total_var,
            porcentaje_var(total_presupuesto, total_var)
        ])
        ws.append(fila)

    # Total general (al final)
    fila = ["TOTAL GENERAL", "", ""]
    # Suma mensual general (asumiendo que puedes sumar por mes)
    for i in range(12):
        mes_presupuesto = sum(grupo['total'][i] for grupo in comparativo)
        mes_gasto = sum(grupo['total_gasto'][i] for grupo in comparativo)
        mes_var = sum(grupo['total_var'][i] for grupo in comparativo)
        fila.extend([
            mes_presupuesto,
            mes_gasto,
            mes_var,
            porcentaje_var(mes_presupuesto, mes_var)
        ])
    fila.extend([
        total_presupuesto_general,
        total_gasto_general,
        total_var_general,
        porcentaje_var(total_presupuesto_general, total_var_general)
    ])
    ws.append(fila)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Comparativo_{empresa}_{anio}.xlsx'
    return response