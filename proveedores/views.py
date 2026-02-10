from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect,get_object_or_404
from unidecode import unidecode
from empresas.models import Empresa
from .forms import ProveedorForm, ProveedorCargaMasivaForm
from django.contrib import messages
import openpyxl
from .models import Proveedor
from django import forms
from django.utils.safestring import mark_safe
from decimal import Decimal
from datetime import date
from django.core.paginator import Paginator


@login_required
def proveedor_crear(request):
    if request.method == 'POST':
        post = request.POST.copy()
        if not request.user.is_superuser:
            post['empresa'] = request.user.perfilusuario.empresa.id  
        form = ProveedorForm(post, user=request.user)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.empresa = request.user.perfilusuario.empresa
            proveedor.save()
            return redirect('proveedor_lista')
    else:
        form = ProveedorForm(user=request.user)
        if not request.user.is_superuser and 'empresa' in form.fields:
            form.fields['empresa'].initial = request.user.perfilusuario.empresa
    return render(request, 'proveedores/crear_proveedor.html', {'form': form})

@login_required
def proveedor_lista(request):
    proveedor_nombre = request.GET.get('proveedor', '').strip()
    if request.user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        if empresa_id:
            proveedores = Proveedor.objects.filter(empresa_id=empresa_id, activo=True)
        else:
            proveedores = Proveedor.objects.filter(activo=True)
    else:
        empresa = request.user.perfilusuario.empresa
        proveedores = Proveedor.objects.filter(empresa=empresa, activo=True)
    
    if proveedor_nombre:
        proveedores = proveedores.filter(nombre__icontains=proveedor_nombre)
    #paginación
    
    paginator = Paginator(proveedores, 10)  # Mostrar 10 proveedores por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context={
        'proveedores': page_obj,
        'today': date.today(),
        'proveedor_nombre': proveedor_nombre,
    }    
    return render(request, 'proveedores/lista.html', context)

@login_required
def proveedor_editar(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    # Solo el superusuario o usuarios de la empresa pueden editar
    if not request.user.is_superuser and proveedor.empresa != request.user.perfilusuario.empresa:
        return redirect('proveedor_lista')
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('proveedor_lista')
    else:
        form = ProveedorForm(instance=proveedor, user=request.user)
    return render(request, 'proveedores/editar.html', {'form': form, 'proveedor': proveedor})

@login_required
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    # Solo el superusuario o usuarios de la empresa pueden eliminar
    if not request.user.is_superuser and proveedor.empresa != request.user.perfilusuario.empresa:
        return redirect('proveedor_lista')
    if request.method == 'POST':
        proveedor.delete()
        return redirect('proveedor_lista')
    return render(request, 'proveedores/confirmar_eliminar.html', {'proveedor': proveedor})

def buscar_por_id_o_nombre(modelo, valor, campo='nombre'):
    if not valor:
        return None
    val = str(valor).strip().replace(',', '')
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        candidatos = [
            obj for obj in todos
            if unidecode(val).lower() in unidecode(str(getattr(obj, campo))).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")

@login_required
def carga_masiva_proveedores(request):
    if request.method == 'POST':
        form = ProveedorCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            errores = []

            creados = 0
            actualizados = 0
            existentes = 0
            creados_list = []
            actualizados_list = []
            existentes_list = []

            # detectar encabezados (opcional)
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers_map = {}
            if header:
                hdrs = [str(h).strip().lower() if h is not None else "" for h in header]
                for idx, h in enumerate(hdrs):
                    if h in ('nombre', 'razon', 'razon social'):
                        headers_map['nombre'] = idx
                    if h in ('rfc', 'rfc proveedor'):
                        headers_map['rfc'] = idx
                    if h in ('telefono', 'tel'):
                        headers_map['telefono'] = idx
                    if h in ('email', 'correo'):
                        headers_map['email'] = idx
                    if h in ('direccion', 'domicilio', 'direccion fiscal'):
                        headers_map['direccion'] = idx
                    if h in ('empresa', 'condominio'):
                        headers_map['empresa'] = idx

            def cell(row, key, pos):
                if key in headers_map:
                    i = headers_map[key]
                    return row[i] if i < len(row) else None
                return row[pos] if pos < len(row) else None

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                    continue
                try:
                    nombre = cell(row, 'nombre', 0)
                    rfc = cell(row, 'rfc', 1)
                    telefono = cell(row, 'telefono', 2)
                    email = cell(row, 'email', 3)
                    direccion = cell(row, 'direccion', 4)
                    empresa_val = cell(row, 'empresa', 5)

                    nombre = str(nombre).strip() if nombre not in (None, "") else ""
                    rfc_norm = str(rfc).strip().upper() if rfc not in (None, "") else None
                    telefono = str(telefono).strip() if telefono not in (None, "") else None
                    email = str(email).strip() if email not in (None, "") else None
                    direccion = str(direccion).strip() if direccion not in (None, "") else None

                    empresa = None
                    if empresa_val:
                        try:
                            empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                        except Exception:
                            try:
                                empresa = Empresa.objects.get(pk=int(empresa_val))
                            except Exception:
                                empresa = Empresa.objects.filter(nombre__iexact=str(empresa_val).strip()).first()

                    proveedor = None
                    updated = False

                    # SI VIENE RFC: buscar globalmente por RFC para evitar duplicados
                    if rfc_norm:
                        proveedor = Proveedor.objects.filter(rfc__iexact=rfc_norm).first()
                        if proveedor:
                            # proveedor ya existe por RFC: no crear duplicado
                            # actualizar campos vacíos y asignar empresa si está vacío
                            if nombre and (not proveedor.nombre or proveedor.nombre.strip() == ""):
                                proveedor.nombre = nombre; updated = True
                            if email and (not proveedor.email or proveedor.email.strip() == ""):
                                proveedor.email = email; updated = True
                            if telefono and (not proveedor.telefono or proveedor.telefono.strip() == ""):
                                proveedor.telefono = telefono; updated = True
                            if direccion and (not proveedor.direccion or proveedor.direccion.strip() == ""):
                                proveedor.direccion = direccion; updated = True
                            if empresa and getattr(proveedor, 'empresa', None) is None:
                                proveedor.empresa = empresa; updated = True
                            if updated:
                                proveedor.save()
                                actualizados += 1
                                actualizados_list.append(proveedor.rfc or proveedor.nombre)
                            else:
                                existentes += 1
                                existentes_list.append(proveedor.rfc or proveedor.nombre)
                        # si no existe por RFC seguimos para buscar por nombre+empresa y crear después
                    # si no hay RFC o no se encontró por RFC, intentar por nombre+empresa
                    if not proveedor:
                        qs = Proveedor.objects.all()
                        if empresa:
                            qs = qs.filter(empresa=empresa)
                        if nombre:
                            proveedor = qs.filter(nombre__iexact=nombre).first()
                            if proveedor:
                                # si tenemos RFC en fila y proveedor no la tiene, actualizar (no crear duplicado)
                                if rfc_norm and (not getattr(proveedor, 'rfc', None) or proveedor.rfc.strip() == ""):
                                    proveedor.rfc = rfc_norm
                                    proveedor.save()
                                    actualizados += 1
                                    actualizados_list.append(proveedor.rfc or proveedor.nombre)
                                else:
                                    existentes += 1
                                    existentes_list.append(proveedor.rfc or proveedor.nombre)

                    # Si sigue sin proveedor, crear nuevo
                    if not proveedor:
                        # Antes de crear: si rfc_norm está presente, ya hicimos la búsqueda global arriba.
                        proveedor = Proveedor.objects.create(
                            nombre=nombre or (f"Proveedor {rfc_norm}" if rfc_norm else "Proveedor"),
                            rfc=rfc_norm,
                            telefono=telefono,
                            email=email,
                            direccion=direccion,
                            empresa=empresa
                        )
                        creados += 1
                        creados_list.append(proveedor.rfc or proveedor.nombre)

                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e)}<br>{traceback.format_exc()}")

            # mensajes resumen
            if creados:
                messages.success(request, f"¡{creados} proveedores creados!")
                if creados_list:
                    messages.info(request, mark_safe("Creados: " + ", ".join(creados_list[:50]) + (f", ...(+{len(creados_list)-50})" if len(creados_list) > 50 else "")))
            if actualizados:
                messages.success(request, f"{actualizados} proveedores actualizados!")
                if actualizados_list:
                    messages.info(request, mark_safe("Actualizados: " + ", ".join(actualizados_list[:50]) + (f", ...(+{len(actualizados_list)-50})" if len(actualizados_list) > 50 else "")))
            if existentes:
                messages.info(request, f"{existentes} proveedores ya existían (sin cambios).")
                if existentes_list:
                    messages.info(request, mark_safe("Existentes: " + ", ".join(existentes_list[:50]) + (f", ...(+{len(existentes_list)-50})" if len(existentes_list) > 50 else "")))

            if errores:
                msg = "<br>".join(errores[:80])
                if len(errores) > 80:
                    msg += f"<br>...y {len(errores)-80} errores más."
                messages.error(request, mark_safe("Algunos proveedores no se procesaron:<br>" + msg))

            return redirect('carga_masiva_proveedores')
    else:
        form = ProveedorCargaMasivaForm()
    return render(request, 'proveedores/carga_masiva_proveedores.html', {'form': form})

@login_required
def plantilla_proveedores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Proveedores"
    ws.append(['nombre', 'rfc', 'telefono', 'email', 'direccion', 'empresa'])
    ws.append(['Proveedor Ejemplo', 'XAXX010101000', '5551234567', 'prove@ejemplo.com', 'Calle Falsa 123', 'Nombre Empresa'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_proveedores.xlsx'
    wb.save(response)
    return response