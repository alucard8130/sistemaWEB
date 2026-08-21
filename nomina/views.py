
import xml.etree.ElementTree as ET

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone as django_timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from conciliaciones.utils import validar_periodo_abierto
from empleados.models import Empleado
from empresas.models import CuentaBancaria
from gastos.models import Gasto, GrupoGasto, PagoGasto, SubgrupoGasto, TipoGasto
from nomina.models import ConceptoNomina, DispersionNomina, MapeoConceptoNomina, ReciboNomina
from nomina.utils import XMLNominaInvalido, parsear_xml_nomina
from sanitarios.views import _empresa_actual


def validar_periodo_nomina(cuenta, periodo_inicio, periodo_fin, user=None):
    for fecha in (periodo_inicio, periodo_fin):
        if not fecha:
            continue

        permitido, error = validar_periodo_abierto(cuenta, fecha, user=user)
        if not permitido:
            return False, error

    return True, None

# ============================================================
# Vista 1 -- Subir los XML de un periodo de nómina
# ============================================================

@login_required
def nueva_dispersion_nomina(request):
    empresa = _empresa_actual(request)  
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect("dashboard_inicio")

    cuentas = CuentaBancaria.objects.filter(empresa=empresa, activa=True).order_by("banco")

    if request.method == "POST":
        cuenta_id = request.POST.get("cuenta_bancaria_id")
        archivos = request.FILES.getlist("archivos_xml")

        cuenta = CuentaBancaria.objects.filter(id=cuenta_id, empresa=empresa).first()
        if not cuenta:
            messages.error(request, "Selecciona la cuenta bancaria desde la que se dispersó la nómina.")
            return redirect("nueva_dispersion_nomina")

        if not archivos:
            messages.error(request, "Sube al menos un archivo XML.")
            return redirect("nueva_dispersion_nomina")

        with transaction.atomic():
            dispersion = DispersionNomina.objects.create(
                empresa=empresa, cuenta_bancaria=cuenta, registrado_por=request.user,
            )

            fechas_pago, periodos_inicio, periodos_fin = [], [], []

            for archivo in archivos:
                recibo = ReciboNomina(dispersion=dispersion, archivo_xml=archivo)
                datos = None  # NUEVO -- se usa despues del save() para saber si hay conceptos que guardar

                try:
                    contenido = archivo.read()
                    datos = parsear_xml_nomina(contenido)

                    recibo.uuid_fiscal = datos["uuid_fiscal"]
                    recibo.rfc_receptor = datos["rfc_receptor"]
                    recibo.nombre_receptor = datos["nombre_receptor"]
                    recibo.fecha_pago = datos["fecha_pago"]
                    recibo.periodo_inicio = datos["periodo_inicio"]
                    recibo.periodo_fin = datos["periodo_fin"]
                    recibo.total_percepciones = datos["total_percepciones"]
                    recibo.total_deducciones = datos["total_deducciones"]
                    recibo.total_otros_pagos = datos["total_otros_pagos"]
                    recibo.neto_pagado = datos["neto_pagado"]

                    if datos["fecha_pago"]:
                        fechas_pago.append(datos["fecha_pago"])
                    if datos["periodo_inicio"]:
                        periodos_inicio.append(datos["periodo_inicio"])
                    if datos["periodo_fin"]:
                        periodos_fin.append(datos["periodo_fin"])

                    # Ya se habia procesado este folio fiscal antes? (en cualquier dispersion)
                    if datos["uuid_fiscal"] and ReciboNomina.objects.filter(
                        uuid_fiscal=datos["uuid_fiscal"]
                    ).exclude(dispersion=dispersion).exists():
                        recibo.estatus = "duplicado"
                    else:
                        # Cruce automatico por RFC, dentro de la misma empresa
                        empleado = Empleado.objects.filter(
                            empresa=empresa, rfc=datos["rfc_receptor"], activo=True
                        ).first()
                        if empleado:
                            recibo.empleado = empleado
                            recibo.estatus = "ok"
                        else:
                            recibo.estatus = "sin_match"

                    # Validacion del periodo
                    permitido, error = validar_periodo_nomina(
                        cuenta,
                        recibo.periodo_inicio,
                        recibo.periodo_fin,
                        user=request.user
                    )

                    if not permitido:
                        recibo.estatus = "error"
                        recibo.error_detalle = (
                            f"Periodo {recibo.periodo_inicio} a {recibo.periodo_fin} "
                            f"no esta abierto para dispersion: {error}"
                        )

                except XMLNominaInvalido as e:
                    recibo.estatus = "error"
                    recibo.error_detalle = str(e)

                recibo.save()

                # NUEVO -- guarda el detalle de conceptos (percepciones,
                # deducciones, otros pagos), si el XML se pudo leer. Se
                # guarda sin importar el estatus final del recibo (ok,
                # sin_match, duplicado) -- solo se omite si el XML ni
                # siquiera se pudo parsear.
                if datos and datos.get("conceptos"):
                    ConceptoNomina.objects.bulk_create([
                        ConceptoNomina(
                            recibo=recibo,
                            tipo=c["tipo"],
                            clave=c["clave"],
                            concepto=c["concepto"],
                            importe=c["importe"],
                        )
                        for c in datos["conceptos"]
                    ])

            # Autocompleta fecha/periodo de la dispersion con lo que traian los XML
            if fechas_pago:
                dispersion.fecha_pago = max(fechas_pago)
            if periodos_inicio:
                dispersion.periodo_inicio = min(periodos_inicio)
            if periodos_fin:
                dispersion.periodo_fin = max(periodos_fin)
            dispersion.save()

        return redirect("revisar_dispersion_nomina", dispersion_id=dispersion.id)

    return render(request, "nomina/nueva_dispersion.html", {"cuentas": cuentas})


#vista borrar dispersionn nomina
@login_required
def borrar_dispersion_nomina(request, dispersion_id):
    empresa = _empresa_actual(request)
    dispersion = get_object_or_404(DispersionNomina, id=dispersion_id, empresa=empresa)

    if dispersion.estatus != "borrador":
        messages.error(request, "Solo puedes eliminar una dispersión en estado borrador.")
        return redirect("lista_dispersiones_nomina")

    if request.method == "POST":
        dispersion.delete()
        messages.success(request, "Dispersión eliminada correctamente.")
        return redirect("lista_dispersiones_nomina")

    return render(request, "nomina/borrar_dispersion.html", {"dispersion": dispersion})

# ============================================================
# Vista 2 -- Revisar la dispersión antes de confirmar
# (permite asignar manualmente los recibos "sin_match")
# ============================================================


@login_required
def revisar_dispersion_nomina(request, dispersion_id):
    empresa = _empresa_actual(request)
    dispersion = get_object_or_404(DispersionNomina, id=dispersion_id, empresa=empresa)

    if dispersion.estatus == "confirmado":
        messages.info(request, "Esta dispersión ya fue confirmada.")
        return redirect("detalle_dispersion_nomina", dispersion_id=dispersion.id)

    if request.method == "POST":
        for recibo in dispersion.recibos.filter(estatus="sin_match"):
            accion = request.POST.get(f"accion_{recibo.id}")  # "existente" o "nuevo"

            if accion == "nuevo":
                # NUEVO -- crear el empleado directo desde los datos del XML,
                # solo pidiendo lo que el XML no trae (puesto y departamento).
                puesto = request.POST.get(f"puesto_{recibo.id}")
                departamento = request.POST.get(f"departamento_{recibo.id}")

                if not puesto or not departamento:
                    messages.warning(
                        request,
                        f"Falta indicar puesto y/o departamento para crear al empleado "
                        f"de \"{recibo.nombre_receptor}\" -- no se creó.",
                    )
                    continue

                nuevo_empleado = Empleado.objects.create(
                    empresa=empresa,
                    nombre=recibo.nombre_receptor or "Sin nombre (revisar)",
                    rfc=recibo.rfc_receptor or "",
                    puesto=puesto,
                    departamento=departamento,
                )
                recibo.empleado = nuevo_empleado
                recibo.estatus = "ok"
                recibo.save()

            else:
                # Vincular a un empleado YA existente (comportamiento original)
                empleado_id = request.POST.get(f"empleado_{recibo.id}")
                if empleado_id:
                    empleado = Empleado.objects.filter(id=empleado_id, empresa=empresa).first()
                    if empleado:
                        # NUEVO -- si ese empleado no tenía RFC capturado, se lo
                        # completamos con el del XML -- así la próxima nómina
                        # de esta misma persona ya hace match solo, sin volver
                        # a pedir asignación manual.
                        if not empleado.rfc and recibo.rfc_receptor:
                            empleado.rfc = recibo.rfc_receptor
                            empleado.save(update_fields=["rfc"])

                        recibo.empleado = empleado
                        recibo.estatus = "ok"
                        recibo.save()

        return redirect("revisar_dispersion_nomina", dispersion_id=dispersion.id)

    recibos = dispersion.recibos.select_related("empleado").order_by(
        "estatus", "nombre_receptor"
    )
    pendientes_sin_match = recibos.filter(estatus="sin_match").count()
    con_error = recibos.filter(estatus="error").count()
    duplicados = recibos.filter(estatus="duplicado").count()
    listos = recibos.filter(estatus="ok").count()

    empleados_empresa = Empleado.objects.filter(empresa=empresa, activo=True).order_by("nombre")

    return render(request, "nomina/revisar_dispersion.html", {
        "dispersion": dispersion,
        "recibos": recibos,
        "empleados_empresa": empleados_empresa,
        "puestos_choices": Empleado.PUESTOS_CHOICES,
        "departamentos_choices": Empleado.DEPARTAMENTO_CHOICES,
        "pendientes_sin_match": pendientes_sin_match,
        "con_error": con_error,
        "duplicados": duplicados,
        "listos": listos,
        "puede_confirmar": pendientes_sin_match == 0 and listos > 0,
    })


# ============================================================
# Vista 3 -- Confirmar: crea Gasto + PagoGasto por cada recibo OK
# ============================================================

@login_required
def confirmar_dispersion_nomina(request, dispersion_id):
    empresa = _empresa_actual(request)
    dispersion = get_object_or_404(DispersionNomina, id=dispersion_id, empresa=empresa)

    if dispersion.estatus == "confirmado":
        messages.info(request, "Esta dispersión ya fue confirmada.")
        return redirect("detalle_dispersion_nomina", dispersion_id=dispersion.id)

    if request.method != "POST":
        return redirect("revisar_dispersion_nomina", dispersion_id=dispersion.id)

    recibos_pendientes = dispersion.recibos.filter(estatus="sin_match")
    if recibos_pendientes.exists():
        messages.error(
            request,
            f"Todavía hay {recibos_pendientes.count()} recibo(s) sin asignar a un empleado. "
            f"Asígnalos antes de confirmar.",
        )
        return redirect("revisar_dispersion_nomina", dispersion_id=dispersion.id)

    recibos_ok = list(
        dispersion.recibos.filter(estatus="ok").select_related("empleado").prefetch_related("conceptos")
    )

    # ============================================================
    # NUEVO -- valida que TODOS los conceptos de deduccion/otro pago
    # de esta dispersion ya tengan un mapeo a una cuenta contable
    # existente. Si falta alguno, se bloquea la confirmacion -- GESAC
    # nunca adivina ni crea una cuenta nueva para estos.
    # ============================================================
    conceptos_a_mapear = set()
    for recibo in recibos_ok:
        for concepto in recibo.conceptos.all():
            if concepto.tipo in ("deduccion", "otro_pago") and concepto.importe > 0:
                conceptos_a_mapear.add(concepto.concepto)

    mapeos_existentes = set(
        MapeoConceptoNomina.objects.filter(empresa=empresa, concepto_xml__in=conceptos_a_mapear)
        .values_list("concepto_xml", flat=True)
    )
    conceptos_sin_mapear = conceptos_a_mapear - mapeos_existentes

    if conceptos_sin_mapear:
        messages.error(
            request,
            "Hay conceptos de deduccion/otro pago sin mapear a una cuenta contable: "
            + ", ".join(sorted(conceptos_sin_mapear))
            + ". Mapealos en \"Mapeo de Conceptos de Nomina\" antes de confirmar.",
        )
        return redirect("revisar_dispersion_nomina", dispersion_id=dispersion.id)

    # Asegura que exista el Grupo/Subgrupo fijo para nomina -- igual que
    # como ya se registran las solicitudes manuales de sueldo.
    grupo_nomina, _ = GrupoGasto.objects.get_or_create(
        nombre="Gastos Nomina", defaults={"es_exento_iva": True}
    )
    if not grupo_nomina.es_exento_iva:
        grupo_nomina.es_exento_iva = True
        grupo_nomina.save(update_fields=["es_exento_iva"])

    subgrupo_nomina, _ = SubgrupoGasto.objects.get_or_create(
        grupo=grupo_nomina, nombre="Sueldos y salarios"
    )

    creados = 0
    with transaction.atomic():
        for recibo in recibos_ok:
            gasto_sueldo_base = None  # el que queda como "recibo.gasto" al final

            for concepto in recibo.conceptos.all():
                if concepto.importe <= 0:
                    continue  # no genera gasto de $0

                if concepto.tipo == "percepcion":
                    # NUEVO -- el sueldo base (clave SAT "001") SIEMPRE
                    # usa el nombre ya establecido "Sueldo {empleado}"
                    # (singular), sin importar como venga redactado el
                    # concepto en el XML ("Sueldos", "Sueldo", etc.) --
                    # evita duplicar la cuenta que ya corregimos.
                    if concepto.clave == "001":
                        nombre_cuenta = f"Sueldo {recibo.empleado.nombre}"
                    else:
                        nombre_cuenta = f"{concepto.concepto} {recibo.empleado.nombre}"

                    tipo_gasto, _ = TipoGasto.objects.get_or_create(
                        empresa=empresa, subgrupo=subgrupo_nomina, nombre=nombre_cuenta,
                    )
                else:
                    # Deduccion u Otro Pago -- SIEMPRE via el mapeo ya
                    # validado arriba, nunca se crea una cuenta nueva aqui.
                    mapeo = MapeoConceptoNomina.objects.filter(
                        empresa=empresa, concepto_xml=concepto.concepto
                    ).first()
                    if not mapeo:
                        continue  # defensivo -- ya se valido que existe
                    tipo_gasto = mapeo.tipo_gasto

                gasto = Gasto.objects.create(
                    empresa=empresa,
                    empleado=recibo.empleado,
                    tipo_gasto=tipo_gasto,
                    descripcion=(
                        f"Nómina {recibo.periodo_inicio} a {recibo.periodo_fin} — "
                        f"{recibo.empleado.nombre} — {concepto.concepto}"
                    ),
                    fecha=recibo.fecha_pago or dispersion.fecha_pago,
                    monto=concepto.importe,
                    folio_comprobante=recibo.uuid_fiscal,
                    estatus="pendiente",
                    observaciones=f"Dispersión de Nómina ({concepto.get_tipo_display()}).",
                )

                # NUEVO -- solo las PERCEPCIONES y OTROS PAGOS representan
                # salida real de efectivo en este momento (van directo en
                # la transferencia al empleado). Las DEDUCCIONES son
                # retenciones -- el dinero YA esta descontado del neto
                # pagado, pero la empresa todavia no se lo entrega al SAT/
                # IMSS/etc., asi que ese Gasto se queda "pendiente" hasta
                # que se registre ese pago por separado, como cualquier
                # otro gasto normal.
                if concepto.tipo in ("percepcion", "otro_pago"):
                    PagoGasto.objects.create(
                        gasto=gasto,
                        fecha_pago=recibo.fecha_pago or dispersion.fecha_pago,
                        monto=concepto.importe,
                        forma_pago="transferencia",
                        referencia=f"Dispersión de nómina — folio fiscal {recibo.uuid_fiscal or 'N/D'}",
                        registrado_por=request.user,
                        cuenta_bancaria=dispersion.cuenta_bancaria,
                    )
                    gasto.actualizar_estatus()

                concepto.gasto = gasto
                concepto.save(update_fields=["gasto"])
                creados += 1

                if concepto.tipo == "percepcion" and concepto.clave == "001":
                    gasto_sueldo_base = gasto

            # recibo.gasto se queda como referencia rapida al gasto del
            # sueldo base -- el detalle completo vive en recibo.conceptos.
            recibo.gasto = gasto_sueldo_base
            recibo.save(update_fields=["gasto"])

        dispersion.estatus = "confirmado"
        dispersion.fecha_confirmacion = django_timezone.now()
        dispersion.save(update_fields=["estatus", "fecha_confirmacion"])

    messages.success(
        request,
        f"✅ Dispersión confirmada -- {creados} gasto(s) generado(s) (sueldos, deducciones y otros pagos detallados).",
    )
    return redirect("detalle_dispersion_nomina", dispersion_id=dispersion.id)


# ============================================================
# Vista 4 -- Detalle de una dispersión (confirmada o en revisión)
# ============================================================

@login_required
def detalle_dispersion_nomina(request, dispersion_id):
    empresa = _empresa_actual(request)
    dispersion = get_object_or_404(DispersionNomina, id=dispersion_id, empresa=empresa)

    recibos = dispersion.recibos.select_related("empleado", "gasto").order_by("nombre_receptor")

    return render(request, "nomina/detalle_dispersion.html", {
        "dispersion": dispersion,
        "recibos": recibos,
    })


# ============================================================
# Vista 5 -- Lista de todas las dispersiones de la empresa
# ============================================================

@login_required
def lista_dispersiones_nomina(request):
    empresa = _empresa_actual(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect("dashboard_inicio")

    dispersiones = DispersionNomina.objects.filter(empresa=empresa).select_related(
        "cuenta_bancaria", "registrado_por"
    ).order_by("-fecha_creacion")

    # NUEVO -- soporta fecha_inicio/fecha_fin, igual que reporte_caja_chica,
    # para poder llegar aquí filtrado desde el enlace de Estado de Resultados.
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    if fecha_inicio:
        dispersiones = dispersiones.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        dispersiones = dispersiones.filter(fecha_pago__lte=fecha_fin)

    return render(request, "nomina/lista_dispersiones.html", {
        "dispersiones": dispersiones,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    })


# ============================================================
# Vista: exportar_dispersiones_nomina_excel
#
# Genera un Excel con el detalle por empleado de todas las
# dispersiones CONFIRMADAS dentro de un rango de fechas -- pensado
# para que el contador lo baje directo para sus registros.
# ============================================================



def extraer_conceptos_nomina(xml_contenido):
    """
    Devuelve una lista de conceptos:
    [
        {"tipo": "percepcion", "clave": "001", "concepto": "Sueldo", "importe_gravado": 10000, "importe_exento": 0},
        {"tipo": "deduccion", "clave": "003", "concepto": "ISR", "importe_gravado": 1500, "importe_exento": 0},
    ]
    """
    conceptos = []

    if not xml_contenido:
        return conceptos

    try:
        root = ET.fromstring(xml_contenido)
    except Exception:  # noqa: BLE001
        try:
            root = ET.fromstring(xml_contenido.encode("utf-8"))
        except Exception:  # noqa: BLE001
            return conceptos

    ns = {
        "nomina12": "http://www.sat.gob.mx/nomina12",
    }

    percepciones = root.findall(".//nomina12:Percepciones/nomina12:Percepcion", ns)
    for p in percepciones:
        conceptos.append({
            "tipo": "percepcion",
            "clave": p.get("Clave", ""),
            "concepto": p.get("Concepto", ""),
            "importe_gravado": float(p.get("ImporteGravado", 0) or 0),
            "importe_exento": float(p.get("ImporteExento", 0) or 0),
        })

    deducciones = root.findall(".//nomina12:Deducciones/nomina12:Deduccion", ns)
    for d in deducciones:
        conceptos.append({
            "tipo": "deduccion",
            "clave": d.get("Clave", ""),
            "concepto": d.get("Concepto", ""),
            "importe_gravado": float(d.get("ImporteGravado", 0) or 0),
            "importe_exento": float(d.get("ImporteExento", 0) or 0),
        })

    return conceptos


@login_required
def exportar_dispersiones_nomina_excel(request):
    empresa = _empresa_actual(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect("dashboard_inicio")

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    dispersiones = DispersionNomina.objects.filter(
        empresa=empresa, estatus="confirmado"
    ).select_related("cuenta_bancaria")

    if fecha_inicio:
        dispersiones = dispersiones.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        dispersiones = dispersiones.filter(fecha_pago__lte=fecha_fin)

    recibos = ReciboNomina.objects.filter(
        dispersion__in=dispersiones, estatus="ok"
    ).select_related(
        "empleado", "dispersion", "dispersion__cuenta_bancaria", "gasto"
    ).order_by("fecha_pago", "empleado__nombre")

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    FUENTE = "Arial"
    fill_header = PatternFill(start_color="1A0533", end_color="1A0533", fill_type="solid")
    font_header = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
    font_normal = Font(name=FUENTE, size=10)
    border_thin = Border(bottom=Side(style="thin", color="DDDDDD"))

    # Hoja de resumen
    ws_resumen.append([empresa.nombre])
    ws_resumen["A1"].font = Font(name=FUENTE, size=13, bold=True)
    ws_resumen.append(["Dispersión de Nómina"])
    ws_resumen["A2"].font = Font(name=FUENTE, size=10, italic=True, color="666666")
    ws_resumen.append([])

    encabezados = [
        "Fecha de pago", "Periodo inicio", "Periodo fin", "Empleado", "RFC",
        "Departamento", "Puesto", "Total Percepciones", "Total Deducciones",
        "Total Otros Pagos", "Neto Pagado", "Folio Fiscal (UUID)",
        "Cuenta Bancaria", "Folio Gasto GESAC",
    ]

    fila_encabezado = 4
    ws_resumen.append(encabezados)
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws_resumen.cell(row=fila_encabezado, column=col_idx)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fila = fila_encabezado + 1
    total_neto_general = 0

    for recibo in recibos:
        cuenta_txt = ""
        if recibo.dispersion and recibo.dispersion.cuenta_bancaria:
            cuenta_txt = f"{recibo.dispersion.cuenta_bancaria.banco} — {recibo.dispersion.cuenta_bancaria.numero_cuenta}"

        valores = [
            recibo.fecha_pago.strftime("%d/%m/%Y") if recibo.fecha_pago else "",
            recibo.periodo_inicio.strftime("%d/%m/%Y") if recibo.periodo_inicio else "",
            recibo.periodo_fin.strftime("%d/%m/%Y") if recibo.periodo_fin else "",
            recibo.empleado.nombre if recibo.empleado else recibo.nombre_receptor,
            recibo.rfc_receptor or "",
            recibo.empleado.get_departamento_display() if recibo.empleado else "",
            recibo.empleado.get_puesto_display() if recibo.empleado else "",
            float(recibo.total_percepciones or 0),
            float(recibo.total_deducciones or 0),
            float(recibo.total_otros_pagos or 0),
            float(recibo.neto_pagado or 0),
            recibo.uuid_fiscal or "",
            cuenta_txt,
            recibo.gasto.id if recibo.gasto else "",
        ]
        ws_resumen.append(valores)

        for col_idx in range(1, len(encabezados) + 1):
            celda = ws_resumen.cell(row=fila, column=col_idx)
            celda.font = font_normal
            celda.border = border_thin
            if col_idx in (8, 9, 10, 11):
                celda.number_format = '$#,##0.00'

        total_neto_general += float(recibo.neto_pagado or 0)
        fila += 1

    ws_resumen.cell(row=fila, column=10, value="Total neto:").font = Font(name=FUENTE, size=10, bold=True)
    celda_total = ws_resumen.cell(row=fila, column=11, value=total_neto_general)
    celda_total.font = Font(name=FUENTE, size=10, bold=True)
    celda_total.number_format = '$#,##0.00'

    # Hoja de detalle de conceptos
    ws_detalle = wb.create_sheet("Detalle conceptos")
    detalle_encabezados = [
        "Fecha de pago", "Periodo inicio", "Periodo fin",
        "Empleado", "RFC", "Cuenta Bancaria",
        "Tipo", "Clave", "Concepto", "Importe Gravado", "Importe Exento", "UUID"
    ]

    ws_detalle.append(detalle_encabezados)
    for col_idx, _ in enumerate(detalle_encabezados, start=1):
        celda = ws_detalle.cell(row=1, column=col_idx)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fila_detalle = 2

    for recibo in recibos:
        try:
            if recibo.archivo_xml:
                xml_bytes = recibo.archivo_xml.read()
                if isinstance(xml_bytes, bytes):
                    xml_text = xml_bytes.decode("utf-8", errors="ignore")
                else:
                    xml_text = str(xml_bytes)
            else:
                xml_text = ""
        except Exception:  # noqa: BLE001
            xml_text = ""

        conceptos = extraer_conceptos_nomina(xml_text)

        for item in conceptos:
            cuenta_txt = ""
            if recibo.dispersion and recibo.dispersion.cuenta_bancaria:
                cuenta_txt = f"{recibo.dispersion.cuenta_bancaria.banco} — {recibo.dispersion.cuenta_bancaria.numero_cuenta}"

            ws_detalle.append([
                recibo.fecha_pago.strftime("%d/%m/%Y") if recibo.fecha_pago else "",
                recibo.periodo_inicio.strftime("%d/%m/%Y") if recibo.periodo_inicio else "",
                recibo.periodo_fin.strftime("%d/%m/%Y") if recibo.periodo_fin else "",
                recibo.empleado.nombre if recibo.empleado else recibo.nombre_receptor,
                recibo.rfc_receptor or "",
                cuenta_txt,
                item["tipo"],
                item["clave"],
                item["concepto"],
                item["importe_gravado"],
                item["importe_exento"],
                recibo.uuid_fiscal or "",
            ])

            for col_idx in range(1, len(detalle_encabezados) + 1):
                celda = ws_detalle.cell(row=fila_detalle, column=col_idx)
                celda.font = font_normal
                celda.border = border_thin
                if col_idx in (10, 11):
                    celda.number_format = '$#,##0.00'
            fila_detalle += 1

    # Anchos
    anchos_resumen = [13, 13, 13, 26, 15, 16, 22, 16, 16, 16, 14, 24, 26, 14]
    for i, ancho in enumerate(anchos_resumen, start=1):
        ws_resumen.column_dimensions[get_column_letter(i)].width = ancho

    anchos_detalle = [13, 13, 13, 26, 15, 26, 12, 12, 28, 16, 16, 24]
    for i, ancho in enumerate(anchos_detalle, start=1):
        ws_detalle.column_dimensions[get_column_letter(i)].width = ancho

    ws_resumen.freeze_panes = f"A{fila_encabezado + 1}"
    ws_detalle.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nombre_archivo = f"dispersion_nomina_{fecha_inicio or 'inicio'}_a_{fecha_fin or 'hoy'}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response



@login_required
def gestionar_mapeo_conceptos_nomina(request):
    empresa = _empresa_actual(request)

    if request.method == "POST":
        concepto_xml = request.POST.get("concepto_xml", "").strip()
        tipo_gasto_id = request.POST.get("tipo_gasto_id")

        if not concepto_xml or not tipo_gasto_id:
            messages.error(request, "Selecciona el concepto y la cuenta contable.")
            return redirect("gestionar_mapeo_conceptos_nomina")

        tipo_gasto = get_object_or_404(TipoGasto, id=tipo_gasto_id, empresa=empresa)

        MapeoConceptoNomina.objects.update_or_create(
            empresa=empresa, concepto_xml=concepto_xml,
            defaults={"tipo_gasto": tipo_gasto},
        )
        messages.success(request, f"\"{concepto_xml}\" ahora se registra en la cuenta \"{tipo_gasto.nombre}\".")
        return redirect("gestionar_mapeo_conceptos_nomina")

    # Mapeos ya capturados
    mapeos = MapeoConceptoNomina.objects.filter(empresa=empresa).select_related("tipo_gasto").order_by("concepto_xml")

    # NUEVO -- detecta automaticamente que conceptos de deduccion/otro
    # pago han aparecido en tus XML pero todavia no tienen mapeo, para
    # que no tengas que adivinar cuales capturar.
    conceptos_vistos = set(
        ConceptoNomina.objects.filter(
            recibo__dispersion__empresa=empresa, tipo__in=["deduccion", "otro_pago"],
        ).values_list("concepto", flat=True).distinct()
    )
    conceptos_mapeados = set(mapeos.values_list("concepto_xml", flat=True))
    conceptos_pendientes = sorted(conceptos_vistos - conceptos_mapeados)

    # Cuentas disponibles para el selector -- todo el catalogo de gastos
    # de la empresa (no solo las de nomina, por si acaso alguna
    # deduccion se quiere mandar a otra cuenta).
    cuentas_disponibles = (
        TipoGasto.objects.filter(empresa=empresa)
        .select_related("subgrupo", "subgrupo__grupo")
        .order_by("subgrupo__grupo__nombre", "subgrupo__nombre", "nombre")
    )

    return render(request, "nomina/mapeo_conceptos.html", {
        "mapeos": mapeos,
        "conceptos_pendientes": conceptos_pendientes,
        "cuentas_disponibles": cuentas_disponibles,
    })


@login_required
def eliminar_mapeo_concepto_nomina(request, mapeo_id):
    empresa = _empresa_actual(request)
    mapeo = get_object_or_404(MapeoConceptoNomina, id=mapeo_id, empresa=empresa)

    if request.method == "POST":
        concepto = mapeo.concepto_xml
        mapeo.delete()
        messages.success(request, f"Se quito el mapeo de \"{concepto}\".")

    return redirect("gestionar_mapeo_conceptos_nomina")