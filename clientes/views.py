
# Create your views here.
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from empresas.models import Empresa
from .models import Cliente
from .forms import ClienteCargaMasivaForm, ClienteForm
from django.contrib.admin.views.decorators import staff_member_required
import openpyxl
from django.contrib import messages

@login_required
def lista_clientes(request):
    if request.user.is_superuser:
        clientes = Cliente.objects.filter(activo=True)
    else:
        empresa = request.user.perfilusuario.empresa
        clientes = Cliente.objects.filter(empresa=empresa, activo=True)
    return render(request, 'clientes/lista_clientes.html', {'clientes': clientes})

@login_required
def crear_cliente(request):
    perfil = getattr(request.user, 'perfilusuario', None)

    if request.method == 'POST':
        form = ClienteForm(request.POST, user=request.user)
        if form.is_valid():
            cliente = form.save(commit=False)
            if not request.user.is_superuser and perfil:
                cliente.empresa = perfil.empresa
            cliente.save()
            return redirect('lista_clientes')
    else:
        form = ClienteForm(user=request.user)
        if perfil:
            form.fields['empresa'].initial = perfil.empresa

    return render(request, 'clientes/crear_cliente.html', {'form': form})

@login_required
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if not request.user.is_superuser and cliente.empresa != request.user.perfilusuario.empresa:
        return redirect('lista_clientes')

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente, user=request.user)

    return render(request, 'clientes/editar_cliente.html', {'form': form, 'cliente': cliente})

@login_required
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if not request.user.is_superuser and cliente.empresa != request.user.perfilusuario.empresa:
        return redirect('lista_clientes')

    if request.method == 'POST':
        cliente.activo = False
        cliente.save()
        return redirect('lista_clientes')

    return render(request, 'clientes/eliminar_cliente.html', {'cliente': cliente})

@staff_member_required
def carga_masiva_clientes(request):
    if request.method == 'POST':
        form = ClienteCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                nombre, rfc, telefono, email, direccion, observaciones = row
                try:
                    if not rfc:
                        raise Exception("RFC vacío")
                    if Cliente.objects.filter(rfc__iexact=rfc.strip()).exists():
                        raise Exception(f"RFC duplicado: {rfc}")
                    Cliente.objects.create(
                        nombre=nombre,
                        rfc=rfc.strip(),
                        telefono=telefono,
                        email=email,
                        direccion=direccion,
                        observaciones=observaciones or ""
                    )
                except Exception as e:
                    errores.append(f"Fila {i}: {e}")

            if errores:
                messages.error(request, "Algunos clientes no se cargaron:<br>" + "<br>".join(errores))
            else:
                messages.success(request, "¡Clientes cargados exitosamente!")
            return redirect('carga_masiva_clientes')
    else:
        form = ClienteCargaMasivaForm()
    return render(request, 'clientes/carga_masiva_clientes.html', {'form': form})

@staff_member_required
def plantilla_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Clientes"
    ws.append(['nombre', 'rfc', 'telefono', 'email', 'direccion', 'observaciones'])
    ws.append(['Juan Pérez', 'JUPE800101ABC', '5551234567', 'juanperez@email.com', 'Av. Reforma 123', 'Cliente frecuente'])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_clientes.xlsx'
    wb.save(response)
    return response