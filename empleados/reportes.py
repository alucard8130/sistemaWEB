"""Reporte de asistencia: resumen por empleado + detalle + exportación a Excel"""
import calendar
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages

from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from empleados.views_asistencia import MINUTOS_CONSIDERADO_VIGENTE

from .models import Empleado, Incidencia, RegistroAsistencia, UbicacionEnVivo
from .tareas_asistencia import detectar_y_registrar_faltas
from django.db.models import Avg, ExpressionWrapper, F, DurationField
from django.utils import timezone as tz


def _rango_fechas_desde_request(request):
    """Lee fecha_inicio/fecha_fin de la URL, o usa el mes actual por default."""
    hoy = date.today()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
 
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
            fecha_fin = date.fromisoformat(fecha_fin_str)
            return fecha_inicio, fecha_fin
        except ValueError:
            pass
 
    fecha_inicio = hoy.replace(day=1)
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    fecha_fin = hoy.replace(day=ultimo_dia)
    return fecha_inicio, fecha_fin
 
 
def _promedio_minutos_comida(empleado, fecha_inicio, fecha_fin):
    """Promedio de minutos de comida del empleado en el periodo, entre los
    días donde sí se registró tanto salida como regreso de comer."""
    resultado = RegistroAsistencia.objects.filter(
        empleado=empleado, fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
        hora_inicio_comida__isnull=False, hora_fin_comida__isnull=False,
    ).annotate(
        duracion=ExpressionWrapper(
            F('hora_fin_comida') - F('hora_inicio_comida'), output_field=DurationField()
        )
    ).aggregate(promedio=Avg('duracion'))
 
    duracion_promedio = resultado['promedio']
    if duracion_promedio is None:
        return None
    return round(duracion_promedio.total_seconds() / 60)

 
def _ubicaciones_en_vivo_por_empleado(empleados_ids):
    """Trae la última ubicación reportada de cada empleado, en un solo query,
    y arma un dict {empleado_id: {'vigente': bool, 'dentro_de_zona': bool|None}}"""
    limite = tz.now() - timedelta(minutes=MINUTOS_CONSIDERADO_VIGENTE)

    ubicaciones = UbicacionEnVivo.objects.filter(empleado_id__in=empleados_ids)

    resultado = {}
    for u in ubicaciones:
        resultado[u.empleado_id] = {
            'vigente': u.actualizado_en >= limite,
            'dentro_de_zona': u.dentro_de_zona,
            'actualizado_en': tz.localtime(u.actualizado_en),
        }
    return resultado


def _resumen_asistencia(empresa, fecha_inicio, fecha_fin, departamento=None):
    """Resumen por empleado: días asistidos, retardos, faltas, % asistencia, comida, ubicación."""
    empleados = Empleado.objects.filter(empresa=empresa, activo=True)
    if departamento:
        empleados = empleados.filter(departamento=departamento)

    empleados = list(empleados.order_by('nombre'))
    ubicaciones_por_empleado = _ubicaciones_en_vivo_por_empleado([e.id for e in empleados])

    resumen = []
    for empleado in empleados:
        dias_asistidos = RegistroAsistencia.objects.filter(
            empleado=empleado, fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
            hora_entrada__isnull=False,
        ).count()

        incidencias_periodo = Incidencia.objects.filter(
            empleado=empleado, fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
        )
        retardos = incidencias_periodo.filter(tipo='retardo').count()
        faltas = incidencias_periodo.filter(tipo='falta').count()

        total_dias_contabilizados = dias_asistidos + faltas
        porcentaje = (
            round(dias_asistidos / total_dias_contabilizados * 100, 1)
            if total_dias_contabilizados > 0 else None
        )

        promedio_minutos_comida = _promedio_minutos_comida(empleado, fecha_inicio, fecha_fin)

        resumen.append({
            'empleado': empleado,
            'dias_asistidos': dias_asistidos,
            'retardos': retardos,
            'faltas': faltas,
            'porcentaje_asistencia': porcentaje,
            'promedio_minutos_comida': promedio_minutos_comida,
            'ubicacion_en_vivo': ubicaciones_por_empleado.get(empleado.id),
        })
    return resumen
 
 
@login_required
def reporte_asistencia(request):
    empresa = request.user.perfilusuario.empresa
    fecha_inicio, fecha_fin = _rango_fechas_desde_request(request)
    departamento = request.GET.get('departamento') or None
 
    resumen = _resumen_asistencia(empresa, fecha_inicio, fecha_fin, departamento)
 
    # Detalle de un empleado especifico, si se selecciono uno
    detalle_empleado = None
    detalle_registros = None
    detalle_incidencias = None
    empleado_id = request.GET.get('empleado_id')
    if empleado_id:
        detalle_empleado = get_object_or_404(Empleado, id=empleado_id, empresa=empresa)
        detalle_registros = RegistroAsistencia.objects.filter(
            empleado=detalle_empleado, fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
        ).order_by('-fecha')
        detalle_incidencias = Incidencia.objects.filter(
            empleado=detalle_empleado, fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
        ).order_by('-fecha')
 
    contexto = {
        'resumen': resumen,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'departamento_actual': departamento,
        'departamentos': Empleado.DEPARTAMENTO_CHOICES,
        'detalle_empleado': detalle_empleado,
        'detalle_registros': detalle_registros,
        'detalle_incidencias': detalle_incidencias,
    }
    return render(request, 'empleados/reporte_asistencia.html', contexto)
 
 
@login_required
@require_POST
def detectar_faltas_manual(request):
    """Corre la detección de faltas para TODO un rango de fechas (el mismo
    periodo que ya se tenga filtrado en el reporte), disparada manualmente
    con un solo clic antes de exportar -- pensado para el cierre mensual
    que se manda a nómina."""
    empresa = request.user.perfilusuario.empresa
    ayer = date.today() - timedelta(days=1)
 
    try:
        fecha_inicio = date.fromisoformat(request.POST.get('fecha_inicio', ''))
        fecha_fin = date.fromisoformat(request.POST.get('fecha_fin', ''))
    except (ValueError, TypeError):
        messages.error(request, "Rango de fechas inválido.")
        return redirect(request.META.get('HTTP_REFERER', 'reporte_asistencia'))
 
    # No tiene sentido marcar falta de un dia que todavia no termina (hoy)
    # o que esta en el futuro -- se limita el rango hasta ayer como maximo.
    fecha_fin_real = min(fecha_fin, ayer)
 
    if fecha_inicio > fecha_fin_real:
        messages.info(request, "No hay días completos en ese rango todavía para revisar (solo se pueden detectar faltas de días ya terminados).")
        return redirect(request.META.get('HTTP_REFERER', 'reporte_asistencia'))
 
    total_faltas = 0
    dias_procesados = 0
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin_real:
        faltas_creadas = detectar_y_registrar_faltas(fecha_actual, empresa=empresa)
        total_faltas += len(faltas_creadas)
        dias_procesados += 1
        fecha_actual += timedelta(days=1)
 
    if total_faltas:
        messages.warning(request, f"Se revisaron {dias_procesados} día(s) ({fecha_inicio:%d/%m} — {fecha_fin_real:%d/%m}) y se registraron {total_faltas} falta(s) nueva(s).")
    else:
        messages.success(request, f"Se revisaron {dias_procesados} día(s) ({fecha_inicio:%d/%m} — {fecha_fin_real:%d/%m}), sin faltas nuevas que registrar.")
 
    return redirect(request.META.get('HTTP_REFERER', 'reporte_asistencia'))
 
 
@login_required
def exportar_reporte_asistencia(request):
    empresa = request.user.perfilusuario.empresa
    fecha_inicio, fecha_fin = _rango_fechas_desde_request(request)
    departamento = request.GET.get('departamento') or None
 
    resumen = _resumen_asistencia(empresa, fecha_inicio, fecha_fin, departamento)
 
    wb = Workbook()
 
    # ---- Hoja 1: resumen (igual que antes, + columna de comida) ----
    ws = wb.active
    ws.title = "Asistencia"
 
    encabezados = ['Empleado', 'Departamento', 'Puesto', 'Días Asistidos', 'Retardos', 'Faltas', '% Asistencia', 'Prom. Min. Comida']
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", start_color="4C4CB0")
        celda.alignment = Alignment(horizontal='center')
 
    for fila in resumen:
        e = fila['empleado']
        porcentaje = f"{fila['porcentaje_asistencia']:.1f}%" if fila['porcentaje_asistencia'] is not None else "N/D"
        promedio_comida = fila['promedio_minutos_comida'] if fila['promedio_minutos_comida'] is not None else "N/D"
        ws.append([
            e.nombre, e.get_departamento_display(), e.get_puesto_display(),
            fila['dias_asistidos'], fila['retardos'], fila['faltas'], porcentaje, promedio_comida,
        ])
 
    anchos = [30, 18, 16, 14, 12, 10, 14, 16]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho
 
    # ---- Hoja 2: detalle de checadas día por día, con comida ----
    ws_detalle = wb.create_sheet(title="Checadas detalladas")
 
    encabezados_detalle = [
        'Empleado', 'Departamento', 'Fecha',
        'Entrada', 'Salida a comer', 'Regreso de comer', 'Minutos comida', 'Salida',
    ]
    ws_detalle.append(encabezados_detalle)
    for col in range(1, len(encabezados_detalle) + 1):
        celda = ws_detalle.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", start_color="4C4CB0")
        celda.alignment = Alignment(horizontal='center')
 
    empleados_qs = Empleado.objects.filter(empresa=empresa, activo=True)
    if departamento:
        empleados_qs = empleados_qs.filter(departamento=departamento)
 
    for empleado in empleados_qs.order_by('nombre'):
        registros = RegistroAsistencia.objects.filter(
            empleado=empleado, fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
        ).order_by('fecha')
        for r in registros:
            ws_detalle.append([
                empleado.nombre,
                empleado.get_departamento_display(),
                r.fecha.strftime('%d/%m/%Y'),
                r.hora_entrada.strftime('%H:%M') if r.hora_entrada else '',
                r.hora_inicio_comida.strftime('%H:%M') if r.hora_inicio_comida else '',
                r.hora_fin_comida.strftime('%H:%M') if r.hora_fin_comida else '',
                r.minutos_comida if r.minutos_comida is not None else '',
                r.hora_salida.strftime('%H:%M') if r.hora_salida else '',
            ])
 
    anchos_detalle = [28, 18, 12, 10, 14, 16, 14, 10]
    for i, ancho in enumerate(anchos_detalle, start=1):
        ws_detalle.column_dimensions[chr(64 + i)].width = ancho
 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"asistencia_{fecha_inicio}_{fecha_fin}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


