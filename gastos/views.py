# from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
import difflib
from io import BytesIO
import json
import unicodedata
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
import openpyxl
from unidecode import unidecode
from caja_chica.models import GastoCajaChica, ValeCaja
from conciliaciones.utils import validar_periodo_abierto
from empleados.models import Empleado
from empresas.models import CuentaBancaria, Empresa
from facturacion.models import CobroOtrosIngresos, Factura, Pago, TipoCuotaHomologacion, TipoOtroIngreso
from presupuestos.models import Presupuesto
from proveedores.models import Proveedor
from .forms import (
    CargaMasivaCuentasForm,
    GastoForm,
    GastosCargaMasivaForm,
    MotivoReversaPagoForm,
    PagoGastoForm,
    SubgrupoGastoForm,
    TipoGastoForm,
)
from .models import CargaCatalogoFila, CargaCatalogoSesion, CuentaContable, Gasto, GrupoGasto, PagoGasto, SubgrupoGasto, TipoGasto
from datetime import datetime
from django.db.models import F, DecimalField, Value
from django.db.models import Q, Sum, Count, Case, When, FloatField
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import ProtectedError
from django.db.models.functions import Coalesce, TruncMonth, TruncYear
from num2words import num2words
from django.utils.dateformat import DateFormat
from datetime import date
from django.contrib import messages as django_messages
from openpyxl.styles import Font, PatternFill


###################CATALOGOS CONTABLES GASTOS E INGRESOS############################
@login_required
def plantilla_catalogo_cuentas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo de Cuentas"
    ws.append(['codigo', 'nombre_cuenta', 'codigo_padre', 'naturaleza', 'grupo_gasto', 'subgrupo_gasto', 'uso_especial'])
    ws.append(['500', 'Gastos de Operación', '', 'deudora', '', '', ''])
    ws.append(['500-01', 'Mantenimiento y Conservación', '500', 'deudora', '', '', ''])
    ws.append(['500-01-001', 'Mantenimiento de Áreas Comunes', '500-01', 'deudora', 'Mantenimiento', 'Áreas Comunes', ''])
    ws.append(['400', 'Ingresos por Cuotas', '', 'acreedora', '', '', ''])
    ws.append(['400-01', 'Cuotas de Mantenimiento', '400', 'acreedora', '', '', 'cuota_mantenimiento'])
    ws.append(['400-02', 'Cuotas de Áreas Comunes', '400', 'acreedora', '', '', 'cuota_renta'])
    ws.append(['400-03', 'Depósitos en Garantía', '400', 'acreedora', '', '', 'cuota_deposito'])
    ws.append(['400-04', 'Cuotas Extraordinarias', '400', 'acreedora', '', '', 'cuota_extraordinaria'])
    ws.append(['700', 'Otros Ingresos', '', 'acreedora', '', '', ''])
    ws.append(['700-01', 'Intereses', '700', 'acreedora', '', '', 'cuota_intereses'])
    ws.append(['700-02', 'Multas', '700', 'acreedora', '', '', 'cuota_penalidad'])
    ws.append(['700-03', 'Renta de Estacionamiento', '700', 'acreedora', '', '', 'otro_ingreso'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_catalogo_cuentas.xlsx'
    wb.save(response)
    return response


def _empresa_del_usuario(request):
    if request.user.is_superuser:
        empresa_id = request.session.get('empresa_id')
        from empresas.models import Empresa
        return Empresa.objects.filter(id=empresa_id).first() if empresa_id else None
    perfil = getattr(request.user, 'perfilusuario', None)
    return perfil.empresa if perfil else None


def _mejor_coincidencia_tipo_gasto(nombre_cuenta, tipos_existentes):
    """Compara el nombre de la cuenta contra los tipos de gasto ya
    existentes de la empresa, y regresa (tipo_gasto, porcentaje) del
    más parecido, o (None, 0) si ninguno se acerca lo suficiente."""
    mejor_tipo = None
    mejor_pct = 0
    nombre_norm = nombre_cuenta.strip().lower()

    for tipo in tipos_existentes:
        ratio = difflib.SequenceMatcher(None, nombre_norm, tipo.nombre.strip().lower()).ratio()
        pct = round(ratio * 100)
        if pct > mejor_pct:
            mejor_pct = pct
            mejor_tipo = tipo

    if mejor_pct >= 70:  # umbral -- ajustable
        return mejor_tipo, mejor_pct
    return None, 0


@login_required
def carga_masiva_catalogo_cuentas(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    if request.method == 'POST':
        form = CargaMasivaCuentasForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            tipos_existentes = list(TipoGasto.objects.filter(empresa=empresa).select_related('subgrupo'))

            sesion = CargaCatalogoSesion.objects.create(empresa=empresa, registrado_por=request.user)
            errores = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                    continue

                try:
                    codigo = str(row[0]).strip() if row[0] not in (None, "") else None
                    nombre_cuenta = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else None
                    codigo_padre = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else None
                    naturaleza = str(row[3]).strip().lower() if len(row) > 3 and row[3] not in (None, "") else 'deudora'
                    grupo_nombre = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "") else None
                    subgrupo_nombre = str(row[5]).strip() if len(row) > 5 and row[5] not in (None, "") else None
                    uso_especial = str(row[6]).strip().lower() if len(row) > 6 and row[6] not in (None, "") else None

                    if naturaleza not in ('deudora', 'acreedora'):
                        naturaleza = 'deudora'

                    if not codigo or not nombre_cuenta:
                        raise Exception("Código o nombre de cuenta vacío.")

                    tipo_sugerido, pct = (None, 0)
                    if grupo_nombre and subgrupo_nombre:
                        tipo_sugerido, pct = _mejor_coincidencia_tipo_gasto(nombre_cuenta, tipos_existentes)

                    CargaCatalogoFila.objects.create(
                        sesion=sesion,
                        fila_excel=i,
                        codigo=codigo,
                        nombre_cuenta=nombre_cuenta,
                        codigo_padre=codigo_padre,
                        naturaleza=naturaleza,
                        grupo_nombre=grupo_nombre,
                        subgrupo_nombre=subgrupo_nombre,
                        uso_especial=uso_especial,
                        tipo_gasto_sugerido=tipo_sugerido,
                        similitud_pct=pct,
                    )

                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")

            if errores:
                from django.utils.safestring import mark_safe
                msg = "<br>".join(errores[:50])
                messages.warning(request, mark_safe("Algunas filas se omitieron:<br>" + msg))

            return redirect('revisar_carga_catalogo', sesion_id=sesion.id)
    else:
        form = CargaMasivaCuentasForm()

    return render(request, 'gastos/carga_masiva_catalogo_cuentas.html', {'form': form})


@login_required
def revisar_carga_catalogo(request, sesion_id):
    empresa = _empresa_del_usuario(request)
    sesion = get_object_or_404(CargaCatalogoSesion, pk=sesion_id, empresa=empresa, estado='pendiente_revision')

    tipos_existentes = TipoGasto.objects.filter(empresa=empresa).select_related('subgrupo', 'subgrupo__grupo').order_by('nombre')

    filas_con_tipo = sesion.filas.exclude(grupo_nombre__isnull=True).exclude(grupo_nombre='')
    filas_solo_cuenta = sesion.filas.filter(grupo_nombre__isnull=True) | sesion.filas.filter(grupo_nombre='')

    return render(request, 'gastos/revisar_carga_catalogo.html', {
        'sesion': sesion,
        'filas_con_tipo': filas_con_tipo,
        'filas_solo_cuenta': filas_solo_cuenta,
        'tipos_existentes': tipos_existentes,
    })


@login_required
def confirmar_carga_catalogo(request, sesion_id):
    empresa = _empresa_del_usuario(request)
    sesion = get_object_or_404(CargaCatalogoSesion, pk=sesion_id, empresa=empresa, estado='pendiente_revision')

    if request.method != 'POST':
        return redirect('revisar_carga_catalogo', sesion_id=sesion_id)

    cuentas_creadas = 0
    tipos_creados = 0
    tipos_homologados = 0

    with transaction.atomic():
        # --- Primera pasada: crear/actualizar TODAS las cuentas contables ---
        cuentas_por_codigo = {}
        for fila in sesion.filas.all():
            cuenta, _ = CuentaContable.objects.update_or_create(
                empresa=empresa, codigo=fila.codigo,
                defaults={'nombre': fila.nombre_cuenta, 'naturaleza': fila.naturaleza, 'activa': True}
            )
            cuentas_por_codigo[fila.codigo] = cuenta
            cuentas_creadas += 1

        # --- Segunda pasada: jerarquía ---
        for fila in sesion.filas.exclude(codigo_padre__isnull=True).exclude(codigo_padre=''):
            padre = cuentas_por_codigo.get(fila.codigo_padre) or CuentaContable.objects.filter(
                empresa=empresa, codigo=fila.codigo_padre
            ).first()
            if padre:
                cuenta = cuentas_por_codigo[fila.codigo]
                cuenta.cuenta_padre = padre
                cuenta.save(update_fields=['cuenta_padre'])

        # --- Tercera pasada: decisiones de tipo de gasto, según lo elegido en el form ---
        for fila in sesion.filas.exclude(grupo_nombre__isnull=True).exclude(grupo_nombre=''):
            accion = request.POST.get(f'accion_{fila.id}')
            cuenta = cuentas_por_codigo[fila.codigo]

            if accion == 'usar_existente':
                tipo_id = request.POST.get(f'tipo_existente_{fila.id}')
                if tipo_id:
                    tipo = TipoGasto.objects.filter(id=tipo_id, empresa=empresa).first()
                    if tipo and not tipo.cuenta_contable_id:
                        tipo.cuenta_contable = cuenta
                        tipo.save(update_fields=['cuenta_contable'])
                        tipos_homologados += 1

            elif accion == 'crear_nuevo':
                grupo, _ = GrupoGasto.objects.get_or_create(nombre=fila.grupo_nombre)
                subgrupo, _ = SubgrupoGasto.objects.get_or_create(grupo=grupo, nombre=fila.subgrupo_nombre)
                tipo, creado = TipoGasto.objects.get_or_create(
                    empresa=empresa, subgrupo=subgrupo, nombre=fila.nombre_cuenta,
                    defaults={'cuenta_contable': cuenta}
                )
                if creado:
                    tipos_creados += 1
                elif not tipo.cuenta_contable_id:
                    tipo.cuenta_contable = cuenta
                    tipo.save(update_fields=['cuenta_contable'])
                    tipos_homologados += 1

            # accion == 'solo_cuenta' -> no se crea ni homologa ningún TipoGasto

        # --- Cuarta pasada: homologaciones especiales (cuotas y otros ingresos) ---
        TIPOS_CUOTA_VALIDOS = dict(Factura.TIPO_CUOTA_CHOICES).keys()
        otros_ingresos_homologados = 0
        cuotas_homologadas = 0

        for fila in sesion.filas.exclude(uso_especial__isnull=True).exclude(uso_especial=''):
            cuenta = cuentas_por_codigo[fila.codigo]

            if fila.uso_especial.startswith('cuota_'):
                tipo_cuota_valor = fila.uso_especial.replace('cuota_', '', 1)
                if tipo_cuota_valor in TIPOS_CUOTA_VALIDOS:
                    homologacion, _ = TipoCuotaHomologacion.objects.get_or_create(
                        empresa=empresa, tipo_cuota=tipo_cuota_valor
                    )
                    if not homologacion.cuenta_contable_id:
                        homologacion.cuenta_contable = cuenta
                        homologacion.save(update_fields=['cuenta_contable'])
                        cuotas_homologadas += 1

            elif fila.uso_especial == 'otro_ingreso':
                tipo_oi, creado = TipoOtroIngreso.objects.get_or_create(
                    empresa=empresa, nombre=fila.nombre_cuenta,
                    defaults={'cuenta_contable': cuenta}
                )
                if not creado and not tipo_oi.cuenta_contable_id:
                    tipo_oi.cuenta_contable = cuenta
                    tipo_oi.save(update_fields=['cuenta_contable'])
                if creado or not tipo_oi.cuenta_contable_id:
                    otros_ingresos_homologados += 1

        sesion.estado = 'aplicada'
        sesion.save(update_fields=['estado'])

    messages.success(
        request,
        f"✅ {cuentas_creadas} cuentas cargadas, {tipos_creados} tipos de gasto nuevos, "
        f"{tipos_homologados} tipos de gasto homologados, {cuotas_homologadas} tipos de cuota "
        f"homologados, {otros_ingresos_homologados} tipos de otro ingreso homologados."
    )
    return redirect('lista_catalogo_cuentas')


@login_required
def lista_catalogo_cuentas(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    cuentas_qs = list(
        CuentaContable.objects.filter(empresa=empresa)
        .prefetch_related('tipos_gasto')
        .order_by('codigo')
    )
    cuentas_por_id = {c.id: c for c in cuentas_qs}

    def calcular_nivel(cuenta):
        nivel = 0
        actual_id = cuenta.cuenta_padre_id
        visitados = set()
        while actual_id and actual_id not in visitados:
            visitados.add(actual_id)
            nivel += 1
            padre = cuentas_por_id.get(actual_id)
            actual_id = padre.cuenta_padre_id if padre else None
        return nivel

    cuentas = [{'obj': c, 'nivel': calcular_nivel(c)} for c in cuentas_qs]

    return render(request, 'gastos/lista_catalogo_cuentas.html', {'cuentas': cuentas})


@login_required
def homologar_tipos_gasto(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    tipos_gasto = TipoGasto.objects.filter(empresa=empresa).select_related(
        'subgrupo', 'subgrupo__grupo', 'cuenta_contable'
    ).order_by('subgrupo__grupo__nombre', 'subgrupo__nombre', 'nombre')

    cuentas_disponibles = CuentaContable.objects.filter(empresa=empresa, activa=True).order_by('codigo')

    return render(request, 'gastos/homologar_tipos_gasto.html', {
        'tipos_gasto': tipos_gasto,
        'cuentas_disponibles': cuentas_disponibles,
    })


@login_required
def asignar_cuenta_contable(request, tipo_gasto_id):
    empresa = _empresa_del_usuario(request)
    tipo = get_object_or_404(TipoGasto, pk=tipo_gasto_id, empresa=empresa)

    if request.method == 'POST':
        cuenta_id = request.POST.get('cuenta_contable_id')
        if cuenta_id:
            cuenta = get_object_or_404(CuentaContable, pk=cuenta_id, empresa=empresa)
            tipo.cuenta_contable = cuenta
        else:
            tipo.cuenta_contable = None
        tipo.save(update_fields=['cuenta_contable'])
        messages.success(request, f"'{tipo.nombre}' homologado correctamente.")

    return redirect('homologar_tipos_gasto')


@login_required
def exportar_poliza_gastos(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    if request.method == 'POST' or request.GET.get('fecha_inicio'):
        fecha_inicio_str = request.POST.get('fecha_inicio') or request.GET.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin') or request.GET.get('fecha_fin')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Fechas inválidas.")
            return redirect('exportar_poliza_gastos')

        pagos = PagoGasto.objects.filter(
            gasto__empresa=empresa,
            fecha_pago__gte=fecha_inicio,
            fecha_pago__lte=fecha_fin,
        ).select_related('gasto', 'gasto__tipo_gasto', 'gasto__tipo_gasto__cuenta_contable', 'gasto__proveedor').order_by('fecha_pago')

        wb = openpyxl.Workbook()

        # --- Hoja 1: Póliza (solo los homologados) ---
        ws = wb.active
        ws.title = "Póliza"
        ws.append(['Fecha', 'Cuenta', 'Concepto', 'Cargo', 'Abono'])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2471A3", end_color="2471A3", fill_type="solid")

        total_exportado = 0
        pagos_sin_homologar = []

        for pago in pagos:
            cuenta_contable = getattr(pago.gasto.tipo_gasto, 'cuenta_contable', None) if pago.gasto.tipo_gasto else None
            if not cuenta_contable:
                pagos_sin_homologar.append(pago)
                continue

            concepto = f"{pago.gasto.tipo_gasto.nombre}"
            if pago.gasto.proveedor:
                concepto += f" — {pago.gasto.proveedor.nombre}"
            if pago.gasto.descripcion:
                concepto += f" — {pago.gasto.descripcion}"

            ws.append([
                pago.fecha_pago.strftime('%d/%m/%Y'),
                f"{cuenta_contable.codigo} — {cuenta_contable.nombre}",
                concepto[:200],
                float(pago.monto),
                '',
            ])
            total_exportado += float(pago.monto)

        ws.append(['', '', 'TOTAL', total_exportado, ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for col_letra, ancho in [('A', 12), ('B', 35), ('C', 45), ('D', 14), ('E', 14)]:
            ws.column_dimensions[col_letra].width = ancho

        # --- Hoja 2: Gastos sin homologar (para que sepa qué falta) ---
        if pagos_sin_homologar:
            ws2 = wb.create_sheet("Sin homologar")
            ws2.append(['Fecha', 'Tipo de Gasto', 'Proveedor', 'Monto'])
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

            for pago in pagos_sin_homologar:
                ws2.append([
                    pago.fecha_pago.strftime('%d/%m/%Y'),
                    pago.gasto.tipo_gasto.nombre if pago.gasto.tipo_gasto else '—',
                    pago.gasto.proveedor.nombre if pago.gasto.proveedor else '—',
                    float(pago.monto),
                ])
            for col_letra, ancho in [('A', 12), ('B', 30), ('C', 30), ('D', 14)]:
                ws2.column_dimensions[col_letra].width = ancho

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=poliza_gastos_{fecha_inicio}_{fecha_fin}.xlsx'
        wb.save(response)
        return response

    return render(request, 'gastos/exportar_poliza.html', {})


@login_required
def exportar_poliza_ingresos(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    if request.method == 'POST' or request.GET.get('fecha_inicio'):
        fecha_inicio_str = request.POST.get('fecha_inicio') or request.GET.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin') or request.GET.get('fecha_fin')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Fechas inválidas.")
            return redirect('exportar_poliza_ingresos')

        # --- Pagos de cuotas ---
        pagos_cuotas = Pago.objects.filter(
            factura__empresa=empresa,
            fecha_pago__gte=fecha_inicio,
            fecha_pago__lte=fecha_fin,
        ).exclude(forma_pago='nota_credito').select_related(
            'factura', 'cuenta_bancaria', 'cuenta_bancaria__cuenta_contable', 'factura__cliente'
        )

        # --- Cobros de otros ingresos ---
        cobros_otros = CobroOtrosIngresos.objects.filter(
            factura__empresa=empresa,
            fecha_cobro__gte=fecha_inicio,
            fecha_cobro__lte=fecha_fin,
        ).select_related(
            'factura', 'factura__tipo_ingreso', 'cuenta_bancaria', 'cuenta_bancaria__cuenta_contable', 'factura__cliente'
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Póliza Ingresos"
        ws.append(['Fecha', 'Cuenta', 'Concepto', 'Cargo', 'Abono'])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")

        total_exportado = 0
        pendientes_homologar = []

        homologaciones_cuota = {
            h.tipo_cuota: h.cuenta_contable
            for h in TipoCuotaHomologacion.objects.filter(empresa=empresa).select_related('cuenta_contable')
        }

        for pago in pagos_cuotas:
            cuenta_banco = getattr(pago.cuenta_bancaria, 'cuenta_contable', None) if pago.cuenta_bancaria else None
            cuenta_ingreso = homologaciones_cuota.get(pago.factura.tipo_cuota)

            if not cuenta_banco or not cuenta_ingreso:
                pendientes_homologar.append({
                    'fecha': pago.fecha_pago, 'tipo': pago.factura.get_tipo_cuota_display(),
                    'monto': pago.monto, 'falta': 'Cuenta bancaria' if not cuenta_banco else 'Tipo de cuota',
                })
                continue

            concepto = f"Cuota {pago.factura.get_tipo_cuota_display()} — {pago.factura.cliente.nombre} — Folio {pago.factura.folio}"[:200]

            ws.append([pago.fecha_pago.strftime('%d/%m/%Y'), f"{cuenta_banco.codigo} — {cuenta_banco.nombre}", concepto, float(pago.monto), ''])
            ws.append([pago.fecha_pago.strftime('%d/%m/%Y'), f"{cuenta_ingreso.codigo} — {cuenta_ingreso.nombre}", concepto, '', float(pago.monto)])

            total_exportado += float(pago.monto)

        for cobro in cobros_otros:
            cuenta_banco = getattr(cobro.cuenta_bancaria, 'cuenta_contable', None) if cobro.cuenta_bancaria else None
            cuenta_ingreso = getattr(cobro.factura.tipo_ingreso, 'cuenta_contable', None) if cobro.factura.tipo_ingreso else None

            if not cuenta_banco or not cuenta_ingreso:
                pendientes_homologar.append({
                    'fecha': cobro.fecha_cobro, 'tipo': cobro.factura.tipo_ingreso.nombre if cobro.factura.tipo_ingreso else '—',
                    'monto': cobro.monto, 'falta': 'Cuenta bancaria' if not cuenta_banco else 'Tipo de otro ingreso',
                })
                continue

            concepto = f"{cobro.factura.tipo_ingreso.nombre} — {cobro.factura.cliente.nombre} — Folio {cobro.factura.folio}"[:200]

            ws.append([cobro.fecha_cobro.strftime('%d/%m/%Y'), f"{cuenta_banco.codigo} — {cuenta_banco.nombre}", concepto, float(cobro.monto), ''])
            ws.append([cobro.fecha_cobro.strftime('%d/%m/%Y'), f"{cuenta_ingreso.codigo} — {cuenta_ingreso.nombre}", concepto, '', float(cobro.monto)])

            total_exportado += float(cobro.monto)

        ws.append(['', '', 'TOTAL (Cargo = Abono)', total_exportado, total_exportado])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for col_letra, ancho in [('A', 12), ('B', 35), ('C', 50), ('D', 14), ('E', 14)]:
            ws.column_dimensions[col_letra].width = ancho

        if pendientes_homologar:
            ws2 = wb.create_sheet("Sin homologar")
            ws2.append(['Fecha', 'Tipo', 'Monto', 'Falta homologar'])
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            for p in pendientes_homologar:
                ws2.append([p['fecha'].strftime('%d/%m/%Y'), p['tipo'], float(p['monto']), p['falta']])
            for col_letra, ancho in [('A', 12), ('B', 30), ('C', 14), ('D', 20)]:
                ws2.column_dimensions[col_letra].width = ancho

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=poliza_ingresos_{fecha_inicio}_{fecha_fin}.xlsx'
        wb.save(response)
        return response

    return render(request, 'gastos/exportar_poliza_ingresos.html', {})



@login_required
def homologar_tipos_cuota(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    # Asegura que exista una fila por cada tipo de cuota posible
    for tipo_valor, _ in Factura.TIPO_CUOTA_CHOICES:
        TipoCuotaHomologacion.objects.get_or_create(empresa=empresa, tipo_cuota=tipo_valor)

    homologaciones = TipoCuotaHomologacion.objects.filter(empresa=empresa).select_related('cuenta_contable')
    cuentas_disponibles = CuentaContable.objects.filter(empresa=empresa, activa=True).order_by('codigo')

    return render(request, 'gastos/homologar_tipos_cuota.html', {
        'homologaciones': homologaciones,
        'cuentas_disponibles': cuentas_disponibles,
    })


@login_required
def asignar_cuenta_contable_cuota(request, homologacion_id):
    empresa = _empresa_del_usuario(request)
    homologacion = get_object_or_404(TipoCuotaHomologacion, pk=homologacion_id, empresa=empresa)

    if request.method == 'POST':
        cuenta_id = request.POST.get('cuenta_contable_id')
        if cuenta_id:
            cuenta = get_object_or_404(CuentaContable, pk=cuenta_id, empresa=empresa)
            homologacion.cuenta_contable = cuenta
        else:
            homologacion.cuenta_contable = None
        homologacion.save(update_fields=['cuenta_contable'])
        messages.success(request, f"'{homologacion.get_tipo_cuota_display()}' homologado correctamente.")

    return redirect('homologar_tipos_cuota')


@login_required
def homologar_tipos_otro_ingreso(request):
    empresa = _empresa_del_usuario(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect('dashboard_inicio')

    tipos_otro_ingreso = TipoOtroIngreso.objects.filter(empresa=empresa).select_related('cuenta_contable').order_by('nombre')
    cuentas_disponibles = CuentaContable.objects.filter(empresa=empresa, activa=True).order_by('codigo')

    return render(request, 'gastos/homologar_tipos_otro_ingreso.html', {
        'tipos_otro_ingreso': tipos_otro_ingreso,
        'cuentas_disponibles': cuentas_disponibles,
    })


@login_required
def asignar_cuenta_contable_otro_ingreso(request, tipo_id):
    empresa = _empresa_del_usuario(request)
    tipo = get_object_or_404(TipoOtroIngreso, pk=tipo_id, empresa=empresa)

    if request.method == 'POST':
        cuenta_id = request.POST.get('cuenta_contable_id')
        if cuenta_id:
            cuenta = get_object_or_404(CuentaContable, pk=cuenta_id, empresa=empresa)
            tipo.cuenta_contable = cuenta
        else:
            tipo.cuenta_contable = None
        tipo.save(update_fields=['cuenta_contable'])
        messages.success(request, f"'{tipo.nombre}' homologado correctamente.")

    return redirect('homologar_tipos_otro_ingreso')


#####################TIPOS DE GASTO, SUBGRUPOS Y GRUPOS###############################################    
@login_required
def subgrupo_gasto_crear(request):
    if request.method == "POST":
        form = SubgrupoGastoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("subgrupos_gasto_lista")
    else:
        form = SubgrupoGastoForm()
    return render(request, "gastos/subgrupo_crear.html", {"form": form})


@login_required
def subgrupo_gasto_eliminar(request, pk):
    subgrupo = get_object_or_404(SubgrupoGasto, pk=pk)
    if request.method == "POST":
        try:
            subgrupo.delete()
            messages.success(request, "Subgrupo de gasto eliminado correctamente.")
            return redirect("subgrupos_gasto_lista")
        except ProtectedError:
            messages.error(
                request,
                "No se puede eliminar este subgrupo porque tiene Tipos de Gasto o Presupuestos relacionados. Elimina o reasigna esos registros primero.",
            )
            return redirect("subgrupos_gasto_lista")
    return render(
        request, "gastos/subgrupo_gasto_confirmar_eliminar.html", {"subgrupo": subgrupo}
    )


@login_required
def subgrupos_gasto_lista(request):
    if request.user.is_superuser:
        subgrupos = SubgrupoGasto.objects.select_related("grupo").order_by(
            "grupo__nombre", "nombre"
        )
    else:
        subgrupos = SubgrupoGasto.objects.select_related("grupo").order_by(
            "grupo__nombre", "nombre"
        )
    return render(request, "gastos/subgrupos_lista.html", {"subgrupos": subgrupos})


@login_required
def tipos_gasto_lista(request):
    if request.user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        if empresa_id:
            tipos = (
                TipoGasto.objects.filter(empresa_id=empresa_id)
                .select_related("subgrupo__grupo")
                .order_by("subgrupo__grupo__nombre")
            )
        else:
            tipos = (
                TipoGasto.objects.select_related("subgrupo__grupo")
                .all()
                .order_by("subgrupo__grupo__nombre")
            )
    else:
        empresa = request.user.perfilusuario.empresa
        tipos = (
            TipoGasto.objects.filter(empresa=empresa)
            .select_related("subgrupo__grupo")
            .order_by("subgrupo__grupo__nombre")
        )
    # filtro buscar
    q = request.GET.get("q")
    if q:
        tipos = tipos.filter(
            Q(nombre__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(subgrupo__nombre__icontains=q)
            | Q(subgrupo__grupo__nombre__icontains=q)
        )

    # Annotate usage count
    tipos_con_uso = tipos.annotate(num_gastos=Count("gasto"))

    # Total types
    total_tipos = tipos.count()

    # Most used
    tipos_mas_usados = tipos_con_uso.order_by("-num_gastos")[:3]

    # Least used (including never used)
    tipos_menos_usados = tipos_con_uso.order_by("num_gastos")[:3]
    tipos_sin_uso = tipos_con_uso.filter(num_gastos=0)

    # Detect possible duplicates (simple normalization)
    nombre_normalizado = {}
    for t in tipos_con_uso:
        norm = unidecode(t.nombre).lower()
        nombre_normalizado.setdefault(norm, []).append(t)
    tipos_duplicados = [v for v in nombre_normalizado.values() if len(v) > 1]

    # paginacion
    paginator = Paginator(tipos, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "gastos/tipos_gasto_lista.html",
        {
            "tipos": page_obj,
            "page_obj": page_obj,
            "q": q,
            "total_tipos": total_tipos,
            "tipos_mas_usados": tipos_mas_usados,
            "tipos_menos_usados": tipos_menos_usados,
            "tipos_sin_uso": tipos_sin_uso,
            "tipos_duplicados": tipos_duplicados,
        },
    )


@login_required
def tipo_gasto_crear(request):
    user = request.user
    perfil = getattr(user, "perfilusuario", None)
    if request.method == "POST":
        post = request.POST.copy()
        if not user.is_superuser and perfil and perfil.empresa:
            post["empresa"] = perfil.empresa.pk
        form = TipoGastoForm(post, user=user)
        if form.is_valid():
            tipo_gasto = form.save(commit=False)
            if not user.is_superuser and perfil and perfil.empresa:
                tipo_gasto.empresa = perfil.empresa
            tipo_gasto.save()
            return redirect("tipos_gasto_lista")
    else:
        form = TipoGastoForm(user=user)
    return render(
        request, "gastos/tipo_gasto_form.html", {"form": form, "modo": "crear"}
    )


@login_required
def tipo_gasto_editar(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    user = request.user
    perfil = getattr(user, "perfilusuario", None)
    if request.method == "POST":
        post = request.POST.copy()
        if not user.is_superuser and perfil and perfil.empresa:
            post["empresa"] = perfil.empresa.pk
        form = TipoGastoForm(post, instance=tipo, user=user)
        if form.is_valid():
            tipo_gasto = form.save(commit=False)
            if not user.is_superuser and perfil and perfil.empresa:
                tipo_gasto.empresa = perfil.empresa
            tipo_gasto.save()
            return redirect("tipos_gasto_lista")
    else:
        form = TipoGastoForm(instance=tipo, user=user)
    return render(
        request,
        "gastos/tipo_gasto_form.html",
        {"form": form, "modo": "editar", "tipo": tipo},
    )


@login_required
@login_required
def tipo_gasto_eliminar(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    if request.method == "POST":
        try:
            tipo.delete()
            messages.success(request, "Tipo de gasto eliminado correctamente.")
            return redirect("tipos_gasto_lista")
        except ProtectedError:
            messages.error(
                request,
                "No se puede eliminar este tipo de gasto porque tiene registros relacionados (por ejemplo, presupuestos o movimientos). "
                "Elimina o reasigna esos registros primero.",
            )
            return redirect("tipos_gasto_lista")
    return render(request, "gastos/tipo_gasto_confirmar_eliminar.html", {"tipo": tipo})


# solicitudes de gasto
# template/gastos/lista.html
@login_required
def gastos_lista(request):
    empresa_id = request.session.get("empresa_id")
    user = request.user

    if user.is_superuser and empresa_id:
        gastos_base = Gasto.objects.filter(empresa_id=empresa_id)
        proveedores = Proveedor.objects.filter(activo=True, empresa_id=empresa_id).order_by("nombre")
        empleados = Empleado.objects.filter(activo=True, empresa_id=empresa_id).order_by("nombre")
        tipos_gasto = TipoGasto.objects.filter(empresa_id=empresa_id).order_by("nombre")
    elif user.is_superuser:
        gastos_base = Gasto.objects.all()
        proveedores = Proveedor.objects.filter(activo=True).order_by("nombre")
        empleados = Empleado.objects.filter(activo=True).order_by("nombre")
        tipos_gasto = TipoGasto.objects.all().order_by("nombre")
    else:
        empresa = user.perfilusuario.empresa
        gastos_base = Gasto.objects.filter(empresa=empresa)
        proveedores = Proveedor.objects.filter(activo=True, empresa=empresa).order_by("nombre")
        empleados = Empleado.objects.filter(activo=True, empresa=empresa).order_by("nombre")
        tipos_gasto = TipoGasto.objects.filter(empresa=empresa).order_by("nombre")

    # Filtros
    proveedor_id = request.GET.get("proveedor")
    empleado_id = request.GET.get("empleado")
    tipo_gasto = request.GET.get("tipo_gasto")
    estatus_filtro = request.GET.get("estatus")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if proveedor_id:
        gastos_base = gastos_base.filter(proveedor_id=proveedor_id)
    if empleado_id:
        gastos_base = gastos_base.filter(empleado_id=empleado_id)
    if tipo_gasto:
        gastos_base = gastos_base.filter(tipo_gasto=tipo_gasto)
    if estatus_filtro:
        gastos_base = gastos_base.filter(estatus=estatus_filtro)
    if fecha_inicio:
        gastos_base = gastos_base.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        gastos_base = gastos_base.filter(fecha__lte=fecha_fin)

    hoy = date.today()
    fecha_inicio_kpi = date(hoy.year, 1, 1)
    fecha_fin_kpi = hoy

    # Si hay filtros de fecha activos, usar esos para los KPIs
    if fecha_inicio:
        fecha_inicio_kpi = date.fromisoformat(fecha_inicio)
    if fecha_fin:
        fecha_fin_kpi = date.fromisoformat(fecha_fin)

    # KPIs filtrados por período
    gastos_kpi = gastos_base.filter(fecha__gte=fecha_inicio_kpi, fecha__lte=fecha_fin_kpi)

    kpi = gastos_kpi.aggregate(
        num_pagadas=Count("id", filter=Q(estatus="pagada")),
        num_pendientes=Count("id", filter=Q(estatus="pendiente")),
        monto_pagadas=Sum(
            Case(When(estatus="pagada", then=F("monto")), default=Value(0), output_field=FloatField())
        ),
        monto_pendientes=Sum(
            Case(When(estatus="pendiente", then=F("monto")), default=Value(0), output_field=FloatField())
        ),
        total_solicitudes=Count("id"),
)

    num_pagadas = kpi["num_pagadas"] or 0
    num_pendientes = kpi["num_pendientes"] or 0
    monto_pagadas = kpi["monto_pagadas"] or 0
    monto_pendientes = kpi["monto_pendientes"] or 0
    total_solicitudes = kpi["total_solicitudes"] or 0
    porc_pagadas = (num_pagadas / total_solicitudes * 100) if total_solicitudes else 0
    porc_pendientes = (num_pendientes / total_solicitudes * 100) if total_solicitudes else 0
    folio_comprobante = request.GET.get("folio_comprobante")
    # # Top 10 proveedores
    # top_proveedores_qs = (
    #     gastos_base.exclude(proveedor__isnull=True)
    #     .values("proveedor__nombre")
    #     .annotate(cantidad=Count("id"))
    #     .order_by("-cantidad")[:10]
    # )
    # top_prov_labels = [item["proveedor__nombre"] for item in top_proveedores_qs]
    # top_prov_data = [item["cantidad"] for item in top_proveedores_qs]

    # Queryset final con anotaciones
    gastos = (
        gastos_base.select_related(
            "empresa", "proveedor", "empleado", "tipo_gasto",
            "tipo_gasto__subgrupo", "tipo_gasto__subgrupo__grupo",
        )
        .annotate(
            total_pagado_ann=Coalesce(Sum('pagos__monto'), Value(0, output_field=DecimalField())),
            saldo_restante_ann=F('monto') - Coalesce(Sum('pagos__monto'), Value(0, output_field=DecimalField())),
        )
        .order_by('-saldo_restante_ann', '-fecha', '-id')  # Ordenar por saldo descendente, luego por fecha descendente y luego por ID descendente
    )

    paginator = Paginator(gastos, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "gastos/lista.html", {
        "gastos": page_obj,
        "proveedores": proveedores,
        "empleados": empleados,
        "tipos_gasto": tipos_gasto,
        "proveedor_id": proveedor_id,
        "empleado_id": empleado_id,
        "tipo_gasto_sel": tipo_gasto,
        "estatus_filtro": estatus_filtro,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "num_pagadas": num_pagadas,
        "num_pendientes": num_pendientes,
        "monto_pagadas": monto_pagadas,
        "monto_pendientes": monto_pendientes,
        "porc_pagadas": porc_pagadas,
        "porc_pendientes": porc_pendientes,
        # "top_prov_labels": top_prov_labels,
        # "top_prov_data": top_prov_data,
        "total_solicitudes": total_solicitudes,
        "fecha_inicio_kpi": fecha_inicio_kpi,
        "fecha_fin_kpi": fecha_fin_kpi,
        "folio_comprobante": folio_comprobante,
        
    })


@login_required
# nueva solicitud pago
def gasto_nuevo(request):
    if request.method == "POST":
        form = GastoForm(request.POST or None, request.FILES, user=request.user)
        if form.is_valid():
            gasto = form.save(commit=False)
            origen = form.cleaned_data["origen_tipo"]
            if origen == "proveedor":
                gasto.empleado = None
            elif origen == "empleado":
                gasto.proveedor = None

            if not request.user.is_superuser:
                gasto.empresa = request.user.perfilusuario.empresa

            # Validar período
            permitido, error = validar_periodo_abierto(None, gasto.fecha, user=request.user)
            if not permitido:
                messages.error(request, error)
                return render(request, "gastos/form.html", {"form": form, "modo": "crear"})    

            gasto.estatus = "pendiente"
            gasto.save()
            messages.success(request, f"Solicitud de gasto folio: {gasto.id}, registrada correctamente.")
            return redirect("gastos_lista")
    else:
        form = GastoForm(
            user=request.user,
        )
        if not request.user.is_superuser:
            form.fields["empresa"].initial = request.user.perfilusuario.empresa
    return render(request, "gastos/form.html", {"form": form, "modo": "crear"})


@login_required
def gasto_editar(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    if (
        not request.user.is_superuser
        and gasto.empresa != request.user.perfilusuario.empresa
    ):
        return redirect("gastos_lista")
    if request.method == "POST":
        form = GastoForm(
            request.POST or None,
            request.FILES,
            instance=gasto,
            user=request.user,
            modo="editar",
        )
        if form.is_valid():
            form.save()
            return redirect("gastos_lista")
    else:
        form = GastoForm(instance=gasto, user=request.user, modo="editar")
    return render(
        request, "gastos/form.html", {"form": form, "modo": "editar", "gasto": gasto}
    )


@login_required
def gasto_eliminar(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    if request.method == "POST":
        gasto.delete()
        return redirect("gastos_lista")
    return render(request, "gastos/confirmar_eliminar.html", {"gasto": gasto})

#pago de solicitudes
@login_required
def registrar_pago_gasto(request, gasto_id):
    gasto = get_object_or_404(Gasto, pk=gasto_id)
    pagos = gasto.pagos.all()
    saldo_restante = gasto.monto - sum([p.monto for p in pagos])
    empresa = gasto.empresa

    if gasto.estatus == "pagada":
        messages.info(
            request, "Este gasto ya está pagado. No se pueden registrar más pagos."
        )
        return redirect("gastos_lista")

    if request.method == "POST":
        form = PagoGastoForm(request.POST, request.FILES, empresa=empresa)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.gasto = gasto
            pago.registrado_por = request.user
            # Validar período cerrado
            fecha_pago = pago.fecha_pago
            cuenta_bancaria = pago.cuenta_bancaria
            if fecha_pago and cuenta_bancaria:
                periodo_valido, error_periodo = validar_periodo_abierto(cuenta_bancaria, fecha_pago)
                if not periodo_valido:
                    form.add_error(None, error_periodo)
                    return render(request, 'gastos/registrar_pago.html', {
                        'form': form,
                        'gasto': gasto,
                        'saldo_restante': saldo_restante,
                    })
        
            if pago.monto > saldo_restante:
                form.add_error(
                    "monto",
                    f"El monto excede el saldo pendiente (${saldo_restante:.2f})",
                )
            else:
                pago.save()
                gasto.actualizar_estatus()
                messages.success(
                    request,
                    f"Pago registrado a la solicitud: {gasto.id}. Saldo restante: ${gasto.saldo_restante:.2f}",
                )
                return redirect("gastos_lista")
    else:
        form = PagoGastoForm(empresa=empresa)

    return render(
        request,
        "gastos/registrar_pago.html",
        {"form": form, "gasto": gasto, "saldo_restante": saldo_restante},
    )


@login_required
def gasto_detalle(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    pagos = gasto.pagos.all().order_by("fecha_pago")
    reversados_ids = set()
    folio_comprobante = gasto.folio_comprobante

    for pago in pagos:
        if (
            pago.monto < 0
            and pago.referencia
            and "Reverso de pago ID" in pago.referencia
        ):
            # Extrae el ID del pago original
            try:
                id_original = int(
                    pago.referencia.split("Reverso de pago ID")[1].split(".")[0].strip()
                )
                reversados_ids.add(id_original)
            except Exception:
                pass
    return render(
        request,
        "gastos/gasto_detalle.html",
        {
            "gasto": gasto,
            "pagos": pagos,
            "reversados_ids": reversados_ids,
            "folio_comprobante": folio_comprobante,
        },
    )


# reversa pago gastos
@login_required
def reversa_pago_gasto(request, pago_id, gasto_id):
    pago = get_object_or_404(PagoGasto, id=pago_id)
    gasto = get_object_or_404(Gasto, id=gasto_id)
    next_url = request.GET.get("next")

     # Validar que la fecha del pago original esté en un período permitido
    if pago.cuenta_bancaria:
        error_periodo = validar_periodo_abierto(pago.cuenta_bancaria, pago.fecha_pago)
        if error_periodo:
            messages.error(request, f"No se puede reversar: {error_periodo}")
            return redirect(next_url or "gasto_detalle", pk=gasto.id)
        
    if request.method == "POST":
        form = MotivoReversaPagoForm(request.POST)
        if form.is_valid():
            motivo = form.cleaned_data["motivo"]
            motivo_display = dict(MotivoReversaPagoForm.MOTIVOS_REVERSA)[motivo]
            PagoGasto.objects.create(
                gasto=gasto,
                monto=-pago.monto,
                forma_pago=pago.forma_pago,
                fecha_pago=pago.fecha_pago,
                cuenta_bancaria=pago.cuenta_bancaria,
                referencia=f"Reverso de pago ID {pago.id}. Motivo: {motivo_display}",
                registrado_por=request.user,
            )
            gasto.actualizar_estatus()
            gasto.save()
            messages.success(request, "Pago de gasto reversado correctamente.")
            return redirect(next_url or "gasto_detalle", pk=gasto.id)
    else:
        form = MotivoReversaPagoForm()

    return render(
        request,
        "gastos/reversa_pago_gasto.html",
        {
            "form": form,
            "pago": pago,
            "gasto": gasto,
            "next": next_url,
        },
    )


@login_required
# template/gastos/reporte_pagos.html actualizado 31/07/26
def reporte_pagos_gastos(request):
    es_super = request.user.is_superuser
    pagos = PagoGasto.objects.select_related(
        "gasto", "gasto__empresa", "gasto__proveedor", "gasto__empleado"
    )

    empresas = (
        Empresa.objects.all()
        if es_super
        else Empresa.objects.filter(pk=request.user.perfilusuario.empresa.id)
    )
    empresa_id = request.GET.get("empresa")
    proveedor_id = request.GET.get("proveedor")
    empleado_id = request.GET.get("empleado")
    forma_pago = request.GET.get("forma_pago")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    cuenta_bancaria = request.GET.get("cuenta_bancaria")

     # Detectar si hay filtros aplicados
    filtros_aplicados = any([
        empresa_id,
        proveedor_id,
        empleado_id,
        forma_pago,
        fecha_inicio,
        fecha_fin,
        cuenta_bancaria,
    ])

    # Por — período actual por defecto (enero a hoy):
    if not filtros_aplicados:
        fecha_inicio = datetime.now().date().replace(month=1, day=1).strftime("%Y-%m-%d")
        fecha_fin = datetime.now().date().strftime("%Y-%m-%d")

    # Cuentas bancarias filtradas por empresa
    if not es_super:
        cuentas_bancarias = CuentaBancaria.objects.filter(
            empresa=request.user.perfilusuario.empresa
        )
    else:
        if empresa_id:
            cuentas_bancarias = CuentaBancaria.objects.filter(empresa_id=empresa_id)
        else:
            cuentas_bancarias = CuentaBancaria.objects.all()

    if not es_super:
        pagos = pagos.filter(gasto__empresa=request.user.perfilusuario.empresa)
        proveedores = Proveedor.objects.filter(
            empresa=request.user.perfilusuario.empresa
        ).order_by("nombre")
        empleados = Empleado.objects.filter(
            empresa=request.user.perfilusuario.empresa
        ).order_by("nombre")
    else:
        if empresa_id:
            pagos = pagos.filter(gasto__empresa_id=empresa_id)
            proveedores = Proveedor.objects.filter(empresa_id=empresa_id).order_by(
                "nombre"
            )
            empleados = Empleado.objects.filter(empresa_id=empresa_id).order_by(
                "nombre"
            )
        else:
            proveedores = Proveedor.objects.all().order_by("nombre")
            empleados = Empleado.objects.all().order_by("nombre")
            
    pagos_base = pagos

    # Solo aplica el filtro si el parámetro es numérico
    if proveedor_id and proveedor_id.isdigit():
        pagos = pagos.filter(gasto__proveedor_id=proveedor_id)
    if empleado_id and empleado_id.isdigit():
        pagos = pagos.filter(gasto__empleado_id=empleado_id)
    if forma_pago:
        pagos = pagos.filter(forma_pago=forma_pago)


    # Corregido -- inicializadas en None para que nunca truene con UnboundLocalError 31/07/26
    fecha_inicio_dt = None
    fecha_fin_dt = None
    if fecha_inicio:
        fecha_inicio_dt = parse_date(fecha_inicio)
    if fecha_fin:
        fecha_fin_dt = parse_date(fecha_fin)

    if fecha_inicio_dt:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio_dt)
    if fecha_fin_dt:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin_dt)

    # Filtro de cuenta bancaria
    if cuenta_bancaria and cuenta_bancaria.isdigit():
        pagos = pagos.filter(cuenta_bancaria_id=cuenta_bancaria)
    

    #actualizar KPIs y totales según el período de referencia 31/07/26
    hoy = datetime.now().date()
    if fecha_fin:
        try:
            fecha_referencia = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except Exception:
            fecha_referencia = hoy
    else:
        fecha_referencia = hoy
    anio_actual = fecha_referencia.year
    anio_anterior = anio_actual - 1
    mes_actual = fecha_referencia.month
    mes_anterior = mes_actual - 1 if mes_actual > 1 else 12
    anio_mes_anterior = anio_actual if mes_actual > 1 else anio_actual - 1

    total_pagos_acumulado = pagos.aggregate(total=Sum("monto"))["total"] or 0

    pagos_anio_actual = pagos_base.filter(fecha_pago__year=anio_actual).aggregate(total=Sum("monto"))["total"] or 0
    pagos_anio_anterior = pagos_base.filter(fecha_pago__year=anio_anterior).aggregate(total=Sum("monto"))["total"] or 0
    pagos_mes_actual = pagos_base.filter(fecha_pago__year=anio_actual, fecha_pago__month=mes_actual).aggregate(total=Sum("monto"))["total"] or 0
    pagos_mes_anterior = pagos_base.filter(fecha_pago__year=anio_mes_anterior, fecha_pago__month=mes_anterior).aggregate(total=Sum("monto"))["total"] or 0

    pagos_gastos_chart = pagos

    # KPIs por tipo de gasto (top 10, etiqueta: subgrupo - tipo)
    tipo_dict = (
        pagos_gastos_chart.values("gasto__tipo_gasto__subgrupo__nombre", "gasto__tipo_gasto__nombre")
        .annotate(total=Sum("monto"))
        .order_by("-total")
    )
    top_tipos = []
    for t in tipo_dict:
        subgrupo = t["gasto__tipo_gasto__subgrupo__nombre"] or ""
        tipo = t["gasto__tipo_gasto__nombre"] or ""
        etiqueta = f"{subgrupo} - {tipo}" if subgrupo else tipo
        top_tipos.append((etiqueta, t["total"]))
    top_tipos = top_tipos[:10]
    tipo_labels = [k for k, v in top_tipos]
    tipo_data = [v for k, v in top_tipos]
    pagos_por_tipo = [{"nombre": k, "total": v} for k, v in top_tipos]

    # KPIs por forma de pago
    forma_dict = (
        pagos_gastos_chart.values("forma_pago").annotate(total=Sum("monto")).order_by("-total")
    )
    pagos_por_forma = []
    forma_labels = []
    forma_data = []
    for f in forma_dict:
        display = dict(PagoGasto._meta.get_field("forma_pago").choices).get(
            f["forma_pago"], f["forma_pago"]
        )
        pagos_por_forma.append({"nombre": display, "total": f["total"]})
        forma_labels.append(display)
        forma_data.append(f["total"])

    def variacion(actual, anterior):
        if anterior == 0:
            return 100 if actual > 0 else 0
        return ((actual - anterior) / anterior) * 100

    var_anio = variacion(pagos_anio_actual, pagos_anio_anterior)
    var_mes = variacion(pagos_mes_actual, pagos_mes_anterior)

    FORMAS_PAGO = PagoGasto._meta.get_field("forma_pago").choices

    pagos = pagos.order_by("-fecha_pago")
    paginator = Paginator(pagos, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "gastos/reporte_pagos.html",
        {
            "pagos": page_obj,
            "empresas": empresas,
            "empresa_id": empresa_id,
            "proveedores": proveedores,
            "empleados": empleados,
            "forma_pago_actual": forma_pago,
            "total_pagos_acumulados": total_pagos_acumulado,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "proveedor_id": proveedor_id,
            "empleado_id": empleado_id,
            "formas_pago": FORMAS_PAGO,
            "pagos_anio_actual": pagos_anio_actual,
            "pagos_anio_anterior": pagos_anio_anterior,
            "pagos_mes_actual": pagos_mes_actual,
            "pagos_mes_anterior": pagos_mes_anterior,
            "pagos_por_tipo": pagos_por_tipo,
            "pagos_por_forma": pagos_por_forma,
            "var_anio": var_anio,
            "var_mes": var_mes,
            "tipo_labels": tipo_labels,
            "tipo_data": tipo_data,
            "forma_labels": forma_labels,
            "forma_data": forma_data,
            "cuentas_bancarias": cuentas_bancarias,
            "cuenta_bancaria_id": cuenta_bancaria,
        },
    )

# exportar pagos de gastos a Excel
@login_required
#actualizado 31/07/26
def exportar_reporte_pagos_gastos_excel(request):
    es_super = request.user.is_superuser
    pagos = PagoGasto.objects.select_related(
        "gasto", "gasto__empresa", "gasto__proveedor", "gasto__empleado"
    )

    empresa_id = request.GET.get("empresa")
    proveedor_id = request.GET.get("proveedor")
    empleado_id = request.GET.get("empleado")
    forma_pago = request.GET.get("forma_pago")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    cuenta_bancaria = request.GET.get("cuenta_bancaria")

    # NUEVO -- mismo default de fechas que la pantalla (enero a hoy si no hay filtros)
    filtros_aplicados = any([
        empresa_id, proveedor_id, empleado_id, forma_pago,
        fecha_inicio, fecha_fin, cuenta_bancaria,
    ])
    if not filtros_aplicados:
        fecha_inicio = datetime.now().date().replace(month=1, day=1).strftime("%Y-%m-%d")
        fecha_fin = datetime.now().date().strftime("%Y-%m-%d")

    if not es_super:
        pagos = pagos.filter(gasto__empresa=request.user.perfilusuario.empresa)
    else:
        if empresa_id:
            pagos = pagos.filter(gasto__empresa_id=empresa_id)

    if proveedor_id and proveedor_id.isdigit():
        pagos = pagos.filter(gasto__proveedor_id=proveedor_id)
    if empleado_id and empleado_id.isdigit():
        pagos = pagos.filter(gasto__empleado_id=empleado_id)
    if forma_pago:
        pagos = pagos.filter(forma_pago=forma_pago)
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=parse_date(fecha_inicio))
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=parse_date(fecha_fin))

    if cuenta_bancaria and cuenta_bancaria.isdigit():
        pagos = pagos.filter(cuenta_bancaria_id=cuenta_bancaria)

    pagos_list = list(pagos)
    for p in pagos_list:
        p.tipo_pago = "normal"
        p.fecha = p.fecha_pago

    pagos_list.sort(key=lambda x: x.fecha, reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos de Gastos"
    ws.append([
        "Folio", "Fecha pago", "Empresa", "Proveedor/Empleado", "Tipo gasto",
        "Concepto", "Forma de pago", "Monto", "Cuenta Bancaria", "Numero Cta", "Estatus",
    ])

    total_general = 0
    for pago in pagos_list:
        gasto = pago.gasto
        origen = (
            gasto.proveedor.nombre if gasto.proveedor
            else (gasto.empleado.nombre if gasto.empleado else "")
        )
        ws.append([
            gasto.id,
            pago.fecha_pago if pago.fecha_pago else "",
            gasto.empresa.nombre if gasto.empresa else "",
            origen,
            gasto.tipo_gasto.nombre if gasto.tipo_gasto else "",
            gasto.descripcion,
            pago.get_forma_pago_display(),
            float(pago.monto),
            pago.cuenta_bancaria.banco if pago.cuenta_bancaria else "",
            pago.cuenta_bancaria.numero_cuenta if pago.cuenta_bancaria else "",
            gasto.estatus,
        ])
        total_general += float(pago.monto)

    # NUEVO -- fila de total, para comparar directo contra el KPI de pantalla
    ws.append(["", "", "", "", "", "", "", "", "", "TOTAL:", total_general])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = "reporte_pagos_gastos.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response

@login_required
#egresos 
def grafico_gastos_mensual(request):
    anio = int(request.GET.get('anio', datetime.now().year))
    mes = request.GET.get('mes')
    mes_int = int(mes) if mes and str(mes).isdigit() else None
    es_super = request.user.is_superuser
    tipo_gasto=request.GET.get('tipo_gasto','todos')

    # Filtro por empresa si no es superusuario
    if es_super:
        pagos = PagoGasto.objects.all()
        gastos_caja_chica = GastoCajaChica.objects.all()
        vales_caja_chica = ValeCaja.objects.all()
        presup_qs = Presupuesto.objects.filter(anio=anio)
    else:
        empresa = request.user.perfilusuario.empresa
        pagos = PagoGasto.objects.filter(gasto__empresa=empresa)
        gastos_caja_chica = GastoCajaChica.objects.filter(fondeo__empresa=empresa)
        vales_caja_chica = ValeCaja.objects.filter(fondeo__empresa=empresa)
        presup_qs = Presupuesto.objects.filter(anio=anio, empresa=empresa)

    # Aplica el filtro por tipo de gasto
    if tipo_gasto == 'pagos':
        gastos_caja_chica = GastoCajaChica.objects.none()
        vales_caja_chica = ValeCaja.objects.none()
    elif tipo_gasto == 'caja':
        pagos = PagoGasto.objects.none()
        vales_caja_chica = ValeCaja.objects.none()
    elif tipo_gasto == 'vales':
        pagos = PagoGasto.objects.none()
        gastos_caja_chica = GastoCajaChica.objects.none()

    # Presupuesto mensual del año seleccionado
    presup_dict = {}
    for p in presup_qs:
        mes_p = int(p.mes)
        presup_dict[mes_p] = presup_dict.get(mes_p, 0.0) + float(p.monto or 0)

    # Filtro por año y mes
    pagos = pagos.filter(fecha_pago__year=anio)
    gastos_caja_chica = gastos_caja_chica.filter(fecha__year=anio)
    vales_caja_chica = vales_caja_chica.filter(fecha__year=anio)

    if mes_int:
        pagos = pagos.filter(fecha_pago__month=mes_int)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__month=mes_int)
        vales_caja_chica = vales_caja_chica.filter(fecha__month=mes_int)

    # Agrupa por mes
    pagos_por_mes = pagos.annotate(mes=TruncMonth('fecha_pago')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    gastos_caja_chica_por_mes = gastos_caja_chica.annotate(mes=TruncMonth('fecha')).values('mes').annotate(total=Sum('importe')).order_by('mes')
    vales_caja_chica_por_mes = vales_caja_chica.annotate(mes=TruncMonth('fecha')).values('mes').annotate(total=Sum('importe')).order_by('mes')

    # Etiquetas y datos para el gráfico
    todos_los_meses = sorted(
    set(
        [x['mes'] for x in pagos_por_mes if x['mes'] is not None] +
        [x['mes'] for x in gastos_caja_chica_por_mes if x['mes'] is not None] +
        [x['mes'] for x in vales_caja_chica_por_mes if x['mes'] is not None]
    )
)
    labels_meses = [DateFormat(m).format('F Y') for m in todos_los_meses]

    #dict para acceso rápido a totales por mes
    pagos_dict = {x["mes"]: float(x["total"] or 0) for x in pagos_por_mes if x["mes"] is not None}
    caja_dict = {x["mes"]: float(x["total"] or 0) for x in gastos_caja_chica_por_mes if x["mes"] is not None}
    vales_dict = {x["mes"]: float(x["total"] or 0) for x in vales_caja_chica_por_mes if x["mes"] is not None}

   
    data_gastos = [pagos_dict.get(m, 0.0) for m in todos_los_meses]
    data_caja = [caja_dict.get(m, 0.0) for m in todos_los_meses]
    data_vales = [vales_dict.get(m, 0.0) for m in todos_los_meses]
    data_presupuesto = [float(presup_dict.get(m.month, 0.0)) for m in todos_los_meses]

    # Suma total de egresos por mes (pagos + caja + vales)
    totales_mes = [data_gastos[i] + data_caja[i] + data_vales[i] for i in range(len(data_gastos))]

    if not data_presupuesto:
        data_presupuesto = [0 for _ in labels_meses]

    # Si no hay datos, muestra un mes vacío
    if not labels_meses:
        labels_meses = ["Sin datos"]
        data_gastos = [0.0]
        data_caja = [0.0]
        data_vales = [0.0]    
        data_presupuesto = [0.0]

    meses_lista = [
    {'num': 1, 'nombre': 'Enero'},
    {'num': 2, 'nombre': 'Febrero'},
    {'num': 3, 'nombre': 'Marzo'},
    {'num': 4, 'nombre': 'Abril'},
    {'num': 5, 'nombre': 'Mayo'},
    {'num': 6, 'nombre': 'Junio'},
    {'num': 7, 'nombre': 'Julio'},
    {'num': 8, 'nombre': 'Agosto'},
    {'num': 9, 'nombre': 'Septiembre'},
    {'num': 10, 'nombre': 'Octubre'},
    {'num': 11, 'nombre': 'Noviembre'},
    {'num': 12, 'nombre': 'Diciembre'},
]
    return render(request, 'gastos/grafico_gastos_mensual.html', {
        'labels_meses': json.dumps(labels_meses),
        #'data_total_egresos': json.dumps(data_total_egresos),
        'data_gastos': json.dumps(data_gastos),
        'data_caja': json.dumps(data_caja),
        'data_vales': json.dumps(data_vales),
        'data_presupuesto': json.dumps(data_presupuesto),
        'anio': anio,
        'mes': mes,
        'meses_lista': meses_lista, 
        'totales_mes': json.dumps(totales_mes),
        'tipo_gasto': tipo_gasto,
    })

@login_required
#grafico gastos comparativo anual
def grafico_gastos_anual(request):
    empresa = request.user.perfilusuario.empresa
    pagos = (
        PagoGasto.objects.filter(gasto__empresa=empresa)
        .annotate(year=TruncYear('fecha_pago'))
        .values('year')
        .annotate(total=Sum('monto'))
        .order_by('year')
    )

    # Gastos caja chica agrupados por año
    gastos_caja = (
        GastoCajaChica.objects.filter(fondeo__empresa=empresa)
        .annotate(year=TruncYear('fecha'))
        .values('year')
        .annotate(total=Sum('importe'))
        .order_by('year')
    )

    # Vales caja agrupados por año
    vales_caja = (
        ValeCaja.objects.filter(fondeo__empresa=empresa)
        .annotate(year=TruncYear('fecha'))
        .values('year')
        .annotate(total=Sum('importe'))
        .order_by('year')
    )

    # Presupuesto agrupado por año
    presupuestos = (
        Presupuesto.objects.filter(empresa=empresa)
        .values('anio')
        .annotate(total=Sum('monto'))
        .order_by('anio')
    )

    # Unificar años
    years = sorted(
        set(
            [p['year'].year for p in pagos] +
            [g['year'].year for g in gastos_caja] +
            [v['year'].year for v in vales_caja] +
            [b['anio'] for b in presupuestos]
        )
    )

    pagos_dict = {p['year'].year: float(p['total']) for p in pagos}
    gastos_caja_dict = {g['year'].year: float(g['total']) for g in gastos_caja}
    vales_caja_dict = {v['year'].year: float(v['total']) for v in vales_caja}
    presupuestos_dict = {b['anio']: float(b['total']) for b in presupuestos}

    pagos_data = [pagos_dict.get(y, 0) for y in years]
    gastos_caja_data = [gastos_caja_dict.get(y, 0) for y in years]
    vales_caja_data = [vales_caja_dict.get(y, 0) for y in years]
    presupuestos_data = [presupuestos_dict.get(y, 0) for y in years]

    totales_anio = [pagos_data[i] + gastos_caja_data[i] + vales_caja_data[i] for i in range(len(years))]
    
    context = {
        'years': json.dumps(years),
        'pagos_data': json.dumps(pagos_data),
        'gastos_caja_data': json.dumps(gastos_caja_data),
        'vales_caja_data': json.dumps(vales_caja_data),
        'presupuestos_data': json.dumps(presupuestos_data),
        'totales_anio': json.dumps(totales_anio),
}
    return render(request, 'gastos/grafico_gastos_anual.html', context)


def buscar_por_id_o_nombre(modelo, valor, campo="nombre"):
    """Busca por ID, si falla busca por nombre (sin acentos, insensible a mayúsculas). Reporta conflicto si hay varias."""
    if not valor:
        return None
    val = str(valor).strip().replace(",", "")
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        # Lista de coincidencias insensibles a acentos y mayúsculas
        candidatos = [
            obj
            for obj in todos
            if unidecode(val).lower() in unidecode(str(getattr(obj, campo))).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join(
                [f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos]
            )
            raise Exception(
                f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}"
            )
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")


def normaliza_texto(texto):
    if not texto:
        return ""
    # texto = texto.upper()
    texto = texto
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join([c for c in texto if not unicodedata.combining(c)])
    return texto


@login_required
def carga_masiva_gastos(request):
    if request.method == "POST":
        form = GastosCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            COLUMNAS_ESPERADAS = 13  # empresa, proveedor, empleado, rfc_empleado, grupo, subgrupo, tipo_gasto, monto, descripcion, fecha, observaciones, retencion_iva, retencion_isr
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(
                        f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})"
                    )
                    continue
                (
                    empresa_val,
                    proveedor_val,
                    empleado_val,
                    rfc_empleado,
                    grupo_val,
                    subgrupo_val,
                    tipo_gasto_val,
                    monto,
                    descripcion,
                    fecha,
                    observaciones,
                    retencion_iva,
                    retencion_isr,
                ) = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(
                            f"Fila {i}: No se encontró la empresa '{empresa_val}'"
                        )
                        continue

                    proveedor = None
                    if proveedor_val:
                        proveedor, _ = Proveedor.objects.get_or_create(
                            nombre=proveedor_val, empresa=empresa
                        )

                    empleado = None
                    if rfc_empleado:
                        empleado, _ = Empleado.objects.get_or_create(
                            rfc=rfc_empleado,
                            defaults={"nombre": empleado_val, "empresa": empresa},
                        )
                    elif empleado_val:
                        empleado, _ = Empleado.objects.get_or_create(
                            nombre=empleado_val, empresa=empresa
                        )

                    # Validar que al menos uno esté presente
                    if not proveedor and not empleado:
                        errores.append(
                            f"Fila {i}: Debe especificar proveedor o empleado."
                        )
                        continue

                    # Validar que el grupo existe
                    grupo_inst = None
                    if grupo_val:
                        grupo_nombre = normaliza_texto(grupo_val)
                        try:
                            grupo_inst = GrupoGasto.objects.get(nombre=grupo_nombre)
                        except GrupoGasto.DoesNotExist:
                            errores.append(
                                f"Fila {i}: El grupo '{grupo_nombre}' no existe en el catálogo."
                            )
                            continue
                    else:
                        errores.append(f"Fila {i}: El grupo es obligatorio.")
                        continue

                    # Validar que el subgrupo existe para ese grupo
                    subgrupo_inst = None
                    if subgrupo_val:
                        subgrupo_nombre = normaliza_texto(subgrupo_val)
                        try:
                            subgrupo_inst = SubgrupoGasto.objects.get(
                                nombre=subgrupo_nombre, grupo=grupo_inst
                            )
                        except SubgrupoGasto.DoesNotExist:
                            errores.append(
                                f"Fila {i}: El subgrupo '{subgrupo_nombre}' no existe en el grupo '{grupo_nombre}'."
                            )
                            continue
                    else:
                        errores.append(f"Fila {i}: El subgrupo es obligatorio.")
                        continue

                    # TIPO DE GASTO por empresa y subgrupo (se puede crear si no existe)
                    tipo_gasto_inst = None
                    if tipo_gasto_val:
                        tipo_gasto_nombre = normaliza_texto(tipo_gasto_val)
                        tipo_gasto_inst = TipoGasto.objects.filter(
                            nombre=tipo_gasto_nombre,
                            subgrupo=subgrupo_inst,
                            empresa=empresa,
                        ).first()
                        if not tipo_gasto_inst:
                            tipo_gasto_inst = TipoGasto.objects.create(
                                nombre=tipo_gasto_nombre,
                                subgrupo=subgrupo_inst,
                                empresa=empresa,
                            )
                    else:
                        errores.append(f"Fila {i}: El tipo de gasto es obligatorio.")
                        continue

                    try:
                        monto_decimal = Decimal(monto)
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(
                            f"Fila {i}: El valor de monto '{monto}' no es válido."
                        )
                        continue

                    try:
                        retencion_iva_decimal = (
                            Decimal(retencion_iva) if retencion_iva else 0
                        )
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(
                            f"Fila {i}: El valor de retención IVA '{retencion_iva}' no es válido."
                        )
                        continue

                    try:
                        retencion_isr_decimal = (
                            Decimal(retencion_isr) if retencion_isr else 0
                        )
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(
                            f"Fila {i}: El valor de retención ISR '{retencion_isr}' no es válido."
                        )
                        continue

                    # Conversión de fecha si es string
                    from datetime import datetime

                    if isinstance(fecha, str):
                        try:
                            fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                        except Exception:
                            errores.append(
                                f"Fila {i}: El formato de fecha '{fecha}' no es válido (debe ser YYYY-MM-DD)."
                            )
                            continue

                    gasto = Gasto.objects.create(
                        empresa=empresa,
                        proveedor=proveedor,
                        empleado=empleado,
                        tipo_gasto=tipo_gasto_inst,
                        monto=monto_decimal,
                        descripcion=descripcion or "",
                        fecha=fecha,
                        observaciones=observaciones or "",
                        retencion_iva=retencion_iva_decimal,
                        retencion_isr=retencion_isr_decimal,
                        estatus="pagada",
                    )

                    # Registrar el pago automáticamente
                    PagoGasto.objects.create(
                        gasto=gasto,
                        fecha_pago=fecha,
                        monto=monto_decimal,
                        forma_pago="transferencia",
                        referencia="Carga masiva",
                        registrado_por=request.user
                        if request.user.is_authenticated
                        else None,
                    )

                    exitos += 1
                except Exception as e:
                    import traceback

                    errores.append(
                        f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}"
                    )

            if exitos:
                messages.success(request, f"¡{exitos} gastos cargados exitosamente!")
            if errores:
                messages.error(
                    request, "Algunos gastos no se cargaron:<br>" + "<br>".join(errores)
                )
            return redirect("carga_masiva_gastos")
    else:
        form = GastosCargaMasivaForm()
    return render(request, "gastos/carga_masiva_gastos.html", {"form": form})


def descargar_plantilla_gastos(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Gastos"

    # Encabezados según el formato de carga masiva
    encabezados = [
        "empresa",
        "proveedor",
        "empleado",
        "rfc_empleado",
        "grupo",
        "subgrupo",
        "tipo_gasto",
        "monto",
        "descripcion",
        "fecha",
        "observaciones",
        "retencion_iva",
        "retencion_isr",
    ]
    ws.append(encabezados)

    # Fila de ejemplo
    ws.append(
        [
            "EMPRESA A.C.",
            "Proveedor Ejemplo",
            "",
            "",
            "Gastos Administracion",
            "Papelería",
            "Copias",
            "1200.50",
            "Compra de hojas",
            "2025-06-19",
            "Carga inicial",
            "0.00",
            "0.00",
        ]
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=plantilla_gastos.xlsx"
    return response


@login_required
def exportar_gastos_lista_excel(request):
    proveedor_id = request.GET.get("proveedor")
    empleado_id = request.GET.get("empleado")
    tipo_gasto = request.GET.get("tipo_gasto")

    if request.user.is_superuser:
        gastos = (
            Gasto.objects.all()
            .select_related("empresa", "proveedor", "empleado", "tipo_gasto")
            .order_by("-fecha")
        )
    else:
        gastos = Gasto.objects.filter(
            empresa=request.user.perfilusuario.empresa
        ).order_by("-fecha")

    if proveedor_id:
        gastos = gastos.filter(proveedor_id=proveedor_id)
    if empleado_id:
        gastos = gastos.filter(empleado_id=empleado_id)
    if tipo_gasto:
        gastos = gastos.filter(tipo_gasto=tipo_gasto)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    ws.append(
        [
            "Folio",
            "Fecha",
            "Empresa",
            "Proveedor",
            "Empleado",
            "Tipo de Gasto",
            "Monto",
            "Saldo",
            "Descripción",
            "Estatus",
            "Observaciones",
        ]
    )

    for gasto in gastos:
        ws.append(
            [
                gasto.id,
                gasto.fecha if gasto.fecha else "",
                gasto.empresa.nombre if gasto.empresa else "",
                gasto.proveedor.nombre if gasto.proveedor else "",
                gasto.empleado.nombre if gasto.empleado else "",
                gasto.tipo_gasto.nombre if gasto.tipo_gasto else "",
                float(gasto.monto),
                float(gasto.saldo_restante)
                if gasto.saldo_restante is not None
                else 0.0,
                gasto.descripcion or "",
                gasto.estatus,
                gasto.observaciones or "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=gastos_lista.xlsx"
    wb.save(response)
    return response


@login_required
def recibo_gasto(request, gasto_id):
    gasto = get_object_or_404(Gasto, pk=gasto_id)
    monto_letra = num2words(
        gasto.monto, lang="es", to="currency", currency="MXN"
    ).capitalize()
    proveedor = gasto.proveedor
    empleado = gasto.empleado
    empresa = gasto.empresa
    return render(
        request,
        "gastos/recibo_gasto.html",
        {
            "gasto": gasto,
            "proveedor": proveedor,
            "empleado": empleado,
            "empresa": empresa,
            "monto_letra": monto_letra,
        },
    )


# consulta retenciones de gastos
@login_required
def reporte_retenciones_gastos(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    # Obtén la empresa del usuario logueado correctamente
    perfil = getattr(request.user, "perfilusuario", None)
    empresa = perfil.empresa if perfil else None
    if not empresa:
        return render(
            request,
            "gastos/reporte_retenciones.html",
            {
                "gastos": [],
                "fecha_inicio": fecha_inicio or "",
                "fecha_fin": fecha_fin or "",
                "error": "No se pudo determinar la empresa del usuario.",
            },
        )

    gastos = Gasto.objects.select_related(
        "proveedor", "empleado", "tipo_gasto", "tipo_gasto__subgrupo"
    ).filter(empresa=empresa)

    if fecha_inicio:
        gastos = gastos.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        gastos = gastos.filter(fecha__lte=parse_date(fecha_fin))

    gastos = gastos.filter(Q(retencion_isr__gt=0) | Q(retencion_iva__gt=0)).order_by(
        "-fecha"
    )

    data = []
    for g in gastos:
        if g.proveedor:
            nombre = g.proveedor.nombre
        elif g.empleado:
            nombre = g.empleado.nombre
        else:
            nombre = ""
        data.append(
            {
                "persona": nombre,
                "subgrupo": g.tipo_gasto.subgrupo.nombre
                if g.tipo_gasto and g.tipo_gasto.subgrupo
                else "",
                "tipo": g.tipo_gasto.nombre if g.tipo_gasto else "",
                "fecha": g.fecha,
                "retencion_isr": getattr(g, "retencion_isr", 0),
                "retencion_iva": getattr(g, "retencion_iva", 0),
            }
        )

    return render(
        request,
        "gastos/reporte_retenciones.html",
        {
            "gastos": data,
            "fecha_inicio": fecha_inicio or "",
            "fecha_fin": fecha_fin or "",
            "error": None,
        },
    )


# descagar reporte de retenciones de gastos en Excel
@login_required
def descargar_reporte_retenciones_gastos(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    gastos = Gasto.objects.select_related(
        "proveedor", "empleado", "tipo_gasto", "tipo_gasto__subgrupo"
    ).all()

    # Filtro por fechas si se proporcionan
    if fecha_inicio:
        gastos = gastos.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        gastos = gastos.filter(fecha__lte=parse_date(fecha_fin))

    # Solo los que tengan retención ISR o IVA mayor a cero
    gastos = gastos.filter(Q(retencion_isr__gt=0) | Q(retencion_iva__gt=0)).order_by(
        "-fecha"
    )

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Retenciones Gastos"

    # Encabezados
    ws.append(
        [
            "Persona",
            "Subgrupo",
            "Tipo de Gasto",
            "Fecha",
            "Retención ISR",
            "Retención IVA",
        ]
    )

    for g in gastos:
        if g.proveedor:
            nombre = g.proveedor.nombre
        elif g.empleado:
            nombre = g.empleado.nombre
        else:
            nombre = ""
        ws.append(
            [
                nombre,
                g.tipo_gasto.subgrupo.nombre
                if g.tipo_gasto and g.tipo_gasto.subgrupo
                else "",
                g.tipo_gasto.nombre if g.tipo_gasto else "",
                g.fecha if g.fecha else "",
                float(getattr(g, "retencion_isr", 0)),
                float(getattr(g, "retencion_iva", 0)),
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = "reporte_retenciones_gastos.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response
