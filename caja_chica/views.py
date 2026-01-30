from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from num2words import num2words
from django.utils import timezone
from empleados.models import Empleado, Incidencia
from proveedores.models import Proveedor
from .models import FondeoCajaChica, GastoCajaChica, ValeCaja
from .forms import ComprobarValeForm, FondeoCajaChicaForm, GastoCajaChicaForm, ValeCajaForm
from gastos.models import TipoGasto
from django.db import transaction
from decimal import Decimal, InvalidOperation
from django.db.models import Sum
from caja_chica import models
from openpyxl import Workbook
from django.core.paginator import Paginator

def imprimir_vale_caja(request, vale_id):
    vale = get_object_or_404(ValeCaja, id=vale_id)
    monto_letra = num2words(vale.importe, lang='es', to='currency', currency='MXN').capitalize()
    empresa = request.user.perfilusuario.empresa
    return render(request, "caja_chica/imprimir_vale_caja.html", {"vale": vale, "monto_letra": monto_letra, "empresa": empresa})


def detalle_fondeo(request, fondeo_id):
    fondeo = get_object_or_404(FondeoCajaChica, id=fondeo_id)
    gastos = fondeo.gastocajachica_set.all()
    vales = fondeo.valecaja_set.all()
    return render(
        request,
        "caja_chica/detalle_fondeo.html",
        {"fondeo": fondeo, "gastos": gastos, "vales": vales},
    )

@login_required
def fondeo_caja_chica(request):
    """
    Alta de fondeo: valida que numero_cheque no se repita dentro de la misma empresa
    del usuario logeado. Si existe, marca error en el form.
    """
    empresa = None
    if request.user.is_authenticated:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil:
            empresa = getattr(perfil, "empresa", None)

    if request.method == "POST":
        form = FondeoCajaChicaForm(request.POST)
        # filtrar campos relacionados por empresa si aplica
        try:
            if empresa:
                if "empleado_asignado" in form.fields:
                    form.fields["empleado_asignado"].queryset = form.fields["empleado_asignado"].queryset.filter(empresa=empresa)
        except Exception:
            pass

        if form.is_valid():
            cheque_val = form.cleaned_data.get("numero_cheque", None)
            if cheque_val:
                cheque_norm = str(cheque_val).strip()
                qs = FondeoCajaChica.objects.filter(numero_cheque__iexact=cheque_norm)
                if empresa:
                    qs = qs.filter(empresa=empresa)
                if qs.exists():
                    form.add_error("numero_cheque", "El número de cheque ya existe para esta empresa.")
                    messages.error(request, "El número de cheque ya existe para esta empresa.")
                    return render(request, "caja_chica/fondeo_caja_chica.html", {"form": form})

            fondeo = form.save(commit=False)
            # asegurar empresa asociada al fondeo (usar empresa del perfil)
            if empresa:
                fondeo.empresa = empresa
            # inicializar saldo con importe_cheque si no se proporcionó saldo
            if (fondeo.saldo is None or fondeo.saldo == 0) and getattr(fondeo, "importe_cheque", None) not in (None, ""):
                try:
                    fondeo.saldo = fondeo.importe_cheque
                except Exception:
                    pass
            fondeo.save()
            messages.success(request, "Fondeo registrado exitosamente.")
            return redirect("lista_fondeos")
    else:
        form = FondeoCajaChicaForm()
        if empresa:
            try:
                if "empleado_asignado" in form.fields:
                    form.fields["empleado_asignado"].queryset = form.fields["empleado_asignado"].queryset.filter(empresa=empresa)
            except Exception:
                pass

    return render(request, "caja_chica/fondeo_caja_chica.html", {"form": form})

@login_required
def registrar_gasto_caja_chica(request):
    empresa = None
    perfil = getattr(request.user, "perfilusuario", None)
    if perfil:
        empresa = getattr(perfil, "empresa", None)
    fondeo_id = request.GET.get("fondeo_id")
    fondeo_instance = None
    if fondeo_id:
        try:
            fondeo_instance = FondeoCajaChica.objects.get(id=fondeo_id)
        except FondeoCajaChica.DoesNotExist:
            fondeo_instance = None
    if request.method == "POST":
        form = GastoCajaChicaForm(request.POST)
        if empresa:
            form.fields["proveedor"].queryset = form.fields["proveedor"].queryset.filter(empresa=empresa)
            form.fields["tipo_gasto"].queryset = form.fields["tipo_gasto"].queryset.filter(empresa=empresa)
            form.fields["fondeo"].queryset = FondeoCajaChica.objects.filter(empresa=empresa,saldo__gt=0)
        if form.is_valid():
            gasto = form.save(commit=False)
            fondeo = gasto.fondeo
            if gasto.importe > fondeo.saldo:
                form.add_error("importe", f"El importe excede el saldo disponible (${fondeo.saldo}) en el fondeo.")
                messages.error(request, f"El importe excede el saldo disponible ${fondeo.saldo} en el fondeo.")
            else:
                fondeo.saldo -= gasto.importe
                fondeo.save()
                gasto.save()
                messages.success(request, "Gasto registrado exitosamente.")
                return redirect("lista_gastos_caja_chica")
    else:
        initial = {}
        if fondeo_instance:
            initial["fondeo"] = fondeo_instance
        form = GastoCajaChicaForm(initial=initial)
        if empresa:
            form.fields["proveedor"].queryset = form.fields["proveedor"].queryset.filter(empresa=empresa)
            form.fields["tipo_gasto"].queryset = form.fields["tipo_gasto"].queryset.filter(empresa=empresa)
            form.fields["fondeo"].queryset = FondeoCajaChica.objects.filter(empresa=empresa,saldo__gt=0)
        if fondeo_instance:
            form.fields["fondeo"].queryset = FondeoCajaChica.objects.filter(id=fondeo_instance.id)
    return render(request, "caja_chica/registrar_gasto_caja_chica.html", {"form": form})

@login_required
def generar_vale_caja(request):
    empresa = None
    if request.user.is_authenticated:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil:
            empresa = getattr(perfil, "empresa", None)
    if request.method == "POST":
        form = ValeCajaForm(request.POST)
        if empresa:
            form.fields["tipo_gasto"].queryset = TipoGasto.objects.filter(empresa=empresa)
            form.fields["fondeo"].queryset = FondeoCajaChica.objects.filter(empresa=empresa, saldo__gt=0)
            form.fields["recibido_por"].queryset = Empleado.objects.filter(empresa=empresa)
        if form.is_valid():
            vale = form.save(commit=False)
            fondeo = vale.fondeo
            if vale.importe > fondeo.saldo:
                form.add_error("importe", f"El importe excede el saldo disponible (${fondeo.saldo}) en el fondeo.")
                messages.error(request, f"El importe excede el saldo disponible ${fondeo.saldo} en el fondeo.")
            else:
                fondeo.saldo -= vale.importe
                fondeo.save()
                vale.save()
                messages.success(request, "Vale generado exitosamente.")
                return redirect("lista_vales_caja_chica")
    else:
        form = ValeCajaForm()
        if empresa:
            form.fields["tipo_gasto"].queryset = TipoGasto.objects.filter(empresa=empresa)
            form.fields["fondeo"].queryset = FondeoCajaChica.objects.filter(empresa=empresa, saldo__gt=0)
            form.fields["recibido_por"].queryset = Empleado.objects.filter(empresa=empresa)
    return render(request, "caja_chica/generar_vale_caja.html", {"form": form})

@login_required
def lista_fondeos(request):
    empresa_id = request.session.get("empresa_id")
    cheque = request.GET.get("cheque")
    empleado_id = request.GET.get("empleado")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if request.user.is_superuser and empresa_id:
        fondeos = FondeoCajaChica.objects.filter(empresa_id=empresa_id).order_by('-fecha')
    elif request.user.is_superuser:
        fondeos = FondeoCajaChica.objects.all().order_by('-fecha')
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil and perfil.empresa:
            fondeos = FondeoCajaChica.objects.filter(empresa=perfil.empresa).order_by('-fecha')
        else:
            fondeos = FondeoCajaChica.objects.none().order_by('-fecha')

    # Filtros
    if cheque:
        # fondeos = fondeos.filter(numero_cheque__icontains=cheque)
        # Coincidencia exacta
        fondeos = fondeos.filter(numero_cheque=cheque)
    if empleado_id and empleado_id.isdigit():
        fondeos = fondeos.filter(empleado_asignado_id=empleado_id)
    if fecha_inicio:
        fondeos = fondeos.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        fondeos = fondeos.filter(fecha__lte=fecha_fin)
    
    fondeos = fondeos.order_by('-fecha')
    empleados = Empleado.objects.filter(id__in=fondeos.values_list('empleado_asignado_id', flat=True)).order_by('nombre')
    cheques_existentes = fondeos.values_list('numero_cheque', flat=True).distinct().order_by('numero_cheque')
    total_importe = fondeos.object_list.aggregate(total=Sum('importe_cheque'))['total'] if hasattr(fondeos, 'object_list') else fondeos.aggregate(total=Sum('importe_cheque'))['total']
    total_saldo = fondeos.object_list.aggregate(total=Sum('saldo'))['total'] if hasattr(fondeos, 'object_list') else fondeos.aggregate(total=Sum('saldo'))['total']

    paginator = Paginator(fondeos, 20)
    page_number = request.GET.get("page")
    fondeos = paginator.get_page(page_number)

    return render(request, "caja_chica/lista_fondeos.html", {
        "fondeos": fondeos,
          "empleados": empleados, 
          "cheque": cheque, 
          "empleado_id": empleado_id,
            "fecha_inicio": fecha_inicio,
              "fecha_fin": fecha_fin,
              "cheques_existentes": cheques_existentes,
                "total_importe": total_importe or 0,
                    "total_saldo": total_saldo or 0,
          })

@login_required
def exportar_fondeos_excel(request):
    empresa_id = request.session.get("empresa_id")
    cheque = request.GET.get("cheque")
    empleado_id = request.GET.get("empleado")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if request.user.is_superuser and empresa_id:
        fondeos = FondeoCajaChica.objects.filter(empresa_id=empresa_id)
    elif request.user.is_superuser:
        fondeos = FondeoCajaChica.objects.all()
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil and perfil.empresa:
            fondeos = FondeoCajaChica.objects.filter(empresa=perfil.empresa)
        else:
            fondeos = FondeoCajaChica.objects.none()

    # Filtros
    if cheque:
        fondeos = fondeos.filter(numero_cheque=cheque)
    if empleado_id and empleado_id.isdigit():
        fondeos = fondeos.filter(empleado_asignado_id=empleado_id)
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fondeos = fondeos.filter(fecha__gte=fecha_inicio_dt)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
            fondeos = fondeos.filter(fecha__lte=fecha_fin_dt)
        except ValueError:
            pass

    fondeos = fondeos.order_by('-fecha')

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Fondeos Caja Chica"
    ws.append([
        "No. Cheque/Transfer", "Fecha", "Importe", "Responsable", "Saldo"
    ])

    for fondeo in fondeos.select_related("empleado_asignado"):
        ws.append([
            fondeo.numero_cheque,
            fondeo.fecha.strftime("%d/%m/%Y") if fondeo.fecha else '',
            float(fondeo.importe_cheque),
            fondeo.empleado_asignado.nombre if fondeo.empleado_asignado else '',
            float(fondeo.saldo)
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=fondeos_caja_chica.xlsx'
    wb.save(response)
    return response


@login_required
def lista_gastos_caja_chica(request):
    empresa_id = request.session.get("empresa_id")
    proveedor_id = request.GET.get("proveedor")
    tipo_gasto_id = request.GET.get("tipo_gasto")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if request.user.is_superuser and empresa_id:
        gastos = GastoCajaChica.objects.select_related("fondeo").filter(
            fondeo__empresa_id=empresa_id
        ).order_by('-fecha')
    elif request.user.is_superuser:
        gastos = GastoCajaChica.objects.select_related("fondeo").all().order_by('-fecha')
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil and perfil.empresa:
            gastos = GastoCajaChica.objects.select_related("fondeo").filter(
                fondeo__empresa=perfil.empresa
            ).order_by('-fecha')
        else:
            gastos = GastoCajaChica.objects.select_related("fondeo").none().order_by('-fecha')

    # Filtros
    if proveedor_id and proveedor_id.isdigit():
        gastos = gastos.filter(proveedor_id=proveedor_id)
    if tipo_gasto_id and tipo_gasto_id.isdigit():
        gastos = gastos.filter(tipo_gasto_id=tipo_gasto_id)
    if fecha_inicio:
        gastos = gastos.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        gastos = gastos.filter(fecha__lte=fecha_fin)

    total_gastos = gastos.aggregate(total=Sum('importe'))['total'] or 0

    # Para los selects en el template
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    tipos_gasto = TipoGasto.objects.all()

    paginator = Paginator(gastos, 20)
    page_number = request.GET.get("page")
    gastos = paginator.get_page(page_number)

    return render(
        request, "caja_chica/lista_gastos_caja_chica.html", {
            "gastos": gastos,
            "proveedores": proveedores,
            "tipos_gasto": tipos_gasto,
            "proveedor_id": proveedor_id,
            "tipo_gasto_id": tipo_gasto_id,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_gastos": total_gastos,
        }
    )

from datetime import datetime

@login_required
def exportar_gastos_caja_chica_excel(request):
    empresa_id = request.session.get("empresa_id")
    proveedor_id = request.GET.get("proveedor")
    tipo_gasto_id = request.GET.get("tipo_gasto")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if request.user.is_superuser and empresa_id:
        gastos = GastoCajaChica.objects.select_related("fondeo").filter(
            fondeo__empresa_id=empresa_id
        )
    elif request.user.is_superuser:
        gastos = GastoCajaChica.objects.select_related("fondeo").all()
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil and perfil.empresa:
            gastos = GastoCajaChica.objects.select_related("fondeo").filter(
                fondeo__empresa=perfil.empresa
            )
        else:
            gastos = GastoCajaChica.objects.select_related("fondeo").none()

    # Filtros
    if proveedor_id and proveedor_id.isdigit():
        gastos = gastos.filter(proveedor_id=proveedor_id)
    if tipo_gasto_id and tipo_gasto_id.isdigit():
        gastos = gastos.filter(tipo_gasto_id=tipo_gasto_id)
    # Validar fechas
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            gastos = gastos.filter(fecha__gte=fecha_inicio_dt)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
            gastos = gastos.filter(fecha__lte=fecha_fin_dt)
        except ValueError:
            pass

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos Caja Chica"
    ws.append([
        "Fondo", "Proveedor", "Tipo de Gasto", "Descripción", "Importe", "Fecha"
    ])

    for gasto in gastos.select_related("fondeo", "proveedor", "tipo_gasto"):
        ws.append([
            gasto.fondeo.numero_cheque if gasto.fondeo else '',
            gasto.proveedor.nombre if gasto.proveedor else '',
            gasto.tipo_gasto.nombre if gasto.tipo_gasto else '',
            gasto.descripcion,
            float(gasto.importe),
            gasto.fecha.strftime("%d/%m/%Y") if gasto.fecha else ''
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=gastos_caja_chica.xlsx'
    wb.save(response)
    return response

@login_required
def lista_vales_caja_chica(request):
    empresa_id = request.session.get("empresa_id")
    empleado_id = request.GET.get("empleado")
    tipo_gasto_id = request.GET.get("tipo_gasto")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if request.user.is_superuser and empresa_id:
        vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").filter(
            fondeo__empresa_id=empresa_id
        ).order_by('-fecha')
    elif request.user.is_superuser:
        vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").all().order_by('-fecha')
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil and perfil.empresa:
            vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").filter(
                fondeo__empresa=perfil.empresa
            ).order_by('-fecha')
        else:
            vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").none().order_by('-fecha')

    # Filtros
    if empleado_id and empleado_id.isdigit():
        vales = vales.filter(recibido_por_id=empleado_id)
    if tipo_gasto_id and tipo_gasto_id.isdigit():
        vales = vales.filter(tipo_gasto_id=tipo_gasto_id)
    if fecha_inicio:
        vales = vales.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        vales = vales.filter(fecha__lte=fecha_fin)

    # total_vales = vales.aggregate(total=Sum('importe'))['total'] or 0
    total_pendientes = vales.filter(status="pendiente").aggregate(total=Sum('importe'))['total'] or 0
    total_comprobados = vales.filter(status="comprobado").aggregate(total=Sum('importe'))['total'] or 0
    gran_total = total_pendientes + total_comprobados

    # Para los selects en el template
    empleados = Empleado.objects.filter(id__in=vales.values_list('recibido_por_id', flat=True)).order_by('nombre')
    tipos_gasto = TipoGasto.objects.all()

    paginator = Paginator(vales, 20)
    page_number = request.GET.get("page")
    vales = paginator.get_page(page_number)

    return render(
        request, "caja_chica/lista_vales_caja_chica.html",
        {
            "vales": vales,
            "empleados": empleados,
            "empleado_id": empleado_id,
            "tipos_gasto": tipos_gasto,
            "tipo_gasto_id": tipo_gasto_id,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_vales": gran_total,
            "total_pendientes": total_pendientes,
            "total_comprobados": total_comprobados,
        }
    )


@login_required
def exportar_vales_caja_chica_excel(request):
    empresa_id = request.session.get("empresa_id")
    empleado_id = request.GET.get("empleado")
    tipo_gasto_id = request.GET.get("tipo_gasto")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if request.user.is_superuser and empresa_id:
        vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").filter(
            fondeo__empresa_id=empresa_id
        )
    elif request.user.is_superuser:
        vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").all()
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if perfil and perfil.empresa:
            vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").filter(
                fondeo__empresa=perfil.empresa
            )
        else:
            vales = ValeCaja.objects.select_related("fondeo", "recibido_por", "tipo_gasto").none()

    # Filtros
    if empleado_id and empleado_id.isdigit():
        vales = vales.filter(recibido_por_id=empleado_id)
    if tipo_gasto_id and tipo_gasto_id.isdigit():
        vales = vales.filter(tipo_gasto_id=tipo_gasto_id)
    # Validar fechas
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            vales = vales.filter(fecha__gte=fecha_inicio_dt)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
            vales = vales.filter(fecha__lte=fecha_fin_dt)
        except ValueError:
            pass

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vales Caja Chica"
    ws.append([
        "Fondo", "Empleado", "Tipo de Gasto", "Descripción", "Importe", "Fecha", "Estatus"
    ])

    for vale in vales:
        ws.append([
            vale.fondeo.numero_cheque if vale.fondeo else '',
            vale.recibido_por.nombre if vale.recibido_por else '',
            vale.tipo_gasto.nombre if vale.tipo_gasto else '',
            vale.descripcion,
            float(vale.importe),
            vale.fecha.strftime("%d/%m/%Y") if vale.fecha else '',
            vale.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=vales_caja_chica.xlsx'
    wb.save(response)
    return response


@login_required
def recibo_fondeo_caja(request, fondeo_id):
    fondeo = get_object_or_404(FondeoCajaChica, pk=fondeo_id)
    empresa = request.user.perfilusuario.empresa
    monto_letra = num2words(fondeo.importe_cheque, lang='es', to='currency', currency='MXN').capitalize()
    return render(request, 'caja_chica/recibo_fondeo.html', {'fondeo': fondeo, 'monto_letra': monto_letra, 'empresa': empresa })



@login_required
def eliminar_vale_caja(request, vale_id):
    """
    Confirmar (GET) y eliminar (POST) un ValeCaja.
    - Superuser: si hay empresa seleccionada en session ('empresa_id'), sólo puede eliminar vales de esa empresa.
    - Usuario normal: sólo puede eliminar vales pertenecientes a su empresa en perfil.
    Al eliminar, devuelve el importe al fondeo asociado.
    """
    vale = get_object_or_404(ValeCaja, pk=vale_id)
    fondeo = getattr(vale, "fondeo", None)
    vale_empresa_id = None
    if fondeo and getattr(fondeo, "empresa", None):
        vale_empresa_id = fondeo.empresa.id

    # permiso multiempresa
    if request.user.is_superuser:
        sess_empresa_id = request.session.get("empresa_id")
        if sess_empresa_id and vale_empresa_id and int(sess_empresa_id) != int(vale_empresa_id):
            messages.error(request, "No tienes permiso para eliminar este vale bajo la empresa seleccionada.")
            return redirect("lista_vales_caja_chica")
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if not perfil or not getattr(perfil, "empresa", None) or perfil.empresa.id != vale_empresa_id:
            messages.error(request, "No tienes permiso para eliminar este vale.")
            return redirect("lista_vales_caja_chica")

    if request.method == "POST":
        with transaction.atomic():
            # devolver importe al fondeo si aplica
            if fondeo:
                fondeo.saldo = (fondeo.saldo or 0) + (vale.importe or 0)
                fondeo.save()
            vale.delete()
        messages.success(request, "Vale eliminado correctamente.")
        return redirect("lista_vales_caja_chica")

    return render(request, "caja_chica/confirm_eliminar_vale.html", {"vale": vale})

@login_required
def eliminar_gasto_caja(request, gasto_id):
    """
    Confirmar (GET) y eliminar (POST) un GastoCaja.
    - Superuser: si hay empresa seleccionada en session ('empresa_id'), sólo puede eliminar gastos de esa empresa.
    - Usuario normal: sólo puede eliminar gastos pertenecientes a su empresa en perfil.
    Al eliminar, devuelve el importe al fondeo asociado si aplica.
    """
    gasto = get_object_or_404(GastoCajaChica, pk=gasto_id)
    fondeo = getattr(gasto, "fondeo", None)
    gasto_empresa_id = None
    if fondeo and getattr(fondeo, "empresa", None):
        gasto_empresa_id = fondeo.empresa.id
    else:
        # intentar obtener empresa desde el gasto o su vale relacionado
        gasto_empresa_id = getattr(gasto, 'empresa_id', None) or getattr(getattr(gasto, 'vale', None), 'empresa_id', None)

    # permisos multiempresa
    if request.user.is_superuser:
        sess_empresa_id = request.session.get("empresa_id")
        if sess_empresa_id and gasto_empresa_id and int(sess_empresa_id) != int(gasto_empresa_id):
            messages.error(request, "No tienes permiso para eliminar este gasto bajo la empresa seleccionada.")
            return redirect("lista_gastos_caja_chica")
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if not perfil or not getattr(perfil, "empresa", None) or perfil.empresa.id != gasto_empresa_id:
            messages.error(request, "No tienes permiso para eliminar este gasto.")
            return redirect("lista_gastos_caja_chica")

    if request.method == "POST":
        with transaction.atomic():
            if fondeo:
                fondeo.saldo = (fondeo.saldo or 0) + (gasto.importe or 0)
                fondeo.save()
            gasto.delete()
        messages.success(request, "Gasto eliminado correctamente.")
        return redirect("lista_gastos_caja_chica")

    return render(request, "caja_chica/confirm_eliminar_gasto.html", {"gasto": gasto})


@login_required
def eliminar_fondeo(request, fondeo_id):
    """
    Elimina un FondeoCajaChica solo si saldo == importe_cheque.
    Respeta permisos multiempresa (session empresa para superuser, perfil para usuario normal).
    """
    fondeo = get_object_or_404(FondeoCajaChica, pk=fondeo_id)
    fondeo_empresa_id = getattr(getattr(fondeo, "empresa", None), "id", None)

    # permisos multiempresa
    if request.user.is_superuser:
        sess_empresa_id = request.session.get("empresa_id")
        if sess_empresa_id and fondeo_empresa_id and int(sess_empresa_id) != int(fondeo_empresa_id):
            messages.error(request, "No tienes permiso para eliminar este fondeo bajo la empresa seleccionada.")
            return redirect("lista_fondeos")
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if not perfil or not getattr(perfil, "empresa", None) or perfil.empresa.id != fondeo_empresa_id:
            messages.error(request, "No tienes permiso para eliminar este fondeo.")
            return redirect("lista_fondeos")

    # obtener importe_cheque y saldo como Decimal
    try:
        importe = Decimal(str(fondeo.importe_cheque or 0))
    except (InvalidOperation, TypeError, ValueError):
        importe = Decimal("0")
    try:
        saldo = Decimal(str(fondeo.saldo or 0))
    except (InvalidOperation, TypeError, ValueError):
        saldo = Decimal("0")

    # GET -> mostrar confirmación
    if request.method == "GET":
        return render(request, "caja_chica/confirm_eliminar_fondeo.html", {"fondeo": fondeo, "importe": importe, "saldo": saldo})

    # POST -> eliminar sólo si saldo == importe_cheque
    if request.method == "POST":
        if saldo != importe:
            messages.error(request, "No se puede eliminar el fondeo: tiene gastos asociados.")
            return redirect("lista_fondeos")
        with transaction.atomic():
            fondeo.delete()
        messages.success(request, "Fondeo eliminado correctamente.")
        return redirect("lista_fondeos")
    

@login_required
def comprobar_vale(request, vale_id):
    vale = get_object_or_404(ValeCaja, pk=vale_id)

    # determinar empresa del vale (intentar fondeo -> fondeo.empresa o vale.empresa)
    vale_empresa_id = None
    if getattr(vale, 'fondeo', None) and getattr(vale.fondeo, 'empresa', None):
        vale_empresa_id = vale.fondeo.empresa.id
    else:
        vale_empresa_id = getattr(vale, 'empresa_id', None)

    # permisos multiempresa
    if request.user.is_superuser:
        sess_empresa_id = request.session.get("empresa_id")
        if sess_empresa_id and vale_empresa_id and int(sess_empresa_id) != int(vale_empresa_id):
            messages.error(request, "No tienes permiso para comprobar este vale bajo la empresa seleccionada.")
            return redirect("lista_vales_caja_chica")
    else:
        perfil = getattr(request.user, "perfilusuario", None)
        if not perfil or not getattr(perfil, "empresa", None) or perfil.empresa.id != vale_empresa_id:
            messages.error(request, "No tienes permiso para comprobar este vale.")
            return redirect("lista_vales_caja_chica")

    if request.method == "POST":
        form = ComprobarValeForm(request.POST)
        if form.is_valid():
            importe_comprobado = form.cleaned_data["importe_comprobado"]
            descripcion = form.cleaned_data.get("descripcion", "").strip()

            try:
                with transaction.atomic():
                    importe_vale = Decimal(str(vale.importe or 0))
                    comprobado = Decimal(str(importe_comprobado or 0))
                    diferencia = comprobado - importe_vale

                    # Si no hay diferencia: solo marcar comprobado y guardar datos mínimos
                    if diferencia == Decimal("0.00"):
                        vale.status = "comprobado"
                        if hasattr(vale, "importe_comprobado"):
                            setattr(vale, "importe_comprobado", comprobado)
                        if descripcion and hasattr(vale, "observaciones"):
                            vale.observaciones = (vale.observaciones or "") + "\nComprobación: " + descripcion
                        vale.save()
                        messages.success(request, "Vale comprobado correctamente.")
                        return redirect("lista_vales_caja_chica")

                    # intentar resolver empleado para registrar incidencia
                    empleado_obj = getattr(vale, "recibido_por", None)
                    if not empleado_obj:
                        perfil = getattr(request.user, "perfilusuario", None)
                        empleado_obj = getattr(perfil, "recibido_por", None) if perfil else None

                    # crear incidencia si hay diferencia y si existe empleado
                    if diferencia != Decimal("0.00"):
                        if not empleado_obj:
                            # no podemos crear Incidencia sin empleado, abortar creación de incidencia
                            messages.error(request, "No se encontró un empleado asociado para registrar la incidencia; por favor asocia un empleado al vale o al usuario.")
                        else:
                            if diferencia < 0:
                                tipo = "descuento"
                                monto = abs(diferencia)
                            else:
                                tipo = "devolucion"
                                monto = diferencia
                            Incidencia.objects.create(
                                empleado=empleado_obj,
                                tipo=tipo,
                                fecha=timezone.now().date(),
                                dias=1,
                                descripcion=descripcion or (f"Diferencia al comprobar vale #{vale.id}: {monto}"),
                                importe=monto
                            )

                    # marcar vale como comprobado y guardar importe comprobado si existe ese campo
                    #vale.status = getattr(vale, "status", "comprobado") or "comprobado"
                    vale.status = "comprobado"
                    if hasattr(vale, "importe_comprobado"):
                        setattr(vale, "importe_comprobado", comprobado)
                    if descripcion and hasattr(vale, "observaciones"):
                        vale.observaciones = (vale.observaciones or "") + "\nComprobación: " + descripcion
                    vale.save()

                messages.success(request, "Vale comprobado correctamente.")
                return redirect("lista_vales_caja_chica")
            except Exception as e:
                messages.error(request, f"Error al comprobar el vale: {e}")
    else:
        initial = {"importe_comprobado": getattr(vale, "importe", None)}
        form = ComprobarValeForm(initial=initial)

    return render(request, "caja_chica/comprobar_vale.html", {"form": form, "vale": vale})    