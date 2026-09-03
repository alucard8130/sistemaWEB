
import io
from decimal import Decimal, InvalidOperation

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Avg, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from unidecode import unidecode

from clientes.models import Cliente
from empresas.models import Empresa
from facturacion.models import PoolVacancia
from locales.utils import generar_facturas_local
from sanitarios.views import _empresa_actual

from .forms import LocalCargaMasivaForm, LocalComercialForm
from .models import LocalComercial


@login_required
def lista_locales(request):
    user = request.user
    query = request.GET.get("q", "")
    if user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        if empresa_id:
            locales = LocalComercial.objects.filter(activo=True, empresa_id=empresa_id).select_related('cliente', 'empresa', 'grupo_facturacion', 'pool_vacancia').order_by('numero')
        else:
            locales = LocalComercial.objects.filter(activo=True).select_related('cliente', 'empresa', 'grupo_facturacion', 'pool_vacancia').order_by('numero')
    else:
        empresa = user.perfilusuario.empresa
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True).select_related('cliente', 'empresa', 'grupo_facturacion', 'pool_vacancia').order_by('numero')

    if query:
        locales = locales.filter(
            Q(numero__icontains=query) | Q(cliente__nombre__icontains=query) | Q(cliente__rfc__icontains=query)
        )

    locales = locales.order_by('numero')
    locales_totales_activos = locales.count()
    total_superficie = locales.aggregate(Sum('superficie_m2'))['superficie_m2__sum'] or 0
    superficie_ocupada = locales.filter(status='ocupado').aggregate(Sum('superficie_m2'))['superficie_m2__sum'] or 0
    superficie_disponible = total_superficie - superficie_ocupada
    total_cuotas = locales.aggregate(Sum('cuota'))['cuota__sum'] or 0
    promedio_cuotas = locales.aggregate(Avg('cuota'))['cuota__avg'] or 0
    promedio_precio_m2 = total_superficie > 0 and (total_cuotas / total_superficie) or 0
    porcentaje_ocupacion = (superficie_ocupada / total_superficie * 100) if total_superficie > 0 else 0

    paginator = Paginator(locales, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'locales/lista_locales.html', {'locales': page_obj, 'q': query,
                                                          'locales_totales_activos': locales_totales_activos,
                                                          'total_superficie': total_superficie,
                                                          'superficie_ocupada': superficie_ocupada,
                                                          'superficie_disponible': superficie_disponible,
                                                          'total_cuotas': total_cuotas,
                                                          'promedio_cuotas': promedio_cuotas,
                                                          'promedio_precio_m2': promedio_precio_m2,
                                                          'porcentaje_ocupacion': porcentaje_ocupacion,
                                                          })


@login_required
def crear_local(request):
    user = request.user
    perfil = getattr(user, 'perfilusuario', None)
    empresa_usuario = perfil.empresa if perfil and not user.is_superuser else None
    es_habitacional = empresa_usuario.segmento == 'habitacional' if empresa_usuario else False
    
    if request.method == 'POST':
        form = LocalComercialForm(request.POST, user=user)
        if form.is_valid():
            local = form.save(commit=False)
            # Si no es superusuario, asignamos su empresa
            if not user.is_superuser and perfil and perfil.empresa:
                local.empresa = perfil.empresa
            local.save()
            # leer datos de adeudo desde POST (ejemplo: campos 'tiene_adeudo', 'adeudo_inicio', 'adeudo_fin', 'adeudo_importe', 'adeudo_descripcion')
            tiene_adeudo = request.POST.get('tiene_adeudo') == 'on'
            with transaction.atomic():
                if tiene_adeudo:
                    inicio = request.POST.get('adeudo_inicio')  # acepta 'YYYY-MM' o 'YYYY-MM-DD'
                    fin = request.POST.get('adeudo_fin') or None
                    importe = request.POST.get('adeudo_importe')
                    observaciones = request.POST.get('adeudo_descripcion', '')
                    try:
                        importe_dec = Decimal(str(importe))
                    except:
                        importe_dec = None
                    if importe_dec is None:
                        messages.error(request, "Importe de adeudo inválido; se omitió la factura de adeudo.")
                        generar_facturas_local(local)  # generar la factura del mes en su defecto
                    else:
                        generar_facturas_local(local, adeudo={
                            'inicio': inicio,
                            'fin': fin,
                            'importe': importe_dec,
                            'descripcion': observaciones
                        })
                else:
                    generar_facturas_local(local)
        
            messages.success(request, "Local creado correctamente.")
            return redirect('lista_locales')
        else:
            messages.error(request, "No se pudo crear el local. Revisa los datos ingresados.")
    else:
        form = LocalComercialForm(user=user)
        # Si no es superusuario, asignamos la empresa inicial al form
        if not user.is_superuser and perfil and perfil.empresa:
            form.fields['empresa'].initial = perfil.empresa

    return render(request, 'locales/crear_local.html', {'form': form, 'es_habitacional': es_habitacional})

@login_required
def editar_local(request, pk):
    user = request.user
    local= get_object_or_404(LocalComercial, pk=pk)
 
    if not user.is_superuser and local.empresa != user.perfilusuario.empresa:
        return redirect('lista_locales')

    if request.method == 'POST':
        form = LocalComercialForm(request.POST, instance=local, user=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Local actualizado correctamente.")
            return redirect('lista_locales')
    else:
        form = LocalComercialForm(instance=local, user=user)

    return render(request, 'locales/editar_local.html', {'form': form, 'local': local})

@login_required
def eliminar_local(request, pk):
    user = request.user
    local= get_object_or_404(LocalComercial, pk=pk)
    if not user.is_superuser and local.empresa != user.perfilusuario.empresa:
        return redirect('lista_locales')

    if request.method == 'POST':
        local.activo = False
        local.save()
        return redirect('lista_locales')

    return render(request, 'locales/eliminar_local.html', {'local': local})


@login_required
def locales_inactivos(request):
    user = request.user
    if user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        if empresa_id:
            locales = LocalComercial.objects.filter(empresa_id=empresa_id, activo=False)
        else:
            locales = LocalComercial.objects.filter(activo=False)
    else:
        empresa = user.perfilusuario.empresa
        locales = LocalComercial.objects.filter(empresa=empresa, activo=False)
    return render(request, 'locales/locales_inactivos.html', {'locales': locales})


def reactivar_local(request, pk):
    local = get_object_or_404(LocalComercial, pk=pk, activo=False)

    if request.method == 'POST':
        local.activo = True
        local.save()
        return redirect('locales_inactivos')

    return render(request, 'locales/reactivar_confirmacion.html', {'local': local})

@login_required
def incrementar_cuotas_locales(request):
    if request.method == 'POST':
        porcentaje = request.POST.get('porcentaje')
        try:
            porcentaje = Decimal(porcentaje)
            empresa = None
            if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
                empresa = request.user.perfilusuario.empresa
                locales = LocalComercial.objects.filter(empresa=empresa, activo=True)
            else:
                locales = LocalComercial.objects.filter(activo=True)

            for local in locales:
                cuota_anterior = local.cuota
                incremento = cuota_anterior * (porcentaje / Decimal('100'))
                local.cuota += incremento
                local.save()

            messages.success(request, f'Se incrementaron las cuotas en un {porcentaje}% para todos los locales activos.')
            return redirect('incrementar_c_locales')
        except:
            messages.error(request, 'Porcentaje inválido.')
    
    return render(request, 'locales/incrementar_c_locales.html')

def buscar_por_id_o_nombre(modelo, valor, campo='nombre'):
    if not valor:
        return None
    val = str(valor).strip().replace(',', '')
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        candidatos = [
            obj for obj in todos
            if unidecode(val).lower() in unidecode(str(getattr(obj, campo))).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")

@login_required
def carga_masiva_locales(request):
    perfil = getattr(request.user, 'perfilusuario', None)
    empresa_usuario = perfil.empresa if perfil and not request.user.is_superuser else None
    es_habitacional = empresa_usuario.segmento == 'habitacional' if empresa_usuario else False

    if request.method == 'POST':
        form = LocalCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            errores = []
            creados = 0
            actualizados = 0

            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers_map = {}
            if header:
                hdrs = [str(h).strip().lower() if h is not None else "" for h in header]
                for idx, h in enumerate(hdrs):
                    if h in ('condominio', 'empresa'):
                        headers_map['empresa'] = idx
                    if h in ('propietario',):
                        headers_map['propietario'] = idx
                    if h in ('cliente', 'cliente nombre', 'nombre'):
                        headers_map['cliente'] = idx
                    if h in ('rfc', 'rfc cliente', 'cliente rfc'):
                        headers_map['rfc'] = idx
                    if h in ('email', 'correo', 'correo electronico'):
                        headers_map['email'] = idx
                    if h in ('numero', 'num', 'número'):
                        headers_map['numero'] = idx
                    if h in ('cuota', 'monto', 'importe'):
                        headers_map['cuota'] = idx
                    if h in ('ubicacion',):
                        headers_map['ubicacion'] = idx
                    if h in ('superficie', 'superficie_m2', 'm2'):
                        headers_map['superficie_m2'] = idx
                    if h in ('proindiviso', 'pro indiviso', '% proindiviso'):
                        headers_map['proindiviso'] = idx    
                    if h in ('giro',):
                        headers_map['giro'] = idx
                    if h in ('status', 'estatus'):
                        headers_map['status'] = idx
                    if h in ('observaciones', 'obs', 'comentarios'):
                        headers_map['observaciones'] = idx
                    if h in ('tipo_propiedad', 'tipo propiedad', 'tipo'):
                        headers_map['tipo_propiedad'] = idx

            def cell(row, key, pos):
                if key in headers_map:
                    i = headers_map[key]
                    return row[i] if i < len(row) else None
                if pos is None:
                    return None
                return row[pos] if pos < len(row) else None

            tipos_validos_habitacional = ('casa', 'departamento', 'terreno')
            tipos_validos_comercial = ('local', 'oficina', 'bodega', 'terreno')
            estatus_validos = ('ocupado', 'disponible', 'mantenimiento')

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                    continue

                try:
                    empresa_val = cell(row, 'empresa', 0)
                    propietario_val = cell(row, 'propietario', 1)
                    nombre_cliente = cell(row, 'cliente', 2)
                    rfc_cliente = cell(row, 'rfc', 3)
                    email_cliente = cell(row, 'email', 4)
                    numero = cell(row, 'numero', 5)
                    cuota = cell(row, 'cuota', 6)
                    ubicacion = cell(row, 'ubicacion', 7)
                    superficie_m2 = cell(row, 'superficie_m2', 8)
                    proindiviso_val = cell(row, 'proindiviso', 9)
                    giro = cell(row, 'giro', 10)
                    status = cell(row, 'status', 11)
                    observaciones = cell(row, 'observaciones', 12)
                    tipo_propiedad_val = cell(row, 'tipo_propiedad', None)

                    if request.user.is_superuser:
                        empresa = buscar_por_id_o_nombre(Empresa, empresa_val) if empresa_val else None
                        if not empresa:
                            raise Exception(f"No se encontró la empresa '{empresa_val}'.")
                        es_habitacional_fila = empresa.segmento == 'habitacional'
                    else:
                        if not perfil or not getattr(perfil, 'empresa', None):
                            raise Exception("No se pudo determinar la empresa del usuario.")
                        empresa = perfil.empresa
                        es_habitacional_fila = es_habitacional

                    if not numero:
                        raise Exception("La columna 'numero' está vacía.")
                    numero_str = str(numero).strip()

                    try:
                        cuota_decimal = Decimal(str(cuota)) if cuota not in (None, "") else None
                    except (InvalidOperation, TypeError, ValueError):
                        raise Exception(f"El valor de cuota '{cuota}' no es un número válido.")

                    try:
                        proindiviso_decimal = Decimal(str(proindiviso_val)) if proindiviso_val not in (None, "") else None
                    except (InvalidOperation, TypeError, ValueError):
                        raise Exception(f"El valor de proindiviso '{proindiviso_val}' no es un número válido.")


                    # --- Validar tipo_propiedad contra el segmento ---
                    tipo_propiedad = str(tipo_propiedad_val).strip().lower() if tipo_propiedad_val not in (None, "") else None
                    tipos_validos = tipos_validos_habitacional if es_habitacional_fila else tipos_validos_comercial

                    if tipo_propiedad and tipo_propiedad not in tipos_validos:
                        raise Exception(
                            f"Tipo de propiedad '{tipo_propiedad}' no válido para empresa "
                            f"{'habitacional' if es_habitacional_fila else 'comercial'}. "
                            f"Usa uno de: {', '.join(tipos_validos)}."
                        )
                    if not tipo_propiedad:
                        tipo_propiedad = 'departamento' if es_habitacional_fila else 'local'

                    # --- Validar status si viene capturado ---
                    status_norm = str(status).strip().lower() if status not in (None, "") else None
                    if status_norm and status_norm not in estatus_validos:
                        raise Exception(
                            f"Estatus '{status_norm}' no válido. Usa uno de: {', '.join(estatus_validos)}."
                        )

                    # --- Cliente: buscar/crear igual que antes ---
                    rfc_norm = str(rfc_cliente).strip().upper() if rfc_cliente not in (None, "") else None
                    nombre_norm = str(nombre_cliente).strip() if nombre_cliente not in (None, "") else ""

                    cliente = None
                    if rfc_norm:
                        cliente = Cliente.objects.filter(empresa=empresa, rfc__iexact=rfc_norm).first()
                        if cliente:
                            updated = False
                            if nombre_norm and (not getattr(cliente, 'nombre', None) or cliente.nombre.strip() == ""):
                                cliente.nombre = nombre_norm; updated = True
                            if email_cliente and (not getattr(cliente, 'email', None) or cliente.email.strip() == ""):
                                cliente.email = email_cliente; updated = True
                            if updated:
                                cliente.save()
                        else:
                            cliente = Cliente.objects.create(
                                empresa=empresa,
                                nombre=nombre_norm or f"Cliente {rfc_norm}",
                                rfc=rfc_norm,
                                email=email_cliente or None,
                                activo=True,
                            )
                    elif nombre_norm:
                        qs = Cliente.objects.filter(empresa=empresa, nombre__iexact=nombre_norm)
                        if qs.exists():
                            cliente = qs.filter(rfc__isnull=False).exclude(rfc='').first() or qs.first()
                        else:
                            cliente = Cliente.objects.create(
                                empresa=empresa, nombre=nombre_norm, email=email_cliente or None, activo=True,
                            )
                    # Si no hay ni RFC ni nombre, cliente queda None -- se permite (propiedad sin asignar)

                    # --- Crear O actualizar la propiedad (idempotente) ---
                    with transaction.atomic():
                        propiedad, creada = LocalComercial.objects.get_or_create(
                            empresa=empresa, numero=numero_str,
                            defaults={
                                'propietario': propietario_val or "",
                                'cliente': cliente,
                                'cuota': cuota_decimal or Decimal('0.00'),
                                'tipo_propiedad': tipo_propiedad,
                                'ubicacion': "" if es_habitacional_fila else (ubicacion or ""),
                                'superficie_m2': Decimal(str(superficie_m2)) if superficie_m2 not in (None, "") else None,
                                'proindiviso': proindiviso_decimal,
                                'giro': "" if es_habitacional_fila else (giro or ""),
                                'status': status_norm or "ocupado",
                                'observaciones': observaciones or "",
                            }
                        )
                        if creada:
                            creados += 1
                        else:
                            # Actualiza solo los campos que vienen con dato en el Excel,
                            # para no borrar información que ya tenías capturada manualmente.
                            if propietario_val:
                                propiedad.propietario = propietario_val
                            if cliente:
                                propiedad.cliente = cliente
                            if cuota_decimal is not None:
                                propiedad.cuota = cuota_decimal
                            if tipo_propiedad_val:
                                propiedad.tipo_propiedad = tipo_propiedad
                            if not es_habitacional_fila and ubicacion:
                                propiedad.ubicacion = ubicacion
                            if superficie_m2 not in (None, ""):
                                propiedad.superficie_m2 = Decimal(str(superficie_m2))
                            if proindiviso_decimal is not None:
                                propiedad.proindiviso = proindiviso_decimal    
                            if not es_habitacional_fila and giro:
                                propiedad.giro = giro
                            if status_norm:
                                propiedad.status = status_norm
                            if observaciones:
                                propiedad.observaciones = observaciones
                            propiedad.save()
                            actualizados += 1

                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")

            if creados or actualizados:
                messages.success(
                    request,
                    f"✅ {creados} propiedades nuevas creadas, {actualizados} propiedades existentes actualizadas."
                )
            if errores:
                msg = "<br>".join(errores[:80])
                if len(errores) > 80:
                    msg += f"<br>...y {len(errores)-80} errores más."
                messages.error(request, mark_safe("Algunas filas tuvieron problemas:<br>" + msg))

            return redirect('carga_masiva_locales')
    else:
        form = LocalCargaMasivaForm()
    return render(request, 'locales/carga_masiva_locales.html', {'form': form, 'es_habitacional': es_habitacional})


@login_required
def plantilla_locales_excel(request):
    perfil = getattr(request.user, 'perfilusuario', None)
    empresa = perfil.empresa if perfil and not request.user.is_superuser else None
    es_habitacional = empresa.segmento == 'habitacional' if empresa else False

    wb = openpyxl.Workbook()
    ws = wb.active

    if es_habitacional:
        ws.title = "Plantilla Viviendas"
        ws.append([
            'condominio', 'propietario', 'cliente', 'rfc', 'email', 'numero',
            'cuota', 'tipo_propiedad', 'superficie_m2', 'proindiviso', 'status', 'observaciones'
        ])
        ws.append([
            'Condominio Las Palmas AC', 'Juan Pérez', 'Juan Pérez', 'XXX-XXX-XXX',
            'juan@ejemplo.com', '101', '1500.00', 'departamento', '85.5', '0.1234',
            'ocupado', 'carga inicial'
        ])
    else:
        ws.title = "Plantilla Locales"
        ws.append([
            'condominio', 'propietario', 'cliente', 'rfc', 'email', 'numero',
            'cuota', 'tipo_propiedad', 'ubicacion', 'superficie_m2', 'proindiviso', 'giro', 'status', 'observaciones'
        ])
        ws.append([
            'plaza en condominio AC', 'Tiendas Soriana SA de CV', 'Juan Pérez',
            'XXX-XXX-XXX', 'email@ejemplo.com', '101', '120.3', 'local', 'planta baja',
            '30.5', '50.56', 'venta ropa', 'ocupado', 'carga inicial'
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = 'plantilla_viviendas.xlsx' if es_habitacional else 'plantilla_locales.xlsx'
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
    wb.save(response)
    return response


def exportar_locales_excel(request):
    if request.user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        empresa = Empresa.objects.filter(id=empresa_id).first()
    else:
        empresa = request.user.perfilusuario.empresa
 
    locales = (
        LocalComercial.objects.filter(empresa=empresa)
        .select_related('cliente', 'grupo_facturacion', 'pool_vacancia')
        .order_by('numero')
    )
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Propiedades"
 
    encabezados = [
        "Número", "Tipo", "Propietario", "Cliente", "Ubicación",
        "Superficie m²", "Giro", "Cuota", "¿Cuota anual?", "Status",
        "Activo", "Proindiviso %", "Grupo de Facturación", "Pool de Vacancia",
        "Referencia de pago", "Observaciones", "Fecha de creación",
    ]
 
    header_fill = PatternFill(start_color="0F2D52", end_color="0F2D52", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
 
    for col_num, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_num, value=encabezado)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = Alignment(horizontal="center", vertical="center")
 
    fila = 2
    for local in locales:
        ws.cell(row=fila, column=1, value=local.numero)
        ws.cell(row=fila, column=2, value=local.get_tipo_propiedad_display())
        ws.cell(row=fila, column=3, value=local.propietario)
        ws.cell(row=fila, column=4, value=local.cliente.nombre if local.cliente else "")
        ws.cell(row=fila, column=5, value=local.ubicacion or "")
        ws.cell(row=fila, column=6, value=float(local.superficie_m2) if local.superficie_m2 is not None else None)
        ws.cell(row=fila, column=7, value=local.giro or "")
        ws.cell(row=fila, column=8, value=float(local.cuota))
        ws.cell(row=fila, column=9, value="Sí" if local.es_cuota_anual else "No")
        ws.cell(row=fila, column=10, value=local.get_status_display())
        ws.cell(row=fila, column=11, value="Sí" if local.activo else "No")
        ws.cell(row=fila, column=12, value=float(local.proindiviso) if local.proindiviso is not None else None)
        ws.cell(row=fila, column=13, value=local.grupo_facturacion.nombre if local.grupo_facturacion else "")
        ws.cell(row=fila, column=14, value=local.pool_vacancia.nombre if local.pool_vacancia else "")
        ws.cell(row=fila, column=15, value=local.referencia_pago or "")
        ws.cell(row=fila, column=16, value=local.observaciones or "")
        ws.cell(row=fila, column=17, value=local.fecha_creacion.strftime("%d/%m/%Y") if local.fecha_creacion else "")
 
        for col_num in range(1, len(encabezados) + 1):
            ws.cell(row=fila, column=col_num).font = Font(name="Arial", size=10)
 
        fila += 1
 
    for row in range(2, fila):
        celda_cuota = ws.cell(row=row, column=8)
        celda_cuota.number_format = '$#,##0.00'
 
    anchos = [12, 14, 22, 22, 22, 13, 16, 13, 12, 12, 9, 13, 20, 20, 18, 26, 15]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
 
    ws.freeze_panes = "A2"
 
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
 
    nombre_empresa = empresa.nombre.replace(" ", "_") if empresa else "GESAC"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="Propiedades_{nombre_empresa}.xlsx"'
    return response




# vistas de pools de vacancia, para asignar locales a un pool y facturar por pool 06/08/26
# ============================================================
@login_required
def lista_pools_vacancia(request):
    empresa = _empresa_actual(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect("dashboard_inicio")

    pools = (
        PoolVacancia.objects.filter(empresa=empresa)
        .select_related("cliente_cobertura")
        .prefetch_related("locales_elegibles")
    )
    return render(request, "locales/lista_pools_vacancia.html", {"pools": pools})


@login_required
def crear_pool_vacancia(request):
    empresa = _empresa_actual(request)
    if not empresa:
        messages.error(request, "No se pudo determinar tu empresa.")
        return redirect("dashboard_inicio")

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        cliente_id = request.POST.get("cliente_id")
        cliente = Cliente.objects.filter(id=cliente_id, empresa=empresa).first()

        if not nombre or not cliente:
            messages.error(request, "Captura un nombre y selecciona el cliente que cubre la vacancia.")
            return redirect("crear_pool_vacancia")

        pool = PoolVacancia.objects.create(empresa=empresa, nombre=nombre, cliente_cobertura=cliente)
        messages.success(request, "Pool de vacancia creado — ahora asigna los locales elegibles.")
        return redirect("editar_pool_vacancia", pool_id=pool.id)

    clientes = Cliente.objects.filter(empresa=empresa, activo=True).order_by("nombre")
    return render(request, "locales/crear_pool_vacancia.html", {"clientes": clientes})


@login_required
def editar_pool_vacancia(request, pool_id):
    empresa = _empresa_actual(request)
    pool = get_object_or_404(PoolVacancia, id=pool_id, empresa=empresa)

    if request.method == "POST":
        ids_seleccionados = {
            int(i) for i in request.POST.getlist("locales") if i.isdigit()
        }

        # IMPORTANTE -- este formulario SOLO muestra/edita locales con
        # status="disponible". Los que ya pertenecen al pool pero están
        # ocupados NO aparecen aquí, y por lo tanto NO deben tocarse
        # aunque no vengan en la lista enviada (si no, se quitarían
        # del pool por accidente solo por estar ocupados hoy).
        candidatos_ids = set(
            LocalComercial.objects.filter(
                empresa=empresa, activo=True, status="disponible"
            ).filter(
                Q(pool_vacancia__isnull=True) | Q(pool_vacancia=pool)
            ).values_list("id", flat=True)
        )

        a_asignar = candidatos_ids & ids_seleccionados
        a_quitar = candidatos_ids - ids_seleccionados

        if a_quitar:
            LocalComercial.objects.filter(id__in=a_quitar).update(pool_vacancia=None)
        if a_asignar:
            LocalComercial.objects.filter(id__in=a_asignar, empresa=empresa).update(pool_vacancia=pool)

        messages.success(request, "Locales elegibles del pool actualizados.")
        return redirect("editar_pool_vacancia", pool_id=pool.id)

    # SOLO locales con status="disponible" se pueden marcar/desmarcar aquí.
    # Un local ocupado no debe poder asignarse al pool directamente --
    # primero hay que liberarlo (ver aviso en el template).
    locales_disponibles = LocalComercial.objects.filter(
        empresa=empresa, activo=True, status="disponible"
    ).filter(
        Q(pool_vacancia__isnull=True) | Q(pool_vacancia=pool)
    ).order_by("numero")

    ids_actuales = set(
        LocalComercial.objects.filter(
            pool_vacancia=pool, status="disponible"
        ).values_list("id", flat=True)
    )

    # Informativo -- locales que ya pertenecen al pool pero AHORA están
    # ocupados (o en otro estado). No se editan desde aquí, solo se listan
    # para que el administrador entienda por qué no aparecen como opción.
    locales_pool_no_disponibles = LocalComercial.objects.filter(
        pool_vacancia=pool
    ).exclude(status="disponible").select_related("cliente").order_by("numero")

    return render(request, "locales/editar_pool_vacancia.html", {
        "pool": pool,
        "locales_disponibles": locales_disponibles,
        "ids_actuales": ids_actuales,
        "locales_pool_no_disponibles": locales_pool_no_disponibles,
    })



