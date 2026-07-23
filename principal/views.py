# import csv
from decimal import ROUND_HALF_UP
import locale
import os
from typing_extensions import OrderedDict

# from urllib import response
from uuid import uuid4

# import uuid
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.db import transaction

# from openai import base_url
import openpyxl

# from areas import models
from core import settings
from empleados.models import Empleado

# import empresas
from empresas.models import CuentaBancaria, Empresa
from clientes.models import Cliente
from facturacion.forms import TimbrarFacturaForm
from facturacion.utils import debe_mostrar_recordatorio_facturacion
from gastos.models import Gasto, TipoGasto
from locales.models import LocalComercial
from areas.models import AreaComun
from facturacion.models import CobroOtrosIngresos, Factura, FacturaOtrosIngresos, Pago
from presupuestos.models import Presupuesto, PresupuestoIngreso

# from principal.admin import VisitanteAccesoForm
from principal.forms import TemaGeneralForm, VisitanteLoginForm
from principal.models import AuditoriaCambio
from proveedores.models import Proveedor
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import send_mail

from .models import (
    Aviso,
    CapturarEmailForm,
    Evento,
    PerfilUsuario,
    TemaGeneral,
    VisitanteAcceso,
    VotacionCorreo,
)
import json
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from datetime import date, datetime, timedelta
import stripe
from .models import TicketMantenimiento

# from django.contrib.auth import get_user_model
from django.utils import timezone
from .models import SeguimientoTicket
from django.contrib.auth.hashers import check_password
from django.urls import reverse

# from django.conf import settings
import requests
from decimal import Decimal
from .forms import (
    AvisoForm,
    CSDUploadForm,
)
import base64
import io
import zipfile
from gastos.models import PagoGasto
from .serializers import FacturaSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from caja_chica.models import FondeoCajaChica, GastoCajaChica, ValeCaja
import logging

# from rest_framework.authtoken.models import Token
# from rest_framework.authentication import TokenAuthentication
# from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import (
    # authentication_classes,
    # permission_classes,
    parser_classes,
)
from .models import VisitanteToken
from functools import wraps
from django.db.models import Sum
from django.db.models import (
    Case,
    When,
    Value,
    CharField,
    Q,
    DecimalField,
    ExpressionWrapper,
    OuterRef,
    Subquery,
    F,
)
from django.db.models.functions import Coalesce
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.functions import ExtractMonth, ExtractYear
from django.contrib.auth.hashers import make_password
from django.utils.crypto import get_random_string
from django.core.paginator import Paginator
import weasyprint
from babel.dates import format_date
import pytz
#from django.db.models.functions import TruncMonth




# pantalla principal del sistema, con indicadores clave de desempeño (KPIs) y gráficos de resumen

# def dashboard_inicio(request):
#     empresa = request.user.perfilusuario.empresa
#     perfil = request.user.perfilusuario
#     # mostrar_wizard = perfil.mostrar_wizard
#     mostrar_wizard = perfil.mostrar_wizard and not request.session.get(
#         "wizard_cerrado", False
#     )

#     hoy = date.today()
#     primer_dia_mes = hoy.replace(day=1)
#     primer_dia_mes_anterior = (primer_dia_mes - timedelta(days=1)).replace(day=1)
#     ultimo_dia_mes_anterior = primer_dia_mes - timedelta(days=1)
#     meses = []
#     for i in range(5, -1, -1):
#         y = (
#             primer_dia_mes.year
#             if primer_dia_mes.month - i > 0
#             else primer_dia_mes.year - 1
#         )
#         m = (primer_dia_mes.month - i - 1) % 12 + 1
#         meses.append(date(y, m, 1))

#     # Ingresos del mes actual y anterior
#     ingresos_mes = (
#         Pago.objects.filter(
#             factura__empresa=empresa,
#             fecha_pago__gte=primer_dia_mes,
#             fecha_pago__lte=hoy,
#         ).aggregate(total=Sum("monto"))["total"]
#         or 0
#     )

#     ingresos_mes_otros = (
#         CobroOtrosIngresos.objects.filter(
#             factura__empresa=empresa,
#             fecha_cobro__gte=primer_dia_mes,
#             fecha_cobro__lte=hoy,
#         ).aggregate(total=Sum("monto"))["total"]
#         or 0
#     )

#     ingresos_mes_total = ingresos_mes + ingresos_mes_otros

#     ingresos_mes_anterior = (
#         Pago.objects.filter(
#             factura__empresa=empresa,
#             fecha_pago__gte=primer_dia_mes_anterior,
#             fecha_pago__lte=ultimo_dia_mes_anterior,
#         ).aggregate(total=Sum("monto"))["total"]
#         or 0
#     )

#     ingresos_mes_anterior_otros = (
#         CobroOtrosIngresos.objects.filter(
#             factura__empresa=empresa,
#             fecha_cobro__gte=primer_dia_mes_anterior,
#             fecha_cobro__lte=ultimo_dia_mes_anterior,
#         ).aggregate(total=Sum("monto"))["total"]
#         or 0
#     )

#     ingresos_mes_anterior_total = ingresos_mes_anterior + ingresos_mes_anterior_otros

#     variacion_ingresos = (
#         (
#             (ingresos_mes_total - ingresos_mes_anterior_total)
#             / ingresos_mes_anterior_total
#         )
#         * 100
#         if ingresos_mes_anterior_total
#         else 0
#     )

#     # Gastos del mes actual y anterior
#     gastos_mes = (
#         PagoGasto.objects.filter(
#             gasto__empresa=empresa, fecha_pago__gte=primer_dia_mes, fecha_pago__lte=hoy
#         ).aggregate(total=Sum("monto"))["total"]
#         or 0
#     )

#     gastos_mes_anterior = (
#         PagoGasto.objects.filter(
#             gasto__empresa=empresa,
#             fecha_pago__gte=primer_dia_mes_anterior,
#             fecha_pago__lte=ultimo_dia_mes_anterior,
#         ).aggregate(total=Sum("monto"))["total"]
#         or 0
#     )

#     variacion_gastos = (
#         ((gastos_mes - gastos_mes_anterior) / gastos_mes_anterior) * 100
#         if gastos_mes_anterior
#         else 0
#     )

#     # Facturas pendientes y vencidas cuotas y áreas comunes
#     facturas_pendientes = (
#         Factura.objects.filter(empresa=empresa, estatus="pendiente")
#         .select_related("cliente")
#         .prefetch_related("pagos")
#     )

#     # facturas pendientes otros ingresos
#     facturas_pendientes_otros = (
#         FacturaOtrosIngresos.objects.filter(empresa=empresa, estatus="pendiente")
#         .select_related("tipo_ingreso")
#         .prefetch_related("cobros")
#     )

#     facturas_vencidas_cuotas = facturas_pendientes.filter(fecha_vencimiento__lt=hoy)
#     facturas_vencidas_otros = facturas_pendientes_otros.filter(
#         fecha_vencimiento__lt=hoy
#     )
#     cartera_vencida = sum(f.saldo_pendiente for f in facturas_vencidas_cuotas) + sum(
#         f.saldo for f in facturas_vencidas_otros
#     )

#     # Cantidad de facturas pendientes
#     cantidad_facturas_pendientes = (
#         facturas_pendientes.count() + facturas_pendientes_otros.count()
#     )

#     # Pagos recibidos hoy
#     pagos_hoy = (
#         Pago.objects.filter(empresa=empresa, fecha_pago=hoy).aggregate(
#             total=Sum("monto")
#         )["total"]
#         or 0
#     )

#     pagos_por_mes = (
#         Pago.objects.filter(
#             factura__empresa=empresa, fecha_pago__gte=meses[0], fecha_pago__lte=hoy
#         )
#         .annotate(mes=TruncMonth("fecha_pago"))
#         .values("mes")
#         .annotate(total=Sum("monto"))
#         .order_by("mes")
#     )

#     # Pagos de gastos generales por mes
#     gastos_por_mes = (
#         PagoGasto.objects.filter(
#             gasto__empresa=empresa, fecha_pago__gte=meses[0], fecha_pago__lte=hoy
#         )
#         .annotate(mes=TruncMonth("fecha_pago"))
#         .values("mes")
#         .annotate(total=Sum("monto"))
#         .order_by("mes")
#     )
#     # gastos y vales caja chica por mes
#     gastos_caja_por_mes = (
#         GastoCajaChica.objects.filter(
#             fondeo__empresa=empresa, fecha__gte=meses[0], fecha__lte=hoy
#         )
#         .annotate(mes=TruncMonth("fecha"))
#         .values("mes")
#         .annotate(total=Sum("importe"))
#         .order_by("mes")
#     )
#     vales_caja_por_mes = (
#         ValeCaja.objects.filter(
#             fondeo__empresa=empresa, fecha__gte=meses[0], fecha__lte=hoy
#         )
#         .annotate(mes=TruncMonth("fecha"))
#         .values("mes")
#         .annotate(total=Sum("importe"))
#         .order_by("mes")
#     )

#     # Convierte los resultados a diccionarios {mes: total}
#     pagos_dict = {}
#     for p in pagos_por_mes:
#         if p["mes"]:
#             pagos_dict[p["mes"].strftime("%Y-%m")] = float(p["total"])

#     pagosgasto_dict = {}
#     for p in gastos_por_mes:
#         if p["mes"]:
#             pagosgasto_dict[p["mes"].strftime("%Y-%m")] = float(p["total"])

#     labels = [m.strftime("%b %Y") for m in meses]
#     labels_keys = [m.strftime("%Y-%m") for m in meses]
#     ingresos_data = [pagos_dict.get(k, 0) for k in labels_keys]
#     gastos_data = [pagosgasto_dict.get(k, 0) for k in labels_keys]

#     # Cartera vencida por rango de días
#     rangos = [
#         (0, 30),
#         (31, 60),
#         (61, 90),
#         (91, 180),
#         (181, 10000),
#     ]
#     cartera_rangos = []
#     for inicio, fin in rangos:
#         facturas_rango = facturas_pendientes.filter(
#             fecha_vencimiento__lt=hoy - timedelta(days=inicio),
#             fecha_vencimiento__gte=hoy - timedelta(days=fin),
#         )
#         total = sum(f.saldo_pendiente for f in facturas_rango)
#         cartera_rangos.append(total)

#     # --- Recordatorio de facturación mensual ---
#     mostrar_recordatorio = debe_mostrar_recordatorio_facturacion(empresa)

#     mensaje_pago = None
#     if request.GET.get("pago") == "ok":
#         mensaje_pago = "¡Tu suscripción se ha activado correctamente! Puedes empezar a usar el sistema."

#     # es_demo=False
#     # es_demo = request.user.perfilusuario.tipo_usuario == "demo"
#     # es_plus = request.user.perfilusuario.tipo_usuario == "plus"
#     # es_premium = request.user.empresa.es_premium == "premium"

#     context = {
#         "ingresos_mes": ingresos_mes,
#         "ingresos_mes_otros": ingresos_mes_otros,
#         "ingresos_mes_total": ingresos_mes_total,
#         "ingresos_mes_anterior": ingresos_mes_anterior,
#         "ingresos_mes_anterior_otros": ingresos_mes_anterior_otros,
#         "ingresos_mes_anterior_total": ingresos_mes_anterior_total,
#         "variacion_ingresos": variacion_ingresos,
#         "gastos_mes": gastos_mes,
#         "gastos_mes_anterior": gastos_mes_anterior,
#         "variacion_gastos": variacion_gastos,
#         "cartera_vencida": cartera_vencida,
#         "facturas_pendientes": cantidad_facturas_pendientes,
#         "pagos_hoy": pagos_hoy,
#         #'top_deudores': top_deudores,
#         "pagos_por_mes": list(pagos_por_mes),
#         "gastos_por_mes": list(gastos_por_mes),
#         "cartera_rangos": cartera_rangos,
#         "labels": labels,
#         "ingresos_data": ingresos_data,
#         "gastos_data": gastos_data,
#         #"es_demo": es_demo,
#         "STRIPE_PUBLIC_KEY": settings.STRIPE_PUBLIC_KEY,
#         "mensaje_pago": mensaje_pago,
#         "mostrar_recordatorio": mostrar_recordatorio,
#         "mostrar_wizard": mostrar_wizard,
#         #"es_plus": es_plus,
#         #"es_premium": es_premium,
#         "regimen_choices": Empresa.REGIMEN_CHOICES,
#         "bancos_choices": CuentaBancaria.BANCOS_CHOICES,
#         "tipo_cuenta_choices": CuentaBancaria.TIPO_CUENTA,
#         "moneda_choices": CuentaBancaria.TIPO_MONEDA,
#         "empresa": empresa,
#     }
#     return render(request, "pantalla_inicio.html", context)
MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

@login_required
def dashboard_inicio(request):
    hoy = date.today()
    mes_actual = hoy.month
    anio_actual = hoy.year
 
    # Empresa
    if request.user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        empresa = Empresa.objects.filter(id=empresa_id).first()
    else:
        empresa = request.user.perfilusuario.empresa
 
    if not empresa:
        return render(request, 'dashboard.html', {'empresa': None})
 
    # ── INGRESOS DEL MES ──
    ingresos_mes = Pago.objects.filter(
        factura__empresa=empresa,
        fecha_pago__year=anio_actual,
        fecha_pago__month=mes_actual
    ).exclude(forma_pago='nota_credito').aggregate(total=Sum('monto'))['total'] or Decimal('0')
 
    # Mes anterior para delta
    mes_ant = mes_actual - 1 if mes_actual > 1 else 12
    anio_ant = anio_actual if mes_actual > 1 else anio_actual - 1
    ingresos_mes_ant = Pago.objects.filter(
        factura__empresa=empresa,
        fecha_pago__year=anio_ant,
        fecha_pago__month=mes_ant
    ).exclude(forma_pago='nota_credito').aggregate(total=Sum('monto'))['total'] or Decimal('0')
    delta_ingresos = round(float((ingresos_mes - ingresos_mes_ant) / ingresos_mes_ant * 100), 1) if ingresos_mes_ant > 0 else None
 
    # Por cobrar del mes (facturas pendientes con vencimiento en el mes)
    por_cobrar_mes = Factura.objects.filter(
        empresa=empresa,
        estatus='pendiente',
        activo=True,
        fecha_vencimiento__year=anio_actual,
        fecha_vencimiento__month=mes_actual
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
 
    total_posible_ingresos = ingresos_mes + por_cobrar_mes
    pct_cobrado = round(float(ingresos_mes / total_posible_ingresos * 100)) if total_posible_ingresos > 0 else 0
    pct_por_cobrar = 100 - pct_cobrado
 
    # ── GASTOS DEL MES ──
    gastos_pagados_mes = PagoGasto.objects.filter(
        gasto__empresa=empresa,
        fecha_pago__year=anio_actual,
        fecha_pago__month=mes_actual,
        monto__gt=0
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
 
    gastos_pendientes_mes = Gasto.objects.filter(
        empresa=empresa,
        fecha__year=anio_actual,
        fecha__month=mes_actual,
        estatus='pendiente'
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
 
    gastos_mes_ant = PagoGasto.objects.filter(
        gasto__empresa=empresa,
        fecha_pago__year=anio_ant,
        fecha_pago__month=mes_ant,
        monto__gt=0
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    delta_gastos = round(float((gastos_pagados_mes - gastos_mes_ant) / gastos_mes_ant * 100), 1) if gastos_mes_ant > 0 else None
 
    total_gastos_mes = gastos_pagados_mes + gastos_pendientes_mes
    resultado_mes = ingresos_mes - gastos_pagados_mes
    pct_gastos = round(float(gastos_pagados_mes / ingresos_mes * 100)) if ingresos_mes > 0 else 0
    pct_margen = 100 - pct_gastos
 
    # ── CARTERA VENCIDA ──
    # facturas_pendientes = Factura.objects.filter(
    #     empresa=empresa, estatus='pendiente', activo=True
    # ).annotate_saldo()  # si no tienes annotate_saldo usa values/aggregate
 
    # Calcular cartera manualmente si no hay annotate_saldo
    
    facturas_pend_qs = Factura.objects.filter(
        empresa=empresa, estatus='pendiente', activo=True
    )
    cartera_vencida = facturas_pend_qs.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    facturas_vencidas_count = facturas_pend_qs.count()
 
    # Antigüedad por días desde fecha_vencimiento
   
    fecha_30 = hoy - timedelta(days=30)
    fecha_60 = hoy - timedelta(days=60)
    fecha_90 = hoy - timedelta(days=90)
    fecha_180 = hoy - timedelta(days=180)
 
    saldo_0_30   = facturas_pend_qs.filter(fecha_vencimiento__gte=fecha_30).aggregate(t=Sum('monto'))['t'] or 0
    saldo_31_60  = facturas_pend_qs.filter(fecha_vencimiento__lt=fecha_30, fecha_vencimiento__gte=fecha_60).aggregate(t=Sum('monto'))['t'] or 0
    saldo_61_90  = facturas_pend_qs.filter(fecha_vencimiento__lt=fecha_60, fecha_vencimiento__gte=fecha_90).aggregate(t=Sum('monto'))['t'] or 0
    saldo_91_180 = facturas_pend_qs.filter(fecha_vencimiento__lt=fecha_90, fecha_vencimiento__gte=fecha_180).aggregate(t=Sum('monto'))['t'] or 0
    saldo_181_mas = facturas_pend_qs.filter(fecha_vencimiento__lt=fecha_180).aggregate(t=Sum('monto'))['t'] or 0
    saldo_31_90  = float(saldo_31_60) + float(saldo_61_90)
    saldo_90_mas = float(saldo_91_180) + float(saldo_181_mas)
 
    # ── TOP DEUDORES ──
    top_deudores = []
    for f in facturas_pend_qs.select_related('cliente', 'local', 'area_comun').order_by('-monto')[:8]:
        dias = (hoy - f.fecha_vencimiento).days if f.fecha_vencimiento < hoy else 0
        top_deudores.append({
            'cliente__nombre': f.cliente.nombre if f.cliente else '—',
            'local__numero': f.local.numero if f.local else None,
            'area_comun__numero': f.area_comun.numero if f.area_comun else None,
            'saldo': float(f.monto),
            'dias_vencido': max(dias, 0),
        })
 
    # ── DATOS 6 MESES ──
    meses_6 = []
    ingresos_cobrados_6 = []
    ingresos_porcobrar_6 = []
    gastos_pagados_6 = []
    gastos_pendientes_6 = []
 
    for i in range(5, -1, -1):
        # Calcular mes/año retrocediendo
        m = (mes_actual - i - 1) % 12 + 1
        a = anio_actual - ((i - mes_actual + 1) // 12 + (1 if (i - mes_actual + 1) % 12 >= 0 and i >= mes_actual else 0))
        # Forma más simple y correcta:
        target = hoy.replace(day=1) - timedelta(days=30*i)
        m, a = target.month, target.year
 
        meses_6.append(MESES[m-1])
 
        cobrado = Pago.objects.filter(
            factura__empresa=empresa, fecha_pago__year=a, fecha_pago__month=m
        ).exclude(forma_pago='nota_credito').aggregate(t=Sum('monto'))['t'] or 0
 
        porcobrar = Factura.objects.filter(
            empresa=empresa, estatus='pendiente', activo=True,
            fecha_vencimiento__year=a, fecha_vencimiento__month=m
        ).aggregate(t=Sum('monto'))['t'] or 0
 
        gpagados = PagoGasto.objects.filter(
            gasto__empresa=empresa, fecha_pago__year=a, fecha_pago__month=m, monto__gt=0
        ).aggregate(t=Sum('monto'))['t'] or 0
 
        gpendientes = Gasto.objects.filter(
            empresa=empresa, fecha__year=a, fecha__month=m, estatus='pendiente'
        ).aggregate(t=Sum('monto'))['t'] or 0
 
        ingresos_cobrados_6.append(float(cobrado))
        ingresos_porcobrar_6.append(float(porcobrar))
        gastos_pagados_6.append(float(gpagados))
        gastos_pendientes_6.append(float(gpendientes))
    
    # Membresía
    membresia_label = 'Premium'
    try:
        nivel = empresa.perfil.nivel if hasattr(empresa, 'perfil') else None
        if nivel:
            membresia_label = nivel.capitalize()
    except Exception:
        pass

    mensaje_pago = None
    if request.GET.get("pago") == "ok":
        mensaje_pago = "¡Tu suscripción se ha activado correctamente! Puedes empezar a usar el sistema."    

    mostrar_recordatorio = debe_mostrar_recordatorio_facturacion(empresa)

    mostrar_wizard = request.user.perfilusuario.mostrar_wizard and not request.session.get("wizard_cerrado", False)

    return render(request, 'pantalla_inicio.html', {
        'empresa': empresa,
        'hoy': hoy,
        'mes_nombre': MESES[mes_actual-1],
        # Ingresos
        'ingresos_mes': ingresos_mes,
        'por_cobrar_mes': por_cobrar_mes,
        'total_posible_ingresos': total_posible_ingresos,
        'pct_cobrado': pct_cobrado,
        'pct_por_cobrar': pct_por_cobrar,
        'delta_ingresos': delta_ingresos,
        # Gastos
        'gastos_pagados_mes': gastos_pagados_mes,
        'gastos_pendientes_mes': gastos_pendientes_mes,
        'total_gastos_mes': total_gastos_mes,
        'delta_gastos': delta_gastos,
        # Resultado
        'resultado_mes': resultado_mes,
        'pct_gastos': pct_gastos,
        'pct_margen': pct_margen,
        # Cartera
        'cartera_vencida': cartera_vencida,
        'facturas_vencidas_count': facturas_vencidas_count,
        'saldo_0_30': saldo_0_30,
        'saldo_31_60': saldo_31_60,
        'saldo_61_90': saldo_61_90,
        'saldo_91_180': saldo_91_180,
        'saldo_181_mas': saldo_181_mas,
        'saldo_31_90': saldo_31_90,
        'saldo_90_mas': saldo_90_mas,
        # Top deudores
        'top_deudores': top_deudores,
        # Gráficas 6 meses
        'meses_6': json.dumps(meses_6),
        'ingresos_cobrados_6': json.dumps(ingresos_cobrados_6),
        'ingresos_porcobrar_6': json.dumps(ingresos_porcobrar_6),
        'gastos_pagados_6': json.dumps(gastos_pagados_6),
        'gastos_pendientes_6': json.dumps(gastos_pendientes_6),
        # Membresía
        'membresia_label': membresia_label,
        # Recordatorio facturación
        'mostrar_recordatorio': mostrar_recordatorio,
        'mostrar_wizard': mostrar_wizard,
        'mensaje_pago': mensaje_pago,
    })



@login_required
def info_plus(request):
    return render(request, 'planes/info_plus.html', {
        'stripe_public_key': settings.STRIPE_PUBLIC_KEY,
    })

@staff_member_required
@login_required
def reiniciar_sistema(request):
    if request.method == "POST":
        empresa_id = request.POST.get("empresa_id") or request.session.get("empresa_id")
        if not empresa_id:
            messages.error(request, "Debes seleccionar una empresa.")
            return redirect("reiniciar_sistema")

        try:
            empresa = Empresa.objects.get(pk=empresa_id)
        except Empresa.DoesNotExist:
            messages.error(request, "Empresa no encontrada.")
            return redirect("reiniciar_sistema")

        try:
            with transaction.atomic():
                # 1) Romper bloqueos PROTECT por cliente
                AreaComun.objects.filter(empresa=empresa).update(cliente=None)
                LocalComercial.objects.filter(empresa=empresa).update(cliente=None)

                # 2) Facturacion otros ingresos (tiene cliente PROTECT)
                CobroOtrosIngresos.objects.filter(factura__empresa=empresa).delete()
                FacturaOtrosIngresos.objects.filter(empresa=empresa).delete()

                # 3) Pagos por identificar (sin factura) de la empresa
                Pago.objects.filter(empresa=empresa, factura__isnull=True).delete()

                # 4) Facturas y pagos ligados a facturas
                Factura.objects.filter(empresa=empresa).delete()

                # 5) Caja chica
                FondeoCajaChica.objects.filter(
                    empresa=empresa
                ).delete()  # cascada a gastos/vales

                # 6) Gastos
                Gasto.objects.filter(empresa=empresa).delete()
                PagoGasto.objects.filter(gasto__empresa=empresa).delete()
                TipoGasto.objects.filter(empresa=empresa).delete()

                # 7) Catalogos y maestros por empresa
                Presupuesto.objects.filter(empresa=empresa).delete()
                PresupuestoIngreso.objects.filter(empresa=empresa).delete()
                Proveedor.objects.filter(empresa=empresa).delete()
                Empleado.objects.filter(empresa=empresa).delete()
                Cliente.objects.filter(empresa=empresa).delete()
                LocalComercial.objects.filter(empresa=empresa).delete()
                AreaComun.objects.filter(empresa=empresa).delete()
                Evento.objects.filter(empresa=empresa).delete()
                TemaGeneral.objects.filter(empresa=empresa).delete()
                Aviso.objects.filter(empresa=empresa).delete()

                # 8) Dejar saldos en cero (empresa sigue activa)
                # empresa.saldo_inicial = 0
                # empresa.saldo_final = 0
                # empresa.save(update_fields=["saldo_inicial", "saldo_final"])

            messages.success(
                request, f"Empresa {empresa.nombre} reiniciada en cero correctamente."
            )
        except Exception as e:
            messages.error(request, f"Error al reiniciar empresa: {e}")

        return redirect("reiniciar_sistema")

    empresas = Empresa.objects.all().order_by("nombre")
    empresa_id = request.session.get("empresa_id")
    return render(
        request,
        "reiniciar_sistema.html",
        {"empresas": empresas, "empresa_id": empresa_id},
    )



@staff_member_required
def respaldo_empresa_excel(request):
    # Si no hay empresa seleccionada, muestra el formulario
    if request.method == "GET" and "empresa_id" not in request.GET:
        empresas = Empresa.objects.all()
        return render(request, "respaldo_empresas.html", {"empresas": empresas})

    empresa_id = request.GET.get("empresa_id")
    try:
        empresa = Empresa.objects.get(pk=empresa_id)
    except Empresa.DoesNotExist:
        return render(
            request,
            "respaldo_empresas.html",
            {"empresas": Empresa.objects.all(), "error": "Empresa no encontrada."},
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Borra hoja por defecto

    # CLIENTES
    ws = wb.create_sheet("Clientes")
    ws.append(["id", "nombre", "rfc", "telefono", "email", "activo"])
    for c in Cliente.objects.filter(empresa=empresa):
        ws.append([c.id, c.nombre, c.rfc, c.telefono, c.email, c.activo])

    # LOCALES
    ws = wb.create_sheet("Locales")
    ws.append(
        [
            "id",
            "numero",
            "cliente",
            "ubicacion",
            "superficie_m2",
            "cuota",
            "status",
            "activo",
            "observaciones",
        ]
    )
    for l in LocalComercial.objects.filter(empresa=empresa):
        ws.append(
            [
                l.id,
                l.numero,
                l.cliente.nombre if l.cliente else "",
                l.ubicacion,
                l.superficie_m2,
                l.cuota,
                l.status,
                l.activo,
                l.observaciones,
            ]
        )

    # ÁREAS COMUNES
    ws = wb.create_sheet("Áreas Comunes")
    ws.append(
        [
            "cliente",
            "numero",
            "cuota",
            "ubicacion",
            "superficie_m2",
            "status",
            "fecha_inicial",
            "fecha_fin",
            "activo",
            "observaciones",
        ]
    )
    for a in AreaComun.objects.filter(empresa=empresa):
        ws.append(
            [
                a.cliente.nombre if a.cliente else "",
                a.numero,
                a.cuota,
                a.ubicacion,
                a.superficie_m2,
                a.status,
                str(a.fecha_inicial) if a.fecha_inicial else "",
                str(a.fecha_fin) if a.fecha_fin else "",
                a.activo,
                a.observaciones,
            ]
        )

    # FACTURAS
    ws = wb.create_sheet("Facturas")
    ws.append(
        [
            "folio",
            "cliente",
            "local",
            "area_comun",
            "monto",
            "fecha_emision",
            "fecha_vencimiento",
            "estatus",
        ]
    )
    for f in Factura.objects.filter(empresa=empresa):
        ws.append(
            [
                f.folio,
                f.cliente.nombre if f.cliente else "",
                f.local.numero if f.local else "",
                f.area_comun.numero if f.area_comun else "",
                float(f.monto),
                str(f.fecha_emision),
                str(f.fecha_vencimiento),
                f.estatus,
            ]
        )

    # PAGOS
    ws = wb.create_sheet("Pagos")
    ws.append(["id", "factura", "fecha_pago", "monto", "registrado_por"])
    for p in Pago.objects.filter(factura__empresa=empresa):
        ws.append(
            [
                p.id,
                p.factura.folio if p.factura else "",
                str(p.fecha_pago),
                float(p.monto),
                p.registrado_por.get_full_name() if p.registrado_por else "",
            ]
        )

    # GASTOS
    ws = wb.create_sheet("Gastos")
    ws.append(
        ["id", "proveedor", "empleado", "descripcion", "monto", "tipo_gasto", "fecha"]
    )
    for g in Gasto.objects.filter(empresa=empresa):
        ws.append(
            [
                g.id,
                str(g.proveedor) if g.proveedor else "",
                str(g.empleado) if g.empleado else "",
                g.descripcion,
                float(g.monto),
                str(g.tipo_gasto) if g.tipo_gasto else "",
                str(g.fecha),
            ]
        )
    # CAJA CHICA - FONDEOS
    ws = wb.create_sheet("Fondeos Caja Chica")
    ws.append(["id", "empresa", "monto", "fecha", "registrado_por"])
    for f in FondeoCajaChica.objects.filter(empresa=empresa):
        ws.append(
            [
                f.id,
                f.empresa.nombre if f.empresa else "",
                float(f.monto),
                str(f.fecha),
                f.registrado_por.get_full_name() if f.registrado_por else "",
            ]
        )
    # CAJA CHICA - GASTOS
    ws = wb.create_sheet("Gastos Caja Chica")
    ws.append(["id", "fondeo", "descripcion", "importe", "fecha", "registrado_por"])
    for g in GastoCajaChica.objects.filter(fondeo__empresa=empresa):
        ws.append(
            [
                g.id,
                g.fondeo.id if g.fondeo else "",
                g.descripcion,
                float(g.importe),
                str(g.fecha),
                g.registrado_por.get_full_name() if g.registrado_por else "",
            ]
        )
    # CAJA CHICA - VALES
    ws = wb.create_sheet("Vales Caja Chica")
    ws.append(["id", "fondeo", "beneficiario", "importe", "fecha", "registrado_por"])
    for v in ValeCaja.objects.filter(fondeo__empresa=empresa):
        ws.append(
            [
                v.id,
                v.fondeo.id if v.fondeo else "",
                v.beneficiario,
                float(v.importe),
                str(v.fecha),
                v.registrado_por.get_full_name() if v.registrado_por else "",
            ]
        )

    # PAGOS GASTOS
    ws = wb.create_sheet("Pagos Gastos")
    ws.append(["id", "referencia", "fecha_pago", "monto", "registrado_por"])
    for g in Gasto.objects.filter(empresa=empresa):
        for pago in g.pagos.all():
            ws.append(
                [
                    pago.id,
                    pago.referencia,
                    str(pago.fecha_pago),
                    float(pago.monto),
                    pago.registrado_por.get_full_name() if pago.registrado_por else "",
                ]
            )

    # PRESUPUESTOS
    ws = wb.create_sheet("Presupuestos")
    ws.append(
        ["id", "empresa", "grupo", "subgrupo", "tipo_gasto", "anio", "mes", "monto"]
    )
    for p in Presupuesto.objects.filter(empresa=empresa):
        ws.append(
            [
                p.id,
                p.empresa.nombre if p.empresa else "",
                str(p.grupo) if p.grupo else "",
                str(p.subgrupo) if p.subgrupo else "",
                str(p.tipo_gasto) if p.tipo_gasto else "",
                p.anio,
                p.mes,
                float(p.monto),
            ]
        )
        # PRESUPUESTOS INGRESOS
    ws = wb.create_sheet("Presupuestos Ingresos")
    ws.append(["id", "empresa", "tipo_ingreso", "anio", "mes", "monto"])
    for p in PresupuestoIngreso.objects.filter(empresa=empresa):
        ws.append(
            [
                p.id,
                p.empresa.nombre if p.empresa else "",
                str(p.tipo_ingreso) if hasattr(p, "tipo_ingreso") else "",
                p.anio,
                p.mes,
                float(p.monto),
            ]
        )

    # EMPLEADOS
    ws = wb.create_sheet("Empleados")
    ws.append(["id", "nombre", "email", "telefono", "puesto", "activo"])
    for e in Empleado.objects.filter(empresa=empresa):
        ws.append([e.id, e.nombre, e.email, e.telefono, e.puesto, e.activo])

    # PROVEEDORES
    ws = wb.create_sheet("Proveedores")
    ws.append(["id", "nombre", "rfc", "telefono", "email", "activo"])
    for p in Proveedor.objects.filter(empresa=empresa):
        ws.append([p.id, str(p.nombre), p.rfc, p.telefono, p.email, p.activo])

    # OTROS INGRESOS
    ws = wb.create_sheet("Otros Ingresos")
    ws.append(
        [
            "id",
            "folio",
            "cliente",
            "tipo_ingreso",
            "monto",
            "saldo",
            "fecha_emision",
            "fecha_vencimiento",
            "estatus",
            "observaciones",
        ]
    )
    for f in FacturaOtrosIngresos.objects.filter(empresa=empresa):
        ws.append(
            [
                f.id,
                f.folio,
                f.cliente.nombre if f.cliente else "",
                str(f.tipo_ingreso) if hasattr(f, "tipo_ingreso") else "",
                float(f.monto),
                float(f.saldo),
                str(f.fecha_emision),
                str(f.fecha_vencimiento) if f.fecha_vencimiento else "",
                f.estatus,
                f.observaciones or "",
            ]
        )

    # PAGOS OTROS INGRESOS
    ws = wb.create_sheet("Pagos Otros Ingresos")
    ws.append(["id", "factura", "fecha_pago", "monto", "registrado_por"])
    for p in CobroOtrosIngresos.objects.filter(factura__empresa=empresa):
        ws.append(
            [
                p.id,
                p.factura.folio if p.factura else "",
                str(p.fecha_cobro),
                float(p.monto),
                p.registrado_por.get_full_name() if p.registrado_por else "",
            ]
        )

    # Responde el archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=respaldo_empresa_{empresa.nombre}.xlsx"
    )
    wb.save(response)
    return response



@staff_member_required
def reporte_auditoria(request):
    modelo = request.GET.get("modelo")
    queryset = AuditoriaCambio.objects.all().order_by("-fecha_cambio")
    if modelo in ["local", "area", "factura"]:
        queryset = queryset.filter(modelo=modelo)
    return render(
        request, "auditoria/reporte.html", {"auditorias": queryset, "modelo": modelo}
    )

@csrf_exempt
@login_required
def crear_evento(request):
    if request.method == "POST":
        empresa = request.user.perfilusuario.empresa
        data = json.loads(request.body)
        evento = Evento.objects.create(
            empresa=empresa,
            titulo=data.get("titulo"),
            fecha=data.get("fecha"),
            descripcion=data.get("descripcion"),
            creado_por=request.user,
        )

        evento.save()
        return JsonResponse({"ok": True, "id": evento.id})
    return JsonResponse({"ok": False}, status=400)



@csrf_exempt
@login_required
def eliminar_evento(request, evento_id):
    if request.method == "POST":
        try:
            evento = Evento.objects.get(
                id=evento_id, empresa=request.user.perfilusuario.empresa
            )
            evento.delete()
            return JsonResponse({"ok": True})
        except Evento.DoesNotExist:
            return JsonResponse({"ok": False, "error": "No encontrado"}, status=404)
    return JsonResponse({"ok": False}, status=400)


@csrf_exempt
@login_required
def enviar_correo_evento(request, evento_id):
    if request.method == "POST":
        correo_destino = request.POST.get("correo")
        archivos = request.FILES.getlist("archivos")
        try:
            evento = Evento.objects.get(
                id=evento_id, empresa=request.user.perfilusuario.empresa
            )
            if correo_destino:
                cuerpo_html = render_to_string(
                    "correo_evento.html", {"evento": evento, "empresa": evento.empresa}
                )
                email = EmailMessage(
                    subject=f"Nuevo evento: {evento.titulo}",
                    body=cuerpo_html,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[correo_destino],
                )
                email.content_subtype = "html"
                for archivo in archivos:
                    email.attach(archivo.name, archivo.read(), archivo.content_type)
                email.send(fail_silently=False)
                evento.enviado_correo = True
                evento.save()
                return JsonResponse({"ok": True})
            else:
                return JsonResponse(
                    {"ok": False, "error": "Correo no proporcionado"}, status=400
                )
        except Evento.DoesNotExist:
            return JsonResponse(
                {"ok": False, "error": "Evento no encontrado"}, status=404
            )
    return JsonResponse({"ok": False}, status=400)


# registro usuario demo, crea empresa demo y asigna perfil de usuario demo GESAC
def registro_usuario(request):
    mensaje = ""
    if request.method == "POST":
        nombre = request.POST["nombre"]
        username = request.POST["username"]
        password = request.POST["password"]
        email = request.POST["email"]
        # telefono = request.POST['telefono']

        if User.objects.filter(username=username).exists():
            mensaje = "El nombre de usuario ya está en uso. Por favor elige otro."
        else:
            user = User.objects.create_user(
                username=username, password=password, email=email, first_name=nombre
            )
            empresa = Empresa.objects.create(
                nombre="EMPRESA DEMO AC", rfc=f"DEMO{uuid4().hex[:8].upper()}"
            )
            # PerfilUsuario.objects.create(
            #     usuario=user, empresa=empresa, tipo_usuario="demo"
            # )
            # Asigna la empresa y tipo_usuario al perfil creado por la señal
            perfil = user.perfilusuario
            perfil.empresa = empresa
            if not user.is_superuser:
                perfil.tipo_usuario = "demo"
            perfil.save()
            messages.success(
                request,
                "¡Registro exitoso! Por favor inicia sesión con tus credenciales.",
            )
            return redirect("login")
    return render(request, "registro.html", {"mensaje": mensaje})


# lista usurios demo
@staff_member_required
@login_required
def usuarios_demo(request):
    usuarios = User.objects.filter(perfilusuario__tipo_usuario="demo")
    usuarios_info = []
    for user in usuarios:
        dias = (date.today() - user.date_joined.date()).days
        usuarios_info.append(
            {
                "id": user.id,
                "username": user.username,
                "first_name": user.first_name,
                "email": user.email,
                "is_active": user.is_active,
                "dias_demo": dias,
            }
        )
    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion == "inactivar":
            ids = request.POST.getlist("inactivar")
            User.objects.filter(id__in=ids).update(is_active=False)
        elif accion == "reactivar":
            ids = request.POST.getlist("reactivar")
            User.objects.filter(id__in=ids).update(is_active=True)
        elif accion == "reactivar_todos":
            User.objects.filter(perfilusuario__tipo_usuario="demo").update(
                is_active=True
            )
        return redirect("usuarios_demo")
    return render(request, "usuarios_demo.html", {"usuarios": usuarios_info})


# --- Helpers membresias Stripe ---
def _sync_membership_state(perfil):
    """
    Regla de negocio:
    - premium activo -> tipo premium
    - plus activo (sin premium) -> tipo plus
    - sin subs -> demo
    """
    tiene_plus = bool(perfil.stripe_plus_subscription_id)
    tiene_premium = bool(perfil.stripe_premium_subscription_id)

    if tiene_premium:
        perfil.tipo_usuario = "premium"
    elif tiene_plus:
        perfil.tipo_usuario = "plus"
    else:
        perfil.tipo_usuario = "demo"

    if perfil.empresa:
        # premium implica acceso plus tambien
        perfil.empresa.es_premium = tiene_premium
        perfil.empresa.es_plus = tiene_plus
        perfil.empresa.save(update_fields=["es_plus", "es_premium"])


def _find_perfil_from_checkout_session(session):
    customer_id = session.get("customer")
    email = session.get("customer_details", {}).get("email")
    client_ref = session.get("client_reference_id")

    perfil = None

    if customer_id:
        perfil = PerfilUsuario.objects.filter(stripe_customer_id=customer_id).first()

    if not perfil and email:
        user = User.objects.filter(email=email).first()
        if user:
            perfil = PerfilUsuario.objects.filter(usuario=user).first()

    if not perfil and client_ref:
        try:
            user = User.objects.filter(id=int(client_ref)).first()
            if user:
                perfil = PerfilUsuario.objects.filter(usuario=user).first()
        except Exception:
            pass

    return perfil


def _extract_price_id_from_session(session_id):
    try:
        line_items = stripe.checkout.Session.list_line_items(session_id)
        if line_items.data:
            return line_items.data[0].price.id
    except Exception as e:
        print("Error al obtener line items:", e)
    return None


def _safe_cancel_subscription(subscription_id):
    """
    Cancela en Stripe; si ya no existe, no rompe el flujo local.
    """
    if not subscription_id:
        return
    try:
        stripe.Subscription.delete(subscription_id)
    except Exception as e:
        if "No such subscription" not in str(e):
            raise


# Webhook de Stripe pago sistema de suscripciones GESAC
def _renovar_vencimiento(perfil, dias=30):
    """
    Extiende perfil.fecha_vencimiento 'dias' días.

    Si la membresía todavía no ha vencido (renovación anticipada, ej. el
    usuario pagó unos días antes de que se le acabara el mes), se extiende
    a partir de la fecha de vencimiento actual, para no perder los días que
    ya tenía pagados. Si ya venció (o es la primera vez), se extiende a
    partir de ahora.
    """
    ahora = timezone.now()
    vencimiento_actual = perfil.fecha_vencimiento
    base = (
        vencimiento_actual
        if vencimiento_actual and vencimiento_actual > ahora
        else ahora
    )
    perfil.fecha_vencimiento = base + timedelta(days=dias)


@csrf_exempt
def stripe_webhook(request):
    stripe.api_key = settings.STRIPE_SECRET_KEY

    payload = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE")
    endpoint_secret = settings.STRIPE_ENDPOINT_SECRET

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
    except Exception as e:
        print("Error Stripe:", e)
        return HttpResponse(status=400)

    plus_prices = {
        "price_1TrLU4PYnlfwKZQHncNKhhvd",  # produccion plus
        #"price_1RnT1IPW7xPgzk0myWccMWtW",  # pruebas plus
    }
    premium_prices = {
        "price_1Tpa52PYnlfwKZQHWUWxDAuE",  # produccion premium
        #"price_1RnSzMPW7xPgzk0mLslR8vT5",  # pruebas premium
    }

    # Alta inicial o upgrade por Checkout
    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        subscription_id = session.get("subscription")
        customer_id = session.get("customer")
        price_id = _extract_price_id_from_session(session.get("id"))

        perfil = _find_perfil_from_checkout_session(session)
        if not perfil:
            print("No se encontró PerfilUsuario para checkout.session.completed")
            return HttpResponse(status=200)
        
        # Si el usuario se registró sin correo, se toma el que Stripe le pidió
        # (y confirmó) durante el checkout. customer_details.email es el dato
        # real que el cliente tecleó en la pantalla de pago; es más confiable
        # que customer_email (que solo refleja lo que nosotros le mandamos,
        # si es que le mandamos algo).
        if perfil.usuario and not perfil.usuario.email:
            email_stripe = (session.get("customer_details") or {}).get("email") or session.get("customer_email")
            if email_stripe:
                perfil.usuario.email = email_stripe
                perfil.usuario.save(update_fields=["email"])
                print(f"Email tomado de Stripe checkout: {email_stripe}")


        tenia_suscripcion = bool(
            perfil.stripe_plus_subscription_id or perfil.stripe_premium_subscription_id
        )

        # persist customer
        if customer_id and not perfil.stripe_customer_id:
            perfil.stripe_customer_id = customer_id

        if not subscription_id:
            print("Session completada sin subscription_id")
            perfil.save(update_fields=["stripe_customer_id"])
            return HttpResponse(status=200)

        if price_id in plus_prices:
            perfil.stripe_plus_subscription_id = subscription_id
            # legacy/fallback temporal
            perfil.stripe_subscription_id = subscription_id

        elif price_id in premium_prices:
        
            perfil.stripe_premium_subscription_id = subscription_id
            # legacy/fallback temporal
            perfil.stripe_subscription_id = subscription_id

        else:
            print("Price ID no reconocido:", price_id)
            return HttpResponse(status=200)

   
        _sync_membership_state(perfil)
        perfil.save()

        if perfil.usuario and not perfil.usuario.is_active:
            perfil.usuario.is_active = True
            perfil.usuario.save(update_fields=["is_active"])

    # Renovaciones recurrentes
    if event["type"] == "invoice.payment_succeeded":
        invoice = event["data"]["object"]
        customer_id = invoice.get("customer")
        subscription_id = invoice.get("subscription")

        if not customer_id:
            return HttpResponse(status=200)

        perfil = PerfilUsuario.objects.filter(stripe_customer_id=customer_id).first()
        if not perfil:
            print("No se encontró perfil para invoice.payment_succeeded")
            return HttpResponse(status=200)

        price_id = None
        if subscription_id:
            try:
                sub = stripe.Subscription.retrieve(
                    subscription_id, expand=["items.data.price"]
                )
                items = sub.get("items", {}).get("data", [])
                if items and items[0].get("price"):
                    price_id = items[0]["price"]["id"]
            except Exception as e:
                print("No se pudo leer subscription en invoice.payment_succeeded:", e)

        if price_id in plus_prices:
            perfil.stripe_plus_subscription_id = subscription_id
            if not perfil.stripe_subscription_id:
                perfil.stripe_subscription_id = subscription_id

        elif price_id in premium_prices:
            perfil.stripe_premium_subscription_id = subscription_id
            perfil.stripe_subscription_id = subscription_id

        else:
            # fallback por coincidencia con ids guardados
            if subscription_id == perfil.stripe_plus_subscription_id:
                pass
            elif subscription_id == perfil.stripe_premium_subscription_id:
                pass
            else:
                print(
                    "invoice.payment_succeeded con subscription no reconocida:",
                    subscription_id,
                )

        # Cada pago exitoso de renovacion (mensual) extiende 30 dias mas
        _renovar_vencimiento(perfil, dias=30)

        _sync_membership_state(perfil)
        perfil.save()
        print("Renovación automática procesada correctamente.")

    return HttpResponse(status=200)


@login_required
def crear_sesion_pago(request):
    stripe.api_key = settings.STRIPE_SECRET_KEY
    perfil = request.user.perfilusuario

    # Si ya tiene plus (o premium, que depende de plus), no duplicar plus
    if perfil.stripe_plus_subscription_id:
        return JsonResponse(
            {"status": "error", "detail": "Ya tienes membresía PLUS activa."},
            status=400
        )

    success = request.build_absolute_uri(reverse("dashboard_inicio") + "?pago=ok")
    cancel = request.build_absolute_uri(reverse("dashboard_inicio"))

    session_kwargs = {
        "payment_method_types": ["card"],
        "line_items": [{"price": "price_1TrLU4PYnlfwKZQHncNKhhvd", "quantity": 1}], #produccion
        #"line_items": [{"price": "price_1RnT1IPW7xPgzk0myWccMWtW", "quantity": 1}], #desarrollo
        "mode": "subscription",
        "success_url": success,
        "cancel_url": cancel,
        "client_reference_id": str(request.user.id),
        "metadata": {"plan": "plus", "user_id": str(request.user.id)},
    }

    if perfil.stripe_customer_id:
        session_kwargs["customer"] = perfil.stripe_customer_id
    elif request.user.email:
        # Solo se manda customer_email si de verdad hay un correo guardado;
        # un string vacío hace que Stripe rechace la sesión con
        # "Invalid email address:". Si no hay correo, se omite el parámetro
        # y Stripe simplemente se lo pide al usuario en la pantalla de pago.
        session_kwargs["customer_email"] = request.user.email

    try:
        session = stripe.checkout.Session.create(**session_kwargs)
    except Exception as e:
        # El customer_id guardado ya no existe en este modo/cuenta de Stripe
        # (típico al alternar entre llaves test/live). Se limpia el id
        # inválido y se reintenta, dejando que Stripe cree un customer nuevo.
        if "No such customer" in str(e) and perfil.stripe_customer_id:
            print(f"Customer inválido {perfil.stripe_customer_id}, se limpia y reintenta")
            perfil.stripe_customer_id = None
            perfil.save(update_fields=["stripe_customer_id"])

            session_kwargs.pop("customer", None)
            if request.user.email:
                session_kwargs["customer_email"] = request.user.email

            session = stripe.checkout.Session.create(**session_kwargs)
        else:
            raise

    return JsonResponse({"id": session.id, "url": session.url})


@login_required
def crear_sesion_pago_premium(request):
    stripe.api_key = settings.STRIPE_SECRET_KEY
    perfil = request.user.perfilusuario

    if perfil.stripe_premium_subscription_id:
        return JsonResponse(
            {"status": "error", "detail": "Ya tienes membresía PREMIUM activa."},
            status=400
        )

    success = request.build_absolute_uri(reverse("dashboard_inicio") + "?pago=ok")
    cancel = request.build_absolute_uri(reverse("dashboard_inicio"))

    session_kwargs = {
        "payment_method_types": ["card"],
        "line_items": [{"price": "price_1Tpa52PYnlfwKZQHWUWxDAuE", "quantity": 1}], #produccion
        #"line_items": [{"price": "price_1RnSzMPW7xPgzk0mLslR8vT5", "quantity": 1}], #desarrollo
        "mode": "subscription",
        "success_url": success,
        "cancel_url": cancel,
        "client_reference_id": str(request.user.id),
        "metadata": {"plan": "premium", "user_id": str(request.user.id)},
    }

    if perfil.stripe_customer_id:
        session_kwargs["customer"] = perfil.stripe_customer_id
    elif request.user.email:
        # Solo se manda customer_email si de verdad hay un correo guardado;
        # un string vacío hace que Stripe rechace la sesión con
        # "Invalid email address:". Si no hay correo, se omite el parámetro
        # y Stripe simplemente se lo pide al usuario en la pantalla de pago.
        session_kwargs["customer_email"] = request.user.email

    try:
        session = stripe.checkout.Session.create(**session_kwargs)
    except Exception as e:
        # El customer_id guardado ya no existe en este modo/cuenta de Stripe
        # (típico al alternar entre llaves test/live). Se limpia el id
        # inválido y se reintenta, dejando que Stripe cree un customer nuevo.
        if "No such customer" in str(e) and perfil.stripe_customer_id:
            print(f"Customer inválido {perfil.stripe_customer_id}, se limpia y reintenta")
            perfil.stripe_customer_id = None
            perfil.save(update_fields=["stripe_customer_id"])

            session_kwargs.pop("customer", None)
            if request.user.email:
                session_kwargs["customer_email"] = request.user.email

            session = stripe.checkout.Session.create(**session_kwargs)
        else:
            raise

    return JsonResponse({"id": session.id, "url": session.url})


@require_POST
@login_required
def cancelar_suscripcion(request):
    """
    Cancela PLUS.
    Regla: si premium está activo, primero debe cancelarse premium.
    """
    stripe.api_key = settings.STRIPE_SECRET_KEY
    perfil = request.user.perfilusuario

    if perfil.stripe_premium_subscription_id:
        return JsonResponse(
            {"status": "error", "detail": "Primero cancela la membresía PREMIUM."},
            status=400,
        )

    plus_subscription_id = (
        perfil.stripe_plus_subscription_id or perfil.stripe_subscription_id
    )
    # if not plus_subscription_id:
    #     return JsonResponse({"status": "no encontrada"}, status=404)

    try:
        _safe_cancel_subscription(plus_subscription_id)

        perfil.stripe_plus_subscription_id = None
        if perfil.stripe_subscription_id == plus_subscription_id:
            perfil.stripe_subscription_id = None

        _sync_membership_state(perfil)
        perfil.save()

        return JsonResponse({"status": "cancelada"})
    except Exception as e:
        return JsonResponse({"status": "error", "detail": str(e)}, status=400)


@require_POST
@login_required
def cancelar_suscripcion_premium(request):
    """
    Cancela PREMIUM y vuelve a PLUS (si plus sigue activo).
    """
    stripe.api_key = settings.STRIPE_SECRET_KEY
    perfil = request.user.perfilusuario

    premium_subscription_id = perfil.stripe_premium_subscription_id
    if not premium_subscription_id and perfil.tipo_usuario == "premium":
        # fallback legacy
        premium_subscription_id = perfil.stripe_subscription_id

    if not premium_subscription_id:
        return JsonResponse({"status": "no encontrada"}, status=404)

    try:
        _safe_cancel_subscription(premium_subscription_id)

        perfil.stripe_premium_subscription_id = None

        # legacy: si el id legacy era el premium cancelado, volverlo al plus (si existe), si no existe que sea None 
        # y que el tipo_usuario se ajuste a plus

        if perfil.stripe_subscription_id == premium_subscription_id:
            perfil.stripe_subscription_id = perfil.stripe_plus_subscription_id
            if not perfil.stripe_subscription_id:
                perfil.tipo_usuario = "plus"

        _sync_membership_state(perfil)
        perfil.save()

        return JsonResponse({"status": "cancelada"})
    except Exception as e:
        return JsonResponse({"status": "error", "detail": str(e)}, status=400)


@login_required
def cerrar_wizard(request):
    # Marca en la sesión que el wizard fue cerrado temporalmente
    request.session["wizard_cerrado"] = True
    return redirect("dashboard_inicio")


@require_POST
@login_required
def guardar_datos_empresa(request):
    perfil = request.user.perfilusuario
    empresa = perfil.empresa
    nuevo_rfc = request.POST.get("rfc_empresa", "").strip()

    if Empresa.objects.filter(rfc=nuevo_rfc).exclude(id=empresa.id).exists():
        messages.error(request, "El RFC ingresado ya está registrado en otra empresa.")
        return redirect("dashboard_inicio")

    # Datos generales de la empresa
    empresa.nombre = request.POST.get("nombre_empresa", "").strip()
    empresa.rfc = nuevo_rfc
    empresa.regimen_fiscal = request.POST.get("regimen_fiscal", "").strip()
    empresa.direccion = request.POST.get("direccion_empresa", "").strip()
    empresa.email = request.POST.get("email_empresa", "").strip()
    empresa.telefono = request.POST.get("telefono_empresa", "").strip()
    empresa.codigo_postal = request.POST.get("codigo_postal", "").strip()
    empresa.save()

    # Datos bancarios — crear cuenta en CuentaBancaria si vienen datos
    banco = request.POST.get("cuenta_bancaria", "").strip()
    numero_cuenta = request.POST.get("numero_cuenta", "").strip()
    clabe = request.POST.get("clabe", "").strip()
    try:
        saldo_inicial = float(request.POST.get("saldo_inicial", 0) or 0)
    except ValueError:
        saldo_inicial = 0.0

    if banco and numero_cuenta:
        # Solo crear si no existe ya una cuenta con ese número para esta empresa
        if not CuentaBancaria.objects.filter(
            empresa=empresa, numero_cuenta=numero_cuenta
        ).exists():
            CuentaBancaria.objects.create(
                empresa=empresa,
                banco=banco,
                numero_cuenta=numero_cuenta,
                clabe=clabe,
                moneda=request.POST.get("moneda", "MXN"),
                tipo_cuenta=request.POST.get("tipo_cuenta", ""),
                saldo_inicial=saldo_inicial,
                activa=True,
            )

    perfil.mostrar_wizard = False
    perfil.save()
    messages.success(request, "¡Datos de empresa configurados correctamente!")
    return redirect("dashboard_inicio")


# MODULO DE TICKETS DE MANTENIMIENTO-- NO ESTA EN FUNCION-SUSTITUIDO POR APP MANTPRO, EN UN FUTURO ENLAZAR CON MANTPRO>
@login_required
def crear_ticket(request):
    empresa = request.user.perfilusuario.empresa
    empleados = Empleado.objects.filter(empresa=empresa, activo=True)
    if request.method == "POST":
        titulo = request.POST["titulo"]
        descripcion = request.POST["descripcion"]
        empleado_id = request.POST["empleado_asignado"]
        empleado = Empleado.objects.get(id=empleado_id)
        ticket = TicketMantenimiento.objects.create(
            titulo=titulo, descripcion=descripcion, empleado_asignado=empleado
        )
        # Enviar correo si el empleado tiene email
        if empleado.email:
            # Preparar email con adjuntos (si hay)
            asunto = "Nuevo reporte de mantenimiento asignado"
            cuerpo = f"Empresa: {empresa.nombre}\nSe te ha asignado el reporte: {titulo}\n\nDescripción: {descripcion}"
            email = EmailMessage(
                subject=asunto,
                body=cuerpo,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[empleado.email],
            )
            # Adjuntar imágenes si fueron subidas
            archivos = request.FILES.getlist("imagenes")
            MAX_SIZE = 5 * 1024 * 1024  # 5MB por archivo
            skipped = []
            attached_count = 0
            for f in archivos:
                try:
                    content_type = getattr(f, "content_type", "") or ""
                    # validar tipo y tamaño
                    if not content_type.startswith("image/"):
                        skipped.append(f"{f.name} (no es imagen)")
                        continue
                    if f.size > MAX_SIZE:
                        skipped.append(f"{f.name} (>{int(MAX_SIZE / 1024 / 1024)}MB)")
                        continue
                    # adjuntar
                    email.attach(f.name, f.read(), content_type)
                    attached_count += 1
                except Exception:
                    skipped.append(f"{f.name} (error al adjuntar)")
                    continue
            try:
                email.send(fail_silently=True)
            except Exception:
                pass

            # Informar al usuario sobre adjuntos omitidos/adjuntados
            if attached_count:
                messages.info(
                    request, f"Se adjuntaron {attached_count} imagen(es) al correo."
                )
            if skipped:
                messages.warning(
                    request,
                    "Se omitieron archivos al enviar el correo: "
                    + ", ".join(skipped[:20])
                    + (f", ...(+{len(skipped) - 20})" if len(skipped) > 20 else ""),
                )

        # # Enviar notificación por WhatsApp si el empleado tiene teléfono
        # #PENDIENTE CONFIGURAR CLOUD API DE WHATSAPP (META)
        # if empleado.telefono:
        #     # Normalizar número (quitar espacios, paréntesis, guiones)
        #     phone_raw = str(empleado.telefono)
        #     phone_digits = "".join(ch for ch in phone_raw if ch.isdigit())
        #     # Quitar prefijo internacional 00 si existe
        #     if phone_digits.startswith("00"):
        #         phone_digits = phone_digits[2:]
        #     # Si aún no tiene código de país intenta no modificar; WhatsApp Cloud API requiere formato internacional
        #     whatsapp_phone = phone_digits

        #     # Intento con WhatsApp Cloud API (Meta) si están configuradas las vars de entorno
        #     whatsapp_phone_id = os.getenv("WHATSAPP_PHONE_NUMBER_ID")  # e.g. "109876543210" (your phone number id)
        #     whatsapp_token = os.getenv("WHATSAPP_TOKEN")  # bearer token (long-lived)

        #     msg_text = f"Se te ha asignado el ticket: {titulo}\nDescripción: {descripcion}"

        #     if whatsapp_phone_id and whatsapp_token:
        #         headers = {"Authorization": f"Bearer {whatsapp_token}"}
        #         base_url = f"https://graph.facebook.com/v15.0/{whatsapp_phone_id}"

        #         # 1) Enviar texto
        #         try:
        #             payload = {
        #                 "messaging_product": "whatsapp",
        #                 "to": whatsapp_phone,
        #                 "type": "text",
        #                 "text": {"body": msg_text}
        #             }
        #             requests.post(f"{base_url}/messages", json=payload, headers=headers, timeout=10)
        #         except Exception:
        #             # no interrumpe el flujo si falla WhatsApp
        #             pass

        #         # 2) Si hay imágenes subidas, subir y enviar cada una como media
        #         archivos = request.FILES.getlist("imagenes")
        #         MAX_SIZE = 5 * 1024 * 1024  # 5MB por archivo
        #         for f in archivos:
        #             try:
        #                 ct = getattr(f, "content_type", "") or ""
        #                 if not ct.startswith("image/"):
        #                     continue
        #                 if f.size > MAX_SIZE:
        #                     continue
        #                 # Subir media -> devuelve id
        #                 files = {"file": (f.name, f.read(), ct)}
        #                 params = {"messaging_product": "whatsapp"}
        #                 r = requests.post(f"{base_url}/media", files=files, params=params, headers=headers, timeout=30)
        #                 media_resp = r.json() if r.ok else {}
        #                 media_id = media_resp.get("id")
        #                 if media_id:
        #                     img_payload = {
        #                         "messaging_product": "whatsapp",
        #                         "to": whatsapp_phone,
        #                         "type": "image",
        #                         "image": {"id": media_id, "caption": titulo}
        #                     }
        #                     requests.post(f"{base_url}/messages", json=img_payload, headers=headers, timeout=10)
        #             except Exception:
        #                 continue
        #     else:
        #         # Fallback mínimo: construir URL wa.me para abrir en cliente (no puedo enviar desde servidor sin API)
        #         # Se guarda en logs/messages para que el operador lo use
        #         wa_link = f"https://wa.me/{whatsapp_phone}?text={requests.utils.requote_uri(msg_text)}"
        #         # opcional: registrar en mensajes de Django para que el admin lo vea
        #         messages.info(request, f"No está configurada la API de WhatsApp. Abre este enlace para notificar manualmente: {wa_link}")

        return redirect("lista_tickets")
    return render(request, "mantenimiento/crear_ticket.html", {"empleados": empleados})


@login_required
def actualizar_ticket(request, ticket_id):
    ticket = TicketMantenimiento.objects.get(id=ticket_id)
    if request.method == "POST":
        ticket.estado = request.POST["estado"]
        ticket.solucion = request.POST.get("solucion", "")
        if ticket.estado == "resuelto":
            ticket.fecha_solucion = timezone.now()
        ticket.save()
        return redirect("detalle_ticket", ticket_id=ticket.id)
    return render(request, "mantenimiento/actualizar_ticket.html", {"ticket": ticket})


@login_required
def agregar_seguimiento(request, ticket_id):
    ticket = get_object_or_404(TicketMantenimiento, id=ticket_id)
    if request.method == "POST":
        comentario = request.POST["comentario"]
        SeguimientoTicket.objects.create(
            ticket=ticket, usuario=request.user, comentario=comentario
        )
    return redirect("detalle_ticket", ticket_id=ticket.id)


@login_required
def detalle_ticket(request, ticket_id):
    ticket = TicketMantenimiento.objects.get(id=ticket_id)
    seguimientos = ticket.seguimientos.order_by("-fecha")
    return render(
        request,
        "mantenimiento/detalle_ticket.html",
        {"ticket": ticket, "seguimientos": seguimientos},
    )

@login_required
def tickets_asignados(request):
    empresa = request.user.perfilusuario.empresa
    empleados = Empleado.objects.filter(empresa=empresa, activo=True)
    empleado_id = request.GET.get("empleado_id")
    tickets = []
    empleado_seleccionado = None
    if empleado_id:
        empleado_seleccionado = Empleado.objects.filter(
            id=empleado_id, empresa=empresa
        ).first()
        tickets = TicketMantenimiento.objects.filter(
            empleado_asignado=empleado_seleccionado
        )
    return render(
        request,
        "mantenimiento/tickets_asignados.html",
        {
            "empleados": empleados,
            "tickets": tickets,
            "empleado_seleccionado": empleado_seleccionado,
        },
    )

@login_required
def lista_tickets(request):
    if request.user.is_superuser:
        empresa_id = request.session.get("empresa_id")
        if empresa_id:
            tickets = TicketMantenimiento.objects.filter(
                empleado_asignado__empresa_id=empresa_id
            ).order_by("-fecha_creacion")
        else:
            tickets = TicketMantenimiento.objects.all().order_by("-fecha_creacion")
    else:
        empresa = request.user.perfilusuario.empresa
        tickets = TicketMantenimiento.objects.filter(
            empleado_asignado__empresa=empresa
        ).order_by("-fecha_creacion")
    return render(request, "mantenimiento/lista_tickets.html", {"tickets": tickets})



# modulo visitantes consulta adeudos y pagos de facturas via WEB

def registro_visitante(request):
    empresas = Empresa.objects.all()
    locales = LocalComercial.objects.all()
    areas = AreaComun.objects.all()

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        email = request.POST.get("email")
        username = request.POST.get("username")
        password = request.POST.get("password")
        password2 = request.POST.get("password2")

        error = False
        if password != password2:
            messages.error(request, "Las contraseñas no coinciden.")
            error = True
        if VisitanteAcceso.objects.filter(email=email).exists():
            messages.error(request, "El correo ya está registrado.")
            error = True
        if VisitanteAcceso.objects.filter(username=username).exists():
            messages.error(request, "El usuario ya está registrado.")
            error = True

        # Para el JS dinámico
        empresas_json = [
            {
                "id": e.id,
                "nombre": e.nombre,
                "locales": [
                    {"id": l.id, "numero": l.numero} for l in locales.filter(empresa=e)
                ],
                "areas": [
                    {"id": a.id, "numero": a.numero} for a in areas.filter(empresa=e)
                ],
            }
            for e in empresas
        ]

        if error:
            # Renderiza el formulario con los datos ya capturados
            return render(
                request,
                "visitantes/registro_visitante.html",
                {
                    "empresas": empresas,
                    "empresas_json": json.dumps(empresas_json),
                    "nombre": nombre,
                    "email": email,
                    "username": username,
                },
            )

        # Procesar bloques dinámicos
        bloques = []
        for key in request.POST:
            if key.startswith("empresa_"):
                idx = key.split("_")[1]
                empresa_id = request.POST.get(f"empresa_{idx}")
                local_ids = request.POST.getlist(f"locales_{idx}")
                area_ids = request.POST.getlist(f"areas_{idx}")
                if empresa_id:
                    bloques.append(
                        {
                            "empresa": Empresa.objects.get(id=empresa_id),
                            "locales": LocalComercial.objects.filter(id__in=local_ids),
                            "areas": AreaComun.objects.filter(id__in=area_ids),
                        }
                    )

        # Crear visitante
        visitante = VisitanteAcceso.objects.create(
            nombre=nombre,
            email=email,
            username=username,
            password=make_password(password),
            activo=False,
        )
        # Email de confirmación al visitante
        mensaje_visitante = (
            f"Hola {nombre},\n\n"
            "Tu registro ha sido recibido y está pendiente de validación por el administrador. "
            "Te notificaremos cuando tu cuenta sea activada.\n\n"
            "Atentamente,\nEl equipo de Softheron.\n\n"
            "Gracias. \n\n"
            "Página de Softheron: https://paginaweb-ro9v.onrender.com \n\n"
        )
        send_mail(
            "Registro recibido - Pendiente de validación",
            mensaje_visitante,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=True,
        )
        # Asignar empresas, locales y áreas
        for bloque in bloques:
            visitante.empresas.add(bloque["empresa"])
            visitante.locales.add(*bloque["locales"])
            visitante.areas.add(*bloque["areas"])

        # Email resumen al admin
        resumen = f"Nuevo usuario condómino registrado:\nNombre: {nombre}\nEmail: {email}\nUsuario: {username}\n\nEmpresas y asignaciones:\n"
        for bloque in bloques:
            resumen += f"- Empresa: {bloque['empresa'].nombre}\n"
            resumen += f"  Locales: {', '.join([l.numero for l in bloque['locales']]) or 'Ninguno'}\n"
            resumen += f"  Áreas: {', '.join([a.numero for a in bloque['areas']]) or 'Ninguna'}\n"
        send_mail(
            "Nuevo registro de usuario condómino",
            resumen,
            settings.DEFAULT_FROM_EMAIL,
            [settings.EMAIL_HOST_USER],
            fail_silently=True,
        )
        return render(request, "visitantes/registro_exitoso.html")

    # Para el JS dinámico
    empresas_json = [
        {
            "id": e.id,
            "nombre": e.nombre,
            "locales": [
                {"id": l.id, "numero": l.numero} for l in locales.filter(empresa=e)
            ],
            "areas": [
                {"id": a.id, "numero": a.numero} for a in areas.filter(empresa=e)
            ],
        }
        for e in empresas
    ]

    return render(
        request,
        "visitantes/registro_visitante.html",
        {
            "empresas": empresas,
            "empresas_json": json.dumps(empresas_json),
        },
    )


def visitante_recuperar_password(request):
    if request.method == "POST":
        email = request.POST.get("email")
        try:
            visitante = VisitanteAcceso.objects.get(email=email)
            nueva_password = get_random_string(8)
            visitante.password = make_password(nueva_password)
            visitante.save()
            mensaje = (
                f"Hola {visitante.nombre},\n\n"
                f"Tu nueva contraseña es: {nueva_password}\n"
                "Accesa con ella en https://adminsoftheron.onrender.com/visitante/login/ \n\n"
                "Atentamente,\nEl equipo de Softheron. \n\n"
                "Pagina de Softheron: https://paginaweb-ro9v.onrender.com \n\n"
            )
            send_mail(
                "Recuperación de contraseña",
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [visitante.email],
                fail_silently=True,
            )
            messages.success(request, "Se ha enviado una nueva contraseña a tu correo.")
        except VisitanteAcceso.DoesNotExist:
            messages.error(request, "No se encontró un usuario con ese correo.")
    return render(request, "visitantes/recuperar_password.html")


def visitante_login(request):
    if request.method == "POST":
        form = VisitanteLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            try:
                visitante = VisitanteAcceso.objects.get(username=username)
                if not visitante.activo:
                    messages.error(
                        request,
                        "Tu cuenta aún no ha sido activada por el administrador.",
                    )
                    return render(request, "visitantes/login.html", {"form": form})
                if check_password(password, visitante.password):
                    request.session["visitante_id"] = visitante.id
                    return redirect("visitante_seleccionar_empresa")
                else:
                    messages.error(request, "Contraseña incorrecta.")
            except VisitanteAcceso.DoesNotExist:
                messages.error(request, "Usuario no encontrado.")
    else:
        form = VisitanteLoginForm()
    return render(request, "visitantes/login.html", {"form": form})


def visitante_seleccionar_empresa(request):
    visitante_id = request.session.get("visitante_id")
    if not visitante_id:
        return redirect("visitante_login")
    visitante = VisitanteAcceso.objects.get(id=visitante_id)
    empresas = visitante.empresas.all()
    if request.method == "POST":
        empresa_id = request.POST.get("empresa_id")
        if empresa_id:
            request.session["empresa_id"] = empresa_id
            return redirect("visitante_consulta_facturas")
    return render(
        request, "visitantes/seleccionar_empresa.html", {"empresas": empresas}
    )


def calcular_interes_mora(fecha_vencimiento, saldo_pendiente, tasa_mensual=0.06):
    if not fecha_vencimiento or saldo_pendiente <= 0:
        return Decimal("0.00")
    hoy = date.today()
    if fecha_vencimiento >= hoy:
        return Decimal("0.00")
    # Calcular meses completos de atraso
    meses = (hoy.year - fecha_vencimiento.year) * 12 + (
        hoy.month - fecha_vencimiento.month
    )
    if hoy.day < fecha_vencimiento.day:
        meses -= 1
    if meses <= 0:
        return Decimal("0.00")
    interes = Decimal(saldo_pendiente) * Decimal(tasa_mensual) * Decimal(meses)
    return interes.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def visitante_consulta_facturas(request):
    visitante_id = request.session.get("visitante_id")
    empresa_id = request.session.get("empresa_id")
    if not visitante_id:
        return redirect("visitante_login")
    visitante = VisitanteAcceso.objects.get(id=visitante_id)

    if not empresa_id:
        # Si no hay empresa seleccionada, redirige a la selección
        return redirect("visitante_seleccionar_empresa")
    empresa = Empresa.objects.get(id=empresa_id)

    # Filtros
    local_id = request.GET.get("local_id")
    area_id = request.GET.get("area_id")
    pagook = request.GET.get("pagook")
    factura_id = request.GET.get("factura_id")
    anio = request.GET.get("anio")

    locales = visitante.locales.filter(empresa=empresa)
    areas = visitante.areas.filter(empresa=empresa)
    nombre_cliente = (
        locales.first().cliente.nombre
        if locales.exists() and locales.first().cliente
        else areas.first().cliente.nombre
        if areas.exists() and areas.first().cliente
        else ""
    )
    facturas = Factura.objects.none()
    if local_id:
        facturas = Factura.objects.filter(
            local_id=local_id,
            local__in=locales,
            empresa=empresa,
            monto__gt=0,  # Solo mostrar facturas con monto mayor a 0
        ).order_by("-fecha_vencimiento")
    elif area_id:
        facturas = Factura.objects.filter(
            area_comun_id=area_id,
            area_comun__in=areas,
            empresa=empresa,
            monto__gt=0,  # Solo mostrar facturas con monto mayor a 0
        ).order_by("-fecha_vencimiento")
    else:
        facturas = Factura.objects.filter(
            Q(local__in=locales) | Q(area_comun__in=areas),
            empresa=empresa,
            monto__gt=0,  # Solo mostrar facturas con monto mayor a 0
        ).order_by("-fecha_vencimiento")

    # Obtén los años únicos de las facturas
    anios_unicos = (
        facturas.order_by().values_list("fecha_vencimiento__year", flat=True).distinct()
    )
    anios_unicos = sorted(set(filter(None, anios_unicos)), reverse=True)

    # Filtro por año
    if anio and anio.isdigit():
        facturas = facturas.filter(fecha_vencimiento__year=int(anio))

    # Calcula total pendiente y total cobrado
    total_pendiente = sum(
        f.saldo_pendiente for f in facturas if f.estatus == "pendiente"
    )
    total_cobrado = sum(f.monto for f in facturas if f.estatus == "cobrada")

    mensaje_pago = None
    if pagook and factura_id:
        try:
            factura_pagada = Factura.objects.get(id=factura_id)
            mensaje_pago = f"¡Pago realizado correctamente!<br>Factura: <b>Folio-{factura_pagada.folio}"
        except Factura.DoesNotExist:
            mensaje_pago = (
                "¡Pago realizado correctamente! En breve se reflejará en el sistema."
            )

    factura_pendiente_mas_antigua = (
        facturas.filter(estatus="pendiente").order_by("fecha_vencimiento").first()
    )
    factura_pendiente_id = (
        factura_pendiente_mas_antigua.id if factura_pendiente_mas_antigua else None
    )
    # paginacion
    page_number = request.GET.get("page", 1)
    paginator = Paginator(facturas, 10)  # Mostrar 10 facturas por página
    facturas_paginadas = paginator.get_page(page_number)
    # Calcula interés de mora para facturas pendientes
    for f in facturas_paginadas:
        if f.estatus == "pendiente":
            f.interes_mora = calcular_interes_mora(
                f.fecha_vencimiento, f.saldo_pendiente
            )
        else:
            f.interes_mora = Decimal("0.00")
    return render(
        request,
        "facturacion/consulta_facturas.html",
        {
            "facturas": facturas,
            "visitante": visitante,
            "locales": locales,
            "areas": areas,
            "local_id": local_id,
            "area_id": area_id,
            "total_pendiente": total_pendiente,
            "total_cobrado": total_cobrado,
            "es_visitante": True,
            "mensaje_pago": mensaje_pago,
            "anio": anio,
            "anios_unicos": anios_unicos,
            "factura_pendiente_id": factura_pendiente_id,
            "empresa": empresa,
            "nombre_cliente": nombre_cliente,
            "facturas": facturas_paginadas,
        },
    )


def descargar_estado_cuenta_pdf(request):
    visitante_id = request.session.get("visitante_id")
    empresa_id = request.session.get("empresa_id")

    if not visitante_id or not empresa_id:
        return redirect("visitante_login")
    visitante = VisitanteAcceso.objects.get(id=visitante_id)
    empresa = Empresa.objects.get(id=empresa_id)

    # Aplica los mismos filtros que en la vista principal
    local_id = request.GET.get("local_id")
    area_id = request.GET.get("area_id")
    anio = request.GET.get("anio")

    locales = visitante.locales.filter(empresa=empresa)
    areas = visitante.areas.filter(empresa=empresa)
    nombre_cliente = (
        locales.first().cliente.nombre
        if locales.exists() and locales.first().cliente
        else ""
    )
    facturas = Factura.objects.none()
    if local_id:
        facturas = Factura.objects.filter(
            local_id=local_id, local__in=locales, empresa=empresa, monto__gt=0
        )
    elif area_id:
        facturas = Factura.objects.filter(
            area_comun_id=area_id, area_comun__in=areas, empresa=empresa, monto__gt=0
        )
    else:
        facturas = Factura.objects.filter(
            Q(local__in=locales) | Q(area_comun__in=areas), empresa=empresa, monto__gt=0
        )
    if anio and anio.isdigit():
        facturas = facturas.filter(fecha_vencimiento__year=int(anio))

    # Calcula intereses si es necesario (igual que en la vista principal)
    for f in facturas:
        if f.estatus == "pendiente":
            f.interes_mora = calcular_interes_mora(
                f.fecha_vencimiento, f.saldo_pendiente
            )
        else:
            f.interes_mora = Decimal("0.00")

    total_saldo = sum(f.saldo_pendiente for f in facturas)
    total_interes = sum(f.interes_mora for f in facturas)
    fecha_generacion = timezone.now()
    html = render_to_string(
        "facturacion/estado_cuenta_pdf.html",
        {
            "facturas": facturas,
            "visitante": visitante,
            "empresa": empresa,
            "local_id": local_id,
            "area_id": area_id,
            "anio": anio,
            "nombre_cliente": nombre_cliente,
            "total_saldo": total_saldo,
            "total_interes": total_interes,
            "fecha_generacion": fecha_generacion,
        },
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="estado_cuenta_{nombre_cliente}.pdf"'
    )
    weasyprint.HTML(string=html).write_pdf(response)
    return response


def visitante_logout(request):
    request.session.flush()
    return redirect("visitante_login")


def visitante_factura_detalle(request, factura_id):
    visitante_id = request.session.get("visitante_id")
    if not visitante_id:
        return redirect("visitante_login")
    visitante = VisitanteAcceso.objects.get(id=visitante_id)
    factura = get_object_or_404(Factura, id=factura_id)
    # Verifica que la factura pertenezca a los locales/áreas del visitante
    if (
        factura.local not in visitante.locales.all()
        and factura.area_comun not in visitante.areas.all()
    ):
        return redirect("visitante_consulta_facturas")
    cobros = factura.pagos.all()
    return render(
        request,
        "facturacion/facturas_detalle.html",
        {"factura": factura, "cobros": cobros, "es_visitante": True},
    )


# Módulo de timbrado para visitantes
def visitante_timbrar_factura(request, pk):
    visitante_id = request.session.get("visitante_id")
    if not visitante_id:
        return redirect("visitante_login")
    visitante = VisitanteAcceso.objects.get(id=visitante_id)
    factura = get_object_or_404(Factura, pk=pk)
    # Verifica que la factura pertenezca a los locales/áreas del visitante
    if (
        factura.local not in visitante.locales.all()
        and factura.area_comun not in visitante.areas.all()
    ):
        messages.error(request, "No tienes permiso para timbrar esta factura.")
        return redirect("visitante_consulta_facturas")

    # Validar membresía
    # if getattr(visitante, "membresia_tipo", "basica") not in ["plus", "premium"]:
    #     return redirect("visitante_membresia_pago")  # Crea esta vista/URL para el pago

    empresa = factura.empresa
    if not empresa.es_premium:
        messages.error(
            request,
            "TIMBRADO de Facturas solo está disponible en la versión PREMIUM. Contacta al administrador del condominio para que contrate la versión PREMIUM.",
        )
        return redirect("visitante_consulta_facturas")

    if factura.uuid:
        messages.info(request, "La factura ya está timbrada.")
        return redirect("visitante_consulta_facturas")

    if request.method == "POST":
        form = TimbrarFacturaForm(request.POST)
        if form.is_valid():
            tax_object = form.cleaned_data["tax_object"]
            payment_method = form.cleaned_data["payment_method"]
            payment_form = form.cleaned_data["payment_form"]
            datos_json = factura_a_json_facturama(
                factura, tax_object, payment_method, payment_form
            )
            resultado = timbrar_factura_facturama(datos_json)
            print("Resultado de timbrado:", resultado)
            if "error" in resultado:
                messages.error(request, f"Error al timbrar: {resultado['error']}")
            else:
                uuid = resultado.get("Uuid") or resultado.get("Complement", {}).get(
                    "TaxStamp", {}
                ).get("Uuid")
                facturama_id = resultado.get("Id")
                if not uuid or not facturama_id:
                    messages.error(request, f"Error inesperado: {resultado}")
                else:
                    factura.uuid = uuid
                    factura.facturama_id = facturama_id
                    factura.save()
                    messages.success(
                        request,
                        "Factura "
                        + factura.folio
                        + " timbrada correctamente. Ahora puedes descargar el PDF y XML.",
                    )
            return redirect("visitante_consulta_facturas")
    else:
        form = TimbrarFacturaForm()

    return render(
        request,
        "facturacion/timbrar_factura.html",
        {
            "form": form,
            "factura": factura,
            "url_cancelar": "visitante_consulta_facturas",
        },
    )


# Pago con Stripe para visitantes multiempresas
def stripe_checkout_visitante(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    empresa = factura.empresa

    # Verifica que la empresa tenga las claves de Stripe configuradas
    if not (
        empresa.stripe_secret_key
        and empresa.stripe_public_key
        and empresa.stripe_webhook_secret
    ):
        messages.error(
            request,
            " PAGO EN LINEA no esta configurado. Contacta al administrador del condominio para su configuración.",
        )
        return redirect("visitante_consulta_facturas")
    elif not empresa.es_premium:
        messages.error(
            request,
            "Disponible en la versión PREMIUM del sistema./n Contacta al administrador del condominio para que contrate la version PREMIUM.",
        )
        return redirect(
            "visitante_consulta_facturas"
        )  # O la vista/lista que corresponda

    stripe.api_key = empresa.stripe_secret_key  # <-- Usa la clave secreta de la empresa

    session = stripe.checkout.Session.create(
        payment_method_types=["card"],
        line_items=[
            {
                "price_data": {
                    "currency": "mxn",
                    "product_data": {
                        "name": f"Pago factura {factura.folio}",
                    },
                    "unit_amount": int(factura.saldo_pendiente * 100),
                },
                "quantity": 1,
            }
        ],
        mode="payment",
        # success_url=request.build_absolute_uri('/visitante/consulta/?pagook=1'),
        success_url=request.build_absolute_uri(
            f"/visitante/consulta/?pagook=1&factura_id={factura.id}"
        ),
        cancel_url=request.build_absolute_uri("/visitante/consulta/?pagocancel=1"),
        metadata={"factura_id": factura.id},
    )
    return redirect(session.url)


@csrf_exempt
def stripe_webhook_visitante(request):
    logger = logging.getLogger("django")
    payload = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE")

    # 1. Carga el evento para saber el tipo
    try:
        event_data = json.loads(payload)
        event_type = event_data.get("type")
    except Exception as e:
        logger.error(f"Error leyendo el evento Stripe: {e}")
        return HttpResponse(status=400)

    # 2. Obtén el objeto y la metadata
    session = event_data.get("data", {}).get("object", {})
    metadata = session.get("metadata", {})

    # 3. Obtén el endpoint_secret (por empresa si aplica, o de settings)
    factura_id = metadata.get("factura_id")
    endpoint_secret = getattr(
        settings, "STRIPE_ENDPOINT_SECRET", None
    )  # Valor por defecto
    if factura_id:
        try:
            factura = Factura.objects.select_related("empresa").get(id=int(factura_id))
            empresa = factura.empresa
            endpoint_secret = getattr(empresa, "stripe_webhook_secret", endpoint_secret)
        except Exception as e:
            logger.error(f"Error extrayendo factura/empresa: {e}")
            return HttpResponse(status=400)

    # 4. Valida la firma del webhook
    try:
        event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
    except Exception as e:
        logger.error(f"Stripe webhook error: {e}")
        return HttpResponse(status=400)

    # 5. Procesa eventos relevantes
    event_type = event["type"]
    obj = event["data"]["object"]
    metadata = obj.get("metadata", {})
    factura_id = metadata.get("factura_id")

    if not factura_id:
        logger.error("No se encontró factura_id en metadata.")
        return HttpResponse(status=400)

    try:
        factura = Factura.objects.get(id=int(factura_id))
        if event_type == "checkout.session.completed":
            monto_pagado = obj.get("amount_total", 0) / 100.0
            observaciones = f"ID transacción: {obj.get('payment_intent')}"
        elif event_type == "payment_intent.succeeded":
            monto_pagado = obj.get("amount", 0) / 100.0
            observaciones = f"ID transacción: {obj.get('id')}"
        else:
            logger.info(f"Ignorando evento Stripe tipo: {event_type}")
            return JsonResponse({"status": "ignored"})

        # Registra el pago y actualiza el estado
        Pago.objects.create(
            factura=factura,
            monto=monto_pagado,
            forma_pago="stripe",
            fecha_pago=timezone.now(),
            registrado_por=None,
            observaciones=observaciones,
        )
        factura.actualizar_estatus()
        factura.save()
    except Factura.DoesNotExist:
        logger.error(f"Factura no encontrada: {factura_id}")
        return HttpResponse(status=400)
    except Exception as e:
        logger.error(f"Error actualizando factura: {e}")
        return HttpResponse(status=400)

    return JsonResponse({"status": "ok"})


##############################################################################
############## Vistas para manejo de membresías de visitantes####################
# Pago de membresía plus con Stripe
@csrf_exempt
def crear_sesion_pago_membresia_plus(request):
    if request.method == "POST":
        visitante_id = request.session.get("visitante_id")
        if not visitante_id:
            return redirect("visitante_login")
        visitante = VisitanteAcceso.objects.get(id=visitante_id)

        plan = request.POST.get("plan", "mensual")
        # Define tus price_id reales aquí
        price_id_mensual = "price_1SxjLCPW7xPgzk0mppaNE8RE"
        price_id_anual = "price_1SxjLCPW7xPgzk0mSUWcWS5k"
        price_id = price_id_mensual if plan == "mensual" else price_id_anual

        stripe.api_key = settings.STRIPE_SECRET_KEY

        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=[
                {
                    "price": price_id,
                    "quantity": 1,
                }
            ],
            mode="subscription",
            success_url=request.build_absolute_uri("/visitante/membresia/success/"),
            cancel_url=request.build_absolute_uri("/visitante/membresia/pago/"),
            metadata={"visitante_id": visitante.id},
            client_reference_id=str(visitante.id),
            customer_email=visitante.email,
        )
        return redirect(session.url)
    return redirect("visitante_membresia_pago")


@csrf_exempt
def crear_sesion_pago_membresia_premium(request):
    if request.method == "POST":
        visitante_id = request.session.get("visitante_id")
        if not visitante_id:
            return redirect("visitante_login")
        visitante = VisitanteAcceso.objects.get(id=visitante_id)

        plan = request.POST.get("plan", "mensual")
        # Define tus price_id reales aquí
        price_id_mensual = "price_1SxjiQPW7xPgzk0m6RimYRkx"
        price_id_anual = "price_1SxjiQPW7xPgzk0mWWn0xHRa"
        price_id = price_id_mensual if plan == "mensual" else price_id_anual

        stripe.api_key = settings.STRIPE_SECRET_KEY

        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=[
                {
                    "price": price_id,
                    "quantity": 1,
                }
            ],
            mode="subscription",
            success_url=request.build_absolute_uri("/visitante/membresia/success/"),
            cancel_url=request.build_absolute_uri("/visitante/membresia/pago/"),
            metadata={"visitante_id": visitante.id},
            client_reference_id=str(visitante.id),
            customer_email=visitante.email,
        )
        return redirect(session.url)
    return redirect("visitante_membresia_pago")


def membresia_pago_exitoso(request):
    visitante_id = request.session.get("visitante_id")
    visitante = VisitanteAcceso.objects.get(id=visitante_id)
    membresia = visitante.membresia_tipo
    return render(
        request, "visitantes/membresia_pago_exitoso.html", {"membresia": membresia}
    )


def visitante_membresia_pago(request):
    return render(request, "visitantes/membresia_pago.html")


@csrf_exempt
def stripe_webhook_membresia(request):
    logger = logging.getLogger("django")
    payload = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE")
    endpoint_secret = getattr(settings, "STRIPE_ENDPOINT_SECRET", None)
    stripe.api_key = settings.STRIPE_SECRET_KEY

    print("Stripe webhook recibido")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
        logger.info(f"Evento Stripe: {event['type']}")
    except Exception as e:
        logger.error(f"Error validando Stripe webhook: {e}")
        return HttpResponse(status=400)

    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        logger.info(f"Session ID: {session.get('id')}")
        visitante_id = session.get("metadata", {}).get("visitante_id") or session.get(
            "client_reference_id"
        )
        visitante = None
        if visitante_id:
            visitante = VisitanteAcceso.objects.filter(id=visitante_id).first()
            logger.info(f"Visitante encontrado por ID: {visitante_id} -> {visitante}")
        else:
            logger.warning(
                "No se encontró visitante_id en metadata ni client_reference_id."
            )

        # Obtén los line items de la sesión
        try:
            line_items = stripe.checkout.Session.list_line_items(session["id"])
            price_id = line_items.data[0].price.id if line_items.data else None
            logger.info(f"Price ID obtenido: {price_id}")
        except Exception as e:
            logger.error(f"Error obteniendo line items: {e}")
            price_id = None

        # Mapeo de price_id a membresía
        price_map = {
            "price_1SxjLCPW7xPgzk0mppaNE8RE": "plus",
            "price_1SxjLCPW7xPgzk0mSUWcWS5k": "plus",
            "price_1SxjiQPW7xPgzk0m6RimYRkx": "premium",
            "price_1SxjiQPW7xPgzk0mWWn0xHRa": "premium",
            # Agrega aquí todos tus price_id reales
        }
        membresia_tipo = price_map.get(price_id)
        logger.info(
            f"Membresía determinada: {membresia_tipo} para Price ID: {price_id}"
        )

        if visitante and membresia_tipo:
            visitante.membresia_tipo = membresia_tipo
            visitante.save()
            logger.info(
                f"Membresía actualizada a {membresia_tipo} para visitante {visitante.id}"
            )
        else:
            logger.warning(
                "No se pudo actualizar la membresía (visitante o price_id faltante)"
            )

    return HttpResponse(status=200)


############################################################################################
####################### Módulo de votaciones por correo electrónico##########################
def enviar_votacion(tema, lista_correos, request):
    empresa = None
    if hasattr(request.user, "perfilusuario"):
        empresa = request.user.perfilusuario.empresa
    else:
        empresa = None

    nombre_empresa = empresa.nombre if empresa else "Tu empresa"

    for correo in lista_correos:
        token = uuid4().hex
        votacion = VotacionCorreo.objects.create(tema=tema, email=correo, token=token)
        url_si = request.build_absolute_uri(
            reverse("votar_tema_correo", args=[token, "si"])
        )
        url_no = request.build_absolute_uri(
            reverse("votar_tema_correo", args=[token, "no"])
        )
        url_abstencion = request.build_absolute_uri(
            reverse("votar_tema_correo", args=[token, "abstencion"])
        )
        asunto = f"Votación: {tema.titulo} - {nombre_empresa}"
        mensaje = (
            f"Buen día,<br><br>"
            f"Estimado miembro del comité, te invitamos a participar en la siguiente votación:<br><br>"
            f"<strong>{tema.titulo}</strong><br>"
            f"{tema.descripcion}<br><br>"
            f"¿Estás de acuerdo?<br><br><br>"
            f"<a href='{url_si}' style='padding:10px 20px; background:#4caf50; color:white; text-decoration:none;'>Sí</a> "
            f"<a href='{url_no}' style='padding:10px 20px; background:#f44336; color:white; text-decoration:none;'>No</a><br><br>"
            f"<a href='{url_abstencion}' style='padding:10px 20px; background:#ffc107; color:black; text-decoration:none;'>Abstención</a><br><br>"
            f"Gracias por tu participación.<br>"
        )
        send_mail(
            subject=asunto,
            message="Te invitamos a votar. Si no ves los botones, copia y pega los enlaces en tu navegador:\nSí: {0}\nNo: {1}".format(
                url_si, url_no
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[correo],
            html_message=mensaje,
        )


def votar_tema_correo(request, token, respuesta):
    votacion = get_object_or_404(VotacionCorreo, token=token)
    if votacion.voto is not None:
        return HttpResponse("Ya has votado.")
    if respuesta not in ["si", "no", "abstencion"]:
        return HttpResponse("Respuesta inválida.")
    votacion.voto = respuesta
    votacion.fecha_voto = timezone.now()
    votacion.save()
    return HttpResponse("¡Gracias por tu voto!")


def resultados_votacion(request, tema_id):
    empresa = request.user.perfilusuario.empresa
    tema = get_object_or_404(TemaGeneral, id=tema_id, empresa=empresa)
    votos = VotacionCorreo.objects.filter(tema=tema)
    total = votos.count()
    si = votos.filter(voto="si").count()
    no = votos.filter(voto="no").count()
    abstencion = votos.filter(voto="abstencion").count()
    pendientes = votos.filter(voto__isnull=True).count()
    return render(
        request,
        "votaciones/resultados_votacion.html",
        {
            "tema": tema,
            "total": total,
            "si": si,
            "no": no,
            "abstencion": abstencion,
            "pendientes": pendientes,
            "votos": votos,
        },
    )


@login_required
def lista_temas(request):
    empresa = request.user.perfilusuario.empresa
    temas = TemaGeneral.objects.filter(empresa=empresa).order_by("-fecha_creacion")
    return render(request, "votaciones/lista_temas.html", {"temas": temas})


@login_required
def crear_tema_y_enviar(request):
    empresa = request.user.perfilusuario.empresa
    tema_id = request.GET.get("tema_id")
    tema = None
    if tema_id:
        tema = get_object_or_404(TemaGeneral, id=tema_id, empresa=empresa)
    if request.method == "POST":
        if tema:
            form = TemaGeneralForm(request.POST, instance=tema)
        else:
            form = TemaGeneralForm(request.POST)
        if form.is_valid():
            tema = form.save(commit=False)
            tema.creado_por = request.user
            tema.empresa = empresa
            tema.save()
            # Procesa los correos
            lista_correos = [
                c.strip() for c in form.cleaned_data["correos"].split(",") if c.strip()
            ]
            enviar_votacion(tema, lista_correos, request)
            messages.success(request, "Asunto creado y correos enviados.")
            return redirect("lista_temas")
    else:
        if tema:
            # Precarga los correos anteriores si existen votaciones previas
            correos_previos = ", ".join(
                VotacionCorreo.objects.filter(tema=tema).values_list("email", flat=True)
            )
            form = TemaGeneralForm(instance=tema, initial={"correos": correos_previos})
        else:
            form = TemaGeneralForm()
    return render(request, "votaciones/crear_tema.html", {"form": form})


@login_required
def eliminar_tema(request, tema_id):
    empresa = request.user.perfilusuario.empresa
    tema = get_object_or_404(TemaGeneral, id=tema_id, empresa=empresa)
    tema.delete()
    messages.success(request, "Asunto eliminado correctamente.")
    return redirect("lista_temas")


# modulo avisos y notificaciones-->
@login_required
def avisos_lista(request):
    empresa = request.user.perfilusuario.empresa
    avisos = Aviso.objects.filter(empresa=empresa).order_by("-fecha_creacion")
    return render(request, "avisos/avisos_lista.html", {"avisos": avisos})


@login_required
def aviso_crear(request):
    empresa = request.user.perfilusuario.empresa
    if request.method == "POST":
        form = AvisoForm(request.POST)
        if form.is_valid():
            aviso = form.save(commit=False)
            aviso.empresa = empresa
            aviso.usuario = request.user
            aviso.save()
            return redirect("avisos_lista")
    else:
        form = AvisoForm()
    return render(request, "avisos/aviso_form.html", {"form": form})


@login_required
def aviso_eliminar(request, aviso_id):
    aviso = get_object_or_404(Aviso, id=aviso_id)
    if request.method == "POST":
        aviso.delete()
        return redirect("avisos_lista")
    return render(request, "avisos/aviso_confirmar_eliminar.html", {"aviso": aviso})

#################################################################################################
# Modulo de timbrado de facturas con FACTURAMA PENDIENTE PROBAR EN PRODUCCION
FACTURAMA_USER = os.getenv("FACTURAMA_USER")
FACTURAMA_PASS = os.getenv("FACTURAMA_PASSWORD")


# configuracion modulo facturama
def timbrar_factura_facturama(datos_factura):
    url = (
        "https://apisandbox.facturama.mx/api-lite/3/cfdis"  # URL de sandbox desarrollo
    )
    # url = 'https://api.facturama.mx/'  # URL de producción
    # print("JSON enviado a Facturama:", json.dumps(datos_factura, indent=2))
    # print("FACTURAMA_USER:", FACTURAMA_USER)
    # print("FACTURAMA_PASS:", FACTURAMA_PASS)
    response = requests.post(
        url,
        auth=(FACTURAMA_USER, FACTURAMA_PASS),
        headers={"Content-Type": "application/json"},
        data=json.dumps(datos_factura),
    )
    # print("Status code:", response.status_code)
    # print("Response text:", response.text)
    try:
        return response.json()
    except Exception:
        return {"error": response.text}


def factura_a_json_facturama(
    factura, tax_object="02", payment_method="PUE", payment_form="99"
):
    monto = Decimal(factura.monto)
    tasa_iva = Decimal("0.16")
    divisor_iva = Decimal("1.16")

    empresa = factura.empresa
    cliente = factura.cliente

    tz_mx = pytz.timezone("America/Mexico_City")
    fecha_timbrado = timezone.now().astimezone(tz_mx).strftime("%Y-%m-%d %H:%M:%S")

    if hasattr(factura, "local") and factura.local:
        # Factura de local comercial
        descripcion = (
            f"Cuota de mantenimiento local {factura.local.numero} "
            f"({format_date(factura.fecha_vencimiento, 'LLLL yyyy', locale='es')})"
        )
    elif hasattr(factura, "area_comun") and factura.area_comun:
        # Factura de área común
        descripcion = (
            f"Cuota área común {factura.area_comun.numero} "
            f"({format_date(factura.fecha_vencimiento, 'LLLL yyyy', locale='es')})"
        )
    elif (
        hasattr(factura, "descripcion")
        and factura.descripcion
        and factura.folio.startswith("FG-")
    ):
        # Factura global (folio inicia con FG-)
        locales = set()
        for f in Factura.objects.filter(factura_global=factura):
            if f.local:
                locales.add(f.local.numero)
        locales_str = ", ".join(sorted(locales))
        descripcion = (
            f"Factura global locales: {locales_str} \n"
            f"({format_date(factura.fecha_vencimiento, 'LLLL yyyy', locale='es')})"
        )
    elif hasattr(factura, "tipo_ingreso"):
        # FacturaOtrosIngresos
        descripcion = (
            "Otro ingreso: "
            + str(getattr(factura, "tipo_ingreso", ""))
            + (
                f" - {factura.observaciones}"
                if getattr(factura, "observaciones", "")
                else ""
            )
        )
    else:
        descripcion = factura.observaciones or "Concepto de factura"

    if tax_object == "01":
        # Sin objeto de impuesto: total y subtotal son iguales al monto
        subtotal = monto.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        item_total = subtotal
        total = subtotal
    else:
        # Con objeto de impuesto: calcula subtotal y desglose de IVA
        subtotal = (monto / divisor_iva).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        iva = (subtotal * tasa_iva).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        item_total = (subtotal + iva).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total = item_total

    item = {
        "ProductCode": "25173108",
        "Description": descripcion,
        "UnitCode": "E48",
        "Quantity": 1.0,
        "UnitPrice": float(subtotal),
        "Subtotal": float(subtotal),
        "TaxObject": tax_object,
        "Total": float(item_total),
    }

    if tax_object == "02":
        item["Taxes"] = [
            {
                "Total": float(iva),
                "Name": "IVA",
                "Base": float(subtotal),
                "Rate": float(tasa_iva),
                "IsRetention": False,
            }
        ]

    return {
        "CfdiType": "I",
        "PaymentForm": payment_form,
        "PaymentMethod": payment_method,
        "ExpeditionPlace": empresa.codigo_postal,
        "Date": fecha_timbrado,
        "Folio": factura.folio,
        "Issuer": {
            "FiscalRegime": empresa.regimen_fiscal,
            "Rfc": empresa.rfc,
            "Name": empresa.nombre,
        },
        "Receiver": {
            "Rfc": cliente.rfc,
            "CfdiUse": cliente.uso_cfdi,
            "Name": cliente.nombre,
            "FiscalRegime": cliente.regimen_fiscal,
            "TaxZipCode": cliente.codigo_postal,
        },
        "Items": [item],
        "Total": float(total),
    }


# timbrar facturas cuotas de mantenimiento y areas comunes
@login_required
def timbrar_factura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    empresa = factura.empresa
    next_url = request.GET.get("next") or request.POST.get("next")
    # # Solo permite timbrar si la empresa es PREMIUM
    # if not empresa.es_premium:
    #     messages.error(
    #         request, "El timbrado solo está disponible en la versión PREMIUM del sistema."
    #     )
    #     return redirect("lista_facturas")

    if (
        not request.user.is_superuser
        and factura.empresa != request.user.perfilusuario.empresa
    ):
        messages.error(request, "No tienes permiso para timbrar esta factura.")
        return redirect("lista_facturas")

    if factura.uuid:
        messages.info(request, "La factura ya está timbrada.")
        return redirect("lista_facturas")

    # --- INICIO FLUJO FACTURA GLOBAL SOLO PARA LOCALES ---
    if factura.cliente.factura_global:
        mes = factura.fecha_vencimiento.month
        anio = factura.fecha_vencimiento.year
        cliente = factura.cliente

        # Solo facturas de locales, no áreas ni otros ingresos
        facturas_mes = Factura.objects.filter(
            cliente=cliente,
            empresa=empresa,
            fecha_vencimiento__year=anio,
            fecha_vencimiento__month=mes,
            uuid__isnull=True,
            local__isnull=False,  # Solo locales
        )

        if facturas_mes.count() > 1:
            total_monto = sum(f.monto for f in facturas_mes)
            descripcion = f"Factura global locales:" + ", ".join(
                [f.local.numero for f in facturas_mes if f.local]
            )

            # Marca las facturas individuales como incluidas en la global
            # facturas_mes.update(estatus="incluida_global", factura_global=factura_global)

            # Timbrar la factura global
            if request.method == "POST":
                form = TimbrarFacturaForm(request.POST)
                if form.is_valid():
                    # Crea la factura global
                    factura_global = Factura.objects.create(
                        empresa=empresa,
                        cliente=cliente,
                        monto=total_monto,
                        fecha_emision=timezone.now(),
                        fecha_vencimiento=factura.fecha_vencimiento,
                        folio="FG-" + timezone.now().strftime("%Y%m%d%H%M%S"),
                        observaciones=descripcion,
                    )
                    tax_object = form.cleaned_data["tax_object"]
                    payment_method = form.cleaned_data["payment_method"]
                    payment_form = form.cleaned_data["payment_form"]
                    datos_json = factura_a_json_facturama(
                        factura_global, tax_object, payment_method, payment_form
                    )
                    resultado = timbrar_factura_facturama(datos_json)

                    if "error" in resultado:
                        messages.error(
                            request, f"Error al timbrar: {resultado['error']}"
                        )
                    else:
                        uuid = resultado.get("Uuid") or resultado.get(
                            "Complement", {}
                        ).get("TaxStamp", {}).get("Uuid")
                        facturama_id = resultado.get("Id")
                        if not uuid or not facturama_id:
                            messages.error(request, f"Error inesperado: {resultado}")
                        else:
                            factura_global.uuid = uuid
                            factura_global.facturama_id = facturama_id
                            factura_global.save()
                            # Solo aquí actualiza las facturas individuales
                            # facturas_mes.update(estatus="incluida_global", factura_global=factura_global)
                            for f in facturas_mes:
                                f.factura_global = factura_global
                                f.uuid = factura_global.uuid
                                f.facturama_id = factura_global.facturama_id
                                f.save(
                                    update_fields=[
                                        "factura_global",
                                        "uuid",
                                        "facturama_id",
                                    ]
                                )
                            messages.success(
                                request,
                                f"Factura global {factura_global.folio} timbrada correctamente. Ahora puedes descargar el PDF y XML.",
                            )
                    if next_url:
                        return redirect(next_url)
                    return redirect("lista_facturas")
            else:
                form = TimbrarFacturaForm()
            return render(
                request,
                "facturacion/timbrar_factura.html",
                {
                    "form": form,
                    "factura": factura,
                    "url_cancelar": next_url,
                },
            )
        # Si solo hay una factura, sigue el flujo normal
    # --- FIN FLUJO FACTURA GLOBAL ---

    # FLUJO NORMAL (una sola factura)
    if request.method == "POST":
        form = TimbrarFacturaForm(request.POST)
        if form.is_valid():
            tax_object = form.cleaned_data["tax_object"]
            payment_method = form.cleaned_data["payment_method"]
            payment_form = form.cleaned_data["payment_form"]
            datos_json = factura_a_json_facturama(
                factura, tax_object, payment_method, payment_form
            )
            resultado = timbrar_factura_facturama(datos_json)

            if "error" in resultado:
                messages.error(request, f"Error al timbrar: {resultado['error']}")
            else:
                uuid = resultado.get("Uuid") or resultado.get("Complement", {}).get(
                    "TaxStamp", {}
                ).get("Uuid")
                facturama_id = resultado.get("Id")
                if not uuid or not facturama_id:
                    messages.error(request, f"Error inesperado: {resultado}")
                else:
                    factura.uuid = uuid
                    factura.facturama_id = facturama_id
                    factura.save()
                    messages.success(
                        request,
                        "Factura "
                        + factura.folio
                        + " timbrada correctamente. Ahora puedes descargar el PDF y XML.",
                    )
            if next_url:
                return redirect(next_url)
            return redirect("lista_facturas")
    else:
        form = TimbrarFacturaForm()
    next_url = request.GET.get("next")
    return render(
        request,
        "facturacion/timbrar_factura.html",
        {
            "form": form,
            "factura": factura,
            "url_cancelar": next_url,
        },
    )


@login_required
def descargar_factura_timbrada(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    if not factura.uuid:
        messages.error(request, "La factura no está timbrada.")
        return redirect("lista_facturas")

    uuid = factura.uuid
    usuario = os.getenv("FACTURAMA_USER")
    password = os.getenv("FACTURAMA_PASSWORD")

    # URLs para descargar XML y PDF
    xml_url = (
        f"https://apisandbox.facturama.mx/api-lite/3/cfdis/{uuid}/xml"  # desarrollo
    )
    pdf_url = (
        f"https://apisandbox.facturama.mx/api-lite/3/cfdis/{uuid}/pdf"  # desarrollo
    )

    # xml_url = f"https://api.facturama.mx/api-lite/3/cfdis/{uuid}/xml"  #producción
    # pdf_url = f"https://api.facturama.mx/api-lite/3/cfdis/{uuid}/pdf"  #producción

    # Descarga XML
    xml_response = requests.get(xml_url, auth=(usuario, password))
    xml_content = xml_response.content if xml_response.status_code == 200 else None
    print("XML status:", xml_response.status_code)
    print("XML response:", xml_response.text)

    # Descarga PDF
    pdf_response = requests.get(pdf_url, auth=(usuario, password))
    pdf_content = pdf_response.content if pdf_response.status_code == 200 else None
    print("PDF status:", pdf_response.status_code)
    print("PDF response:", pdf_response.text)

    if not xml_content and not pdf_content:
        messages.error(
            request, "No se pudo descargar el PDF ni el XML desde Facturama."
        )
        return redirect("lista_facturas")

    # Prepara ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        if xml_content:
            zip_file.writestr(f"factura_{factura.folio}.xml", xml_content)
        if pdf_content:
            zip_file.writestr(f"factura_{factura.folio}.pdf", pdf_content)

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = (
        f"attachment; filename=factura_{factura.folio}.zip"
    )
    return response


@staff_member_required
def subir_csd_facturama(request):
    import os

    mensaje = ""
    usuario = os.getenv("FACTURAMA_USER")
    password = os.getenv("FACTURAMA_PASSWORD")
    csds = obtener_csds_facturama(usuario, password)

    if request.method == "POST":
        form = CSDUploadForm(request.POST, request.FILES)
        if form.is_valid():
            empresa = form.cleaned_data["empresa"]
            rfc = empresa.rfc.upper()
            cer_file = form.cleaned_data["cer_file"]
            key_file = form.cleaned_data["key_file"]
            key_password = form.cleaned_data["key_password"]

            # Valida si ya existe un CSD para ese RFC
            if any(csd.get("Rfc", "").upper() == rfc for csd in csds):
                mensaje = f"Ya existe un CSD cargado para el RFC {rfc}."
            else:
                # Convierte archivos a base64
                cert_b64 = base64.b64encode(cer_file.read()).decode("utf-8")
                key_b64 = base64.b64encode(key_file.read()).decode("utf-8")

                # Llama a la API de Facturama
                url = "https://apisandbox.facturama.mx/api-lite/csds"  # en desarrollo
                # url = "https://api.facturama.mx/api-lite/csds" # en producción
                data = {
                    "Rfc": rfc,
                    "Certificate": cert_b64,
                    "PrivateKey": key_b64,
                    "PrivateKeyPassword": key_password,
                }
                response = requests.post(
                    url,
                    auth=(usuario, password),
                    headers={"Content-Type": "application/json"},
                    data=json.dumps(data),
                )
                if response.status_code == 200:
                    mensaje = (
                        f"CSD cargado correctamente para la empresa {empresa.nombre}."
                    )
                    csds = obtener_csds_facturama(usuario, password)  # Actualiza lista
                else:
                    print("Status code:", response.status_code)
                    print("Response text:", response.text)
                    mensaje = f"Error al cargar CSD: {response.text}"
    else:
        form = CSDUploadForm()
    return render(
        request,
        "facturacion/subir_csd.html",
        {"form": form, "mensaje": mensaje, "csds": csds},
    )


def obtener_csds_facturama(usuario, password):
    url = "https://apisandbox.facturama.mx/api-lite/csds"  # URL de sandbox desarrollo
    # url = "https://api.facturama.mx/api-lite/csds"  # URL de producción
    response = requests.get(url, auth=(usuario, password))
    if response.status_code == 200:
        return response.json()  # Lista de CSDs
    return []


def consultar_cfdis_facturama(
    rfc_issuer=None,
    uuid=None,
    folio_start=None,
    folio_end=None,
    date_start=None,
    date_end=None,
    status="active",
    page=0,
):
    url = "https://apisandbox.facturama.mx/cfdi"  # URL de sandbox desarrollo
    # url = "https://api.facturama.mx/cfdi"  # URL de producción
    params = {"type": "issuedLite", "status": status, "page": page}
    if rfc_issuer:
        params["rfcIssuer"] = rfc_issuer
    if uuid:
        params["uuid"] = uuid
    if folio_start:
        params["folioStart"] = folio_start
    if folio_end:
        params["folioEnd"] = folio_end
    if date_start:
        params["dateStart"] = date_start  # formato: dd/mm/yyyy
    if date_end:
        params["dateEnd"] = date_end  # formato: dd/mm/yyyy

    usuario = os.getenv("FACTURAMA_USER")
    password = os.getenv("FACTURAMA_PASSWORD")
    response = requests.get(url, params=params, auth=(usuario, password))
    if response.status_code == 200:
        return response.json()
    return []


@staff_member_required
def consulta_cfdis_facturama(request):
    empresas = Empresa.objects.all()
    resultados = []
    mensaje = ""
    if request.method == "POST":
        empresa_id = request.POST.get("empresa_id")
        uuid = request.POST.get("uuid")
        folio_start = request.POST.get("folio_start")
        folio_end = request.POST.get("folio_end")
        date_start = request.POST.get("date_start")
        date_end = request.POST.get("date_end")

        empresa = Empresa.objects.filter(id=empresa_id).first()
        rfc_issuer = empresa.rfc if empresa else None

        resultados = consultar_cfdis_facturama(
            rfc_issuer=rfc_issuer,
            uuid=uuid,
            folio_start=folio_start,
            folio_end=folio_end,
            date_start=date_start,
            date_end=date_end,
        )
        if not resultados:
            mensaje = "No se encontraron CFDIs con esos filtros o hubo un error en la consulta."

    return render(
        request,
        "facturacion/consulta_cfdis.html",
        {"empresas": empresas, "resultados": resultados, "mensaje": mensaje},
    )


def descargar_cfdi_facturama(request, id):
    usuario = os.getenv("FACTURAMA_USER")
    password = os.getenv("FACTURAMA_PASSWORD")
    base_url = "https://apisandbox.facturama.mx/cfdi"  # URL de sandbox desarrollo
    # base_url = "https://api.facturama.mx/cfdi"  # URL de producción

    # Descarga XML
    xml_url = f"{base_url}/xml/issuedLite/{id}"
    xml_response = requests.get(xml_url, auth=(usuario, password))
    xml_content = None
    if xml_response.status_code == 200:
        xml_json = xml_response.json()
        xml_content = base64.b64decode(xml_json.get("Content", ""))

    # Descarga PDF
    pdf_url = f"{base_url}/pdf/issuedLite/{id}"
    pdf_response = requests.get(pdf_url, auth=(usuario, password))
    pdf_content = None
    if pdf_response.status_code == 200:
        pdf_json = pdf_response.json()
        pdf_content = base64.b64decode(pdf_json.get("Content", ""))

    if not xml_content and not pdf_content:
        messages.error(
            request, "No se pudo descargar el PDF ni el XML desde Facturama."
        )
        return redirect("consulta_cfdis_facturama")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        if xml_content:
            zip_file.writestr(f"{id}.xml", xml_content)
        if pdf_content:
            zip_file.writestr(f"{id}.pdf", pdf_content)

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = f"attachment; filename=cfdi_{id}.zip"
    return response


# timbrar facturas de otros ingresos
@login_required
def timbrar_factura_otros_ingresos(request, pk):
    factura = get_object_or_404(FacturaOtrosIngresos, pk=pk)
    # empresa = factura.empresa

    # # Solo permite timbrar si la empresa es PLUS
    # if not empresa.es_plus:
    #     messages.error(
    #         request, "El timbrado solo está disponible en la versión PLUS del sistema."
    #     )
    #     return redirect("lista_facturas_otros_ingresos")

    if factura.uuid:
        messages.info(request, "La factura ya está timbrada.")
        return redirect("lista_facturas_otros_ingresos")

    if request.method == "POST":
        form = TimbrarFacturaForm(request.POST)
        if form.is_valid():
            tax_object = form.cleaned_data["tax_object"]
            payment_method = form.cleaned_data["payment_method"]
            payment_form = form.cleaned_data["payment_form"]
            datos_json = factura_a_json_facturama(
                factura, tax_object, payment_method, payment_form
            )
            resultado = timbrar_factura_facturama(datos_json)
            print("Resultado de timbrado:", resultado)
            if "error" in resultado:
                messages.error(request, f"Error al timbrar: {resultado['error']}")
            else:
                uuid = resultado.get("Uuid") or resultado.get("Complement", {}).get(
                    "TaxStamp", {}
                ).get("Uuid")
                facturama_id = resultado.get("Id")
                if not uuid or not facturama_id:
                    messages.error(request, f"Error inesperado: {resultado}")
                else:
                    factura.uuid = uuid
                    factura.facturama_id = facturama_id
                    factura.save()
                    messages.success(
                        request,
                        "Factura "
                        + factura.folio
                        + " timbrada correctamente. Ahora puedes descargar el PDF y XML.",
                    )
            return redirect("lista_facturas_otros_ingresos")
    else:
        form = TimbrarFacturaForm()

    return render(
        request,
        "facturacion/timbrar_factura.html",
        {
            "form": form,
            "factura": factura,
            "url_cancelar": "lista_facturas_otros_ingresos",
        },
    )


# Modulo APIS visitantes via APLICACIONES MOVILES

# Decorador para verificar token de visitante
def visitante_token_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        token_key = request.headers.get("Authorization", "").replace("Token ", "")
        try:
            token = VisitanteToken.objects.get(key=token_key)
            visitante = token.visitante
            request.visitante = visitante
            return view_func(request, *args, **kwargs)
        except VisitanteToken.DoesNotExist:
            return Response({"error": "Token inválido"}, status=401)

    return _wrapped_view


# APIS registro visitante
# Lista de empresas
@api_view(["GET"])
def api_empresas_lista(request):
    empresas = Empresa.objects.all().order_by("nombre")
    data = [{"id": e.id, "nombre": e.nombre} for e in empresas]
    return Response(data)


# Lista de locales por empresa
@api_view(["GET"])
def api_locales_por_empresa(request, empresa_id):
    locales = LocalComercial.objects.filter(empresa_id=empresa_id).order_by("numero")
    data = [{"id": l.id, "numero": l.numero} for l in locales]
    return Response(data)


# Lista de áreas comunes por empresa
@api_view(["GET"])
def api_areas_por_empresa(request, empresa_id):
    areas = AreaComun.objects.filter(empresa_id=empresa_id).order_by("numero")
    data = [{"id": a.id, "numero": a.numero} for a in areas]
    return Response(data)


# API registro visitante
@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def visitante_registro_api(request):
    empresa_ids = request.data.getlist("empresa_ids[]")
    locales_numeros = request.data.getlist("locales_numeros[]")
    areas_numeros = request.data.getlist("areas_numeros[]")
    nombre_completo = request.data.get("nombre_completo")
    username = request.data.get("username")
    password = request.data.get("password")
    email = request.data.get("email")
    ine_file = request.FILES.get("ine_file")

    if not empresa_ids:
        return Response(
            {"ok": False, "error": "Debes seleccionar una empresa."}, status=400
        )
    if not ine_file:
        return Response({"ok": False, "error": "Debes subir tu INE."}, status=400)

    if not locales_numeros and not areas_numeros:
        return Response(
            {
                "ok": False,
                "error": "Debes seleccionar al menos un local o un área común.",
            },
            status=400,
        )

    if VisitanteAcceso.objects.filter(username=username).exists():
        return Response({"ok": False, "error": "El usuario ya existe."}, status=400)

    visitante = VisitanteAcceso.objects.create(
        nombre=nombre_completo,
        username=username,
        email=email,
        activo=False,
        # empresa=empresas.set(empresa)
    )

    visitante.set_password(password)
    visitante.save()

    empresa_objs = Empresa.objects.filter(id__in=empresa_ids)
    empresas_dict = {str(e.id): e.nombre for e in empresa_objs}

    # Agrupa locales por empresa
    locales_por_empresa = {}
    for item in locales_numeros:
        if "-" in item:
            empresa_id, numero = item.split("-", 1)
            nombre_empresa = empresas_dict.get(empresa_id, f"ID {empresa_id}")
            locales_por_empresa.setdefault(nombre_empresa, []).append(numero)

    # Agrupa áreas por empresa
    areas_por_empresa = {}
    for item in areas_numeros:
        if "-" in item:
            empresa_id, numero = item.split("-", 1)
            nombre_empresa = empresas_dict.get(empresa_id, f"ID {empresa_id}")
            areas_por_empresa.setdefault(nombre_empresa, []).append(numero)

    # Construye el mensaje agrupado
    detalle_empresas = ""
    for empresa in empresa_objs:
        nombre = empresa.nombre
        locales_list = locales_por_empresa.get(nombre, [])
        areas_list = areas_por_empresa.get(nombre, [])
        detalle_empresas += f"\nEmpresa: {nombre}\n"
        detalle_empresas += (
            f"  Locales: {', '.join(locales_list) if locales_list else '-'}\n"
        )
        detalle_empresas += f"  Áreas: {', '.join(areas_list) if areas_list else '-'}\n"

    # Notifica solo al superusuario/admin para validación manual
    admin_email = settings.EMAIL_HOST_USER
    asunto = "Solicitud de registro de visitante"
    mensaje = (
        f"Nuevo visitante solicita acceso:\n\n"
        f"Nombre: {nombre_completo}\n"
        f"Usuario: {username}\n"
        f"Email: {email}\n"
        f"{detalle_empresas}\n"
        f"Revisar y asignar relaciones en el admin.\n"
    )
    email_msg = EmailMessage(
        subject=asunto,
        body=mensaje,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[admin_email] if admin_email else [],
    )
    if ine_file:
        email_msg.attach(ine_file.name, ine_file.read(), ine_file.content_type)
    email_msg.send(fail_silently=True)

    # print(list(destinatarios))
    return Response(
        {
            "ok": True,
            "mensaje": "Tu registro fue enviado. El sistema validará tus datos y te notificará por correo.",
        }
    )


# API login visitante
@api_view(["POST"])
def visitante_login_api(request):
    username = request.data.get("username")
    password = request.data.get("password")
    try:
        visitante = VisitanteAcceso.objects.get(username=username)
        if visitante.check_password(password):
            # Obtén todas las empresas asociadas al visitante
            empresas = visitante.empresas.all().distinct()
            empresas_data = [
                {
                    "id": e.id,
                    "nombre": e.nombre,
                    "email": e.email,
                    "stripe_public_key": e.stripe_public_key,
                }
                for e in empresas
            ]
            es_admin = getattr(visitante, "es_admin", False)
            # Crea o recupera el token
            token, _ = VisitanteToken.objects.get_or_create(visitante=visitante)
            return Response(
                {
                    "ok": True,
                    "visitante_id": visitante.id,
                    "empresas": empresas_data,
                    "token": token.key,
                    "es_admin": es_admin,
                }
            )
        else:
            return Response({"ok": False, "error": "Contraseña incorrecta"}, status=400)
    except VisitanteAcceso.DoesNotExist:
        return Response({"ok": False, "error": "Usuario no encontrado"}, status=404)


# API obtener facturas del visitante
@api_view(["GET"])
@visitante_token_required
def visitante_facturas_api(request):
    visitante = request.visitante  # El usuario autenticado por token
    empresa_id = request.GET.get("empresa_id")
    if not empresa_id:
        return Response({"error": "Debe seleccionar una empresa"}, status=400)
    locales = visitante.locales.filter(empresa_id=empresa_id)
    areas = visitante.areas.filter(empresa_id=empresa_id)
    facturas = Factura.objects.filter(
        Q(local__in=locales) | Q(area_comun__in=areas),
        empresa_id=empresa_id,
    ).select_related("cliente", "empresa", "local", "area_comun")
    serializer = FacturaSerializer(facturas, many=True)
    return Response({"facturas": serializer.data})


# API crear Payment Intent con Stripe
@api_view(["POST"])
def create_payment_intent(request):
    amount = request.data.get("amount")
    factura_id = request.data.get("factura_id")
    if not amount:
        return Response({"error": "Monto requerido"}, status=400)
    try:
        factura = Factura.objects.get(id=factura_id)
        empresa = factura.empresa
        stripe.api_key = empresa.stripe_secret_key
        intent = stripe.PaymentIntent.create(
            amount=int(float(amount) * 100),
            currency="mxn",
            payment_method_types=["card"],
            metadata={"factura_id": factura_id},
        )
        return Response({"client_secret": intent.client_secret})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# API reporte ingresos vs gastos
@api_view(["GET"])
@visitante_token_required
def api_reporte_ingresos_vs_gastos(request):
    visitante = request.visitante
    if not getattr(visitante, "acceso_api_reporte", False):
        return Response({"error": "Acceso denegado"}, status=403)

    if getattr(visitante, "es_admin", False):
        # if getattr(visitante, "es_admin", False) or request.GET.get("empresa_id"):
        empresa_id = request.GET.get("empresa_id")
        if not empresa_id:
            return Response({"error": "Debe seleccionar una empresa"}, status=400)
        # solo empresas vinculadas al visitante administrador
        empresa = Empresa.objects.filter(id=empresa_id).first()
        if not empresa:
            return Response({"error": "Empresa no encontrada"}, status=404)
    else:
        empresa = None
        empresa_id = None
        if visitante.locales.exists():
            empresa = visitante.locales.first().empresa
            empresa_id = str(empresa.id) if empresa else None
        elif visitante.areas.exists():
            empresa = visitante.areas.first().empresa
            empresa_id = str(empresa.id) if empresa else None

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    periodo = request.GET.get("periodo")
    hoy = date.today()

    # Si no hay ningún filtro, mostrar periodo actual por default
    if not periodo and not fecha_inicio and not fecha_fin and not mes and not anio:
        periodo = "periodo_actual"

    # Prioridad: periodo > mes/año > fechas manuales
    if periodo == "mes_actual":
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = (hoy.replace(day=1) + timedelta(days=32)).replace(
            day=1
        ) - timedelta(days=1)
        mes = hoy.month
        anio = hoy.year
    elif periodo == "periodo_actual":
        fecha_inicio = hoy.replace(month=1, day=1)
        fecha_fin = hoy
        mes = ""
        anio = ""
    elif mes and anio:
        try:
            mes = int(mes)
            anio = int(anio)
            fecha_inicio = date(anio, mes, 1)
            # if mes == 12:
            #     fecha_fin = date(anio, 12, 31)
            # else:
            #     fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
            fecha_fin = (
                date(anio, mes + 1, 1) - timedelta(days=1)
                if mes < 12
                else date(anio, 12, 31)
            )
        except Exception:
            fecha_inicio = None
            fecha_fin = None
    elif fecha_inicio and fecha_fin:
        # Ya vienen del formulario
        pass
    else:
        fecha_inicio = None
        fecha_fin = None

    # Convierte a date si es string
    if isinstance(fecha_inicio, str):
        try:
            # fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_inicio = date.fromisoformat(fecha_inicio)
        except Exception:
            fecha_inicio = None
    # else:
    #     fecha_inicio_dt = fecha_inicio

    if isinstance(fecha_fin, str):
        try:
            # fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            fecha_fin = date.fromisoformat(fecha_fin)
        except Exception:
            fecha_fin = None
    # else:
    #     fecha_fin_dt = fecha_fin

    # Para mostrar el mes y año en letras
    try:
        locale.setlocale(locale.LC_TIME, "es_MX.UTF-8")
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
        except locale.Error:
            locale.setlocale(locale.LC_TIME, "C")  # Fallback seguro

    # mes_letra = ""
    # if (
    #     fecha_inicio_dt
    #     and fecha_fin_dt
    #     and fecha_inicio_dt == fecha_fin_dt.replace(day=1)
    # ):
    #     mes_letra = fecha_inicio_dt.strftime("%B %Y").capitalize()
    # elif fecha_inicio_dt and fecha_fin_dt:
    #     mes_letra = f"{fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')}"

    mes_letra = ""
    if fecha_inicio and fecha_fin:
        if (
            fecha_inicio.month == fecha_fin.month
            and fecha_inicio.year == fecha_fin.year
        ):
            mes_letra = fecha_inicio.strftime("%B %Y").capitalize()
        else:
            mes_letra = f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"

    # Filtra todos los objetos por la empresa del visitante
    pagos = Pago.objects.exclude(forma_pago="nota_credito").filter(
        factura__empresa=empresa
    )
    pagos_gastos = PagoGasto.objects.filter(gasto__empresa=empresa)
    cobros_otros = CobroOtrosIngresos.objects.select_related(
        "factura", "factura__empresa"
    ).filter(factura__empresa=empresa)
    gastos_caja_chica = GastoCajaChica.objects.filter(fondeo__empresa=empresa)
    vales_caja_chica = ValeCaja.objects.filter(fondeo__empresa=empresa)

    # PAGOS POR IDENTIFICAR
    pagos_por_identificar = Pago.objects.filter(
        factura__isnull=True, identificado=False
    )

    # Aplica filtros de fecha
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio)
        pagos_gastos = pagos_gastos.filter(fecha_pago__gte=fecha_inicio)
        cobros_otros = cobros_otros.filter(fecha_cobro__gte=fecha_inicio)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__gte=fecha_inicio)
        vales_caja_chica = vales_caja_chica.filter(fecha__gte=fecha_inicio)
        pagos_por_identificar = pagos_por_identificar.filter(
            fecha_pago__gte=fecha_inicio
        )
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin)
        pagos_gastos = pagos_gastos.filter(fecha_pago__lte=fecha_fin)
        cobros_otros = cobros_otros.filter(fecha_cobro__lte=fecha_fin)
        gastos_caja_chica = gastos_caja_chica.filter(fecha__lte=fecha_fin)
        vales_caja_chica = vales_caja_chica.filter(fecha__lte=fecha_fin)
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__lte=fecha_fin)

    total_ingresos = pagos.aggregate(total=Sum("monto"))["total"] or 0
    total_otros_ingresos = cobros_otros.aggregate(total=Sum("monto"))["total"] or 0
    total_pagos_por_identificar = (
        pagos_por_identificar.aggregate(total=Sum("monto"))["total"] or 0
    )
    total_ingresos_cobrados = (
        total_ingresos + total_otros_ingresos + total_pagos_por_identificar
    )

    total_gastos_pagados = pagos_gastos.aggregate(total=Sum("monto"))["total"] or 0
    total_gastos_caja_chica = (
        gastos_caja_chica.aggregate(total=Sum("importe"))["total"] or 0
    )
    total_vales_caja_chica = (
        vales_caja_chica.aggregate(total=Sum("importe"))["total"] or 0
    )
    total_egresos = (
        total_gastos_pagados + total_gastos_caja_chica + total_vales_caja_chica
    )

    # Agrupar por tipo de origen (Local/Área)
    ingresos_qs = (
        pagos.annotate(
            origen=Case(
                When(factura__local__isnull=False, then=Value("Propiedades")),
                When(factura__area_comun__isnull=False, then=Value("Áreas Comunes")),
                default=Value("Sin origen"),
                output_field=CharField(),
            )
        )
        .values("origen")
        .annotate(total=Sum("monto"))
        .order_by("origen")
    )

    otros_ingresos_qs = (
        cobros_otros.select_related("factura__tipo_ingreso")
        .values("factura__tipo_ingreso__nombre")
        .annotate(total=Sum("monto"))
        .order_by("factura__tipo_ingreso")
    )

    ingresos_por_origen = OrderedDict()
    for x in ingresos_qs:
        ingresos_por_origen[x["origen"]] = float(x["total"])
    for x in otros_ingresos_qs:
        tipo = x["factura__tipo_ingreso__nombre"] or "Otros ingresos"
        ingresos_por_origen[f"{tipo}"] = float(x["total"])
        ingresos_por_origen["Depositos no identificados"] = float(
            total_pagos_por_identificar
        )

    gastos_agregados = {}
    gastos_qs = PagoGasto.objects.select_related("gasto__tipo_gasto__subgrupo__grupo")
    gastos_qs = gastos_qs.filter(gasto__empresa=empresa)

    if fecha_inicio:
        gastos_qs = gastos_qs.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        gastos_qs = gastos_qs.filter(fecha_pago__lte=fecha_fin)

    for g in gastos_qs.values(
        "gasto__tipo_gasto__subgrupo__grupo__nombre",
        "gasto__tipo_gasto__subgrupo__nombre",
        "gasto__tipo_gasto__nombre",
    ).annotate(total=Sum("monto")):
        grupo = (
            (g["gasto__tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
            .strip()
            .title()
        )
        subgrupo = (
            (g["gasto__tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
        )
        tipo = (g["gasto__tipo_gasto__nombre"] or "Sin tipo").strip().title()
        key = (grupo, subgrupo, tipo)
        gastos_agregados[key] = gastos_agregados.get(key, 0) + float(g["total"])

    for g in gastos_caja_chica.values(
        "tipo_gasto__subgrupo__grupo__nombre",
        "tipo_gasto__subgrupo__nombre",
        "tipo_gasto__nombre",
    ).annotate(total=Sum("importe")):
        grupo = (
            (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo").strip().title()
        )
        subgrupo = (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
        tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
        key = (grupo, subgrupo, tipo)
        gastos_agregados[key] = gastos_agregados.get(key, 0) + float(g["total"])

    for g in vales_caja_chica.values(
        "tipo_gasto__subgrupo__grupo__nombre",
        "tipo_gasto__subgrupo__nombre",
        "tipo_gasto__nombre",
    ).annotate(total=Sum("importe")):
        grupo = (
            (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo").strip().title()
        )
        subgrupo = (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
        tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
        key = (grupo, subgrupo, tipo)
        gastos_agregados[key] = gastos_agregados.get(key, 0) + float(g["total"])

    estructura_gastos = OrderedDict()
    for (grupo, subgrupo, tipo), total in gastos_agregados.items():
        if grupo not in estructura_gastos:
            estructura_gastos[grupo] = OrderedDict()
        if subgrupo not in estructura_gastos[grupo]:
            estructura_gastos[grupo][subgrupo] = []
        estructura_gastos[grupo][subgrupo].append({"tipo": tipo, "total": total})

    saldo = total_ingresos_cobrados - total_egresos

    resultado = {
        "total_ingresos": total_ingresos_cobrados,
        "total_otros_ingresos": total_otros_ingresos,
        "total_pagos_por_identificar": total_pagos_por_identificar,
        "total_gastos_pagados": total_gastos_pagados,
        "total_gastos_caja_chica": total_gastos_caja_chica,
        "total_vales_caja_chica": total_vales_caja_chica,
        "total_egresos": total_egresos,
        "ingresos_por_origen": ingresos_por_origen,
        "saldo": saldo,
        "periodo": periodo,
        "mes_letra": mes_letra,
        "empresa_nombre": empresa.nombre if empresa else "",
        "empresa_email": empresa.email if empresa else "",
        "estructura_gastos": estructura_gastos,
        "fecha_inicio": str(fecha_inicio) if fecha_inicio else "",
        "fecha_fin": str(fecha_fin) if fecha_fin else "",
        "mes": mes,
        "anio": anio,
    }
    return Response(resultado)


# API dashboard cartera vencida
@api_view(["GET"])
@visitante_token_required
def api_dashboard_saldos_visitante(request):
    visitante = request.visitante
    if not getattr(visitante, "acceso_api_reporte", False):
        return Response({"error": "Acceso denegado"}, status=403)

    empresa_id = request.GET.get("empresa_id")
    empresa = None
    if getattr(visitante, "es_admin", False) or request.GET.get("empresa_id"):
        if not empresa_id:
            return Response({"error": "Debe seleccionar una empresa"}, status=400)
        empresa = Empresa.objects.filter(id=empresa_id).first()
        if not empresa:
            return Response({"error": "Empresa no encontrada"}, status=404)
    else:
        if visitante.locales.exists():
            empresa = visitante.locales.first().empresa
        elif visitante.areas.exists():
            empresa = visitante.areas.first().empresa

    hoy = timezone.now().date()
    cliente_id = request.GET.get("cliente")
    origen = request.GET.get("origen", "todos")
    tipo_cuota = request.GET.get("tipo_cuota")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")

    # Filtro por empresa del visitante
    filtro_empresa = Q(empresa=empresa)

    facturas = Factura.objects.filter(estatus="pendiente").filter(filtro_empresa)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if origen == "local":
        facturas = facturas.filter(local__isnull=False)
    elif origen == "area":
        facturas = facturas.filter(area_comun__isnull=False)
    if tipo_cuota:
        facturas = facturas.filter(tipo_cuota=tipo_cuota)
    if anio:
        try:
            anio = int(anio)
            facturas = facturas.filter(fecha_vencimiento__year=anio)
        except ValueError:
            pass
    if mes:
        try:
            mes = int(mes)
            facturas = facturas.filter(fecha_vencimiento__month=mes)
        except ValueError:
            pass

    pagos_subquery = (
        Pago.objects.filter(factura=OuterRef("pk"))
        .values("factura")
        .annotate(
            total_pagado_dash=Coalesce(
                Sum("monto"), Value(0, output_field=DecimalField())
            )
        )
        .values("total_pagado_dash")
    )
    facturas = facturas.annotate(
        total_pagado_dash=Coalesce(
            Subquery(pagos_subquery), Value(0, output_field=DecimalField())
        ),
        saldo_pendiente_dash=ExpressionWrapper(
            F("monto")
            - Coalesce(Subquery(pagos_subquery), Value(0, output_field=DecimalField())),
            output_field=DecimalField(),
        ),
    )
    # Anota el rango de vencimiento en cada factura
    facturas = facturas.annotate(
        rango=Case(
            When(fecha_vencimiento__gt=hoy - timedelta(days=30), then=Value("0_30")),
            When(
                fecha_vencimiento__gt=hoy - timedelta(days=60),
                fecha_vencimiento__lte=hoy - timedelta(days=30),
                then=Value("31_60"),
            ),
            When(
                fecha_vencimiento__gt=hoy - timedelta(days=90),
                fecha_vencimiento__lte=hoy - timedelta(days=60),
                then=Value("61_90"),
            ),
            When(
                fecha_vencimiento__gt=hoy - timedelta(days=180),
                fecha_vencimiento__lte=hoy - timedelta(days=90),
                then=Value("91_180"),
            ),
            When(
                fecha_vencimiento__lte=hoy - timedelta(days=180), then=Value("181_mas")
            ),
            default=Value("otro"),
            output_field=CharField(),
        )
    )

    # Agrupa y suma en una sola consulta
    saldos = facturas.values("rango").annotate(total=Sum("saldo_pendiente_dash"))
    saldos_dict = {x["rango"]: float(x["total"]) for x in saldos}
    for key in ["0_30", "31_60", "61_90", "91_180", "181_mas"]:
        saldos_dict.setdefault(key, 0.0)

    # Facturas otros ingresos
    facturas_otros = FacturaOtrosIngresos.objects.filter(
        estatus="pendiente", activo=True
    ).filter(filtro_empresa)
    if cliente_id:
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)
    if anio:
        try:
            anio = int(anio)
            facturas_otros = facturas_otros.filter(fecha_vencimiento__year=anio)
        except ValueError:
            pass
    if mes:
        try:
            mes = int(mes)
            facturas_otros = facturas_otros.filter(fecha_vencimiento__month=mes)
        except ValueError:
            pass

    cobros_subquery = (
        CobroOtrosIngresos.objects.filter(factura=OuterRef("pk"))
        .values("factura")
        .annotate(
            total_cobrado_dash=Coalesce(
                Sum("monto"), Value(0, output_field=DecimalField())
            )
        )
        .values("total_cobrado_dash")
    )
    facturas_otros = facturas_otros.annotate(
        total_cobrado_dash=Coalesce(
            Subquery(cobros_subquery), Value(0, output_field=DecimalField())
        ),
        saldo_pendiente_dash=ExpressionWrapper(
            F("monto")
            - Coalesce(
                Subquery(cobros_subquery), Value(0, output_field=DecimalField())
            ),
            output_field=DecimalField(),
        ),
    )

    # Anota el rango de vencimiento en cada factura otros ingresos
    facturas_otros = facturas_otros.annotate(
        rango=Case(
            When(fecha_vencimiento__gt=hoy - timedelta(days=30), then=Value("0_30")),
            When(
                fecha_vencimiento__gt=hoy - timedelta(days=60),
                fecha_vencimiento__lte=hoy - timedelta(days=30),
                then=Value("31_60"),
            ),
            When(
                fecha_vencimiento__gt=hoy - timedelta(days=90),
                fecha_vencimiento__lte=hoy - timedelta(days=60),
                then=Value("61_90"),
            ),
            When(
                fecha_vencimiento__gt=hoy - timedelta(days=180),
                fecha_vencimiento__lte=hoy - timedelta(days=90),
                then=Value("91_180"),
            ),
            When(
                fecha_vencimiento__lte=hoy - timedelta(days=180), then=Value("181_mas")
            ),
            default=Value("otro"),
            output_field=CharField(),
        )
    )
    saldos_otros = facturas_otros.values("rango").annotate(
        total=Sum("saldo_pendiente_dash")
    )
    saldos_otros_dict = {x["rango"]: float(x["total"]) for x in saldos_otros}
    for key in ["0_30", "31_60", "61_90", "91_180", "181_mas"]:
        saldos_otros_dict.setdefault(key, 0.0)

    # Top 10 adeudos
    top_adeudos = (
        facturas.annotate(
            nombre_local_area=Coalesce(
                F("local__numero"), F("area_comun__numero"), output_field=CharField()
            ),
            tipo_origen=Case(
                When(local__isnull=False, then=Value("Local")),
                When(area_comun__isnull=False, then=Value("Área")),
                default=Value(""),
                output_field=CharField(),
            ),
            nombre_cliente=F("cliente__nombre"),
        )
        .values("nombre_local_area", "tipo_origen", "nombre_cliente")
        .annotate(total=Sum("saldo_pendiente_dash"))
        .order_by("-total")[:10]
    )

    # Serializa resultados
    facturas_data = [
        {
            "folio": f.folio,
            "cliente": f.cliente.nombre if f.cliente else "",
            "empresa": f.empresa.nombre if f.empresa else "",
            "local": f.local.numero if f.local else "",
            "area_comun": f.area_comun.numero if f.area_comun else "",
            "monto": float(f.monto),
            "saldo_pendiente": float(f.saldo_pendiente_dash),
            "fecha_vencimiento": f.fecha_vencimiento,
            "estatus": f.estatus,
        }
        for f in facturas
    ]

    otros_data = [
        {
            "folio": f.folio,
            "cliente": f.cliente.nombre if f.cliente else "",
            "empresa": f.empresa.nombre if f.empresa else "",
            "tipo_ingreso": f.tipo_ingreso.nombre if f.tipo_ingreso else "",
            "monto": float(f.monto),
            "saldo_pendiente": float(f.saldo_pendiente_dash),
            "fecha_vencimiento": f.fecha_vencimiento,
            "estatus": f.estatus,
        }
        for f in facturas_otros
    ]

    return Response(
        {
            "saldos": {
                "0_30": saldos_dict["0_30"] + saldos_otros_dict["0_30"],
                "31_60": saldos_dict["31_60"] + saldos_otros_dict["31_60"],
                "61_90": saldos_dict["61_90"] + saldos_otros_dict["61_90"],
                "91_180": saldos_dict["91_180"] + saldos_otros_dict["91_180"],
                "181_mas": saldos_dict["181_mas"] + saldos_otros_dict["181_mas"],
            },
            "top_adeudos": list(top_adeudos),
            "facturas": facturas_data,
            "facturas_otros": otros_data,
        }
    )


# vista API estado de resultados
@api_view(["GET"])
@visitante_token_required
def api_estado_resultados(request):
    visitante = request.visitante
    if not getattr(visitante, "acceso_api_reporte", False):
        return Response({"error": "Acceso denegado"}, status=403)

    if getattr(visitante, "es_admin", False) or request.GET.get("empresa_id"):
        empresa_id = request.GET.get("empresa_id")
        if not empresa_id:
            return Response({"error": "Debe seleccionar una empresa"}, status=400)
        empresa = Empresa.objects.filter(id=empresa_id).first()
        if not empresa:
            return Response({"error": "Empresa no encontrada"}, status=404)
    else:
        empresa = None
        if visitante.locales.exists():
            empresa = visitante.locales.first().empresa
        elif visitante.areas.exists():
            empresa = visitante.areas.first().empresa

    # empresas = Empresa.objects.all()
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    periodo = request.GET.get("periodo")
    modo = request.GET.get("modo", "flujo")
    hoy = date.today()

    # Obtener meses y años existentes en la base de datos
    if empresa_id:
        meses_anios = (
            Factura.objects.filter(empresa_id=empresa_id)
            .annotate(
                mes=ExtractMonth("fecha_vencimiento"),
                anio=ExtractYear("fecha_vencimiento"),
            )
            .values("mes", "anio")
            .distinct()
        )
        meses_anios_otros = (
            FacturaOtrosIngresos.objects.filter(empresa_id=empresa_id)
            .annotate(
                mes=ExtractMonth("fecha_vencimiento"),
                anio=ExtractYear("fecha_vencimiento"),
            )
            .values("mes", "anio")
            .distinct()
        )
    else:
        meses_anios = (
            Factura.objects.annotate(
                mes=ExtractMonth("fecha_vencimiento"),
                anio=ExtractYear("fecha_vencimiento"),
            )
            .values("mes", "anio")
            .distinct()
        )
        meses_anios_otros = (
            FacturaOtrosIngresos.objects.annotate(
                mes=ExtractMonth("fecha_vencimiento"),
                anio=ExtractYear("fecha_vencimiento"),
            )
            .values("mes", "anio")
            .distinct()
        )

    meses_anios_set = set(
        (x["mes"], x["anio"]) for x in list(meses_anios) + list(meses_anios_otros)
    )
    # Filtra tuplas donde mes o año sean None
    meses_anios_list = sorted(
        [t for t in meses_anios_set if t[0] is not None and t[1] is not None],
        key=lambda x: (x[1], x[0]),
    )
    # meses_anios_list = sorted(list(meses_anios_set), key=lambda x: (x[1], x[0]))
    meses_unicos = sorted(set(m for m, y in meses_anios_list if m))
    anios_unicos = sorted(set(y for m, y in meses_anios_list if y))

    if not periodo and not fecha_inicio and not fecha_fin and not mes and not anio:
        periodo = "periodo_actual"

    if periodo == "mes_actual":
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = (hoy.replace(day=1) + timedelta(days=32)).replace(
            day=1
        ) - timedelta(days=1)
        mes = hoy.month
        anio = hoy.year
    elif periodo == "periodo_actual":
        fecha_inicio = hoy.replace(month=1, day=1)
        fecha_fin = hoy
        mes = ""
        anio = ""
    elif mes and anio:
        try:
            mes = int(mes)
            anio = int(anio)
            fecha_inicio = date(anio, mes, 1)
            if mes == 12:
                fecha_fin = date(anio, 12, 31)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
        except Exception:
            fecha_inicio = None
            fecha_fin = None
    elif fecha_inicio and fecha_fin:
        pass
    else:
        fecha_inicio = None
        fecha_fin = None
    # Convierte a date si es string
    if isinstance(fecha_inicio, str):
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        except Exception:
            fecha_inicio_dt = None
    else:
        fecha_inicio_dt = fecha_inicio

    if isinstance(fecha_fin, str):
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except Exception:
            fecha_fin_dt = None
    else:
        fecha_fin_dt = fecha_fin

    # Para mostrar el mes y año en letras
    try:
        locale.setlocale(locale.LC_TIME, "es_MX.UTF-8")
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
        except locale.Error:
            locale.setlocale(locale.LC_TIME, "C")  # Fallback seguro

    mes_letra = ""
    if (
        fecha_inicio_dt
        and fecha_fin_dt
        and fecha_inicio_dt == fecha_fin_dt.replace(day=1)
    ):
        mes_letra = fecha_inicio_dt.strftime("%B %Y").capitalize()
    elif fecha_inicio_dt and fecha_fin_dt:
        mes_letra = f"{fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')}"

    # --- CONSULTAS BASE ---
    pagos = Pago.objects.exclude(forma_pago="nota_credito")
    cobros_otros = CobroOtrosIngresos.objects.select_related(
        "factura", "factura__empresa"
    )
    gastos = Gasto.objects.select_related(
        "tipo_gasto", "tipo_gasto__subgrupo", "tipo_gasto__subgrupo__grupo"
    ).all()
    gastos_caja_chica = GastoCajaChica.objects.select_related(
        "tipo_gasto", "tipo_gasto__subgrupo", "tipo_gasto__subgrupo__grupo"
    ).all()
    vales_caja_chica = ValeCaja.objects.select_related(
        "tipo_gasto", "tipo_gasto__subgrupo", "tipo_gasto__subgrupo__grupo"
    ).all()

    empresa = None
    saldo_inicial = 0
    saldo_final = 0

    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id)
        cobros_otros = cobros_otros.filter(factura__empresa_id=empresa_id)
        gastos = gastos.filter(empresa_id=empresa_id)
        gastos_caja_chica = gastos_caja_chica.filter(fondeo__empresa_id=empresa_id)
        vales_caja_chica = vales_caja_chica.filter(fondeo__empresa_id=empresa_id)
    try:
        empresa = Empresa.objects.get(id=empresa_id)
        saldo_inicial_empresa = float(
            CuentaBancaria.objects.filter(empresa_id=empresa_id, activa=True).aggregate(
                t=Sum("saldo_inicial")
            )["t"]
            or 0
        )
        saldo_final = 0.00
    except Empresa.DoesNotExist:
        saldo_inicial_empresa = 0
        saldo_final = 0
    # --- FILTRO POR FECHA ---
    if fecha_inicio:
        gastos_caja_chica = gastos_caja_chica.filter(fecha__gte=fecha_inicio)
        vales_caja_chica = vales_caja_chica.filter(fecha__gte=fecha_inicio)

    if fecha_fin:
        gastos_caja_chica = gastos_caja_chica.filter(fecha__lte=fecha_fin)
        vales_caja_chica = vales_caja_chica.filter(fecha__lte=fecha_fin)

    # --- Saldo inicial dinámico acumulado ---
    saldo_inicial = saldo_inicial_empresa
    if empresa and empresa_id and fecha_inicio:
        anio_inicio = (
            empresa.fecha_creacion.year
            if hasattr(empresa, "fecha_creacion") and empresa.fecha_creacion
            else fecha_inicio.year
        )

        if mes and anio:
            anio_tope = int(anio)
            mes_tope = int(mes) - 1
            if mes_tope == 0:
                anio_tope -= 1
                mes_tope = 12
        else:
            if fecha_inicio.month == 1:
                anio_tope = fecha_inicio.year - 1
                mes_tope = 12
            else:
                anio_tope = fecha_inicio.year
                mes_tope = fecha_inicio.month - 1

        if anio_tope >= anio_inicio:
            for y in range(anio_inicio, anio_tope + 1):
                mes_fin_loop = mes_tope if y == anio_tope else 12
                for m in range(1, mes_fin_loop + 1):
                    fi = date(y, m, 1)
                    ff = (
                        date(y, m + 1, 1) - timedelta(days=1)
                        if m < 12
                        else date(y, 12, 31)
                    )
                    ing = float(
                        Pago.objects.exclude(forma_pago="nota_credito")
                        .filter(
                            factura__empresa_id=empresa_id,
                            fecha_pago__gte=fi,
                            fecha_pago__lte=ff,
                        )
                        .aggregate(t=Sum("monto"))["t"]
                        or 0
                    ) + float(
                        CobroOtrosIngresos.objects.filter(
                            factura__empresa_id=empresa_id,
                            fecha_cobro__gte=fi,
                            fecha_cobro__lte=ff,
                        ).aggregate(t=Sum("monto"))["t"]
                        or 0
                    )
                    gto = float(
                    PagoGasto.objects
                    .filter(gasto__empresa_id=empresa_id, fecha_pago__gte=fi, fecha_pago__lte=ff)
                    .aggregate(t=Sum("monto"))["t"] or 0
                    ) + float(
                        FondeoCajaChica.objects  # ← usar fondeos, NO gastos individuales
                        .filter(empresa_id=empresa_id, fecha__gte=fi, fecha__lte=ff)
                        .aggregate(t=Sum("importe_cheque"))["t"] or 0
                    )
                    saldo_inicial += ing - gto

        # --- Fin de saldo inicial dinámico ---
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio)
        cobros_otros = cobros_otros.filter(fecha_cobro__gte=fecha_inicio)
        gastos = gastos.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin)
        cobros_otros = cobros_otros.filter(fecha_cobro__lte=fecha_fin)
        gastos = gastos.filter(fecha__lte=fecha_fin)

    # depositos por identificar
    pagos_por_identificar = Pago.objects.filter(
        factura__isnull=True, identificado=False
    ).filter(factura__empresa_id=empresa_id)
    if empresa_id:
        pagos_por_identificar = pagos_por_identificar.filter(empresa_id=empresa_id)
    if fecha_inicio:
        pagos_por_identificar = pagos_por_identificar.filter(
            fecha_pago__gte=fecha_inicio
        )
    if fecha_fin:
        pagos_por_identificar = pagos_por_identificar.filter(fecha_pago__lte=fecha_fin)
    total_por_identificar = float(
        pagos_por_identificar.aggregate(total=Sum("monto"))["total"] or 0
    )

    saldo_final_flujo = None
    total_gastos = 0.0
    gastos_por_grupo = []

    if modo == "flujo":
        pagos_modo = pagos
        cobros_otros_modo = cobros_otros
        gastos_modo = PagoGasto.objects.filter(
            fecha_pago__range=[fecha_inicio, fecha_fin]
        )
        if empresa_id:
            gastos_modo = gastos_modo.filter(gasto__empresa_id=empresa_id)
        ingresos_qs = (
            pagos_modo.annotate(
                origen=Case(
                    When(factura__local__isnull=False, then=Value("Propiedades")),
                    When(
                        factura__area_comun__isnull=False, then=Value("Áreas Comunes")
                    ),
                    default=Value("Sin origen"),
                    output_field=CharField(),
                )
            )
            .values("origen")
            .annotate(total=Sum("monto"))
            .order_by("origen")
        )
        otros_ingresos_qs = (
            cobros_otros_modo.select_related("factura__tipo_ingreso")
            .values("factura__tipo_ingreso__nombre")
            .annotate(total=Sum("monto"))
            .order_by("factura__tipo_ingreso__nombre")
        )
        ingresos_por_origen = OrderedDict()
        for x in ingresos_qs:
            origen = (x["origen"] or "Sin origen").strip().title()
            ingresos_por_origen[origen] = float(x["total"])
        for x in otros_ingresos_qs:
            tipo = (
                (x["factura__tipo_ingreso__nombre"] or "Otros ingresos").strip().title()
            )
            ingresos_por_origen[f"Otros ingresos - {tipo}"] = float(x["total"])
            ingresos_por_origen["Depositos no identificados"] = float(
                total_por_identificar
            )
        total_ingresos = float(sum(ingresos_por_origen.values()))

        # Agrupar y sumar todos los gastos por tipo real (gastos normales, caja chica y vales)
        gastos_por_tipo_dict = {}
        tipos_gasto = set()
        tipos_gasto.update(
            [
                g["gasto__tipo_gasto__nombre"]
                for g in gastos_modo.values("gasto__tipo_gasto__nombre")
                if g["gasto__tipo_gasto__nombre"]
            ]
        )
        # Agrupa y suma en una sola consulta para cada fuente
        gastos_modo_agrupados = gastos_modo.values(
            "gasto__tipo_gasto__nombre"
        ).annotate(suma=Sum("monto"))
        gastos_caja_agrupados = gastos_caja_chica.values("tipo_gasto__nombre").annotate(
            suma=Sum("importe")
        )
        gastos_vales_agrupados = vales_caja_chica.values("tipo_gasto__nombre").annotate(
            suma=Sum("importe")
        )

        # Unifica todos los tipos
        tipos_gasto = set()
        tipos_gasto.update(
            g["gasto__tipo_gasto__nombre"]
            for g in gastos_modo_agrupados
            if g["gasto__tipo_gasto__nombre"]
        )
        tipos_gasto.update(
            g["tipo_gasto__nombre"]
            for g in gastos_caja_agrupados
            if g["tipo_gasto__nombre"]
        )
        tipos_gasto.update(
            g["tipo_gasto__nombre"]
            for g in gastos_vales_agrupados
            if g["tipo_gasto__nombre"]
        )

        # Crea diccionarios para acceso rápido
        dict_modo = {
            g["gasto__tipo_gasto__nombre"]: g["suma"] for g in gastos_modo_agrupados
        }
        dict_caja = {g["tipo_gasto__nombre"]: g["suma"] for g in gastos_caja_agrupados}
        dict_vales = {
            g["tipo_gasto__nombre"]: g["suma"] for g in gastos_vales_agrupados
        }

        for tipo in tipos_gasto:
            nombre_tipo = (tipo or "Sin tipo").strip().title()
            total_gastos_modo = float(dict_modo.get(tipo, 0) or 0)
            total_gastos_caja = float(dict_caja.get(tipo, 0) or 0)
            total_vales = float(dict_vales.get(tipo, 0) or 0)
            total = total_gastos_modo + total_gastos_caja + total_vales
            if total > 0:
                gastos_por_tipo_dict[nombre_tipo] = total

        gastos_por_tipo = [
            {"tipo": tipo, "total": total}
            for tipo, total in gastos_por_tipo_dict.items()
        ]

        estructura_gastos = OrderedDict()
        # Gastos modo flujo
        for g in gastos_modo.values(
            "gasto__tipo_gasto__subgrupo__grupo__nombre",
            "gasto__tipo_gasto__subgrupo__nombre",
            "gasto__tipo_gasto__nombre",
        ).annotate(total=Sum("monto")):
            grupo = (
                (g["gasto__tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
                .strip()
                .title()
            )
            subgrupo = (
                (g["gasto__tipo_gasto__subgrupo__nombre"] or "Sin subgrupo")
                .strip()
                .title()
            )
            tipo = (g["gasto__tipo_gasto__nombre"] or "Sin tipo").strip().title()
            total = float(g["total"])
            if grupo not in estructura_gastos:
                estructura_gastos[grupo] = OrderedDict()
            if subgrupo not in estructura_gastos[grupo]:
                estructura_gastos[grupo][subgrupo] = {}
            estructura_gastos[grupo][subgrupo][tipo] = (
                estructura_gastos[grupo][subgrupo].get(tipo, 0) + total
            )

        # Caja chica modo flujo
        for g in gastos_caja_chica.values(
            "tipo_gasto__subgrupo__grupo__nombre",
            "tipo_gasto__subgrupo__nombre",
            "tipo_gasto__nombre",
        ).annotate(total=Sum("importe")):
            grupo = (
                (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
                .strip()
                .title()
            )
            subgrupo = (
                (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
            )
            tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
            total = float(g["total"])
            if grupo not in estructura_gastos:
                estructura_gastos[grupo] = OrderedDict()
            if subgrupo not in estructura_gastos[grupo]:
                estructura_gastos[grupo][subgrupo] = {}
            estructura_gastos[grupo][subgrupo][tipo] = (
                estructura_gastos[grupo][subgrupo].get(tipo, 0) + total
            )

        # Vales de caja chica modo flujo
        for g in vales_caja_chica.values(
            "tipo_gasto__subgrupo__grupo__nombre",
            "tipo_gasto__subgrupo__nombre",
            "tipo_gasto__nombre",
        ).annotate(total=Sum("importe")):
            grupo = (
                (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
                .strip()
                .title()
            )
            subgrupo = (
                (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
            )
            tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
            total = float(g["total"])
            if grupo not in estructura_gastos:
                estructura_gastos[grupo] = OrderedDict()
            if subgrupo not in estructura_gastos[grupo]:
                estructura_gastos[grupo][subgrupo] = {}
            estructura_gastos[grupo][subgrupo][tipo] = (
                estructura_gastos[grupo][subgrupo].get(tipo, 0) + total
            )

        # Convertir los dicts de tipos a listas para compatibilidad con el template
        for grupo in estructura_gastos:
            for subgrupo in estructura_gastos[grupo]:
                tipos_dict = estructura_gastos[grupo][subgrupo]
                estructura_gastos[grupo][subgrupo] = [
                    {"tipo": tipo, "total": total} for tipo, total in tipos_dict.items()
                ]
        total_gastos = sum([g["total"] for g in gastos_por_tipo])
        saldo_final_flujo = (
            float(saldo_inicial) + float(total_ingresos) - float(total_gastos)
        )
    else:
        facturas_cuotas = Factura.objects.filter(
            fecha_vencimiento__range=[fecha_inicio, fecha_fin]
        )
        facturas_otros = FacturaOtrosIngresos.objects.filter(
            fecha_vencimiento__range=[fecha_inicio, fecha_fin]
        )
        if empresa_id:
            facturas_cuotas = facturas_cuotas.filter(empresa_id=empresa_id)
            facturas_otros = facturas_otros.filter(empresa_id=empresa_id)
        ingresos_por_origen = OrderedDict()
        origenes = (
            facturas_cuotas.annotate(
                origen=Case(
                    When(local__isnull=False, then=Value("Propiedades")),
                    When(area_comun__isnull=False, then=Value("Áreas Comunes")),
                    default=Value("Sin origen"),
                    output_field=CharField(),
                )
            )
            .values("origen")
            .annotate(total=Sum("monto"))
            .order_by("origen")
        )
        for x in origenes:
            origen = (x["origen"] or "Sin origen").strip().title()
            ingresos_por_origen[origen] = float(x["total"])
        otros = (
            facturas_otros.values("tipo_ingreso__nombre")
            .annotate(total=Sum("monto"))
            .order_by("tipo_ingreso__nombre")
        )
        for x in otros:
            tipo = (x["tipo_ingreso__nombre"] or "Otros ingresos").strip().title()
            ingresos_por_origen[f"Otros ingresos - {tipo}"] = float(x["total"])
            ingresos_por_origen["Depositos no identificados"] = float(
                total_por_identificar
            )
        total_ingresos = float(sum(ingresos_por_origen.values()))

        # Agrupar y sumar todos los gastos por tipo real (gastos normales, caja chica y vales)
        gastos_por_tipo_dict = {}
        tipos_gasto = set()
        tipos_gasto.update(
            [g["tipo_gasto__nombre"] for g in gastos.values("tipo_gasto__nombre")]
        )
        tipos_gasto.update(
            [
                g["tipo_gasto__nombre"]
                for g in gastos_caja_chica.values("tipo_gasto__nombre")
            ]
        )
        tipos_gasto.update(
            [
                g["tipo_gasto__nombre"]
                for g in vales_caja_chica.values("tipo_gasto__nombre")
            ]
        )
        for tipo in tipos_gasto:
            if tipo and tipo not in ["Gastos de caja chica", "Vales de caja chica"]:
                nombre_tipo = (tipo or "Sin tipo").strip().title()
                total = 0.0
                total_gastos = (
                    gastos.filter(tipo_gasto__nombre=tipo).aggregate(suma=Sum("monto"))[
                        "suma"
                    ]
                    or 0
                )
                total_gastos_caja = (
                    gastos_caja_chica.filter(tipo_gasto__nombre=tipo).aggregate(
                        suma=Sum("importe")
                    )["suma"]
                    or 0
                )
                total_vales = (
                    vales_caja_chica.filter(tipo_gasto__nombre=tipo).aggregate(
                        suma=Sum("importe")
                    )["suma"]
                    or 0
                )
                total = (
                    float(total_gastos) + float(total_gastos_caja) + float(total_vales)
                )
                if total > 0:
                    gastos_por_tipo_dict[nombre_tipo] = total

        gastos_por_tipo = []
        for tipo, total in gastos_por_tipo_dict.items():
            gastos_por_tipo.append({"tipo": tipo, "total": total})

        # Unificar gastos normales, caja chica y vales por grupo, subgrupo y tipo
        estructura_gastos = OrderedDict()
        # Gastos normales
        for g in gastos.values(
            "tipo_gasto__subgrupo__grupo__nombre",
            "tipo_gasto__subgrupo__nombre",
            "tipo_gasto__nombre",
        ).annotate(total=Sum("monto")):
            grupo = (
                (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
                .strip()
                .title()
            )
            subgrupo = (
                (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
            )
            tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
            total = float(g["total"])
            if grupo not in estructura_gastos:
                estructura_gastos[grupo] = OrderedDict()
            if subgrupo not in estructura_gastos[grupo]:
                estructura_gastos[grupo][subgrupo] = {}
            estructura_gastos[grupo][subgrupo][tipo] = (
                estructura_gastos[grupo][subgrupo].get(tipo, 0) + total
            )

        # Caja chica
        for g in gastos_caja_chica.values(
            "tipo_gasto__subgrupo__grupo__nombre",
            "tipo_gasto__subgrupo__nombre",
            "tipo_gasto__nombre",
        ).annotate(total=Sum("importe")):
            grupo = (
                (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
                .strip()
                .title()
            )
            subgrupo = (
                (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
            )
            tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
            total = float(g["total"])
            if grupo not in estructura_gastos:
                estructura_gastos[grupo] = OrderedDict()
            if subgrupo not in estructura_gastos[grupo]:
                estructura_gastos[grupo][subgrupo] = {}
            estructura_gastos[grupo][subgrupo][tipo] = (
                estructura_gastos[grupo][subgrupo].get(tipo, 0) + total
            )

        # Vales de caja chica
        for g in vales_caja_chica.values(
            "tipo_gasto__subgrupo__grupo__nombre",
            "tipo_gasto__subgrupo__nombre",
            "tipo_gasto__nombre",
        ).annotate(total=Sum("importe")):
            grupo = (
                (g["tipo_gasto__subgrupo__grupo__nombre"] or "Sin grupo")
                .strip()
                .title()
            )
            subgrupo = (
                (g["tipo_gasto__subgrupo__nombre"] or "Sin subgrupo").strip().title()
            )
            tipo = (g["tipo_gasto__nombre"] or "Sin tipo").strip().title()
            total = float(g["total"])
            if grupo not in estructura_gastos:
                estructura_gastos[grupo] = OrderedDict()
            if subgrupo not in estructura_gastos[grupo]:
                estructura_gastos[grupo][subgrupo] = {}
            estructura_gastos[grupo][subgrupo][tipo] = (
                estructura_gastos[grupo][subgrupo].get(tipo, 0) + total
            )

        # Convertir los dicts de tipos a listas para compatibilidad con el template
        for grupo in estructura_gastos:
            for subgrupo in estructura_gastos[grupo]:
                tipos_dict = estructura_gastos[grupo][subgrupo]
                estructura_gastos[grupo][subgrupo] = [
                    {"tipo": tipo, "total": total} for tipo, total in tipos_dict.items()
                ]
        total_gastos = sum([g["total"] for g in gastos_por_tipo])
        saldo_final_flujo = None

    saldo = float(total_ingresos) - float(total_gastos)

    return Response(
        {
            "ingresos_por_origen": ingresos_por_origen,
            "gastos_por_tipo": gastos_por_tipo,
            "estructura_gastos": estructura_gastos,
            "total_ingresos": total_ingresos,
            "total_gastos": total_gastos,
            "saldo": saldo,
            "empresa_id": str(empresa_id or ""),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "mes": str(mes or ""),
            "anio": str(anio or ""),
            "periodo": periodo,
            "modo": modo,
            "saldo_inicial": saldo_inicial,
            "saldo_final": saldo_final,
            "saldo_final_flujo": saldo_final_flujo,
            "meses_unicos": meses_unicos,
            "anios_unicos": anios_unicos,
            "mes_letra": mes_letra,
        },
    )


# vista API para avisos
@api_view(["GET"])
@visitante_token_required
def api_avisos_empresa(request):
    visitante = request.visitante
    empresa_id = request.GET.get("empresa_id")
    empresa = None
    if empresa_id:
        empresa = Empresa.objects.filter(id=empresa_id).first()
    else:
        if visitante.locales.exists():
            empresa = visitante.locales.first().empresa
        elif visitante.areas.exists():
            empresa = visitante.areas.first().empresa

    if not empresa:
        return Response(
            {"error": "No se encontró empresa asociada al visitante."}, status=400
        )

    avisos = Aviso.objects.filter(empresa=empresa).order_by("-fecha_creacion")
    data = [
        {
            "id": aviso.id,
            "titulo": aviso.titulo,
            "mensaje": aviso.mensaje,
            "fecha_creacion": aviso.fecha_creacion.strftime("%Y-%m-%d %H:%M"),
        }
        for aviso in avisos
    ]
    return Response(data)


####################################RECORDATORIOS MOROSIDAD DEUDORES#######################################################3
# recordatorios morosidad
@login_required
def enviar_recordatorio_morosidad(request):
    local_id = request.GET.get("local_id")
    area_id = request.GET.get("area_id")
    next_url = request.GET.get("next")

    # Validación de parámetros
    if local_id and not local_id.isdigit():
        messages.error(request, "ID de local inválido.")
        return redirect(next_url or "lista_facturas")
    if area_id and not area_id.isdigit():
        messages.error(request, "ID de área común inválido.")
        return redirect(next_url or "lista_facturas")

    def formato_importe(importe):
        return "${:,.2f}".format(round(importe, 2))

    def construir_correo_html(
        cliente, facturas, empresa_nombre, email_empresa, tipo="local"
    ):
        total = sum(f.saldo_pendiente for f in facturas)
        tipo_label = (
            "cuotas de local comercial" if tipo == "local" else "cuotas de área común"
        )

        filas_facturas = ""
        for factura in facturas:
            if tipo == "local":
                ubicacion = (
                    f"Local: {factura.local.numero}"
                    if factura.local
                    else "Sin ubicación"
                )
            else:
                ubicacion = (
                    f"Área común: {factura.area_comun.numero}"
                    if factura.area_comun
                    else "Sin ubicación"
                )

            filas_facturas += f"""
            <tr>
                <td style="padding: 10px 12px; border-bottom: 1px solid #eee; color: #444;">{factura.folio}</td>
                <td style="padding: 10px 12px; border-bottom: 1px solid #eee; color: #444;">{ubicacion}</td>
                <td style="padding: 10px 12px; border-bottom: 1px solid #eee; color: #444;">{factura.fecha_vencimiento}</td>
                <td style="padding: 10px 12px; border-bottom: 1px solid #eee; color: #c0392b; font-weight: bold; text-align: right;">{formato_importe(factura.saldo_pendiente)}</td>
            </tr>
            """

        html = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
        <body style="margin:0; padding:0; background:#f4f6f8; font-family: Arial, sans-serif;">
            <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8; padding: 32px 0;">
                <tr><td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff; border-radius:10px; overflow:hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">

                        <!-- Header -->
                        <tr>
                            <td style="background: #152A52; padding: 28px 32px; text-align: center;">
                                <h1 style="margin:0; color:#ffffff; font-size:22px; letter-spacing:1px;">{empresa_nombre}</h1>
                                <p style="margin:6px 0 0 0; color:#AFC0DC; font-size:13px;">Aviso de adeudo pendiente</p>
                            </td>
                        </tr>

                        <!-- Saludo -->
                        <tr>
                            <td style="padding: 28px 32px 16px 32px;">
                                <p style="margin:0; color:#333; font-size:15px;">Estimado(a) <strong>{cliente.nombre}</strong>,</p>
                                <p style="margin:12px 0 0 0; color:#555; font-size:14px; line-height:1.6;">
                                    Le informamos que tiene <strong>adeudos pendientes</strong> correspondientes a {tipo_label}.
                                    Le pedimos atender este aviso a la brevedad posible para evitar cargos adicionales.
                                </p>
                            </td>
                        </tr>

                        <!-- Tabla de adeudos -->
                        <tr>
                            <td style="padding: 0 32px 24px 32px;">
                                <table width="100%" cellpadding="0" cellspacing="0" style="border: 1px solid #e0e0e0; border-radius:6px; overflow:hidden;">
                                    <tr style="background:#f0f4f8;">
                                        <th style="padding: 10px 12px; text-align:left; color:#152A52; font-size:12px; text-transform:uppercase; letter-spacing:0.5px;">Folio</th>
                                        <th style="padding: 10px 12px; text-align:left; color:#152A52; font-size:12px; text-transform:uppercase; letter-spacing:0.5px;">Ubicación</th>
                                        <th style="padding: 10px 12px; text-align:left; color:#152A52; font-size:12px; text-transform:uppercase; letter-spacing:0.5px;">Vencimiento</th>
                                        <th style="padding: 10px 12px; text-align:right; color:#152A52; font-size:12px; text-transform:uppercase; letter-spacing:0.5px;">Saldo</th>
                                    </tr>
                                    {filas_facturas}
                                    <tr style="background:#fdf3f3;">
                                        <td colspan="3" style="padding: 12px; text-align:right; font-weight:bold; color:#333;">Total pendiente:</td>
                                        <td style="padding: 12px; text-align:right; font-weight:bold; color:#c0392b; font-size:16px;">{formato_importe(total)}</td>
                                    </tr>
                                </table>
                            </td>
                        </tr>

                        <!-- Botón pago en línea -->
                        <tr>
                            <td style="padding: 0 32px 24px 32px; text-align:center;">
                                <p style="margin:0 0 16px 0; color:#555; font-size:14px;">Puede realizar su pago de forma rápida y segura:</p>
                                <a href="{settings.PORTAL_PAGOS_URL}"
                                   style="display:inline-block; background:#152A52; color:#ffffff; text-decoration:none; padding:14px 32px; border-radius:6px; font-size:15px; font-weight:bold; letter-spacing:0.5px;">
                                   💳 Pagar en línea
                                </a>
                            </td>
                        </tr>

                        <!-- Descarga de la app -->
                        <tr>
                            <td style="padding: 0 32px 28px 32px;">
                                <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8; border-radius:8px; padding:20px;">
                                    <tr>
                                        <td style="padding:16px; text-align:center;">
                                            <p style="margin:0 0 12px 0; color:#152A52; font-weight:bold; font-size:14px;">📱 También puede pagar desde nuestra app</p>
                                            <p style="margin:0 0 16px 0; color:#555; font-size:13px;">Descárguela gratis y consulte su saldo, pague sus cuotas y envíe comprobantes desde su celular.</p>
                                            <a href="https://apps.apple.com/us/app/gesac-condominos/id6756532273"
                                               style="display:inline-block; background:#000000; color:#ffffff; text-decoration:none; padding:10px 20px; border-radius:6px; font-size:13px; font-weight:bold; margin: 0 6px;">
                                                🍎 App Store
                                            </a>
                                            <a href="https://paginaweb-ro9v.onrender.com/beta"
                                               style="display:inline-block; background:#2d7a27; color:#ffffff; text-decoration:none; padding:10px 20px; border-radius:6px; font-size:13px; font-weight:bold; margin: 0 6px; opacity:0.7;">
                                                🤖 Google Play (testers)
                                            </a>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>

                        <!-- Comprobante -->
                        <tr>
                            <td style="padding: 0 32px 28px 32px;">
                                <p style="margin:0; color:#555; font-size:13px; line-height:1.6; background:#fff8e1; border-left:4px solid #f9a825; padding:12px 16px; border-radius:4px;">
                                    ¿Ya realizó su pago? Envíe su comprobante a
                                    <a href="mailto:{email_empresa}" style="color:#152A52; font-weight:bold;">{email_empresa}</a>
                                    o súbalo directamente desde la app.
                                </p>
                            </td>
                        </tr>

                        <!-- Footer -->
                        <tr>
                            <td style="background:#f0f4f8; padding:18px 32px; text-align:center; border-top:1px solid #e0e0e0;">
                                <p style="margin:0; color:#888; font-size:12px;">
                                    Este es un mensaje automático de <strong>{empresa_nombre}</strong>.<br>
                                    Por favor no responda directamente a este correo.
                                </p>
                            </td>
                        </tr>

                    </table>
                </td></tr>
            </table>
        </body>
        </html>
        """
        return html, formato_importe(total)

    if local_id:
        local_id = int(local_id)
        facturas = Factura.objects.filter(local_id=local_id, estatus="pendiente")
        facturas = [f for f in facturas if f.saldo_pendiente > 0]
        if not facturas:
            messages.warning(
                request, "No hay adeudos pendientes para el local seleccionado."
            )
            return redirect(next_url or "lista_facturas")
        cliente = facturas[0].cliente
        email = cliente.email
        email_empresa = facturas[0].empresa.email if facturas[0].empresa else ""
        if not email:
            if request.method == "POST":
                form = CapturarEmailForm(request.POST)
                next_url = request.POST.get("next")
                if form.is_valid():
                    email = form.cleaned_data["email"]
                    if cliente.nombre.strip().lower() != "venta al publico en general":
                        cliente.email = email
                        cliente.save()
                else:
                    return render(
                        request,
                        "facturacion/capturar_email.html",
                        {"form": form, "cliente": cliente, "next": next_url},
                    )
            else:
                form = CapturarEmailForm()
                return render(
                    request,
                    "facturacion/capturar_email.html",
                    {"form": form, "cliente": cliente, "next": next_url},
                )

        empresa_nombre = facturas[0].empresa.nombre if facturas[0].empresa else ""
        html_message, total_str = construir_correo_html(
            cliente, facturas, empresa_nombre, email_empresa, tipo="local"
        )
        texto_plano = f"{empresa_nombre}\n\nEstimado {cliente.nombre}, tiene adeudos pendientes por un total de {total_str}.\n\nPague en línea: {settings.PORTAL_PAGOS_URL}/\nDescargue la app: https://apps.apple.com/us/app/gesac-condominos/id6756532273"

        send_mail(
            subject=f"⚠️ Recordatorio de adeudo pendiente — {empresa_nombre}",
            message=texto_plano,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            html_message=html_message,
            fail_silently=True,
        )
        messages.success(
            request,
            f"Recordatorio enviado correctamente al cliente {cliente.nombre} del local {facturas[0].local.numero if facturas[0].local else ''}.",
        )
        return redirect(next_url or "lista_facturas")

    elif area_id:
        area_id = int(area_id)
        facturas = Factura.objects.filter(area_comun_id=area_id, estatus="pendiente")
        facturas = [f for f in facturas if f.saldo_pendiente > 0]
        if not facturas:
            messages.warning(
                request, "No hay adeudos pendientes para el área común seleccionada."
            )
            return redirect(next_url or "lista_facturas")
        cliente = facturas[0].cliente
        email = cliente.email
        email_empresa = facturas[0].empresa.email if facturas[0].empresa else ""
        if not email:
            if request.method == "POST":
                form = CapturarEmailForm(request.POST)
                next_url = request.POST.get("next")
                if form.is_valid():
                    email = form.cleaned_data["email"]
                    if cliente.nombre.strip().lower() != "venta al publico en general":
                        cliente.email = email
                        cliente.save()
                else:
                    return render(
                        request,
                        "facturacion/capturar_email.html",
                        {"form": form, "cliente": cliente, "next": next_url},
                    )
            else:
                form = CapturarEmailForm()
                return render(
                    request,
                    "facturacion/capturar_email.html",
                    {"form": form, "cliente": cliente, "next": next_url},
                )

        empresa_nombre = facturas[0].empresa.nombre if facturas[0].empresa else ""
        html_message, total_str = construir_correo_html(
            cliente, facturas, empresa_nombre, email_empresa, tipo="area"
        )
        texto_plano = f"{empresa_nombre}\n\nEstimado {cliente.nombre}, tiene adeudos pendientes por un total de {total_str}.\n\nPague en línea: {settings.PORTAL_PAGOS_URL}/\nDescargue la app: https://apps.apple.com/us/app/gesac-condominos/id6756532273"

        send_mail(
            subject=f"⚠️ Recordatorio de adeudo pendiente — {empresa_nombre}",
            message=texto_plano,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            html_message=html_message,
            fail_silently=True,
        )
        messages.success(
            request,
            f"Recordatorio enviado correctamente al cliente {cliente.nombre} del área común {facturas[0].area_comun.numero if facturas[0].area_comun else ''}.",
        )
        return redirect(next_url or "lista_facturas")

    else:
        messages.error(
            request,
            "Debes seleccionar un local o un área común antes de enviar el recordatorio.",
        )
        return redirect(next_url or "lista_facturas")
